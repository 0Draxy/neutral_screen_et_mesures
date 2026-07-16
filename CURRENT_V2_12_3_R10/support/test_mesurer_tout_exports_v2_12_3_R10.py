from __future__ import annotations

import ast
import json
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
PY_FILE = BASE / "main_ihm_relais_rp2040_v2_12_3.py"
UI_FILE = BASE / "ihm_relais_rp2040_28vdc_precision_v2_12_3.ui"


def load_harness(method_names):
    source = PY_FILE.read_text(encoding="utf-8")
    tree = ast.parse(source)
    class_node = next(node for node in tree.body if isinstance(node, ast.ClassDef) and node.name == "IhmRelaisRp2040")
    methods = [node for node in class_node.body if isinstance(node, ast.FunctionDef) and node.name in method_names]
    found = {node.name for node in methods}
    missing = set(method_names) - found
    if missing:
        raise AssertionError(f"Méthodes absentes : {sorted(missing)}")
    module = ast.parse("class Harness:\n    pass\n")
    module.body[0].body = methods
    ast.fix_missing_locations(module)
    namespace = {
        "json": json,
        "Path": Path,
        "time": __import__("time"),
        "CHRONO_TRANSFER_LIMIT_MS": 1.0,
    }
    exec(compile(module, str(PY_FILE), "exec"), namespace)
    return namespace["Harness"]


def chrono_record(action, result, details):
    return {
        "lot": "LOT-R8",
        "date_test": "16/07/2026",
        "relais": "REL-1",
        "ambiance_c": "20",
        "nom_test": "Mesure complète",
        "sn": "001",
        "relay_type": "BISTABLE",
        "action": action,
        "nb_inverseurs": 2,
        "capture_ms": 50,
        "pulse_ms": 10,
        "limite_temps_ms": 1.5,
        "limite_rebond_ms": 2.0,
        "resultat": result,
        "overflow": 0,
        "details_json": json.dumps({"lignes_action": details}),
        "events_json": "{}",
        "timestamp": "2026-07-16 10:00:00",
    }


def metric(inv, name, value):
    return {"inverseur": inv, "metric": name, "temps_ms": value}


def test_two_sheet_layout_and_exact_labels():
    methods = {
        "chrono_json_dict", "chrono_export_sanction_label", "chrono_export_limit_label",
        "chrono_export_value_ms", "chrono_export_metric_order", "chrono_export_metric_label",
        "chrono_export_sanction_for_metric", "chrono_export_group_key", "chrono_export_group_data",
        "chrono_export_find_voltage", "chrono_export_base_rows", "chrono_export_time_rows",
        "chrono_export_global_voltage_rows", "chrono_export_individual_voltage_rows",
        "chrono_export_measure_sheets", "chrono_export_measure_cards", "chrono_export_voltage_value",
        "write_chrono_lot_xlsx",
    }
    Harness = load_harness(methods)
    obj = Harness()

    be_rows = []
    br_rows = []
    for inv in (1, 2):
        be_rows.extend([
            metric(inv, "enclenchement", 1.0 + inv / 10),
            metric(inv, "transfert_travail", 0.2 + inv / 100),
            metric(inv, "rebond_travail", 0.3 + inv / 100),
        ])
        br_rows.extend([
            metric(inv, "declenchement", 1.2 + inv / 10),
            metric(inv, "transfert_repos", 0.25 + inv / 100),
            metric(inv, "rebond_repos", 0.35 + inv / 100),
        ])
    records = [chrono_record("BE", "OK", be_rows), chrono_record("BR", "OK", br_rows)]
    voltage = {
        "id": 1, "lot": "LOT-R8", "date_test": "16/07/2026", "relais": "REL-1",
        "ambiance_c": "20", "nom_test": "Mesure complète", "sn": "001",
        "relay_type": "BISTABLE", "nb_inverseurs": 2,
        "pickup_global_v": 12.345, "dropout_global_v": 8.765,
        "pickup_json": json.dumps({"1": 12.300, "2": 12.340}),
        "dropout_json": json.dumps({"1": 8.700, "2": 8.750}),
        "pickup_plausibility_status": "OK", "dropout_plausibility_status": "OK",
        "resultat": "OK", "timestamp": "2026-07-16 10:05:00",
    }

    summary, detail = obj.chrono_export_measure_sheets(records, [voltage])
    assert len(summary) == len(detail) == 1
    summary_labels = [row[0] for row in summary[0]]
    detail_labels = [row[0] for row in detail[0]]
    assert summary_labels[8:10] == ["Tension d'Enclenchement", "Tension de Rappel"]
    assert detail_labels[8:14] == [
        "Tension d'Enclenchement (globale)",
        "Tension de Rappel (globale)",
        "Tension d'Enclenchement inverseur 1",
        "Tension de Rappel inverseur 1",
        "Tension d'Enclenchement inverseur 2",
        "Tension de Rappel inverseur 2",
    ]
    expected_time_labels = [
        "Temps d'Enclenchement 1 (ms)",
        "Temps de transfère 1 (ms)",
        "Temps Rebond Travail Fermeture 1 (ms)",
        "Temps de Déclenchement 1 (ms)",
        "Temps de transfère 1 retour (ms)",
        "Temps Rebond Repos Fermeture 1 (ms)",
        "Temps d'Enclenchement 2 (ms)",
        "Temps de transfère 2 (ms)",
        "Temps Rebond Travail Fermeture 2 (ms)",
        "Temps de Déclenchement 2 (ms)",
        "Temps de transfère 2 retour (ms)",
        "Temps Rebond Repos Fermeture 2 (ms)",
    ]
    assert summary_labels[10:] == expected_time_labels
    assert detail_labels[14:] == expected_time_labels
    assert not any("Ouverture" in label for label in summary_labels + detail_labels)

    with tempfile.TemporaryDirectory() as tmp:
        path = Path(tmp) / "rapport.xlsx"
        obj.write_chrono_lot_xlsx(path, "LOT-R8", summary, detail)
        wb = load_workbook(path, data_only=True)
        assert wb.sheetnames == ["Synthèse globale", "Détail tensions"]
        assert wb["Synthèse globale"]["A1"].value.startswith("Synthèse globale")
        assert wb["Détail tensions"]["A1"].value.startswith("Détail tensions")
        values1 = [cell.value for cell in wb["Synthèse globale"]["A"]]
        values2 = [cell.value for cell in wb["Détail tensions"]["A"]]
        assert "Tension d'Enclenchement" in values1
        assert "Tension d'Enclenchement inverseur 2" in values2


def test_global_fallback_values():
    methods = {"voltage_timeout_global_fallback", "voltage_apply_timeout_global_fallback"}
    Harness = load_harness(methods)

    class Spin:
        def value(self):
            return 30.0

    obj = Harness()
    obj.doubleSpinBox_voltage_vmax = Spin()
    obj.voltage_run_setting = lambda key, widget: 30.0
    obj.voltage_refresh_results_table = lambda: None
    obj.voltage_requested_mode = "CYCLE"
    obj.voltage_ramp_info = {"mode": "PICKUP"}
    obj.voltage_active_scan = "PICKUP"
    obj.voltage_results = {"PICKUP": {}, "DROPOUT": {}}
    obj.voltage_raw_results = {"PICKUP": {}, "DROPOUT": {}}
    obj.voltage_time_results = {"PICKUP": {}, "DROPOUT": {}}
    obj.voltage_plausibility = {"PICKUP": {}, "DROPOUT": {}}
    obj.voltage_first_passage = {"PICKUP": {}, "DROPOUT": {}}
    obj.voltage_relay_type = lambda: "BISTABLE"
    mode, value, _, status = obj.voltage_apply_timeout_global_fallback()
    assert mode == "PICKUP" and value == 30.0 and status == "DEFAUT_CONTACTS"
    assert obj.voltage_results["PICKUP"]["GLOBAL"] == 30.0
    assert obj.voltage_result_override == "DEFAUT_CONTACTS"

    obj.voltage_active_scan = "DROPOUT"
    obj.voltage_ramp_info = {"mode": "DROPOUT"}
    obj.voltage_relay_type = lambda: "MONOSTABLE"
    mode, value, _, status = obj.voltage_apply_timeout_global_fallback()
    assert mode == "DROPOUT" and value == 0.0 and status == "DEFAUT_CONTACTS"
    assert obj.voltage_results["DROPOUT"]["GLOBAL"] == 0.0

    obj.voltage_relay_type = lambda: "BISTABLE"
    _mode, value, _, status = obj.voltage_apply_timeout_global_fallback()
    assert value == 30.0 and status == "DEFAUT_CONTACTS"

    obj.voltage_active_scan = "PICKUP"
    obj.voltage_first_passage["PICKUP"]["GLOBAL"] = {"mv": 12345, "raw": 1000, "t_us": 1000000}
    mode, value, _, status = obj.voltage_apply_timeout_global_fallback()
    assert mode == "PICKUP" and value == 12.345 and status == "DEFAUT_STABILITE"
    assert obj.voltage_raw_results["PICKUP"]["GLOBAL"] == 1000
    assert obj.voltage_time_results["PICKUP"]["GLOBAL"] == 1000000


def test_ui_and_static_sequences():
    root = ET.parse(UI_FILE).getroot()
    title = root.find('.//widget[@class="QMainWindow"]/property[@name="windowTitle"]/string')
    assert title is not None and "R10" in (title.text or "")
    button = root.find('.//widget[@name="pushButton_voltage_measure_all"]')
    assert button is not None
    assert button.findtext('./property[@name="text"]/string') == "MESURER TOUT"

    source = PY_FILE.read_text(encoding="utf-8")
    ast.parse(source)
    forbidden = ["METTRE LE SÉLECTEUR SUR EA", "METTRE LE SÉLECTEUR SUR FIXE"]
    assert not any(token in source for token in forbidden)
    required = [
        "configure_static_output_and_confirm",
        "EA_CHRONO_NO_GP26",
        'source_suffix = ";EA"',
        "allow_measure_all_continue=True",
        'status = "DEFAUT_CONTACTS"',
        'self.voltage_result_override = status',
        'draw_section("FEUILLE 1 - SYNTHÈSE GLOBALE"',
        'draw_section("FEUILLE 2 - DÉTAIL TENSIONS"',
    ]
    missing = [token for token in required if token not in source]
    assert not missing, missing
    chrono_spin = root.find('.//widget[@name="doubleSpinBox_voltage_chrono_v"]')
    assert chrono_spin is not None
    assert chrono_spin.findtext('./property[@name="value"]/double') == "28.000000000000000"


def main():
    test_two_sheet_layout_and_exact_labels()
    test_global_fallback_values()
    test_ui_and_static_sequences()
    print("MESURER_TOUT_EXPORTS_R10_TESTS_OK")


if __name__ == "__main__":
    main()
