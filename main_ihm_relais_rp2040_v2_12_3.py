import sys
import time
import json
import copy
import csv
import math
import os
import re
import sqlite3
import shutil
import gc
import zipfile
from pathlib import Path

from licence_manager import require_license

import serial
import serial.tools.list_ports

from PySide6.QtCore import QThread, Signal, QTimer, QFile, Qt, QDate, QObject, QEvent, QLineF
from PySide6.QtGui import QColor, QFont, QKeySequence, QPageLayout, QPageSize, QPainter, QPdfWriter, QPen, QShortcut
from PySide6.QtWidgets import QApplication, QMessageBox, QWidget, QPushButton, QComboBox, QLineEdit, QLabel, QTextEdit, QTableWidget, QTableWidgetItem, QFileDialog, QSpinBox, QHeaderView, QCheckBox, QTabWidget, QDateEdit, QInputDialog, QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox, QToolTip, QDoubleSpinBox
from PySide6.QtUiTools import QUiLoader


APP_DIR = Path(__file__).resolve().parent


def runtime_output_dir():
    """Dossier imposé pour les exports : répertoire de l'EXE, ou du script en mode développement."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return APP_DIR


EXPORT_DIR = runtime_output_dir()
UI_FILE = APP_DIR / "ihm_relais_rp2040_28vdc_precision_v2_12_3.ui"

UINT32_MAX_MS = 4_294_967_295
MAX_US_TOTAL = UINT32_MAX_MS * 1000
EDITOR_TABLE_WIDTHS = [92, 112, 145, 105, 112, 300, 1]
LOCK_ACCESS_CODE = os.environ.get("IHM_ADMIN_CODE", "1234")
LOCK_RECOVERY_TRIGGER = os.environ.get("IHM_RECOVERY_TRIGGER", "marechal")
CHRONO_SPREAD_INFO_US = 50
CHRONO_LOOP_WARN_US = 200
CHRONO_TRANSFER_LIMIT_MS = 1.0
EA_STOP_CONFIRM_MAX_V = 0.200
EA_STOP_CONFIRM_SETTLE_S = 0.100
EA_STOP_CONFIRM_ATTEMPTS = 2
EA_STOP_CONFIRM_QUERY_TIMEOUT_S = 0.350
EA_STOP_CONFIRM_TIMEOUT_S = 5.000
EA_STOP_CONFIRM_POLL_INTERVAL_S = 0.150
EA_STATIC_CONFIRM_TIMEOUT_S = 3.000
EA_STATIC_CONFIRM_POLL_INTERVAL_S = 0.150
EA_STATIC_CONFIRM_MIN_TOL_V = 0.150
EA_STATIC_CONFIRM_REL_TOL = 0.010
VOLTAGE_PLAUSIBILITY_MIN_TOL_S = 0.750
VOLTAGE_PLAUSIBILITY_REL_TOL = 0.100
CHRONO_CONTACT_NAMES = ["R1", "R2", "R3", "R4", "T1", "T2", "T3", "T4"]
OSCILLO_COMBINED_GAP_US = 2000
OSCILLO_DISPLAY_ELECTRIC = "ELECTRIC"
OSCILLO_DISPLAY_LOGIC = "LOGIC"
OSCILLO_DISPLAY_SYNTHESIS = "SYNTHESIS"


def oscillo_pre_t0_gap_us(view_start_us, view_end_us):
    """Marge uniquement visuelle avant T0 pour dégager la commande bobine."""
    try:
        view_start_us = int(view_start_us or 0)
        view_end_us = int(view_end_us or 0)
    except Exception:
        return 0
    if view_start_us > 0:
        return 0
    span_us = max(1, view_end_us - view_start_us)
    return max(1, min(2000, int(math.ceil(span_us * 0.04))))


def oscillo_t0_command_label(current=None, phase_markers=None):
    """Libellé lisible du top départ, sans modifier le T0 de mesure."""
    action = ""
    if isinstance(current, dict):
        action = str(current.get("ACTION") or current.get("action") or "").upper()
    labels = {
        "BE": "BE",
        "BR": "BR",
        "MONO_ON": "MONO ON",
        "MONO_OFF": "MONO OFF",
        "CYCLE_BE_BR": "BE",
        "CYCLE_MONO": "MONO ON",
    }
    command = labels.get(action, "")
    if not command:
        for marker in phase_markers or []:
            try:
                if int(marker.get("t_us", -1)) == 0:
                    command = str(marker.get("label", "")).strip()
                    break
            except Exception:
                pass
    return f"T0 commande {command}".strip()


UNIT_FACTORS_US = {
    "µs": 1, "us": 1, "ms": 1_000, "s": 1_000_000,
    "min": 60_000_000, "h": 3_600_000_000,
}

UNIT_LIMITS = {
    "µs": (1, 1000), "us": (1, 1000), "ms": (1, 4_294_967_295),
    "s": (1, 4_294_967), "min": (1, 71_582), "h": (1, 1193),
}


class ClosingSQLiteConnection(sqlite3.Connection):
    """Connexion SQLite qui se ferme réellement à la sortie du bloc with."""

    def __exit__(self, exc_type, exc, tb):
        try:
            if exc_type is None:
                self.commit()
            else:
                self.rollback()
        finally:
            self.close()
        return False



class DelayedButtonHelp(QObject):
    """Aide opérateur différée : infobulle après immobilité volontaire sur un bouton."""

    def __init__(self, owner, delay_ms=3000, duration_ms=12000):
        super().__init__()
        self.owner = owner
        self.delay_ms = int(delay_ms)
        self.duration_ms = int(duration_ms)
        self.current_button = None
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)
        self.timer.timeout.connect(self.show_current_help)

    def install_on(self, button, text):
        if button is None or not text:
            return
        button.setToolTip("")
        button.setProperty("operatorHelp", str(text))
        button.installEventFilter(self)

    def help_blocked(self):
        return bool(getattr(self.owner, "auto_neutral_running", False))

    def eventFilter(self, obj, event):
        if isinstance(obj, QPushButton) and obj.property("operatorHelp"):
            event_type = event.type()
            if event_type == QEvent.Enter:
                QToolTip.hideText()
                self.current_button = obj
                self.timer.start(self.delay_ms)
            elif event_type in (QEvent.Leave, QEvent.MouseButtonPress, QEvent.Hide):
                if self.current_button is obj:
                    self.timer.stop()
                    self.current_button = None
                    QToolTip.hideText()
        return super().eventFilter(obj, event)

    def show_current_help(self):
        button = self.current_button
        if button is None or not button.isVisible() or not button.isEnabled():
            return
        if self.help_blocked():
            QToolTip.hideText()
            return
        text = str(button.property("operatorHelp") or "").strip()
        if not text:
            return
        QToolTip.showText(button.mapToGlobal(button.rect().center()), text, button, button.rect(), self.duration_ms)

    def hide(self):
        self.timer.stop()
        self.current_button = None
        QToolTip.hideText()


class SerialReader(QThread):
    line_received = Signal(str)
    error_received = Signal(str)

    def __init__(self, ser):
        super().__init__()
        self.ser = ser
        self.running = True

    def run(self):
        buffer = b""
        while self.running:
            try:
                if self.ser is None or not self.ser.is_open:
                    time.sleep(0.005)
                    continue

                nb = self.ser.in_waiting
                if nb <= 0:
                    time.sleep(0.001)
                    continue

                data = self.ser.read(min(nb, 512))
                if not data:
                    continue

                buffer += data.replace(b"\r", b"\n")

                while b"\n" in buffer:
                    raw_line, buffer = buffer.split(b"\n", 1)
                    line = raw_line.decode("utf-8", errors="replace").strip()
                    if line:
                        self.line_received.emit(line)

                if len(buffer) > 1200:
                    buffer = b""
                    self.error_received.emit("Buffer série trop long, ligne ignorée.")

            except Exception as exc:
                self.error_received.emit(str(exc))
                break

    def stop(self):
        self.running = False



class EAPSU:
    """Pilote SCPI de l'alimentation EA avec contrôle des rampes arbitraires."""

    # EA-PSI 9200-04 T : tension nominale 200 V.
    # Le manuel PSI 9000 T donne une pente minimale de 0,000725 * valeur nominale / s.
    NOMINAL_VOLTAGE_V = 200.0
    MIN_VOLTAGE_SLOPE_V_PER_S = 0.000725 * NOMINAL_VOLTAGE_V

    _NUMBER_RE = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?")

    def __init__(self):
        self.ser = None
        self.connected = False
        self.port = ""
        self.baudrate = 9600
        self.identity = ""
        self.last_ramp_readback = {}
        self.last_scpi_error = ""

    def connect(self, port, baudrate, timeout_s=1.0):
        self.disconnect()
        self.ser = serial.Serial(
            port=port,
            baudrate=int(baudrate),
            timeout=float(timeout_s),
            write_timeout=float(timeout_s),
            bytesize=serial.EIGHTBITS,
            parity=serial.PARITY_NONE,
            stopbits=serial.STOPBITS_ONE,
        )
        self.connected = True
        self.port = str(port)
        self.baudrate = int(baudrate)

    def disconnect(self):
        if self.ser is not None:
            try:
                if self.ser.is_open:
                    self.ser.close()
            except Exception:
                pass
        self.ser = None
        self.connected = False
        self.port = ""
        self.identity = ""
        self.last_ramp_readback = {}
        self.last_scpi_error = ""

    def send(self, command):
        if not self.connected or self.ser is None or not self.ser.is_open:
            raise RuntimeError("Alimentation EA non connectée.")
        payload = (str(command).strip() + "\n").encode("ascii", errors="strict")
        self.ser.write(payload)
        self.ser.flush()

    def query(self, command):
        if not self.connected or self.ser is None or not self.ser.is_open:
            raise RuntimeError("Alimentation EA non connectée.")
        try:
            self.ser.reset_input_buffer()
        except Exception:
            pass
        self.send(command)
        return self.ser.readline().decode("ascii", errors="replace").strip()

    def query_required(self, command, name="réponse"):
        response = self.query(command)
        text = str(response or "").strip()
        if not text:
            raise RuntimeError(f"Aucune réponse EA pour {name} ({command}).")
        return text

    @classmethod
    def parse_number(cls, response, name="valeur"):
        match = cls._NUMBER_RE.search(str(response or ""))
        if not match:
            raise RuntimeError(f"Réponse EA invalide pour {name} : {response!r}")
        return float(match.group(0))

    def query_float(self, command, name="valeur"):
        return self.parse_number(self.query(command), name)

    def set_remote(self):
        # Commande déjà validée sur l'EA-PSI 9200-04 T du projet.
        self.send("SYST:LOCK 1")

    def set_local(self):
        try:
            self.send("SYST:LOC")
        except Exception:
            try:
                self.send("SYST:LOCK 0")
            except Exception:
                pass

    def output(self, enabled):
        self.send("OUTP ON" if enabled else "OUTP OFF")

    def stop_generator(self, leave_mode=False):
        try:
            self.send("FUNC:GEN:WAVE:STAT STOP")
        except Exception:
            pass
        if leave_mode:
            try:
                self.send("FUNC:GEN:SEL NONE")
            except Exception:
                pass

    def read_scpi_error(self):
        """Retourne la file d'erreur SCPI. Une réponse vide ou un timeout est une erreur."""
        response = self.query_required("SYST:ERR?", "file d'erreurs SCPI")
        self.last_scpi_error = str(response).strip()
        return self.last_scpi_error

    @staticmethod
    def scpi_error_is_clear(response):
        text = str(response or "").strip()
        if not text:
            return False
        match = re.match(r"\s*([-+]?\d+)", text)
        return bool(match and int(match.group(1)) == 0)

    def clear_scpi_errors(self, max_reads=8):
        """Vide les anciennes erreurs afin que l'erreur relue après SUBMIT concerne la rampe courante."""
        last = ""
        for _ in range(max(1, int(max_reads))):
            last = self.read_scpi_error()
            if self.scpi_error_is_clear(last):
                return last
        raise RuntimeError(f"File d'erreurs SCPI EA non vidée : {last}")

    def read_wave_value(self, level, index):
        self.send(f"FUNC:GEN:WAVE:LEVEL {int(level)}")
        self.send(f"FUNC:GEN:WAVE:IND {int(index)}")
        return self.query_float("FUNC:GEN:WAVE:DATA?", f"point {level}, index {index}")

    def verify_voltage_ramp(self, start_v, end_v, duration_s):
        """Relit les paramètres réellement mémorisés par l'EA après le délai SUBMIT."""
        error = self.read_scpi_error()
        if error and not self.scpi_error_is_clear(error):
            raise RuntimeError(f"Erreur SCPI après SUBMIT : {error}")

        readback = {
            "start_v": self.read_wave_value(1, 5),
            "end_v": self.read_wave_value(1, 6),
            "duration_s": self.read_wave_value(1, 7),
        }
        voltage_tol = 0.010
        time_tol = max(0.010, abs(float(duration_s)) * 0.002)
        errors = []
        if abs(readback["start_v"] - float(start_v)) > voltage_tol:
            errors.append(f"départ relu {readback['start_v']:.6f} V")
        if abs(readback["end_v"] - float(end_v)) > voltage_tol:
            errors.append(f"fin relue {readback['end_v']:.6f} V")
        if abs(readback["duration_s"] - float(duration_s)) > time_tol:
            errors.append(f"durée relue {readback['duration_s']:.6f} s")
        self.last_ramp_readback = readback
        if errors:
            raise RuntimeError(
                "Rampe EA non conforme à la saisie : " + ", ".join(errors)
                + f" (demandé {float(start_v):.6f} → {float(end_v):.6f} V en {float(duration_s):.6f} s)."
            )
        return readback

    def configure_voltage_ramp(self, start_v, end_v, duration_s, hold_s, current_limit_a):
        """Configure deux points : rampe linéaire puis maintien de la valeur finale."""
        start_v = float(start_v)
        end_v = float(end_v)
        duration_s = float(duration_s)
        hold_s = max(0.1, float(hold_s))
        current_limit_a = float(current_limit_a)
        if not (0.0001 <= duration_s <= 36000.0):
            raise ValueError("Durée de rampe EA hors plage 0,0001...36000 s.")
        if start_v < 0 or end_v < 0:
            raise ValueError("Les tensions de rampe doivent être positives.")
        if current_limit_a <= 0:
            raise ValueError("La limite de courant doit être supérieure à 0 A.")
        span_v = abs(end_v - start_v)
        if span_v > 0:
            requested_slope = span_v / duration_s
            if requested_slope + 1e-12 < self.MIN_VOLTAGE_SLOPE_V_PER_S:
                max_duration_s = span_v / self.MIN_VOLTAGE_SLOPE_V_PER_S
                raise ValueError(
                    f"Rampe trop lente pour l'EA-PSI 9200-04 T : {requested_slope:.6f} V/s, "
                    f"minimum {self.MIN_VOLTAGE_SLOPE_V_PER_S:.6f} V/s. "
                    f"Pour un écart de {span_v:.3f} V, durée maximale ≈ {max_duration_s:.3f} s."
                )

        self.set_remote()
        self.stop_generator(leave_mode=True)
        self.clear_scpi_errors()
        self.send("FUNC:GEN:SEL VOLTAGE")
        for level, va, vb, seconds in ((1, start_v, end_v, duration_s), (2, end_v, end_v, hold_s)):
            self.send(f"FUNC:GEN:WAVE:LEVEL {level}")
            # Les paramètres AC inutilisés sont explicitement remis à zéro.
            for index in range(5):
                self.send(f"FUNC:GEN:WAVE:IND {index}")
                self.send("FUNC:GEN:WAVE:DATA 0")
            self.send("FUNC:GEN:WAVE:IND 5")
            self.send(f"FUNC:GEN:WAVE:DATA {va:.6f}")
            self.send("FUNC:GEN:WAVE:IND 6")
            self.send(f"FUNC:GEN:WAVE:DATA {vb:.6f}")
            # La durée doit être écrite en dernier : elle déclenche le contrôle de pente EA.
            self.send("FUNC:GEN:WAVE:IND 7")
            self.send(f"FUNC:GEN:WAVE:DATA {seconds:.6f}")
        self.send("FUNC:GEN:WAVE:END 2")
        self.send("FUNC:GEN:WAVE:START 1")
        self.send("FUNC:GEN:WAVE:NUM 1")

        # Valeurs statiques/limites placées avant SUBMIT afin que SUBMIT reste la
        # dernière commande, puis aucune commande n'est envoyée pendant ≥2 s.
        self.send(f"CURR {current_limit_a:.6f}")
        self.send("POW MAX")
        self.send(f"VOLT {start_v:.6f}")
        self.send("FUNC:GEN:WAVE:SUBMIT")
        self.last_ramp_readback = {}

    def start_generator(self):
        self.send("OUTP ON")
        self.send("FUNC:GEN:WAVE:STAT RUN")

    @staticmethod
    def _state_tokens(response):
        text = str(response or "").strip().upper()
        return text, [token for token in re.split(r"[^A-Z0-9.+-]+", text) if token]

    @classmethod
    def generator_state_is_running(cls, response):
        text, tokens = cls._state_tokens(response)
        if text == "ON" or any(token.startswith("RUN") for token in tokens):
            return True
        try:
            return cls.parse_number(text, "état générateur") >= 0.5
        except Exception:
            return False

    @classmethod
    def generator_state_is_stopped(cls, response):
        text, tokens = cls._state_tokens(response)
        if text == "OFF" or any(token.startswith(("STOP", "IDLE")) for token in tokens):
            return True
        try:
            return cls.parse_number(text, "état générateur") < 0.5
        except Exception:
            return False

    @classmethod
    def output_state_is_off(cls, response):
        text, tokens = cls._state_tokens(response)
        if text in ("0", "OFF", "FALSE") or "OFF" in tokens:
            return True
        try:
            return abs(cls.parse_number(text, "état sortie")) < 0.5
        except Exception:
            return False

    @classmethod
    def output_state_is_on(cls, response):
        text, tokens = cls._state_tokens(response)
        if text in ("1", "ON", "TRUE") or "ON" in tokens:
            return True
        try:
            return abs(cls.parse_number(text, "état sortie")) >= 0.5
        except Exception:
            return False

    def configure_static_output_and_confirm(
        self,
        target_voltage_v,
        current_limit_a,
        timeout_s=EA_STATIC_CONFIRM_TIMEOUT_S,
        poll_interval_s=EA_STATIC_CONFIRM_POLL_INTERVAL_S,
    ):
        """Prépare l'EA en source continue et confirme la tension avant chronométrie.

        Le générateur arbitraire est quitté, la sortie est réglée en mode statique,
        puis SEL=NONE, OUTP=ON, MEAS:VOLT et SYST:ERR? sont contrôlés. Une réponse
        vide ou un timeout reste une erreur fermée.
        """
        target_voltage_v = float(target_voltage_v)
        current_limit_a = float(current_limit_a)
        if not (0.0 < target_voltage_v <= self.NOMINAL_VOLTAGE_V):
            raise ValueError(
                f"Tension chronométrie hors plage : {target_voltage_v:.3f} V."
            )
        if current_limit_a <= 0.0:
            raise ValueError("La limite de courant chronométrie doit être supérieure à 0 A.")
        if not self.connected or self.ser is None or not self.ser.is_open:
            raise RuntimeError("Alimentation EA non connectée.")

        tolerance_v = max(
            EA_STATIC_CONFIRM_MIN_TOL_V,
            abs(target_voltage_v) * EA_STATIC_CONFIRM_REL_TOL,
        )
        result = {
            "confirmed": False,
            "target_voltage_v": target_voltage_v,
            "tolerance_v": tolerance_v,
            "generator_selection": "",
            "output_state": "",
            "measured_voltage_v": None,
            "scpi_error": "",
            "preexisting_scpi_errors": [],
            "errors": [],
            "poll_count": 0,
            "confirmation_elapsed_s": 0.0,
        }
        started = time.monotonic()
        deadline = started + max(0.25, float(timeout_s))
        old_timeout = getattr(self.ser, "timeout", None)
        timeout_changed = False
        try:
            if old_timeout is not None:
                self.ser.timeout = min(float(old_timeout), EA_STOP_CONFIRM_QUERY_TIMEOUT_S)
                timeout_changed = True
        except Exception:
            timeout_changed = False

        try:
            self.set_remote()
            previous = self.drain_scpi_errors()
            if previous:
                result["preexisting_scpi_errors"] = list(previous)
            selection = self.generator_selection()
            if self.generator_selection_is_arbitrary(selection):
                self.send("FUNC:GEN:WAVE:STAT STOP")
            self.send("FUNC:GEN:SEL NONE")
            self.send(f"CURR {current_limit_a:.6f}")
            self.send("POW MAX")
            self.send(f"VOLT {target_voltage_v:.6f}")
            self.send("OUTP ON")

            last_errors = []
            while time.monotonic() <= deadline:
                result["poll_count"] += 1
                poll_errors = []
                try:
                    result["generator_selection"] = self.generator_selection()
                except Exception as exc:
                    poll_errors.append(f"sélection générateur: {exc}")
                try:
                    result["output_state"] = self.output_state()
                except Exception as exc:
                    poll_errors.append(f"état sortie: {exc}")
                try:
                    result["measured_voltage_v"] = float(self.measured_voltage())
                except Exception as exc:
                    poll_errors.append(f"tension sortie: {exc}")

                selection_ok = self.generator_selection_is_none(result["generator_selection"])
                output_ok = self.output_state_is_on(result["output_state"])
                measured = result["measured_voltage_v"]
                voltage_ok = (
                    measured is not None
                    and abs(float(measured) - target_voltage_v) <= tolerance_v
                )
                if selection_ok and output_ok and voltage_ok and not poll_errors:
                    try:
                        result["scpi_error"] = self.read_scpi_error()
                    except Exception as exc:
                        poll_errors.append(f"file SCPI: {exc}")
                    if self.scpi_error_is_clear(result["scpi_error"]) and not poll_errors:
                        result["confirmed"] = True
                        result["confirmation_elapsed_s"] = time.monotonic() - started
                        return result
                    if not self.scpi_error_is_clear(result["scpi_error"]):
                        poll_errors.append(
                            f"file SCPI non claire: {result['scpi_error'] or 'réponse vide'}"
                        )
                        last_errors = poll_errors
                        break
                last_errors = poll_errors
                remaining = max(0.0, deadline - time.monotonic())
                if remaining <= 0:
                    break
                time.sleep(min(max(0.01, float(poll_interval_s)), remaining))

            selection_ok = self.generator_selection_is_none(result["generator_selection"])
            output_ok = self.output_state_is_on(result["output_state"])
            measured = result["measured_voltage_v"]
            voltage_ok = (
                measured is not None
                and abs(float(measured) - target_voltage_v) <= tolerance_v
            )
            validation_errors = list(last_errors)
            if not selection_ok:
                validation_errors.append(
                    f"mode générateur non statique: {result['generator_selection'] or 'réponse vide'}"
                )
            if not output_ok:
                validation_errors.append(
                    f"sortie EA non active: {result['output_state'] or 'réponse vide'}"
                )
            if not voltage_ok:
                measured_text = "non relue" if measured is None else f"{float(measured):.3f} V"
                validation_errors.append(
                    f"tension statique {measured_text}, cible {target_voltage_v:.3f} V "
                    f"± {tolerance_v:.3f} V"
                )
            if not result["scpi_error"]:
                try:
                    result["scpi_error"] = self.read_scpi_error()
                except Exception as exc:
                    validation_errors.append(f"file SCPI finale: {exc}")
            if not self.scpi_error_is_clear(result["scpi_error"]):
                validation_errors.append(
                    f"file SCPI non claire: {result['scpi_error'] or 'réponse vide'}"
                )
            result["errors"] = validation_errors
            result["confirmation_elapsed_s"] = time.monotonic() - started
            return result
        except Exception as exc:
            result["errors"] = list(result.get("errors", [])) + [str(exc)]
            result["confirmation_elapsed_s"] = time.monotonic() - started
            return result
        finally:
            if timeout_changed:
                try:
                    self.ser.timeout = old_timeout
                except Exception:
                    pass

    def generator_state(self):
        return self.query_required("FUNC:GEN:WAVE:STAT?", "état générateur").upper()

    def generator_selection(self):
        """Retourne le mode générateur actif (VOLTAGE/CURRENT/NONE...)."""
        return self.query_required("FUNC:GEN:SEL?", "sélection générateur").upper()

    @classmethod
    def generator_selection_is_none(cls, response):
        text, tokens = cls._state_tokens(response)
        return text == "NONE" or "NONE" in tokens

    @classmethod
    def generator_selection_is_arbitrary(cls, response):
        text, tokens = cls._state_tokens(response)
        return any(token.startswith(("VOLT", "CURR")) for token in tokens) or text.startswith(("VOLT", "CURR"))

    def drain_scpi_errors(self, max_reads=8):
        """Vide les erreurs anciennes et les retourne pour diagnostic, sans masquer une absence de réponse."""
        previous = []
        for _ in range(max(1, int(max_reads))):
            response = self.read_scpi_error()
            if self.scpi_error_is_clear(response):
                return previous
            previous.append(response)
        raise RuntimeError("File SCPI ancienne impossible à vider : " + " | ".join(previous))

    def output_state(self):
        return self.query_required("OUTP?", "état sortie").upper()

    def measured_voltage(self):
        return self.parse_number(self.query_required("MEAS:VOLT?", "tension de sortie"), "tension de sortie")

    def safe_stop_and_confirm(
        self,
        max_voltage_v=EA_STOP_CONFIRM_MAX_V,
        settle_s=EA_STOP_CONFIRM_SETTLE_S,
        attempts=EA_STOP_CONFIRM_ATTEMPTS,
        confirm_timeout_s=EA_STOP_CONFIRM_TIMEOUT_S,
        poll_interval_s=EA_STOP_CONFIRM_POLL_INTERVAL_S,
    ):
        """Met l'EA dans un état sûr et attend sa décharge réelle avant validation.

        R3 : la confirmation n'est plus faite après un délai fixe de 0,2 s. La sortie,
        la sélection du générateur et MEAS:VOLT? sont surveillées jusqu'à 5 s. Le
        contrôle réussit dès que SEL=NONE, OUTP=OFF et |V| <= seuil. Le diagnostic
        conserve les dernières réponses exactes en cas d'échec.
        """
        result = {
            "confirmed": False,
            "generator_selection_before": "",
            "generator_selection": "",
            "generator_state": "",
            "output_state": "",
            "measured_voltage_v": None,
            "scpi_error": "",
            "preexisting_scpi_errors": [],
            "errors": [],
            "poll_count": 0,
            "confirmation_elapsed_s": 0.0,
        }
        if not self.connected or self.ser is None or not self.ser.is_open:
            result["errors"].append("Alimentation EA non connectée.")
            return result

        old_timeout = getattr(self.ser, "timeout", None)
        timeout_changed = False
        try:
            if old_timeout is not None:
                self.ser.timeout = min(float(old_timeout), EA_STOP_CONFIRM_QUERY_TIMEOUT_S)
                timeout_changed = True
        except Exception:
            timeout_changed = False

        started = time.monotonic()
        total_timeout_s = max(0.05, float(confirm_timeout_s))
        deadline = started + total_timeout_s
        max_attempts = max(1, int(attempts))
        last_command_errors = []
        last_check_errors = []

        try:
            for attempt in range(1, max_attempts + 1):
                command_errors = []
                check_errors = []

                # Les erreurs antérieures sont purgées avant la séquence actuelle,
                # mais restent conservées dans le diagnostic.
                try:
                    old_errors = self.drain_scpi_errors()
                    if old_errors:
                        result["preexisting_scpi_errors"].extend(old_errors)
                except Exception as exc:
                    check_errors.append(f"purge file SCPI: {exc}")

                selection_before = ""
                if not check_errors:
                    try:
                        selection_before = self.generator_selection()
                        result["generator_selection_before"] = selection_before
                    except Exception as exc:
                        check_errors.append(f"sélection générateur initiale: {exc}")

                # WAVE:STAT n'est envoyé que si un générateur arbitraire est actif.
                if not check_errors and self.generator_selection_is_arbitrary(selection_before):
                    try:
                        self.send("FUNC:GEN:WAVE:STAT STOP")
                    except Exception as exc:
                        command_errors.append(f"FUNC:GEN:WAVE:STAT STOP: {exc}")

                for command in ("OUTP OFF", "VOLT 0", "FUNC:GEN:SEL NONE"):
                    try:
                        self.send(command)
                    except Exception as exc:
                        command_errors.append(f"{command}: {exc}")

                if settle_s:
                    time.sleep(max(0.0, min(float(settle_s), max(0.0, deadline - time.monotonic()))))

                # Surveillance adaptative : la sortie d'une alimentation peu chargée
                # peut nécessiter plusieurs secondes pour descendre sous 0,200 V.
                # Le délai total est partagé entre les tentatives afin de pouvoir
                # réémettre une fois les commandes d'arrêt sans dépasser 5 s.
                attempt_deadline = started + (total_timeout_s * attempt / max_attempts)
                attempt_deadline = min(deadline, attempt_deadline)
                while time.monotonic() <= attempt_deadline:
                    result["poll_count"] += 1
                    poll_errors = []
                    try:
                        result["generator_selection"] = self.generator_selection()
                        result["generator_state"] = result["generator_selection"]
                    except Exception as exc:
                        poll_errors.append(f"sélection générateur finale: {exc}")
                    try:
                        result["output_state"] = self.output_state()
                    except Exception as exc:
                        poll_errors.append(f"état sortie: {exc}")
                    try:
                        result["measured_voltage_v"] = float(self.measured_voltage())
                    except Exception as exc:
                        poll_errors.append(f"tension sortie: {exc}")

                    generator_ok = self.generator_selection_is_none(result["generator_selection"])
                    output_ok = self.output_state_is_off(result["output_state"])
                    voltage = result["measured_voltage_v"]
                    voltage_ok = voltage is not None and abs(float(voltage)) <= abs(float(max_voltage_v))

                    if generator_ok and output_ok and voltage_ok and not poll_errors:
                        try:
                            result["scpi_error"] = self.read_scpi_error()
                        except Exception as exc:
                            poll_errors.append(f"file SCPI: {exc}")
                        scpi_ok = self.scpi_error_is_clear(result["scpi_error"])
                        if scpi_ok and not poll_errors and not command_errors and not check_errors:
                            result["confirmed"] = True
                            result["errors"] = []
                            result["confirmation_elapsed_s"] = time.monotonic() - started
                            return result
                        if not scpi_ok:
                            poll_errors.append(
                                f"file SCPI non claire: {result['scpi_error'] or 'réponse vide'}"
                            )
                            last_check_errors = check_errors + poll_errors
                            break

                    last_check_errors = check_errors + poll_errors
                    if time.monotonic() >= attempt_deadline:
                        break
                    remaining = max(0.0, attempt_deadline - time.monotonic())
                    time.sleep(min(max(0.01, float(poll_interval_s)), remaining))

                last_command_errors = command_errors
                if result["confirmed"] or time.monotonic() >= deadline:
                    break

            # Une dernière lecture de la file d'erreur complète le diagnostic si
            # l'état sûr n'a pas été atteint avant le délai maximal.
            if not result["scpi_error"]:
                try:
                    result["scpi_error"] = self.read_scpi_error()
                except Exception as exc:
                    last_check_errors.append(f"file SCPI finale: {exc}")

            generator_ok = self.generator_selection_is_none(result["generator_selection"])
            output_ok = self.output_state_is_off(result["output_state"])
            voltage = result["measured_voltage_v"]
            voltage_ok = voltage is not None and abs(float(voltage)) <= abs(float(max_voltage_v))
            scpi_ok = self.scpi_error_is_clear(result["scpi_error"])
            validation_errors = []
            if not generator_ok:
                validation_errors.append(
                    f"mode générateur non quitté: {result['generator_selection'] or 'réponse vide'}"
                )
            if not output_ok:
                validation_errors.append(
                    f"sortie non coupée: {result['output_state'] or 'réponse vide'}"
                )
            if not voltage_ok:
                voltage_text = "non relue" if voltage is None else f"{float(voltage):.3f} V"
                validation_errors.append(
                    f"tension résiduelle {voltage_text} > {abs(float(max_voltage_v)):.3f} V "
                    f"après {total_timeout_s:.1f} s"
                )
            if not scpi_ok:
                validation_errors.append(
                    f"file SCPI non claire: {result['scpi_error'] or 'réponse vide'}"
                )
            result["errors"] = last_command_errors + last_check_errors + validation_errors
            result["confirmation_elapsed_s"] = time.monotonic() - started
            return result
        finally:
            if timeout_changed:
                try:
                    self.ser.timeout = old_timeout
                except Exception:
                    pass



class SortableTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        left = self.data(Qt.UserRole)
        right = other.data(Qt.UserRole) if other is not None else None
        if left is not None and right is not None:
            return left < right
        return super().__lt__(other)



def draw_logic_oscillogram(
    painter, x, y, width, height, events, start_bits, capture_us, title="",
    view_start_us=0, view_end_us=None, cursor_a_us=None, cursor_b_us=None,
    display_mode=OSCILLO_DISPLAY_ELECTRIC, phase_markers=None, current=None,
):
    """Dessine un oscillogramme vectoriel à l'échelle temps réelle."""
    try:
        painter.setRenderHint(QPainter.Antialiasing, True)
    except Exception:
        pass

    painter.fillRect(int(x), int(y), int(width), int(height), QColor(255, 255, 255))
    painter.setPen(QPen(QColor(20, 20, 20), 1))
    painter.drawRect(int(x), int(y), int(width), int(height))

    compact = height < 360
    margin_left = 54
    margin_right = 16
    margin_top = 32 if compact else 38
    margin_bottom = 24 if compact else 32
    plot_x = float(x + margin_left)
    plot_y = float(y + margin_top)
    plot_w = float(max(10, width - margin_left - margin_right))
    plot_h = float(max(10, height - margin_top - margin_bottom))
    row_h = plot_h / max(1, len(CHRONO_CONTACT_NAMES))
    capture_end_us = max(1, int(capture_us or 0))
    if events:
        capture_end_us = max(capture_end_us, max(int(event.get("t_us", 0)) for event in events) + 1)
    view_start_us = max(0, min(capture_end_us - 1, int(view_start_us or 0)))
    view_end_us = capture_end_us if view_end_us is None else int(view_end_us)
    view_end_us = max(view_start_us + 1, min(capture_end_us, view_end_us))
    view_span_us = max(1, view_end_us - view_start_us)
    pre_t0_gap_us = oscillo_pre_t0_gap_us(view_start_us, view_end_us)
    visual_start_us = view_start_us - pre_t0_gap_us
    visual_end_us = view_end_us
    visual_span_us = max(1, visual_end_us - visual_start_us)
    us_per_px = visual_span_us / max(1.0, plot_w)

    def x_from_us(value):
        clipped = max(visual_start_us, min(visual_end_us, int(value)))
        return plot_x + ((clipped - visual_start_us) / visual_span_us) * plot_w

    def fmt_time(us_value):
        us_value = float(us_value)
        if us_value < 0:
            return f"-{abs(us_value):.0f} µs"
        if visual_span_us <= 2000:
            return f"{us_value:.0f} µs"
        if visual_span_us <= 200000:
            return f"{us_value / 1000.0:.3f} ms"
        return f"{us_value / 1000.0:.1f} ms"

    painter.setFont(QFont("Arial", 8 if compact else 9, QFont.Bold))
    painter.setPen(QColor(20, 20, 20))
    painter.drawText(int(x + 8), int(y + (17 if compact else 20)), title or "Oscillogramme contacts")
    painter.setFont(QFont("Arial", 7 if compact else 8))
    painter.setPen(QColor(80, 80, 80))
    mode_label = "Électrique GPIO : 1 = 3,3 V ouvert / 0 = 0 V fermé"
    if display_mode == OSCILLO_DISPLAY_LOGIC:
        mode_label = "Logique contact : 1 = fermé / 0 = ouvert"
    pre_t0_txt = f"  |  pré-T0 visuel {pre_t0_gap_us} µs" if pre_t0_gap_us else ""
    painter.drawText(
        int(x + 8), int(y + (29 if compact else 34)),
        f"{mode_label}  |  Vue {view_start_us} → {view_end_us} µs{pre_t0_txt}  |  {us_per_px:.3f} µs/pixel",
    )

    painter.setFont(QFont("Arial", 8))
    painter.setPen(QPen(QColor(225, 225, 225), 1))
    for i in range(11):
        gx = plot_x + plot_w * i / 10.0
        painter.drawLine(QLineF(gx, plot_y, gx, plot_y + plot_h))
    for row in range(len(CHRONO_CONTACT_NAMES) + 1):
        gy = plot_y + row_h * row
        painter.drawLine(QLineF(plot_x, gy, plot_x + plot_w, gy))

    if pre_t0_gap_us:
        t0_x = x_from_us(0)
        painter.fillRect(int(plot_x), int(plot_y), max(1, int(t0_x - plot_x)), int(plot_h), QColor(245, 248, 255, 95))
        painter.setPen(QPen(QColor(30, 95, 210), 1, Qt.DashLine))
        painter.drawLine(QLineF(t0_x, plot_y, t0_x, plot_y + plot_h))
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.setPen(QColor(30, 95, 210))
        painter.drawText(int(t0_x + 4), int(plot_y + plot_h - 8), oscillo_t0_command_label(current, phase_markers))

    for marker in phase_markers or []:
        try:
            marker_t = int(marker.get("t_us", 0))
        except Exception:
            continue
        if not (visual_start_us <= marker_t <= visual_end_us):
            continue
        mx = x_from_us(marker_t)
        painter.setPen(QPen(QColor(120, 80, 0), 1, Qt.DashLine))
        painter.drawLine(QLineF(mx, plot_y, mx, plot_y + plot_h))
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.setPen(QColor(120, 80, 0))
        painter.drawText(int(mx + 4), int(plot_y + 12), str(marker.get("label", "")))

    painter.setPen(QColor(80, 80, 80))
    for i in range(6):
        tx = visual_start_us + visual_span_us * i / 5.0
        gx = plot_x + plot_w * i / 5.0
        painter.drawText(int(gx - 28), int(y + height - 8), fmt_time(tx))

    events_by_contact = {name: [] for name in CHRONO_CONTACT_NAMES}
    for event in events:
        contact = str(event.get("contact", "")).upper()
        if contact in events_by_contact:
            try:
                state = int(event.get("state", 0))
                if display_mode == OSCILLO_DISPLAY_ELECTRIC:
                    state = 0 if state else 1
                events_by_contact[contact].append((int(event.get("t_us", 0)), state))
            except Exception:
                pass
    for contact in CHRONO_CONTACT_NAMES:
        events_by_contact[contact].sort(key=lambda item: item[0])

    for idx, contact in enumerate(CHRONO_CONTACT_NAMES):
        top = plot_y + idx * row_h
        low_y = top + row_h * 0.72
        high_y = top + row_h * 0.28
        label_y = top + row_h * 0.55
        state = 1 if start_bits is not None and ((int(start_bits) >> idx) & 0x01) else 0
        if display_mode == OSCILLO_DISPLAY_ELECTRIC:
            state = 0 if state else 1
        for t_us, next_state in events_by_contact[contact]:
            if t_us < view_start_us:
                state = next_state
            else:
                break
        prev_y = high_y if state else low_y
        previous_x = x_from_us(view_start_us)
        color = QColor(0, 120, 55) if contact.startswith("R") else QColor(180, 20, 20)
        painter.setFont(QFont("Arial", 7 if compact else 8, QFont.Bold))
        painter.setPen(QPen(QColor(20, 20, 20), 1))
        painter.drawText(int(x + 10), int(label_y + 4), contact)
        marker_count_by_x = {}
        for t_us, next_state in events_by_contact[contact]:
            if t_us < view_start_us:
                continue
            if t_us > view_end_us:
                break
            x_event = x_from_us(t_us)
            painter.setPen(QPen(color, 2))
            painter.drawLine(QLineF(previous_x, prev_y, x_event, prev_y))
            next_y = high_y if next_state else low_y
            painter.drawLine(QLineF(x_event, prev_y, x_event, next_y))
            px_key = int(round(x_event))
            marker_index = marker_count_by_x.get(px_key, 0)
            marker_count_by_x[px_key] = marker_index + 1
            slot = ((marker_index + 1) // 2) * (1 if marker_index % 2 else -1)
            marker_y = max(top + 6, min(top + row_h - 6, next_y + slot * 4))
            marker_radius = 2 if compact or us_per_px <= 1.0 else 3
            painter.setPen(QPen(QColor(25, 25, 25), 1))
            painter.drawLine(QLineF(x_event, top + 2, x_event, top + row_h - 2))
            painter.drawEllipse(int(round(x_event)) - marker_radius, int(round(marker_y)) - marker_radius, marker_radius * 2, marker_radius * 2)
            if us_per_px <= 0.25:
                painter.setFont(QFont("Arial", 7))
                painter.drawText(int(round(x_event)) + 3, int(round(marker_y)) - 3, str(t_us))
            previous_x = x_event
            prev_y = next_y
        painter.setPen(QPen(color, 2))
        painter.drawLine(QLineF(previous_x, prev_y, x_from_us(view_end_us), prev_y))

    for cursor, color, label in ((cursor_a_us, QColor(0, 90, 200), "A"), (cursor_b_us, QColor(210, 120, 0), "B")):
        if cursor is None:
            continue
        cursor = int(cursor)
        if view_start_us <= cursor <= view_end_us:
            cx = x_from_us(cursor)
            painter.setPen(QPen(color, 2))
            painter.drawLine(QLineF(cx, plot_y, cx, plot_y + plot_h))
            painter.setFont(QFont("Arial", 8, QFont.Bold))
            painter.drawText(int(cx + 4), int(plot_y + 12), f"{label} {cursor} µs")


def draw_synthesis_oscillogram(
    painter, x, y, width, height, events, start_bits, capture_us, current=None, title="",
    view_start_us=0, view_end_us=None, cursor_a_us=None, cursor_b_us=None,
):
    """Vue synthèse métier : une seule trace composite par inverseur."""
    current = dict(current or {})
    action = str(current.get("ACTION") or current.get("action") or "").upper()
    try:
        nb_inv = int(current.get("NB_INV") or current.get("nb_inv") or 4)
    except Exception:
        nb_inv = 4
    nb_inv = max(1, min(4, nb_inv))

    try:
        painter.setRenderHint(QPainter.Antialiasing, True)
    except Exception:
        pass
    painter.fillRect(int(x), int(y), int(width), int(height), QColor(255, 255, 255))
    painter.setPen(QPen(QColor(20, 20, 20), 1))
    painter.drawRect(int(x), int(y), int(width), int(height))

    margin_left = 92
    margin_right = 18
    margin_top = 46
    margin_bottom = 24
    plot_x = float(x + margin_left)
    plot_y = float(y + margin_top)
    plot_w = float(max(10, width - margin_left - margin_right))
    plot_h = float(max(10, height - margin_top - margin_bottom))
    capture_end_us = max(1, int(capture_us or 0))
    if events:
        capture_end_us = max(capture_end_us, max(int(event.get("t_us", 0)) for event in events) + 1)
    view_start_us = max(0, min(capture_end_us - 1, int(view_start_us or 0)))
    view_end_us = capture_end_us if view_end_us is None else int(view_end_us)
    view_end_us = max(view_start_us + 1, min(capture_end_us, view_end_us))
    pre_t0_gap_us = oscillo_pre_t0_gap_us(view_start_us, view_end_us)
    visual_start_us = view_start_us - pre_t0_gap_us
    visual_end_us = view_end_us
    visual_span_us = max(1, visual_end_us - visual_start_us)
    block_h = plot_h / nb_inv

    def x_from_us(value):
        clipped = max(visual_start_us, min(visual_end_us, int(value)))
        return plot_x + ((clipped - visual_start_us) / visual_span_us) * plot_w

    def events_for(contact, phase=None):
        data = []
        for event in events:
            if str(event.get("contact", "")).upper() == contact:
                event_phase = str(event.get("phase", "")).upper()
                if phase and event_phase != str(phase).upper():
                    continue
                try:
                    data.append((int(event.get("t_us", 0)), int(event.get("state", 0)), event_phase))
                except Exception:
                    pass
        return sorted(data, key=lambda item: item[0])

    def first_state(contact, state, phase=None, after_us=None):
        for t_us, value, _phase in events_for(contact, phase):
            if after_us is not None and t_us < after_us:
                continue
            if value == state:
                return t_us
        return None

    def last_state(contact, state, phase=None, after_us=None):
        values = [t_us for t_us, value, _phase in events_for(contact, phase) if value == state and (after_us is None or t_us >= after_us)]
        return values[-1] if values else None

    def draw_marker(t_us, label, color, top, bottom, above=True, dashed=True):
        if t_us is None or not (visual_start_us <= t_us <= visual_end_us):
            return
        mx = x_from_us(t_us)
        painter.setPen(QPen(color, 1, Qt.DashLine if dashed else Qt.SolidLine))
        painter.drawLine(QLineF(mx, top, mx, bottom))
        painter.setFont(QFont("Arial", 7, QFont.Bold))
        painter.setPen(color)
        text_y = top - 4 if above else bottom + 12
        painter.drawText(int(mx + 3), int(text_y), label)

    def bit_state(bit_index):
        return 1 if start_bits is not None and ((int(start_bits) >> bit_index) & 0x01) else 0

    def composite_y(r_state, t_state, repos_y, transfer_y, travail_y):
        if int(t_state) == 1 and int(r_state) == 0:
            return travail_y
        if int(r_state) == 1 and int(t_state) == 0:
            return repos_y
        return transfer_y

    def draw_zone(start_us, end_us, top, bottom, label, color):
        if start_us is None or end_us is None or end_us <= start_us:
            return
        left = x_from_us(start_us)
        right = x_from_us(end_us)
        if right > left:
            painter.fillRect(int(left), int(top + 4), max(1, int(right - left)), int(bottom - top - 8), color)
            painter.setFont(QFont("Arial", 7))
            painter.setPen(QColor(80, 80, 80))
            painter.drawText(int((left + right) / 2 - 28), int(top + 13), label)

    painter.setFont(QFont("Arial", 9, QFont.Bold))
    painter.setPen(QColor(20, 20, 20))
    painter.drawText(int(x + 8), int(y + 20), title or "Synthèse transfert / rebonds")
    painter.setFont(QFont("Arial", 8))
    painter.setPen(QColor(80, 80, 80))
    painter.drawText(int(x + 8), int(y + 36), f"Vue synthèse : une trace par inverseur, niveaux Repos / Transfert / Travail | {view_start_us} → {view_end_us} µs" + (f" | pré-T0 visuel {pre_t0_gap_us} µs" if pre_t0_gap_us else ""))

    painter.setPen(QPen(QColor(230, 230, 230), 1))
    for i in range(11):
        gx = plot_x + plot_w * i / 10.0
        painter.drawLine(QLineF(gx, plot_y, gx, plot_y + plot_h))
    for row in range(nb_inv + 1):
        gy = plot_y + block_h * row
        painter.drawLine(QLineF(plot_x, gy, plot_x + plot_w, gy))

    if pre_t0_gap_us:
        t0_x = x_from_us(0)
        painter.fillRect(int(plot_x), int(plot_y), max(1, int(t0_x - plot_x)), int(plot_h), QColor(245, 248, 255, 95))
        painter.setPen(QPen(QColor(30, 95, 210), 1, Qt.DashLine))
        painter.drawLine(QLineF(t0_x, plot_y, t0_x, plot_y + plot_h))
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.setPen(QColor(30, 95, 210))
        painter.drawText(int(t0_x + 4), int(plot_y + plot_h - 8), oscillo_t0_command_label(current, current.get("phase_markers", [])))

    forward = action in ("BE", "MONO_ON")
    reverse = action in ("BR", "MONO_OFF")
    combined = bool(current.get("combined_cycle"))
    mode_text = "Cycle complet" if combined else "Repos → Travail" if forward else "Travail → Repos" if reverse else "Mesure"
    painter.setFont(QFont("Arial", 8, QFont.Bold))
    painter.setPen(QColor(30, 30, 30))
    painter.drawText(int(plot_x + 4), int(plot_y - 8), mode_text)

    for marker in current.get("phase_markers", []) if isinstance(current.get("phase_markers"), list) else []:
        try:
            marker_t = int(marker.get("t_us", 0))
        except Exception:
            continue
        draw_marker(marker_t, str(marker.get("label", "")), QColor(120, 80, 0), plot_y, plot_y + plot_h, above=True)

    for inv in range(1, nb_inv + 1):
        top = plot_y + (inv - 1) * block_h
        bottom = top + block_h
        travail_y = top + block_h * 0.25
        repos_y = top + block_h * 0.72
        transfer_y = top + block_h * 0.50
        contact_r = f"R{inv}"
        contact_t = f"T{inv}"

        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.setPen(QColor(20, 20, 20))
        painter.drawText(int(x + 10), int(travail_y + 4), f"TRAVAIL {inv}")
        painter.drawText(int(x + 10), int(transfer_y + 4), "transfert")
        painter.drawText(int(x + 10), int(repos_y + 4), f"REPOS {inv}")

        painter.setPen(QPen(QColor(235, 120, 120), 1, Qt.DashLine))
        painter.drawLine(QLineF(plot_x, travail_y, plot_x + plot_w, travail_y))
        painter.setPen(QPen(QColor(130, 130, 130), 1, Qt.DashLine))
        painter.drawLine(QLineF(plot_x, transfer_y, plot_x + plot_w, transfer_y))
        painter.setPen(QPen(QColor(210, 210, 210), 1))
        painter.drawLine(QLineF(plot_x, repos_y, plot_x + plot_w, repos_y))

        all_events = []
        for t_us, state, phase in events_for(contact_r):
            all_events.append((t_us, contact_r, state, phase))
        for t_us, state, phase in events_for(contact_t):
            all_events.append((t_us, contact_t, state, phase))
        all_events.sort(key=lambda item: item[0])

        r_state = bit_state(inv - 1)
        t_state = bit_state(inv + 3)
        for t_us, contact, state, _phase in all_events:
            if t_us < view_start_us:
                if contact == contact_r:
                    r_state = state
                else:
                    t_state = state
            else:
                break

        prev_x = x_from_us(view_start_us)
        prev_y = composite_y(r_state, t_state, repos_y, transfer_y, travail_y)
        painter.setPen(QPen(QColor(15, 15, 15), 1))
        for t_us, contact, state, _phase in all_events:
            if t_us < view_start_us:
                continue
            if t_us > view_end_us:
                break
            x_event = x_from_us(t_us)
            painter.setPen(QPen(QColor(15, 15, 15), 1))
            painter.drawLine(QLineF(prev_x, prev_y, x_event, prev_y))
            if contact == contact_r:
                r_state = state
            else:
                t_state = state
            next_y = composite_y(r_state, t_state, repos_y, transfer_y, travail_y)
            painter.drawLine(QLineF(x_event, prev_y, x_event, next_y))
            painter.setPen(QPen(QColor(35, 35, 35), 1))
            painter.drawEllipse(int(round(x_event)) - 2, int(round(next_y)) - 2, 4, 4)
            prev_x = x_event
            prev_y = next_y
        painter.setPen(QPen(QColor(15, 15, 15), 1))
        painter.drawLine(QLineF(prev_x, prev_y, x_from_us(view_end_us), prev_y))

        phases_to_draw = []
        if combined:
            phases_to_draw = [("BE", "forward"), ("MONO ON", "forward"), ("BR", "reverse"), ("MONO OFF", "reverse")]
        elif forward:
            phases_to_draw = [(None, "forward")]
        elif reverse:
            phases_to_draw = [(None, "reverse")]
        for phase, direction in phases_to_draw:
            if direction == "forward":
                open_first = first_state(contact_r, 0, phase=phase)
                open_last = last_state(contact_r, 0, phase=phase, after_us=open_first)
                close_first = first_state(contact_t, 1, phase=phase, after_us=open_first)
                close_last = last_state(contact_t, 1, phase=phase, after_us=close_first)
                draw_zone(open_first, close_first, top, bottom, "transfert", QColor(225, 225, 225, 55))
                draw_zone(open_first, open_last, top, bottom, "rebond ouv.", QColor(210, 230, 255, 70))
                draw_zone(close_first, close_last, top, bottom, "rebond ferm.", QColor(255, 230, 120, 70))
                draw_marker(open_first, "R ouvre", QColor(80, 80, 80), top + 4, bottom - 4, above=False)
                draw_marker(close_first, "T ferme", QColor(40, 110, 210), top + 4, bottom - 4, above=True)
                draw_marker(close_last, "fin rebond", QColor(180, 105, 0), top + 4, bottom - 4, above=True)
            else:
                open_first = first_state(contact_t, 0, phase=phase)
                open_last = last_state(contact_t, 0, phase=phase, after_us=open_first)
                close_first = first_state(contact_r, 1, phase=phase, after_us=open_first)
                close_last = last_state(contact_r, 1, phase=phase, after_us=close_first)
                draw_zone(open_first, close_first, top, bottom, "transfert", QColor(225, 225, 225, 55))
                draw_zone(open_first, open_last, top, bottom, "rebond ouv.", QColor(210, 230, 255, 70))
                draw_zone(close_first, close_last, top, bottom, "rebond ferm.", QColor(255, 230, 120, 70))
                draw_marker(open_first, "T ouvre", QColor(80, 80, 80), top + 4, bottom - 4, above=False)
                draw_marker(close_first, "R ferme", QColor(40, 110, 210), top + 4, bottom - 4, above=True)
                draw_marker(close_last, "fin rebond", QColor(180, 105, 0), top + 4, bottom - 4, above=True)

    for cursor, color, label in ((cursor_a_us, QColor(0, 90, 200), "A"), (cursor_b_us, QColor(210, 120, 0), "B")):
        if cursor is None:
            continue
        cursor = int(cursor)
        if view_start_us <= cursor <= view_end_us:
            cx = x_from_us(cursor)
            painter.setPen(QPen(color, 2))
            painter.drawLine(QLineF(cx, plot_y, cx, plot_y + plot_h))
            painter.setFont(QFont("Arial", 8, QFont.Bold))
            painter.drawText(int(cx + 4), int(plot_y + 12), f"{label} {cursor} µs")


class ChronoOscilloCanvas(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.events = []
        self.start_bits = None
        self.capture_us = 0
        self.view_start_us = 0
        self.view_end_us = None
        self.cursor_a_us = None
        self.cursor_b_us = None
        self.title = "Oscillogramme contacts"
        self.display_mode = OSCILLO_DISPLAY_ELECTRIC
        self.current = {}
        self.on_zoom_window = None
        self._drag_mode = ""
        self._drag_start_x = None
        self._drag_start_y = None
        self._drag_current_x = None
        self._drag_current_y = None
        self._drag_start_view = (0, 1)
        self._measure_key_down = False
        self._measure_start_t = None
        self._measure_current_t = None
        self.setMinimumSize(760, 300)
        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.StrongFocus)
        self.installEventFilter(self)
        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(self)

    def set_data(self, events, start_bits, capture_us, title, view_start_us=0, view_end_us=None, cursor_a_us=None, cursor_b_us=None, display_mode=OSCILLO_DISPLAY_ELECTRIC, current=None):
        self.events = list(events or [])
        self.start_bits = start_bits
        self.capture_us = int(capture_us or 0)
        self.view_start_us = int(view_start_us or 0)
        self.view_end_us = None if view_end_us is None else int(view_end_us)
        self.cursor_a_us = cursor_a_us
        self.cursor_b_us = cursor_b_us
        self.title = str(title or "Oscillogramme contacts")
        self.display_mode = display_mode or OSCILLO_DISPLAY_ELECTRIC
        self.current = dict(current or {})
        self.update()

    def plot_geometry(self):
        rect = self.rect()
        if self.display_mode == OSCILLO_DISPLAY_SYNTHESIS:
            margin_left, margin_right, margin_top, margin_bottom = 92, 18, 46, 24
        else:
            compact = rect.height() < 360
            margin_left, margin_right = 54, 16
            margin_top = 32 if compact else 38
            margin_bottom = 24 if compact else 32
        return (
            float(rect.x() + margin_left),
            float(rect.y() + margin_top),
            float(max(10, rect.width() - margin_left - margin_right - 1)),
            float(max(10, rect.height() - margin_top - margin_bottom - 1)),
        )

    def time_from_x(self, x_pos):
        plot_x, _plot_y, plot_w, _plot_h = self.plot_geometry()
        capture_end = max(1, int(self.capture_us or 0))
        view_start = max(0, min(capture_end - 1, int(self.view_start_us or 0)))
        view_end = capture_end if self.view_end_us is None else int(self.view_end_us)
        view_end = max(view_start + 1, min(capture_end, view_end))
        pre_t0_gap = oscillo_pre_t0_gap_us(view_start, view_end)
        visual_start = view_start - pre_t0_gap
        visual_span = max(1, view_end - visual_start)
        ratio = (float(x_pos) - plot_x) / max(1.0, plot_w)
        ratio = max(0.0, min(1.0, ratio))
        value = int(round(visual_start + ratio * visual_span))
        return max(0, min(capture_end, value))

    def x_from_time(self, t_us):
        plot_x, _plot_y, plot_w, _plot_h = self.plot_geometry()
        capture_end = max(1, int(self.capture_us or 0))
        view_start = max(0, min(capture_end - 1, int(self.view_start_us or 0)))
        view_end = capture_end if self.view_end_us is None else int(self.view_end_us)
        view_end = max(view_start + 1, min(capture_end, view_end))
        pre_t0_gap = oscillo_pre_t0_gap_us(view_start, view_end)
        visual_start = view_start - pre_t0_gap
        visual_span = max(1, view_end - visual_start)
        return plot_x + ((max(visual_start, min(view_end, int(t_us))) - visual_start) / visual_span) * plot_w

    def request_window(self, start_us, end_us):
        if callable(self.on_zoom_window):
            self.on_zoom_window(int(start_us), int(end_us))

    def eventFilter(self, obj, event):
        if event.type() in (QEvent.KeyPress, QEvent.KeyRelease) and event.key() == Qt.Key_M:
            if not event.isAutoRepeat():
                self._measure_key_down = event.type() == QEvent.KeyPress
                if self._measure_key_down and self._drag_mode == "pan":
                    self._drag_mode = "measure"
                    self._measure_start_t = self.time_from_x(self._drag_start_x or 0)
                    self._measure_current_t = self.time_from_x(self._drag_current_x or self._drag_start_x or 0)
                    self.update()
                elif not self._measure_key_down and self._drag_mode == "measure":
                    self.update()
        # Un filtre installé au niveau QApplication doit simplement indiquer
        # que l'événement n'est pas consommé. Appeler super().eventFilter() ici
        # peut récursiver pendant la destruction des objets Qt.
        return False

    def wheelEvent(self, event):
        capture_end = max(1, int(self.capture_us or 0))
        view_start = max(0, min(capture_end - 1, int(self.view_start_us or 0)))
        view_end = capture_end if self.view_end_us is None else int(self.view_end_us)
        view_end = max(view_start + 1, min(capture_end, view_end))
        span = max(1, view_end - view_start)
        cursor_t = self.time_from_x(event.position().x())
        factor = 1.25
        if event.angleDelta().y() > 0:
            new_span = max(1, int(math.ceil(span / factor)))
            if new_span >= span and span > 1:
                new_span = span - 1
        else:
            new_span = min(capture_end, max(span + 1, int(math.ceil(span * factor))))
            if new_span >= capture_end or capture_end - new_span <= 1:
                self.request_window(0, capture_end)
                event.accept()
                return
        left_ratio = (cursor_t - view_start) / max(1, span)
        new_start = int(cursor_t - new_span * left_ratio)
        new_end = new_start + new_span
        if new_start < 0:
            new_end -= new_start
            new_start = 0
        if new_end > capture_end:
            new_start = max(0, new_start - (new_end - capture_end))
            new_end = capture_end
        self.request_window(new_start, new_end)
        event.accept()

    def mousePressEvent(self, event):
        self.setFocus(Qt.MouseFocusReason)
        if event.button() == Qt.LeftButton:
            self._drag_mode = "select"
            self._drag_start_x = event.position().x()
            self._drag_start_y = event.position().y()
            self._drag_current_x = self._drag_start_x
            self._drag_current_y = self._drag_start_y
        elif event.button() == Qt.RightButton:
            self._drag_start_x = event.position().x()
            self._drag_start_y = event.position().y()
            self._drag_current_x = self._drag_start_x
            self._drag_current_y = self._drag_start_y
            self._drag_start_view = (int(self.view_start_us or 0), int(self.view_end_us or self.capture_us or 1))
            if self._measure_key_down:
                self._drag_mode = "measure"
                self._measure_start_t = self.time_from_x(self._drag_start_x)
                self._measure_current_t = self._measure_start_t
            else:
                self._drag_mode = "pan"
        self.update()
        event.accept()

    def mouseMoveEvent(self, event):
        if not self._drag_mode:
            return
        self._drag_current_x = event.position().x()
        self._drag_current_y = event.position().y()
        if self._drag_mode == "pan":
            capture_end = max(1, int(self.capture_us or 0))
            start0, end0 = self._drag_start_view
            span = max(1, end0 - start0)
            t0 = self.time_from_x(self._drag_start_x or 0)
            t1 = self.time_from_x(self._drag_current_x or 0)
            delta = t0 - t1
            new_start = max(0, min(capture_end - span, start0 + delta))
            new_end = min(capture_end, new_start + span)
            self.request_window(new_start, new_end)
        elif self._drag_mode == "measure":
            self._measure_current_t = self.time_from_x(self._drag_current_x)
        self.update()
        event.accept()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and self._drag_mode == "select":
            t0 = self.time_from_x(self._drag_start_x or 0)
            t1 = self.time_from_x(self._drag_current_x or self._drag_start_x or 0)
            if abs(t1 - t0) >= 1:
                self.request_window(min(t0, t1), max(t0, t1))
        if event.button() == Qt.RightButton and self._drag_mode == "measure":
            self._measure_start_t = None
            self._measure_current_t = None
        if event.button() in (Qt.LeftButton, Qt.RightButton):
            self._drag_mode = ""
            self._drag_start_x = self._drag_current_x = None
            self._drag_start_y = self._drag_current_y = None
            self.update()
        event.accept()

    def draw_selection_overlay(self, painter):
        if self._drag_start_x is None or self._drag_current_x is None:
            return
        plot_x, plot_y, plot_w, plot_h = self.plot_geometry()
        left = max(plot_x, min(float(self._drag_start_x), float(self._drag_current_x)))
        right = min(plot_x + plot_w, max(float(self._drag_start_x), float(self._drag_current_x)))
        if right <= left:
            return
        painter.setPen(QPen(QColor(30, 100, 200), 1, Qt.DashLine))
        painter.setBrush(QColor(30, 100, 200, 35))
        painter.drawRect(int(left), int(plot_y), int(right - left), int(plot_h))

    def draw_measure_overlay(self, painter):
        if self._measure_start_t is None or self._measure_current_t is None:
            return
        plot_x, plot_y, plot_w, plot_h = self.plot_geometry()
        start_x = self.x_from_time(self._measure_start_t)
        end_x = self.x_from_time(self._measure_current_t)
        y_mid = plot_y + plot_h * 0.50
        red = QColor(210, 0, 0)
        painter.setPen(QPen(red, 2))
        painter.drawLine(QLineF(start_x, y_mid, end_x, y_mid))
        angle = 0 if end_x >= start_x else math.pi
        head_len = 10.0
        head_angle = math.radians(25)
        x1 = end_x - head_len * math.cos(angle - head_angle)
        y1 = y_mid - head_len * math.sin(angle - head_angle)
        x2 = end_x - head_len * math.cos(angle + head_angle)
        y2 = y_mid - head_len * math.sin(angle + head_angle)
        painter.drawLine(QLineF(end_x, y_mid, x1, y1))
        painter.drawLine(QLineF(end_x, y_mid, x2, y2))
        delta_us = int(self._measure_current_t) - int(self._measure_start_t)
        label = f"Δ {delta_us:+d} µs = {delta_us / 1000.0:+.3f} ms"
        mid_x = (start_x + end_x) / 2.0
        text_w = max(120, len(label) * 7)
        text_h = 20
        text_x = max(plot_x, min(plot_x + plot_w - text_w, mid_x + 8))
        text_y = max(plot_y, min(plot_y + plot_h - text_h, y_mid - 28))
        painter.setPen(QPen(red, 1))
        painter.setBrush(QColor(255, 245, 245, 230))
        painter.drawRect(int(text_x), int(text_y), int(text_w), int(text_h))
        painter.setFont(QFont("Arial", 8, QFont.Bold))
        painter.drawText(int(text_x + 6), int(text_y + 14), label)
        painter.setPen(QPen(QColor(210, 0, 0, 120), 1, Qt.DashLine))
        painter.drawLine(QLineF(start_x, plot_y, start_x, plot_y + plot_h))
        painter.drawLine(QLineF(end_x, plot_y, end_x, plot_y + plot_h))

    def paintEvent(self, _event):
        painter = QPainter(self)
        try:
            rect = self.rect()
            if self.display_mode == OSCILLO_DISPLAY_SYNTHESIS:
                draw_synthesis_oscillogram(
                    painter, rect.x(), rect.y(), rect.width() - 1, rect.height() - 1,
                    self.events, self.start_bits, self.capture_us, self.current, self.title,
                    self.view_start_us, self.view_end_us, self.cursor_a_us, self.cursor_b_us,
                )
            else:
                display_mode = self.display_mode if self.display_mode != OSCILLO_DISPLAY_SYNTHESIS else OSCILLO_DISPLAY_LOGIC
                draw_logic_oscillogram(
                    painter, rect.x(), rect.y(), rect.width() - 1, rect.height() - 1,
                    self.events, self.start_bits, self.capture_us, self.title,
                    self.view_start_us, self.view_end_us, self.cursor_a_us, self.cursor_b_us,
                    display_mode, self.current.get("phase_markers") if isinstance(self.current, dict) else None, self.current,
                )
            if self._drag_mode == "select":
                self.draw_selection_overlay(painter)
            if self._drag_mode == "measure" or self._measure_start_t is not None:
                self.draw_measure_overlay(painter)
        finally:
            painter.end()


class IhmRelaisRp2040:
    def __init__(self):
        self.ser = None
        self.reader = None

        # Les chemins persistants doivent exister avant la création des onglets.
        # L'onglet collage/décollage initialise sa table SQLite dès sa création.
        self.scenario_file = self.resolve_scenario_file()
        self.production_db_file = self.resolve_production_db_file()
        self.chrono_db_file = self.resolve_chrono_db_file()
        self.production_legacy_json_file = self.resolve_legacy_production_json_file()

        self.window = self.load_ui()

        self.tabWidget_principal = self.get_widget(QTabWidget, "tabWidget_principal")
        self.tab_production_accueil = self.get_widget(QWidget, "tab_production_accueil")
        self.comboBox_prod_scenario = self.get_widget(QComboBox, "comboBox_prod_scenario")
        self.lineEdit_prod_lot = self.get_widget(QLineEdit, "lineEdit_prod_lot")
        self.lineEdit_prod_designation = self.get_widget(QLineEdit, "lineEdit_prod_designation")
        self.spinBox_prod_nb_inverseurs = self.get_widget(QSpinBox, "spinBox_prod_nb_inverseurs")
        self.comboBox_prod_operateur = self.get_widget(QComboBox, "comboBox_prod_operateur")
        self.dateEdit_prod_date = self.get_widget(QDateEdit, "dateEdit_prod_date")
        self.pushButton_prod_save_context = self.get_widget(QPushButton, "pushButton_prod_save_context")
        self.pushButton_prod_reload_base = self.get_widget(QPushButton, "pushButton_prod_reload_base")
        self.pushButton_prod_export_pdf_lot = self.get_widget(QPushButton, "pushButton_prod_export_pdf_lot")
        self.lineEdit_prod_search_lot = self.get_widget(QLineEdit, "lineEdit_prod_search_lot")
        self.pushButton_prod_search_clear = self.get_widget(QPushButton, "pushButton_prod_search_clear")
        self.tableWidget_prod_records = self.get_widget(QTableWidget, "tableWidget_prod_records")
        self.label_prod_status = self.get_widget(QLabel, "label_prod_status")
        self.tab_database_admin = self.get_widget(QWidget, "tab_database_admin")
        self.comboBox_db_target = self.get_widget(QComboBox, "comboBox_db_target")
        self.label_db_admin_status = self.get_widget(QLabel, "label_db_admin_status")
        self.label_db_admin_file = self.get_widget(QLabel, "label_db_admin_file")
        self.pushButton_db_refresh = self.get_widget(QPushButton, "pushButton_db_refresh")
        self.pushButton_db_backup = self.get_widget(QPushButton, "pushButton_db_backup")
        self.pushButton_db_restore = self.get_widget(QPushButton, "pushButton_db_restore")
        self.pushButton_db_export_csv = self.get_widget(QPushButton, "pushButton_db_export_csv")
        self.pushButton_db_export_xlsx = self.get_widget(QPushButton, "pushButton_db_export_xlsx")
        self.pushButton_db_export_pdf = self.get_widget(QPushButton, "pushButton_db_export_pdf")
        self.pushButton_db_vacuum = self.get_widget(QPushButton, "pushButton_db_vacuum")
        self.pushButton_db_recreate_default = self.get_widget(QPushButton, "pushButton_db_recreate_default")
        self.pushButton_db_merge = self.get_widget(QPushButton, "pushButton_db_merge")
        self.lineEdit_db_operator = self.get_widget(QLineEdit, "lineEdit_db_operator")
        self.pushButton_db_operator_add = self.get_widget(QPushButton, "pushButton_db_operator_add")
        self.pushButton_db_operator_delete = self.get_widget(QPushButton, "pushButton_db_operator_delete")
        self.tableWidget_db_operators = self.get_widget(QTableWidget, "tableWidget_db_operators")
        self.lineEdit_db_lot_filter = self.get_widget(QLineEdit, "lineEdit_db_lot_filter")
        self.pushButton_db_lot_open = self.get_widget(QPushButton, "pushButton_db_lot_open")
        self.pushButton_db_lot_pdf = self.get_widget(QPushButton, "pushButton_db_lot_pdf")
        self.pushButton_db_lot_xlsx = self.get_widget(QPushButton, "pushButton_db_lot_xlsx")
        self.pushButton_db_lot_delete = self.get_widget(QPushButton, "pushButton_db_lot_delete")
        self.tableWidget_db_lots = self.get_widget(QTableWidget, "tableWidget_db_lots")
        self.groupBox_db_operators = self.window.findChild(QGroupBox, "groupBox_db_operators")
        self.groupBox_db_lots = self.window.findChild(QGroupBox, "groupBox_db_lots")

        self.comboBox_ports = self.get_widget(QComboBox, "comboBox_ports")
        self.comboBox_baudrate = self.get_widget(QComboBox, "comboBox_baudrate")
        self.comboBox_ports_production = self.get_widget(QComboBox, "comboBox_ports_2")
        self.comboBox_baudrate_production = self.get_widget(QComboBox, "comboBox_baudrate_2")

        self.pushButton_rafraichir_ports = self.get_widget(QPushButton, "pushButton_rafraichir_ports")
        self.pushButton_connecter = self.get_widget(QPushButton, "pushButton_connecter")
        self.pushButton_deconnecter = self.get_widget(QPushButton, "pushButton_deconnecter")
        self.pushButton_rafraichir_ports_production = self.get_widget(QPushButton, "pushButton_rafraichir_ports_2")
        self.pushButton_connecter_production = self.get_widget(QPushButton, "pushButton_connecter_2")
        self.pushButton_deconnecter_production = self.get_widget(QPushButton, "pushButton_deconnecter_2")

        self.pushButton_mode_us = self.get_widget(QPushButton, "pushButton_mode_us")
        self.pushButton_mode_ms = self.get_widget(QPushButton, "pushButton_mode_ms")
        self.pushButton_mode_s = self.get_widget(QPushButton, "pushButton_mode_s")
        self.pushButton_mode_min = self.get_widget(QPushButton, "pushButton_mode_min")
        self.pushButton_mode_h = self.get_widget(QPushButton, "pushButton_mode_h")
        self.label_mode_temps_actif = self.get_widget(QLabel, "label_mode_temps_actif")

        self.radioButton_monostable = self.get_widget(QWidget, "radioButton_monostable")
        self.radioButton_bistable = self.get_widget(QWidget, "radioButton_bistable")

        self.lineEdit_temps_on = self.get_widget(QLineEdit, "lineEdit_temps_on")
        self.lineEdit_temps_off = self.get_widget(QLineEdit, "lineEdit_temps_off")
        self.lineEdit_impulsion_set = self.get_widget(QLineEdit, "lineEdit_impulsion_set")
        self.lineEdit_impulsion_reset = self.get_widget(QLineEdit, "lineEdit_impulsion_reset")
        self.lineEdit_nombre_cycles = self.get_widget(QLineEdit, "lineEdit_nombre_cycles")

        self.comboBox_unite_on = self.get_widget(QComboBox, "comboBox_unite_on")
        self.comboBox_unite_off = self.get_widget(QComboBox, "comboBox_unite_off")
        self.comboBox_unite_set = self.get_widget(QComboBox, "comboBox_unite_set")
        self.comboBox_unite_reset = self.get_widget(QComboBox, "comboBox_unite_reset")

        self.pushButton_demarrer = self.get_widget(QPushButton, "pushButton_demarrer")
        self.pushButton_pause = self.get_widget(QPushButton, "pushButton_pause")
        self.pushButton_reprendre = self.get_widget(QPushButton, "pushButton_reprendre")
        self.pushButton_arret = self.get_widget(QPushButton, "pushButton_arret")
        self.pushButton_status = self.get_widget(QPushButton, "pushButton_status")

        self.label_etat_connexion = self.get_widget(QLabel, "label_etat_connexion")
        self.label_etat_connexion_production = self.get_widget(QLabel, "label_etat_connexion_2")
        self.label_etat_sortie1 = self.get_widget(QLabel, "label_etat_sortie1")
        self.label_etat_sortie2 = self.get_widget(QLabel, "label_etat_sortie2")
        self.label_etat_essai = self.get_widget(QLabel, "label_etat_essai")
        self.label_cycle_actuel = self.get_widget(QLabel, "label_cycle_actuel")
        self.label_derniere_commande_us = self.get_widget(QLabel, "label_derniere_commande_us")

        self.textEdit_log = self.get_widget(QTextEdit, "textEdit_log")

        self.lineEdit_neutral_be = self.get_widget(QLineEdit, "lineEdit_neutral_be")
        self.lineEdit_neutral_br = self.get_widget(QLineEdit, "lineEdit_neutral_br")
        self.lineEdit_neutral_bebr = self.get_widget(QLineEdit, "lineEdit_neutral_bebr")

        self.comboBox_unite_neutral_be = self.get_widget(QComboBox, "comboBox_unite_neutral_be")
        self.comboBox_unite_neutral_br = self.get_widget(QComboBox, "comboBox_unite_neutral_br")
        self.comboBox_unite_neutral_bebr = self.get_widget(QComboBox, "comboBox_unite_neutral_bebr")

        self.pushButton_neutral_be = self.get_widget(QPushButton, "pushButton_neutral_be")
        self.pushButton_neutral_br = self.get_widget(QPushButton, "pushButton_neutral_br")
        self.pushButton_neutral_bebr = self.get_widget(QPushButton, "pushButton_neutral_bebr")
        self.pushButton_neutral_stop = self.get_widget(QPushButton, "pushButton_neutral_stop")

        self.label_neutral_derniere_commande = self.get_widget(QLabel, "label_neutral_derniere_commande")
        self.label_neutral_sortie1 = self.get_widget(QLabel, "label_neutral_sortie1")
        self.label_neutral_sortie2 = self.get_widget(QLabel, "label_neutral_sortie2")
        self.label_neutral_tension_selection = self.window.findChild(QLabel, "label_neutral_tension_selection")
        self.lineEdit_tension_basse_info = self.window.findChild(QLineEdit, "lineEdit_tension_basse_info")
        self.lineEdit_tension_haute_info = self.window.findChild(QLineEdit, "lineEdit_tension_haute_info")

        self.label_led_reset_contact_1 = self.get_widget(QLabel, "label_led_reset_contact_1")
        self.label_led_reset_contact_2 = self.get_widget(QLabel, "label_led_reset_contact_2")
        self.label_led_reset_contact_5 = self.get_widget(QLabel, "label_led_reset_contact_5")
        self.label_led_reset_contact_6 = self.get_widget(QLabel, "label_led_reset_contact_6")
        self.label_led_latch_contact_1 = self.get_widget(QLabel, "label_led_latch_contact_1")
        self.label_led_latch_contact_2 = self.get_widget(QLabel, "label_led_latch_contact_2")
        self.label_led_latch_contact_5 = self.get_widget(QLabel, "label_led_latch_contact_5")
        self.label_led_latch_contact_6 = self.get_widget(QLabel, "label_led_latch_contact_6")

        self.label_led_reset_contact_3 = self.get_widget(QLabel, "label_led_reset_contact_3")
        self.label_led_reset_contact_4 = self.get_widget(QLabel, "label_led_reset_contact_4")
        self.label_led_reset_contact_7 = self.get_widget(QLabel, "label_led_reset_contact_7")
        self.label_led_reset_contact_8 = self.get_widget(QLabel, "label_led_reset_contact_8")
        self.label_led_latch_contact_3 = self.get_widget(QLabel, "label_led_latch_contact_3")
        self.label_led_latch_contact_4 = self.get_widget(QLabel, "label_led_latch_contact_4")
        self.label_led_latch_contact_7 = self.get_widget(QLabel, "label_led_latch_contact_7")
        self.label_led_latch_contact_8 = self.get_widget(QLabel, "label_led_latch_contact_8")
        self.label_neutral_contact_summary = self.get_widget(QLabel, "label_neutral_contact_summary")

        # Onglet Neutral Screen Automatique V2.12.3
        self.pushButton_auto_neutral_marche = self.get_widget(QPushButton, "pushButton_auto_neutral_marche")
        self.pushButton_auto_neutral_arret = self.get_widget(QPushButton, "pushButton_auto_neutral_arret")
        self.lineEdit_auto_delai_ms = self.get_widget(QLineEdit, "lineEdit_auto_delai_ms")
        self.lineEdit_auto_nb_inverseurs = self.get_widget(QLineEdit, "lineEdit_auto_nb_inverseurs")
        self.checkBox_auto_pulses_particuliers = self.get_widget(QCheckBox, "checkBox_auto_pulses_particuliers")
        self.lineEdit_auto_pulse_bebr_ms = self.get_widget(QLineEdit, "lineEdit_auto_pulse_bebr_ms")
        self.lineEdit_auto_pulse_be_ms = self.get_widget(QLineEdit, "lineEdit_auto_pulse_be_ms")
        self.lineEdit_auto_pulse_br_ms = self.get_widget(QLineEdit, "lineEdit_auto_pulse_br_ms")
        self.comboBox_auto_scenario = self.get_widget(QComboBox, "comboBox_auto_scenario")
        self.pushButton_auto_scenario_recharger = self.get_widget(QPushButton, "pushButton_auto_scenario_recharger")
        self.pushButton_auto_scenario_editer = self.get_widget(QPushButton, "pushButton_auto_scenario_editer")
        self.lineEdit_SN = self.get_widget(QLineEdit, "lineEdit_SN")
        self.label_auto_lot = self.window.findChild(QLabel, "label_LOT")
        self.label_auto_design_relais = self.window.findChild(QLabel, "label_design_relais")
        self.pushButton_auto_lot_fini = self.get_widget(QPushButton, "pushButton_auto_lot_fini")
        self.tableWidget_auto_logigramme = self.get_widget(QTableWidget, "tableWidget_auto_logigramme")
        self.label_auto_tension_basse = self.get_widget(QLabel, "label_auto_tension_basse")
        self.label_auto_tension_haute = self.get_widget(QLabel, "label_auto_tension_haute")
        self.label_auto_status = self.get_widget(QLabel, "label_auto_status")
        self.label_auto_resultat = self.get_widget(QLabel, "label_auto_resultat")

        self.label_auto_led_r1 = self.get_widget(QLabel, "label_auto_led_r1")
        self.label_auto_led_r2 = self.get_widget(QLabel, "label_auto_led_r2")
        self.label_auto_led_r3 = self.get_widget(QLabel, "label_auto_led_r3")
        self.label_auto_led_r4 = self.get_widget(QLabel, "label_auto_led_r4")
        self.label_auto_led_t1 = self.get_widget(QLabel, "label_auto_led_t1")
        self.label_auto_led_t2 = self.get_widget(QLabel, "label_auto_led_t2")
        self.label_auto_led_t3 = self.get_widget(QLabel, "label_auto_led_t3")
        self.label_auto_led_t4 = self.get_widget(QLabel, "label_auto_led_t4")

        # Onglet Éditeur Scénarios Neutral Screen V2.12.3
        self.comboBox_editor_scenarios = self.get_widget(QComboBox, "comboBox_editor_scenarios")
        self.lineEdit_editor_nom = self.get_widget(QLineEdit, "lineEdit_editor_nom")
        self.textEdit_editor_description = self.get_widget(QTextEdit, "textEdit_editor_description")
        self.tableWidget_editor_steps = self.get_widget(QTableWidget, "tableWidget_editor_steps")
        self.label_editor_fichier = self.get_widget(QLabel, "label_editor_fichier")
        self.label_editor_status = self.get_widget(QLabel, "label_editor_status")
        self.pushButton_editor_nouveau = self.get_widget(QPushButton, "pushButton_editor_nouveau")
        self.pushButton_editor_dupliquer = self.get_widget(QPushButton, "pushButton_editor_dupliquer")
        self.pushButton_editor_supprimer = self.get_widget(QPushButton, "pushButton_editor_supprimer")
        self.pushButton_editor_sauvegarder = self.get_widget(QPushButton, "pushButton_editor_sauvegarder")
        self.pushButton_editor_recharger = self.get_widget(QPushButton, "pushButton_editor_recharger")
        self.pushButton_editor_ajouter_etape = self.get_widget(QPushButton, "pushButton_editor_ajouter_etape")
        self.pushButton_editor_supprimer_etape = self.get_widget(QPushButton, "pushButton_editor_supprimer_etape")
        self.pushButton_editor_monter_etape = self.get_widget(QPushButton, "pushButton_editor_monter_etape")
        self.pushButton_editor_descendre_etape = self.get_widget(QPushButton, "pushButton_editor_descendre_etape")
        self.pushButton_editor_importer = self.get_widget(QPushButton, "pushButton_editor_importer")
        self.pushButton_editor_exporter = self.get_widget(QPushButton, "pushButton_editor_exporter")
        self.pushButton_editor_mot_de_passe = self.get_widget(QPushButton, "pushButton_editor_mot_de_passe")
        self.create_chronometrie_contacts_tab()
        self.create_voltage_operation_tab()
        self.create_voltage_calibration_tab()

        self.auto_step_names = [
            "label_auto_step_start",
            "label_auto_step_bebr1",
            "label_auto_step_check_neutral1",
            "label_auto_step_bebr2",
            "label_auto_step_check_neutral2",
            "label_auto_step_bebr3",
            "label_auto_step_check_neutral3",
            "label_auto_step_accept_no_neutral",
            "label_auto_step_be",
            "label_auto_step_check_latch",
            "label_auto_step_bebr_after_latch",
            "label_auto_step_br",
            "label_auto_step_check_reset",
            "label_auto_step_accept_ok",
            "label_auto_step_reject",
        ]
        # Ancien logigramme fixe conservé uniquement pour compatibilité historique.
        # En V2.12.3 le vrai affichage est tableWidget_auto_logigramme.
        self.auto_steps = {}
        for name in self.auto_step_names:
            w = self.window.findChild(QLabel, name)
            if w is not None:
                self.auto_steps[name] = w

        # État sorties courant (source de vérité de l'affichage).
        self.current_out1 = "0"
        self.current_out2 = "0"
        self.contacts_last_values = (None, None, None, None, None, None, None, None)
        self.contacts_known_values = [None, None, None, None, None, None, None, None]
        self.contacts_force_refresh = False

        # Automate Neutral Screen Automatique V2.12.3
        self.auto_neutral_running = False
        self.auto_neutral_attempt = 0
        self.auto_next_action = None
        self.auto_neutral_timer = QTimer()
        self.auto_neutral_timer.setSingleShot(True)
        self.auto_neutral_timer.timeout.connect(self.auto_timer_timeout)
        self.AUTO_PULSE_US = 10_000  # secours, les pulses auto sont maintenant réglables dans l’IHM

        # Flash pulse : non bloquant. On garde l'état réel à jour en arrière-plan
        # et on ne masque QUE le texte des labels pendant la fenêtre de flash.
        self.flash_timer = QTimer()
        self.flash_timer.setSingleShot(True)
        self.flash_timer.timeout.connect(self.end_flash)
        self.flash_active = False

        # Scénarios Neutral Screen V2.12.3
        self.scenarios_data = {"version": "2.12.3", "scenarios": []}
        self.current_runtime_steps = []
        self.current_runtime_scenario_name = ""
        self.runtime_step_index = 0
        self.runtime_attempt = 0
        self._refreshing_scenario_combos = False
        self.production_data = {"version": "2.12.3", "last_context": {}, "records": [], "access_code": LOCK_ACCESS_CODE}
        self.current_access_code = LOCK_ACCESS_CODE
        self._production_autofill_running = False
        self._auto_start_prompts_done = False
        self.auto_end_validation_pending = False
        self.last_finished_sn = ""
        self.last_finished_result = ""
        self.interrupted_auto_sn = ""
        self.interrupted_auto_scenario = ""
        self.interrupted_auto_reason = ""
        self._auto_connect_done = False
        self._auto_connect_in_progress = False
        self._connection_alert_phase = 0
        self._auto_reconnect_counter = 0
        self._manual_disconnect = False
        self._last_connection_error = ""
        self._last_connected_port = ""
        self._tab_guard_internal = False
        self._last_allowed_tab_index = 0
        self._protected_access_granted = False
        self._lot_session_active = False
        self._active_lot = ""
        self._active_lot_finished = True
        self.protected_tab_names = {"tab_database_admin", "tab_neutral_screen", "tab_neutral_scenario_editor", "tab_cyclage", "tab_chronometrie_contacts", "tab_chrono_oscillo", "tab_tension_operation", "tab_voltage_calibration"}
        self.button_help_filter = DelayedButtonHelp(self, delay_ms=3000, duration_ms=12000)

        self.initialiser_widgets_production()
        self.initialiser_widgets_database_admin()
        self.initialiser_verrouillage_onglets()
        self.initialiser_validation_fin_essai()
        self.initialiser_aide_boutons_operateur()

        self.initialiser_unites()
        self.connect_signals()
        self.production_load_db()
        self.chrono_init_db()
        self.production_refresh_table()
        self.database_admin_refresh()
        self.refresh_ports()
        self.update_button_states()
        self.update_mode_fields()
        self.initialiser_leds_contacts()
        self.initialiser_auto_neutral()
        self.scenarios_load_or_create()
        self.scenarios_refresh_all()
        self.tabWidget_principal.setCurrentWidget(self.tab_production_accueil)

        self.port_timer = QTimer()
        self.port_timer.timeout.connect(self.refresh_ports_keep_selection)
        self.port_timer.start(2000)

        self.connection_alert_timer = QTimer()
        self.connection_alert_timer.timeout.connect(self.update_connection_status_visual)
        self.connection_alert_timer.start(700)
        self.update_connection_status_visual()
        QTimer.singleShot(600, self.auto_connect_rp2040)

    def create_chronometrie_contacts_tab(self):
        self.tab_chronometrie_contacts = self.get_widget(QWidget, "tab_chronometrie_contacts")
        self.lineEdit_chrono_lot = self.get_widget(QLineEdit, "lineEdit_chrono_lot")
        self.lineEdit_chrono_date = self.get_widget(QLineEdit, "lineEdit_chrono_date")
        self.lineEdit_chrono_relais = self.get_widget(QLineEdit, "lineEdit_chrono_relais")
        self.lineEdit_chrono_ambiance = self.get_widget(QLineEdit, "lineEdit_chrono_ambiance")
        self.lineEdit_chrono_nom_test = self.get_widget(QLineEdit, "lineEdit_chrono_nom_test")
        self.lineEdit_chrono_sn = self.get_widget(QLineEdit, "lineEdit_chrono_sn")
        self.lineEdit_chrono_resultat = self.get_widget(QLineEdit, "lineEdit_chrono_resultat")
        self.comboBox_chrono_type_relais = self.get_widget(QComboBox, "comboBox_chrono_type_relais")
        self.spinBox_chrono_nb_inverseurs = self.get_widget(QSpinBox, "spinBox_chrono_nb_inverseurs")
        self.spinBox_chrono_capture_ms = self.get_widget(QSpinBox, "spinBox_chrono_capture_ms")
        self.spinBox_chrono_pulse_ms = self.get_widget(QSpinBox, "spinBox_chrono_pulse_ms")
        self.lineEdit_chrono_limite_temps_ms = self.get_widget(QLineEdit, "lineEdit_chrono_limite_temps_ms")
        self.lineEdit_chrono_limite_rebond_ms = self.get_widget(QLineEdit, "lineEdit_chrono_limite_rebond_ms")
        self.pushButton_chrono_mesure_be = self.get_widget(QPushButton, "pushButton_chrono_mesure_be")
        self.pushButton_chrono_mesure_br = self.get_widget(QPushButton, "pushButton_chrono_mesure_br")
        self.pushButton_chrono_mesure_be_br = self.get_widget(QPushButton, "pushButton_chrono_mesure_be_br")
        self.pushButton_chrono_export_xlsx_lot = self.get_widget(QPushButton, "pushButton_chrono_export_xlsx_lot")
        self.pushButton_chrono_export_pdf_lot = self.get_widget(QPushButton, "pushButton_chrono_export_pdf_lot")
        self.pushButton_chrono_effacer = self.get_widget(QPushButton, "pushButton_chrono_effacer")
        self.label_chrono_led_r1 = self.get_widget(QLabel, "label_chrono_led_r1")
        self.label_chrono_led_r2 = self.get_widget(QLabel, "label_chrono_led_r2")
        self.label_chrono_led_r3 = self.get_widget(QLabel, "label_chrono_led_r3")
        self.label_chrono_led_r4 = self.get_widget(QLabel, "label_chrono_led_r4")
        self.label_chrono_led_t1 = self.get_widget(QLabel, "label_chrono_led_t1")
        self.label_chrono_led_t2 = self.get_widget(QLabel, "label_chrono_led_t2")
        self.label_chrono_led_t3 = self.get_widget(QLabel, "label_chrono_led_t3")
        self.label_chrono_led_t4 = self.get_widget(QLabel, "label_chrono_led_t4")
        self.label_chrono_contact_summary = self.get_widget(QLabel, "label_chrono_contact_summary")
        self.label_chrono_status = self.get_widget(QLabel, "label_chrono_status")
        self.tableWidget_chrono_results = self.get_widget(QTableWidget, "tableWidget_chrono_results")
        self.tableWidget_chrono_events = self.get_widget(QTableWidget, "tableWidget_chrono_events")
        self.tab_chrono_oscillo = self.get_widget(QWidget, "tab_chrono_oscillo")
        self.widget_oscillo_canvas = self.get_widget(QWidget, "widget_oscillo_canvas")
        self.pushButton_oscillo_export_xlsx = self.get_widget(QPushButton, "pushButton_oscillo_export_xlsx")
        self.pushButton_oscillo_export_pdf = self.get_widget(QPushButton, "pushButton_oscillo_export_pdf")
        self.pushButton_oscillo_vue_complete = self.get_widget(QPushButton, "pushButton_oscillo_vue_complete")
        self.pushButton_oscillo_zoom_fronts = self.get_widget(QPushButton, "pushButton_oscillo_zoom_fronts")
        self.comboBox_oscillo_capture = self.get_widget(QComboBox, "comboBox_oscillo_capture")
        self.comboBox_oscillo_contact = self.get_widget(QComboBox, "comboBox_oscillo_contact")
        self.comboBox_oscillo_display_mode = self.get_widget(QComboBox, "comboBox_oscillo_display_mode")
        self.pushButton_oscillo_load_saved = self.get_widget(QPushButton, "pushButton_oscillo_load_saved")
        self.pushButton_oscillo_zoom_contact = self.get_widget(QPushButton, "pushButton_oscillo_zoom_contact")
        self.pushButton_oscillo_zoom_rebonds = self.get_widget(QPushButton, "pushButton_oscillo_zoom_rebonds")
        self.pushButton_oscillo_zoom_in = self.get_widget(QPushButton, "pushButton_oscillo_zoom_in")
        self.pushButton_oscillo_zoom_out = self.get_widget(QPushButton, "pushButton_oscillo_zoom_out")
        self.spinBox_oscillo_zoom_factor = self.get_widget(QSpinBox, "spinBox_oscillo_zoom_factor")
        self.label_oscillo_status = self.get_widget(QLabel, "label_oscillo_status")
        self.label_oscillo_delta = self.get_widget(QLabel, "label_oscillo_delta")
        self.spinBox_oscillo_zoom_start_us = self.get_widget(QSpinBox, "spinBox_oscillo_zoom_start_us")
        self.spinBox_oscillo_zoom_end_us = self.get_widget(QSpinBox, "spinBox_oscillo_zoom_end_us")
        self.spinBox_oscillo_cursor_a_us = self.get_widget(QSpinBox, "spinBox_oscillo_cursor_a_us")
        self.spinBox_oscillo_cursor_b_us = self.get_widget(QSpinBox, "spinBox_oscillo_cursor_b_us")
        self.tableWidget_oscillo_points = self.get_widget(QTableWidget, "tableWidget_oscillo_points")
        self.oscillo_canvas = ChronoOscilloCanvas(self.widget_oscillo_canvas)
        self.oscillo_canvas.on_zoom_window = lambda start, end: self.oscillo_set_time_window(start, end, update=True)
        oscillo_layout = QVBoxLayout(self.widget_oscillo_canvas)
        oscillo_layout.setContentsMargins(0, 0, 0, 0)
        oscillo_layout.addWidget(self.oscillo_canvas)

        if not self.lineEdit_chrono_date.text().strip():
            self.lineEdit_chrono_date.setText(QDate.currentDate().toString("dd/MM/yyyy"))
        if not self.lineEdit_chrono_ambiance.text().strip():
            self.lineEdit_chrono_ambiance.setText("20")
        if not self.lineEdit_chrono_nom_test.text().strip():
            self.lineEdit_chrono_nom_test.setText("init")
        self.lineEdit_chrono_resultat.setReadOnly(True)
        if self.comboBox_chrono_type_relais.count() == 0:
            self.comboBox_chrono_type_relais.addItems(["Bistable", "Monostable"])
        self.spinBox_chrono_nb_inverseurs.setRange(1, 4)
        if self.spinBox_chrono_nb_inverseurs.value() < 1:
            self.spinBox_chrono_nb_inverseurs.setValue(2)
        self.spinBox_chrono_capture_ms.setRange(1, 100)
        if self.spinBox_chrono_capture_ms.value() < 1:
            self.spinBox_chrono_capture_ms.setValue(50)
        self.spinBox_chrono_pulse_ms.setRange(1, 100)
        if self.spinBox_chrono_pulse_ms.value() < 1:
            self.spinBox_chrono_pulse_ms.setValue(10)

        self.tableWidget_chrono_results.setColumnCount(6)
        self.tableWidget_chrono_results.setHorizontalHeaderLabels([
            "Inverseur", "Mesure", "Début", "Fin", "Temps (ms)", "Sanction"
        ])
        self.tableWidget_chrono_results.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_chrono_results.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_chrono_results.verticalHeader().setVisible(False)

        self.tableWidget_chrono_events.setColumnCount(4)
        self.tableWidget_chrono_events.setHorizontalHeaderLabels(["#", "Temps (µs)", "Contact", "État"])
        self.tableWidget_chrono_events.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_chrono_events.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_chrono_events.verticalHeader().setVisible(False)
        self.tableWidget_oscillo_points.setColumnCount(7)
        self.tableWidget_oscillo_points.setHorizontalHeaderLabels(["#", "Temps (µs)", "Phase", "Contact", "État logique", "État électrique", "État contact"])
        self.tableWidget_oscillo_points.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_oscillo_points.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_oscillo_points.verticalHeader().setVisible(False)
        self.comboBox_oscillo_capture.clear()
        self.comboBox_oscillo_capture.addItem("Dernière mesure", "LAST")
        self.comboBox_oscillo_contact.clear()
        self.comboBox_oscillo_contact.addItems(["AUTO"] + CHRONO_CONTACT_NAMES)
        self.comboBox_oscillo_display_mode.clear()
        self.comboBox_oscillo_display_mode.addItem("Électrique GPIO", OSCILLO_DISPLAY_ELECTRIC)
        self.comboBox_oscillo_display_mode.addItem("Logique contact", OSCILLO_DISPLAY_LOGIC)
        self.comboBox_oscillo_display_mode.addItem("Synthèse transfert / rebonds", OSCILLO_DISPLAY_SYNTHESIS)
        self.comboBox_oscillo_display_mode.setCurrentIndex(0)
        self.spinBox_oscillo_zoom_factor.setRange(2, 100)
        if self.spinBox_oscillo_zoom_factor.value() < 2:
            self.spinBox_oscillo_zoom_factor.setValue(5)
        self.pushButton_oscillo_zoom_in.setText("ZOOM +")
        self.pushButton_oscillo_zoom_out.setText("DÉZOOM")
        self.oscillo_update_zoom_button_labels()
        for spin in (
            self.spinBox_oscillo_zoom_start_us,
            self.spinBox_oscillo_zoom_end_us,
            self.spinBox_oscillo_cursor_a_us,
            self.spinBox_oscillo_cursor_b_us,
        ):
            spin.setRange(0, 10_000_000)

        self.chrono_events = []
        self.chrono_current = {}
        self.oscillo_captures = {}
        self.oscillo_selected_capture_key = "LAST"
        self.chrono_measure_running = False
        self.chrono_auto_sequence_active = False
        self.chrono_auto_sequence_queue = []
        self.chrono_auto_prereset_pending = False
        self.chrono_result_display_key = None
        self.chrono_result_rows_by_action = {}
        self.chrono_update_relay_type_ui()


    def create_voltage_operation_tab(self):
        self.tab_tension_operation = self.get_widget(QWidget, "tab_tension_operation")
        self.lineEdit_voltage_lot = self.get_widget(QLineEdit, "lineEdit_voltage_lot")
        self.lineEdit_voltage_relais = self.get_widget(QLineEdit, "lineEdit_voltage_relais")
        self.lineEdit_voltage_sn = self.get_widget(QLineEdit, "lineEdit_voltage_sn")
        self.lineEdit_voltage_date = self.get_widget(QLineEdit, "lineEdit_voltage_date")
        self.lineEdit_voltage_ambiance = self.get_widget(QLineEdit, "lineEdit_voltage_ambiance")
        self.lineEdit_voltage_test = self.get_widget(QLineEdit, "lineEdit_voltage_test")
        self.pushButton_voltage_copy_chrono = self.get_widget(QPushButton, "pushButton_voltage_copy_chrono")
        self.comboBox_voltage_ea_port = self.get_widget(QComboBox, "comboBox_voltage_ea_port")
        self.comboBox_voltage_ea_baudrate = self.get_widget(QComboBox, "comboBox_voltage_ea_baudrate")
        self.pushButton_voltage_ea_refresh = self.get_widget(QPushButton, "pushButton_voltage_ea_refresh")
        self.pushButton_voltage_ea_connect = self.get_widget(QPushButton, "pushButton_voltage_ea_connect")
        self.pushButton_voltage_ea_disconnect = self.get_widget(QPushButton, "pushButton_voltage_ea_disconnect")
        self.label_voltage_ea_status = self.get_widget(QLabel, "label_voltage_ea_status")
        self.doubleSpinBox_voltage_vmax = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_vmax")
        self.doubleSpinBox_voltage_ramp_up_s = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_ramp_up_s")
        self.doubleSpinBox_voltage_ramp_down_s = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_ramp_down_s")
        self.doubleSpinBox_voltage_current_limit = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_current_limit")
        self.doubleSpinBox_voltage_chrono_v = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_chrono_v")
        self.doubleSpinBox_voltage_interphase_s = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_interphase_s")
        for spin in (
            self.doubleSpinBox_voltage_vmax,
            self.doubleSpinBox_voltage_ramp_up_s,
            self.doubleSpinBox_voltage_ramp_down_s,
            self.doubleSpinBox_voltage_current_limit,
            self.doubleSpinBox_voltage_chrono_v,
            self.doubleSpinBox_voltage_interphase_s,
        ):
            # La saisie clavier n'est validée qu'à la fin de l'édition. Cela évite
            # qu'une valeur partiellement saisie soit corrigée vers l'ancienne valeur.
            spin.setKeyboardTracking(False)
        self.label_voltage_interphase = self.get_widget(QLabel, "label_voltage_interphase")
        self.label_voltage_interphase_note = self.get_widget(QLabel, "label_voltage_interphase_note")
        self.label_voltage_ramp_up = self.get_widget(QLabel, "label_voltage_ramp_up")
        self.label_voltage_ramp_down = self.get_widget(QLabel, "label_voltage_ramp_down")
        self.spinBox_voltage_nb_inverseurs = self.get_widget(QSpinBox, "spinBox_voltage_nb_inverseurs")
        self.spinBox_voltage_stable_ms = self.get_widget(QSpinBox, "spinBox_voltage_stable_ms")
        self.doubleSpinBox_voltage_divider_ratio = self.get_widget(QDoubleSpinBox, "doubleSpinBox_voltage_divider_ratio")
        self.spinBox_voltage_offset_mv = self.get_widget(QSpinBox, "spinBox_voltage_offset_mv")
        self.comboBox_voltage_relay_type = self.get_widget(QComboBox, "comboBox_voltage_relay_type")
        self.pushButton_voltage_open_calibration = self.get_widget(QPushButton, "pushButton_voltage_open_calibration")
        self.label_voltage_calibration_summary = self.get_widget(QLabel, "label_voltage_calibration_summary")
        self.pushButton_voltage_pickup = self.get_widget(QPushButton, "pushButton_voltage_pickup")
        self.pushButton_voltage_dropout = self.get_widget(QPushButton, "pushButton_voltage_dropout")
        self.pushButton_voltage_cycle = self.get_widget(QPushButton, "pushButton_voltage_cycle")
        self.pushButton_voltage_measure_all = self.get_widget(QPushButton, "pushButton_voltage_measure_all")
        self.pushButton_voltage_stop = self.get_widget(QPushButton, "pushButton_voltage_stop")
        self.pushButton_voltage_export_xlsx = self.get_widget(QPushButton, "pushButton_voltage_export_xlsx")
        self.pushButton_voltage_export_pdf = self.get_widget(QPushButton, "pushButton_voltage_export_pdf")
        self.label_voltage_live = self.get_widget(QLabel, "label_voltage_live")
        self.label_voltage_status = self.get_widget(QLabel, "label_voltage_status")
        self.label_voltage_accuracy = self.get_widget(QLabel, "label_voltage_accuracy")
        self.label_voltage_led_r1 = self.get_widget(QLabel, "label_voltage_led_r1")
        self.label_voltage_led_r2 = self.get_widget(QLabel, "label_voltage_led_r2")
        self.label_voltage_led_r3 = self.get_widget(QLabel, "label_voltage_led_r3")
        self.label_voltage_led_r4 = self.get_widget(QLabel, "label_voltage_led_r4")
        self.label_voltage_led_t1 = self.get_widget(QLabel, "label_voltage_led_t1")
        self.label_voltage_led_t2 = self.get_widget(QLabel, "label_voltage_led_t2")
        self.label_voltage_led_t3 = self.get_widget(QLabel, "label_voltage_led_t3")
        self.label_voltage_led_t4 = self.get_widget(QLabel, "label_voltage_led_t4")
        self.tableWidget_voltage_results = self.get_widget(QTableWidget, "tableWidget_voltage_results")

        if not self.lineEdit_voltage_date.text().strip():
            self.lineEdit_voltage_date.setText(QDate.currentDate().toString("dd/MM/yyyy"))
        self.tableWidget_voltage_results.setColumnCount(4)
        self.tableWidget_voltage_results.setHorizontalHeaderLabels([
            "Inverseur", "Tension collage (V)", "Tension décollage (V)", "État"
        ])
        self.tableWidget_voltage_results.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_voltage_results.verticalHeader().setVisible(False)
        self.tableWidget_voltage_results.horizontalHeader().setStretchLastSection(True)

        self.ea_psu = EAPSU()
        self.voltage_test_running = False
        self.voltage_requested_mode = ""
        self.voltage_active_scan = ""
        self.voltage_results = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_raw_results = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_time_results = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_first_passage = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_capture_policy = ""
        self.voltage_effective_ramp_s = {"PICKUP": None, "DROPOUT": None}
        self.voltage_ramp_readbacks = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_plausibility = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_result_override = ""
        self.voltage_ea_stop_confirmation = self.voltage_empty_stop_confirmation()
        self.voltage_last_stop_diagnostic = ""
        self.voltage_last_adc_mv = None
        self.voltage_last_adc_raw = None
        self.voltage_ads_ok = False
        self.active_voltage_calibration = None
        self.voltage_ramp_started_monotonic = None
        self.voltage_ramp_info = None
        self.voltage_interphase_target_monotonic = None
        self.voltage_interphase_origin_monotonic = None
        self.voltage_interphase_actual_s = None
        self.voltage_run_settings = {}
        # Séquence R8 : l'EA réalise les rampes puis fournit automatiquement
        # la tension continue de chronométrie. Aucun passage manuel EA/FIXE.
        self.measure_all_active = False
        self.measure_all_phase = ""
        self.measure_all_context = {}
        self.measure_all_chrono_results = {}
        self.measure_all_static_confirmation = {}
        self.chrono_external_supply_mode = False
        self.rp2040_ea_chrono_capable = False
        self.voltage_last_saved_result = ""
        self._voltage_loading_measure_settings = False
        self.voltage_measure_settings_file = runtime_output_dir() / "voltage_measure_settings.json"
        self.voltage_waiting_for_rp_arm = False
        self.voltage_timeout_timer = QTimer(self.window)
        self.voltage_timeout_timer.setSingleShot(True)
        self.voltage_timeout_timer.timeout.connect(self.voltage_test_timeout)
        self.voltage_phase_timer = QTimer(self.window)
        self.voltage_phase_timer.setSingleShot(True)
        self.voltage_phase_timer.timeout.connect(self.voltage_validate_configured_ramp)
        self.voltage_arm_timeout_timer = QTimer(self.window)
        self.voltage_arm_timeout_timer.setSingleShot(True)
        self.voltage_arm_timeout_timer.timeout.connect(self.voltage_arm_timeout)
        self.voltage_progress_timer = QTimer(self.window)
        self.voltage_progress_timer.setInterval(100)
        self.voltage_progress_timer.timeout.connect(self.voltage_update_ramp_progress)
        self.voltage_ea_monitor_timer = QTimer(self.window)
        self.voltage_ea_monitor_timer.setInterval(750)
        self.voltage_ea_monitor_timer.timeout.connect(self.voltage_check_generator_running)
        self.voltage_pending_ramp = None
        self.voltage_refresh_ea_ports()
        self.voltage_refresh_results_table()
        self.voltage_refresh_contact_leds()
        self.voltage_init_db()
        self.voltage_load_measure_settings()
        self.voltage_update_relay_type_ui()
        self.voltage_update_ramp_limits()

    def create_voltage_calibration_tab(self):
        self.tab_voltage_calibration = self.get_widget(QWidget, "tab_voltage_calibration")
        self.lineEdit_calibration_operator = self.get_widget(QLineEdit, "lineEdit_calibration_operator")
        self.lineEdit_calibration_meter = self.get_widget(QLineEdit, "lineEdit_calibration_meter")
        self.dateEdit_calibration_date = self.get_widget(QDateEdit, "dateEdit_calibration_date")
        self.spinBox_calibration_valid_days = self.get_widget(QSpinBox, "spinBox_calibration_valid_days")
        self.doubleSpinBox_calibration_tolerance_v = self.get_widget(QDoubleSpinBox, "doubleSpinBox_calibration_tolerance_v")
        self.pushButton_calibration_request_ads = self.get_widget(QPushButton, "pushButton_calibration_request_ads")
        self.label_calibration_live = self.get_widget(QLabel, "label_calibration_live")
        self.doubleSpinBox_calibration_low_actual_v = self.get_widget(QDoubleSpinBox, "doubleSpinBox_calibration_low_actual_v")
        self.doubleSpinBox_calibration_high_actual_v = self.get_widget(QDoubleSpinBox, "doubleSpinBox_calibration_high_actual_v")
        self.doubleSpinBox_calibration_check_actual_v = self.get_widget(QDoubleSpinBox, "doubleSpinBox_calibration_check_actual_v")
        self.lineEdit_calibration_low_raw = self.get_widget(QLineEdit, "lineEdit_calibration_low_raw")
        self.lineEdit_calibration_high_raw = self.get_widget(QLineEdit, "lineEdit_calibration_high_raw")
        self.lineEdit_calibration_check_raw = self.get_widget(QLineEdit, "lineEdit_calibration_check_raw")
        self.pushButton_calibration_capture_low = self.get_widget(QPushButton, "pushButton_calibration_capture_low")
        self.pushButton_calibration_capture_high = self.get_widget(QPushButton, "pushButton_calibration_capture_high")
        self.pushButton_calibration_capture_check = self.get_widget(QPushButton, "pushButton_calibration_capture_check")
        self.pushButton_calibration_calculate = self.get_widget(QPushButton, "pushButton_calibration_calculate")
        self.pushButton_calibration_save_activate = self.get_widget(QPushButton, "pushButton_calibration_save_activate")
        self.pushButton_calibration_invalidate = self.get_widget(QPushButton, "pushButton_calibration_invalidate")
        self.pushButton_calibration_clear = self.get_widget(QPushButton, "pushButton_calibration_clear")
        self.label_calibration_coefficients = self.get_widget(QLabel, "label_calibration_coefficients")
        self.label_calibration_status = self.get_widget(QLabel, "label_calibration_status")
        self.tableWidget_calibration_history = self.get_widget(QTableWidget, "tableWidget_calibration_history")

        self.dateEdit_calibration_date.setDate(QDate.currentDate())
        self.tableWidget_calibration_history.setColumnCount(8)
        self.tableWidget_calibration_history.setHorizontalHeaderLabels([
            "ID", "Date", "Opérateur", "Multimètre", "Rapport", "Offset mV", "Erreur contrôle V", "Statut"
        ])
        self.tableWidget_calibration_history.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_calibration_history.setSelectionBehavior(QTableWidget.SelectRows)
        self.tableWidget_calibration_history.verticalHeader().setVisible(False)
        self.tableWidget_calibration_history.horizontalHeader().setStretchLastSection(True)

        self.calibration_capture_raw = {"LOW": None, "HIGH": None, "CHECK": None}
        self.calibration_pending_capture = ""
        self.calibration_calculated = None
        self.voltage_calibration_init_db()
        self.voltage_calibration_load_active()
        self.voltage_calibration_refresh_history()
        self.voltage_calibration_update_live()

    def load_ui(self):
        if not UI_FILE.exists():
            QMessageBox.critical(None, "Erreur", f"Fichier UI introuvable :\n{UI_FILE}")
            raise SystemExit(1)
        loader = QUiLoader()
        ui_file = QFile(str(UI_FILE))
        if not ui_file.open(QFile.ReadOnly):
            QMessageBox.critical(None, "Erreur", f"Impossible d'ouvrir le fichier UI :\n{UI_FILE}")
            raise SystemExit(1)
        window = loader.load(ui_file)
        ui_file.close()
        if window is None:
            QMessageBox.critical(None, "Erreur", "Chargement du fichier UI impossible.")
            raise SystemExit(1)
        return window

    def get_widget(self, widget_type, name):
        widget = self.window.findChild(widget_type, name)
        if widget is None:
            QMessageBox.critical(
                None, "Erreur UI",
                f"Widget introuvable : {name}\n"
                f"Ne pas renommer les objectName dans Qt Designer sans modifier le Python."
            )
            raise SystemExit(1)
        return widget

    def initialiser_widgets_production(self):
        self.dateEdit_prod_date.setCalendarPopup(True)
        self.dateEdit_prod_date.setDisplayFormat("dd/MM/yyyy")
        self.dateEdit_prod_date.setDate(QDate.currentDate())
        self.spinBox_prod_nb_inverseurs.setRange(1, 4)
        if self.spinBox_prod_nb_inverseurs.value() < 1:
            self.spinBox_prod_nb_inverseurs.setValue(2)
        self.tableWidget_prod_records.setColumnCount(10)
        self.tableWidget_prod_records.setHorizontalHeaderLabels(["Lot", "Nb essais", "SN distincts", "Acceptés", "Refusés", "Premier SN", "Dernier SN", "Dernier essai", "Inv.", "Désignation"])
        self.tableWidget_prod_records.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_prod_records.setSelectionBehavior(QTableWidget.SelectRows)
        self.tableWidget_prod_records.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tableWidget_prod_records.horizontalHeader().setStretchLastSection(True)
        self.label_etat_connexion_production.setMinimumHeight(34)
        self.label_etat_connexion_production.setAlignment(Qt.AlignCenter)
        self.label_etat_connexion.setAlignment(Qt.AlignCenter)

    def initialiser_widgets_database_admin(self):
        self.label_db_admin_status.setAlignment(Qt.AlignCenter)
        self.label_db_admin_file.setText(str(self.production_db_file))
        if self.comboBox_db_target.count() == 0:
            self.comboBox_db_target.addItems(["Neutral Screen automatique", "Chronométrie contacts + tensions"])
        self.tableWidget_db_operators.setColumnCount(2)
        self.tableWidget_db_operators.setHorizontalHeaderLabels(["Opérateur", "Créé le"])
        self.tableWidget_db_operators.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_db_operators.setSelectionBehavior(QTableWidget.SelectRows)
        self.tableWidget_db_operators.setSelectionMode(QTableWidget.SingleSelection)
        self.tableWidget_db_operators.setSortingEnabled(True)
        self.tableWidget_db_operators.horizontalHeader().setStretchLastSection(True)
        self.tableWidget_db_operators.setColumnWidth(0, 120)
        self.tableWidget_db_lots.setColumnCount(7)
        self.tableWidget_db_lots.setHorizontalHeaderLabels(["Lot", "Désignation", "Inv.", "Date création essai", "Bon / mauvais", "Nb essais", "SN distincts"])
        self.tableWidget_db_lots.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget_db_lots.setSelectionBehavior(QTableWidget.SelectRows)
        self.tableWidget_db_lots.setSelectionMode(QTableWidget.SingleSelection)
        self.tableWidget_db_lots.setSortingEnabled(True)
        self.tableWidget_db_lots.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tableWidget_db_lots.horizontalHeader().setStretchLastSection(True)

    def initialiser_validation_fin_essai(self):
        # La validation opérateur est assurée exclusivement par la fenêtre modale TEST FINI.
        self.set_auto_finish_validation_state(False)

    def initialiser_aide_boutons_operateur(self):
        """Installe une aide explicite après 3 secondes d'immobilité sur chaque bouton."""
        aides = {
            "pushButton_prod_save_context": "Vérifie les informations de production, charge le prochain SN si le lot existe, enregistre le contexte puis ouvre Neutral Screen Automatique.",
            "pushButton_prod_reload_base": "Relit la base Production depuis le disque puis actualise les tableaux Production et Gestion Base.",
            "pushButton_prod_export_pdf_lot": "Crée le rapport PDF du lot sélectionné dans l'historique Production. Le PDF contient le détail des relais testés et le résumé accepté/refusé.",
            "pushButton_prod_search_clear": "Efface le filtre de recherche lot et réaffiche l'ensemble des lots conservés dans la base.",
            "pushButton_rafraichir_ports_2": "Recherche à nouveau les ports COM disponibles pour retrouver le RP2040 si le câble USB vient d'être branché.",
            "pushButton_connecter_2": "Ouvre la communication série avec le RP2040 sélectionné. À utiliser si la connexion automatique n'a pas trouvé le bon port.",
            "pushButton_deconnecter_2": "Ferme proprement la communication série avec le RP2040 sans quitter le logiciel.",
            "pushButton_auto_neutral_marche": "Lance le scénario Neutral Screen Automatique sélectionné pour le relais SN affiché. Les pulses sont exécutés par le RP2040.",
            "pushButton_auto_neutral_arret": "Arrête immédiatement le scénario automatique en cours et envoie STOP au RP2040 si la connexion est active.",
            "pushButton_auto_scenario_recharger": "Recharge le fichier neutral_scenarios.json depuis le disque pour prendre en compte les scénarios modifiés.",
            "pushButton_auto_scenario_editer": "Ouvre l'onglet Editeur Scénario Neutral après déverrouillage, afin de modifier les scénarios de test.",
            "pushButton_auto_lot_fini": "Clôture obligatoirement le lot en cours. Tant que ce bouton n'est pas validé, l'opérateur ne peut pas préparer un autre lot.",
            "pushButton_neutral_be": "Mode manuel : envoie un pulse BE sur la sortie 1. À utiliser uniquement pour un contrôle manuel maîtrisé.",
            "pushButton_neutral_br": "Mode manuel : envoie un pulse BR sur la sortie 2. À utiliser uniquement pour un contrôle manuel maîtrisé.",
            "pushButton_neutral_bebr": "Mode manuel : envoie un pulse simultané BE/BR. C'est l'action manuelle correspondant à la recherche de position Neutral Screen.",
            "pushButton_neutral_stop": "Coupe immédiatement les sorties de commande côté RP2040. Bouton de sécurité manuel.",
            "pushButton_editor_nouveau": "Crée un nouveau scénario Neutral Screen dans l'éditeur.",
            "pushButton_editor_dupliquer": "Copie le scénario actuellement sélectionné pour créer une variante sans modifier l'original.",
            "pushButton_editor_supprimer": "Supprime le scénario sélectionné du fichier neutral_scenarios.json après confirmation.",
            "pushButton_editor_sauvegarder": "Valide les champs de l'éditeur puis sauvegarde le scénario dans neutral_scenarios.json.",
            "pushButton_editor_recharger": "Recharge neutral_scenarios.json depuis le disque et abandonne les modifications non sauvegardées.",
            "pushButton_editor_ajouter_etape": "Ajoute une ligne d'étape au scénario : action, durée, vérification, nombre d'essais et comportement si échec.",
            "pushButton_editor_supprimer_etape": "Supprime l'étape sélectionnée dans le tableau du scénario.",
            "pushButton_editor_monter_etape": "Déplace l'étape sélectionnée vers le haut pour changer l'ordre d'exécution du scénario.",
            "pushButton_editor_descendre_etape": "Déplace l'étape sélectionnée vers le bas pour changer l'ordre d'exécution du scénario.",
            "pushButton_editor_importer": "Importe un fichier JSON de scénarios après validation de sa structure.",
            "pushButton_editor_exporter": "Exporte les scénarios actuels dans un fichier JSON de sauvegarde ou de transfert.",
            "pushButton_editor_mot_de_passe": "Permet de modifier le mot de passe de déverrouillage des onglets protégés.",
            "pushButton_db_refresh": "Relit la base SQLite et met à jour les opérateurs, les lots affichés, le statut et la taille de la base.",
            "pushButton_db_backup": "Crée une sauvegarde complète de la base SQLite active avant une manipulation ou un archivage.",
            "pushButton_db_restore": "Remplace la base active par une sauvegarde SQLite sélectionnée, avec sauvegarde de sécurité avant remplacement.",
            "pushButton_db_export_csv": "Exporte tous les essais enregistrés dans un fichier CSV lisible avec un tableur.",
            "pushButton_db_vacuum": "Optimise la base SQLite : nettoyage du journal, compactage du fichier et analyse des index.",
            "pushButton_db_export_xlsx": "Exporte en XLSX le contenu de la base actuellement sélectionnée : Production ou Chronométrie contacts + tensions.",
            "pushButton_db_export_pdf": "Exporte en PDF le contenu de la base actuellement sélectionnée : Production ou Chronométrie contacts + tensions.",
            "pushButton_db_recreate_default": "Crée une base SQLite neuve et vide. Une sauvegarde de l'ancien fichier est faite si possible avant remplacement.",
            "pushButton_db_merge": "Importe les essais et opérateurs manquants depuis une autre base SQLite sans écraser la base active.",
            "pushButton_db_operator_add": "Ajoute l'opérateur saisi à la liste proposée dans l'onglet Production.",
            "pushButton_db_operator_delete": "Supprime définitivement l'opérateur sélectionné de la liste proposée. Les anciens essais gardent son nom en historique.",
            "pushButton_db_lot_open": "Ouvre la fenêtre détail du lot sélectionné avec la liste complète des relais testés.",
            "pushButton_db_lot_pdf": "Crée le PDF du lot sélectionné depuis la base actuellement sélectionnée.",
            "pushButton_db_lot_xlsx": "Crée le XLSX du lot sélectionné depuis la base actuellement sélectionnée.",
            "pushButton_db_lot_delete": "Supprime définitivement le lot sélectionné et tous les relais associés après confirmation. Une sauvegarde de sécurité est faite avant suppression.",
            "pushButton_chrono_export_xlsx_lot": "Exporte en XLSX les mesures de chronométrie et de tension enregistrées pour le lot saisi.",
            "pushButton_chrono_export_pdf_lot": "Crée un rapport PDF regroupant chronométrie et tensions pour le lot saisi.",
            "pushButton_voltage_open_calibration": "Ouvre l'onglet d'étalonnage ADS1115. Une calibration valide est obligatoire avant toute mesure officielle de tension.",
            "pushButton_voltage_measure_all": "Lance les rampes de tension puis règle automatiquement l'EA à la tension fixe de chronométrie. Aucun changement manuel de source. Les réglages de capture, pulse et sanctions proviennent de l'onglet Chronométrie contacts.",
            "pushButton_calibration_request_ads": "Demande une lecture fraîche du RAW ADS1115 au RP2040.",
            "pushButton_calibration_capture_low": "Capture le RAW ADS1115 au point bas. Saisir la valeur réellement lue au multimètre, pas la consigne EA.",
            "pushButton_calibration_capture_high": "Capture le RAW ADS1115 au point haut, idéalement vers 30 V.",
            "pushButton_calibration_capture_check": "Capture le RAW du point de contrôle intermédiaire, idéalement vers 15 V.",
            "pushButton_calibration_calculate": "Calcule le rapport réel du pont et l'offset à partir des deux points, puis vérifie l'erreur au point intermédiaire.",
            "pushButton_calibration_save_activate": "Enregistre et active la calibration uniquement si le contrôle intermédiaire respecte la tolérance réglée.",
            "pushButton_calibration_invalidate": "Désactive la calibration active. Les mesures officielles de tension seront alors bloquées.",
            "pushButton_oscillo_export_xlsx": "Exporte les points 0/1 de l'oscillogramme courant en XLSX pour analyse sous tableur.",
            "pushButton_oscillo_export_pdf": "Exporte l'oscillogramme courant en PDF avec les signaux carrés des contacts.",
            "pushButton_oscillo_vue_complete": "Affiche toute la fenêtre de capture de 0 µs jusqu'à la fin de capture.",
            "pushButton_oscillo_zoom_fronts": "Zoome automatiquement autour de la première et de la dernière transition capturée.",
            "pushButton_oscillo_zoom_contact": "Zoome autour des fronts du contact sélectionné, par exemple T1 pour voir un rebond de fermeture Travail.",
            "pushButton_oscillo_load_saved": "Recharge depuis la base chronométrie un ancien oscillogramme en deux étapes : choix du lot, puis choix du SN du lot.",
            "pushButton_oscillo_zoom_rebonds": "Zoome automatiquement sur la zone métier de rebond la plus longue détectée : fermeture ou ouverture selon la phase et le contact.",
            "pushButton_oscillo_zoom_in": "Divise la fenêtre de temps par le facteur choisi autour du centre actuel pour agrandir les fronts avec précision.",
            "pushButton_oscillo_zoom_out": "Multiplie la fenêtre de temps par le facteur choisi autour du centre actuel. Le dézoom est borné proprement à la vue complète.",
            "pushButton_rafraichir_ports": "Recherche à nouveau les ports COM disponibles dans l'onglet Cyclage.",
            "pushButton_connecter": "Connecte l'IHM au RP2040 depuis l'onglet Cyclage.",
            "pushButton_deconnecter": "Déconnecte proprement le RP2040 depuis l'onglet Cyclage.",
            "pushButton_mode_us": "Passe les champs de cyclage en microsecondes pour régler des durées très courtes.",
            "pushButton_mode_ms": "Passe les champs de cyclage en millisecondes.",
            "pushButton_mode_s": "Passe les champs de cyclage en secondes.",
            "pushButton_mode_min": "Passe les champs de cyclage en minutes.",
            "pushButton_mode_h": "Passe les champs de cyclage en heures pour des essais longs.",
            "pushButton_demarrer": "Démarre le cyclage selon les paramètres affichés dans l'onglet Cyclage.",
            "pushButton_pause": "Met le cyclage en pause côté RP2040 sans fermer la connexion série.",
            "pushButton_reprendre": "Reprend le cyclage après une pause.",
            "pushButton_arret": "Arrête le cyclage et coupe les sorties commandées.",
            "pushButton_status": "Demande un STATUS au RP2040 pour rafraîchir l'état des sorties, de la sélection tension et des contacts.",
        }
        for button in self.window.findChildren(QPushButton):
            name = button.objectName()
            text = aides.get(name)
            if not text:
                label = button.text().strip() or name
                text = f"Bouton {label} : action de commande de l'IHM. Vérifier le contexte avant utilisation."
            self.button_help_filter.install_on(button, text)

    def initialiser_verrouillage_onglets(self):
        self._last_allowed_tab_index = 0
        self.tabWidget_principal.currentChanged.connect(self.on_main_tab_changed)

    def on_main_tab_changed(self, index):
        if self._tab_guard_internal or index < 0:
            return
        widget = self.tabWidget_principal.widget(index)
        object_name = widget.objectName() if widget is not None else ""

        if self.lot_session_blocks_tab(object_name):
            self.show_lot_session_locked_message(object_name)
            self._tab_guard_internal = True
            self.tabWidget_principal.setCurrentIndex(self._last_allowed_tab_index)
            self._tab_guard_internal = False
            return

        if object_name in ("tab_production_accueil", "tab_neutral_auto"):
            self._protected_access_granted = False
            self._last_allowed_tab_index = index
            return
        if object_name not in self.protected_tab_names:
            self._last_allowed_tab_index = index
            return
        if self._protected_access_granted:
            self._last_allowed_tab_index = index
            return

        label = self.tabWidget_principal.tabText(index).replace(" (verrouillé)", "")
        code, ok = QInputDialog.getText(
            self.window,
            "Accès sécurisé",
            f"Code requis pour accéder à:\n{label}",
            QLineEdit.Password
        )
        if ok and code == self.current_access_code:
            self._protected_access_granted = True
            self._last_allowed_tab_index = index
            return
        if ok and self.reinitialiser_mot_de_passe_par_cle(str(code or "")):
            self._protected_access_granted = True
            self._last_allowed_tab_index = index
            return

        QMessageBox.warning(self.window, "Accès refusé", "Code incorrect ou accès annulé.")
        self._tab_guard_internal = True
        self.tabWidget_principal.setCurrentIndex(self._last_allowed_tab_index)
        self._tab_guard_internal = False

    def lot_session_blocks_tab(self, object_name):
        if not getattr(self, "_lot_session_active", False):
            return False
        if getattr(self, "_active_lot_finished", True):
            return False
        return object_name != "tab_neutral_auto"

    def show_lot_session_locked_message(self, target_object_name=""):
        lot = str(getattr(self, "_active_lot", "") or self.lineEdit_prod_lot.text().strip() or "-")
        self.big_message_box(
            "Lot en cours",
            "LOT NON CLOTURE",
            f"Le lot {lot} est encore en cours.\n\n"
            "Pour éviter de mélanger les SN ou de démarrer un autre lot par erreur, vous devez d'abord appuyer sur LOT FINI dans l'onglet Neutral Screen Automatique.\n\n"
            "Ensuite seulement vous pourrez revenir à Production et préparer un autre lot.",
            ok_text="COMPRIS",
            icon=QMessageBox.Warning,
        )
        self.label_auto_status.setText(f"Lot {lot} en cours - appuyer sur LOT FINI pour revenir à Production")
        self.label_auto_status.setStyleSheet("background-color: rgb(255,235,120); color: black; font-weight: bold; border: 2px solid rgb(180,120,0);")

    def set_tab_internal(self, widget):
        self._tab_guard_internal = True
        try:
            self.tabWidget_principal.setCurrentWidget(widget)
            index = self.tabWidget_principal.indexOf(widget)
            if index >= 0:
                self._last_allowed_tab_index = index
        finally:
            self._tab_guard_internal = False

    def changer_mot_de_passe_acces(self):
        code_actuel, ok = QInputDialog.getText(
            self.window,
            "Modifier mot de passe",
            "Code actuel :",
            QLineEdit.Password
        )
        if not ok:
            return
        if code_actuel != self.current_access_code:
            QMessageBox.warning(self.window, "Mot de passe", "Code actuel incorrect.")
            return
        nouveau, ok = QInputDialog.getText(
            self.window,
            "Modifier mot de passe",
            "Nouveau code :",
            QLineEdit.Password
        )
        if not ok:
            return
        nouveau = str(nouveau or "").strip()
        if len(nouveau) < 4:
            QMessageBox.warning(self.window, "Mot de passe", "Le nouveau code doit contenir au moins 4 caractères.")
            return
        confirmation, ok = QInputDialog.getText(
            self.window,
            "Modifier mot de passe",
            "Confirmer le nouveau code :",
            QLineEdit.Password
        )
        if not ok:
            return
        if confirmation != nouveau:
            QMessageBox.warning(self.window, "Mot de passe", "La confirmation ne correspond pas.")
            return
        self.current_access_code = nouveau
        self.production_data["access_code"] = nouveau
        try:
            self.production_save_db()
        except Exception as exc:
            QMessageBox.warning(self.window, "Mot de passe", f"Code changé mais sauvegarde impossible : {exc}")
            return
        QMessageBox.information(self.window, "Mot de passe", "Code d'accès modifié et sauvegardé.")

    def reinitialiser_mot_de_passe_par_cle(self, code_saisi):
        if str(code_saisi or "").strip().lower() != str(LOCK_RECOVERY_TRIGGER).strip().lower():
            return False
        if not self.big_message_box(
            "Récupération accès",
            "RÉINITIALISER LE MOT DE PASSE ?",
            "La clé de secours a été reconnue.\n\n"
            "Le mot de passe des onglets protégés sera remis à 1234.\n"
            "La base d'essais et les résultats ne seront pas modifiés.",
            ok_text="REMETTRE À 1234",
            cancel_text="ANNULER",
            icon=QMessageBox.Warning,
        ):
            return False
        self.current_access_code = LOCK_ACCESS_CODE
        self.production_data["access_code"] = LOCK_ACCESS_CODE
        try:
            self.production_save_db()
        except Exception as exc:
            QMessageBox.warning(self.window, "Mot de passe", f"Réinitialisation impossible : {exc}")
            return False
        self.big_message_box(
            "Récupération accès",
            "MOT DE PASSE RÉINITIALISÉ",
            "Le mot de passe est maintenant :\n\n1234",
            ok_text="COMPRIS",
            icon=QMessageBox.Information,
        )
        return True

    def clamp_nb_inverseurs(self, value, default=2):
        try:
            nb_inverseurs = int(value)
        except Exception:
            nb_inverseurs = int(default)
        return max(1, min(4, nb_inverseurs))

    def production_nb_inverseurs(self):
        valeur = int(self.spinBox_prod_nb_inverseurs.value())
        if valeur < 1 or valeur > 4:
            raise ValueError("Le nombre d'inverseurs doit être compris entre 1 et 4.")
        return valeur

    def production_sync_nb_inverseurs_to_auto(self):
        if hasattr(self, "lineEdit_auto_nb_inverseurs"):
            self.lineEdit_auto_nb_inverseurs.setText(str(self.production_nb_inverseurs()))
            self.on_auto_nb_inverseurs_changed()

    def production_context(self):
        return {
            "scenario": self.comboBox_prod_scenario.currentText().strip(),
            "lot": self.lineEdit_prod_lot.text().strip(),
            "designation": self.lineEdit_prod_designation.text().strip(),
            "nb_inverseurs": self.production_nb_inverseurs(),
            "operateur": self.comboBox_prod_operateur.currentText().strip(),
            "date": self.dateEdit_prod_date.date().toString("yyyy-MM-dd"),
            "sn": self.lineEdit_SN.text().strip(),
        }

    def production_apply_context(self, context):
        context = dict(context or {})
        scenario = str(context.get("scenario", "") or "")
        if scenario and self.comboBox_prod_scenario.findText(scenario) >= 0:
            self.comboBox_prod_scenario.setCurrentText(scenario)
        self.lineEdit_prod_lot.setText(str(context.get("lot", "") or ""))
        self.lineEdit_prod_designation.setText(str(context.get("designation", "") or ""))
        self.spinBox_prod_nb_inverseurs.setValue(self.clamp_nb_inverseurs(context.get("nb_inverseurs", 2)))
        operateur = str(context.get("operateur", "") or "")
        if operateur and self.comboBox_prod_operateur.findText(operateur) < 0:
            self.comboBox_prod_operateur.addItem(operateur)
        self.comboBox_prod_operateur.setCurrentText(operateur)
        date_txt = str(context.get("date", "") or "")
        date = QDate.fromString(date_txt, "yyyy-MM-dd")
        self.dateEdit_prod_date.setDate(date if date.isValid() else QDate.currentDate())
        self.lineEdit_SN.setText(str(context.get("sn", self.lineEdit_SN.text()) or ""))
        self.production_sync_nb_inverseurs_to_auto()
        self.sync_auto_production_labels()

    def production_clear_entry_fields(self):
        self._production_autofill_running = True
        try:
            self.lineEdit_prod_lot.clear()
            self.lineEdit_prod_designation.clear()
            self.spinBox_prod_nb_inverseurs.setValue(2)
            self.lineEdit_SN.clear()
            self.comboBox_prod_operateur.setCurrentIndex(-1)
            self.dateEdit_prod_date.setDate(QDate.currentDate())
        finally:
            self._production_autofill_running = False
        self.production_sync_nb_inverseurs_to_auto()
        self.sync_auto_production_labels()

    def sync_auto_production_labels(self):
        lot = self.lineEdit_prod_lot.text().strip()
        designation = self.lineEdit_prod_designation.text().strip()
        if self.label_auto_lot is not None:
            self.label_auto_lot.setText(f"Lot : {lot}" if lot else "Lot : --")
        if self.label_auto_design_relais is not None:
            self.label_auto_design_relais.setText(f"Désignation : {designation}" if designation else "Désignation : --")

    def production_autofill_from_lot(self):
        if getattr(self, "_production_autofill_running", False):
            return
        lot = self.lineEdit_prod_lot.text().strip()
        if not lot:
            self.lineEdit_SN.clear()
            self._auto_start_prompts_done = False
            return
        try:
            self.production_init_db()
            with self.production_connect_db() as con:
                record = con.execute(
                    """
                    SELECT lot, designation, nb_inverseurs, operateur, date, scenario, sn
                    FROM essais
                    WHERE lot = ?
                    ORDER BY timestamp DESC, id DESC
                    LIMIT 1
                    """,
                    (lot,),
                ).fetchone()
        except Exception as exc:
            self.label_prod_status.setText(f"Recherche lot impossible : {exc}")
            return
        if record is None:
            self.lineEdit_SN.clear()
            self._auto_start_prompts_done = False
            self.set_auto_finish_validation_state(False)
            self.sync_auto_production_labels()
            self.label_prod_status.setText(
                f"Nouveau lot détecté : {lot} - saisir le premier SN au démarrage du test."
            )
            return

        last_sn = str(record["sn"] or "").strip()
        next_sn = self.next_sn_value(last_sn)
        sn_status = ""
        self._production_autofill_running = True
        try:
            scenario = str(record["scenario"] or "")
            if scenario and self.comboBox_prod_scenario.findText(scenario) >= 0:
                self.comboBox_prod_scenario.setCurrentText(scenario)
            self.lineEdit_prod_designation.setText(str(record["designation"] or ""))
            self.spinBox_prod_nb_inverseurs.setValue(self.clamp_nb_inverseurs(record["nb_inverseurs"]))
            operateur = str(record["operateur"] or "")
            if operateur and self.comboBox_prod_operateur.findText(operateur) < 0:
                self.comboBox_prod_operateur.addItem(operateur)
            self.comboBox_prod_operateur.setCurrentText(operateur)
            date = QDate.fromString(str(record["date"] or ""), "yyyy-MM-dd")
            self.dateEdit_prod_date.setDate(date if date.isValid() else QDate.currentDate())
            if last_sn and next_sn and next_sn != last_sn:
                self.lineEdit_SN.setText(next_sn)
                sn_status = f" - prochain SN prêt : {next_sn}"
            elif last_sn:
                self.lineEdit_SN.clear()
                sn_status = f" - dernier SN {last_sn} non incrémentable, saisir le SN suivant"
            else:
                self.lineEdit_SN.clear()
                sn_status = " - aucun dernier SN trouvé, saisir le SN"
        finally:
            self._production_autofill_running = False
        self.production_sync_nb_inverseurs_to_auto()
        self.sync_auto_production_labels()
        self.label_prod_status.setText(f"Lot existant chargé : {lot}{sn_status}")

    def production_on_lot_edited(self, _text=""):
        if getattr(self, "_production_autofill_running", False):
            return
        self.lineEdit_SN.clear()
        self._auto_start_prompts_done = False
        if not getattr(self, "auto_neutral_running", False):
            self.set_auto_finish_validation_state(False)
        self.sync_auto_production_labels()

    def production_connect_db(self):
        con = sqlite3.connect(str(self.production_db_file), timeout=5.0, factory=ClosingSQLiteConnection)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.execute("PRAGMA foreign_keys=ON")
        return con

    def database_connect_file(self, path):
        con = sqlite3.connect(str(path), timeout=5.0, factory=ClosingSQLiteConnection)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA foreign_keys=ON")
        return con

    def production_init_db(self):
        with self.production_connect_db() as con:
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
                """
            )
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS essais (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    lot TEXT NOT NULL DEFAULT '',
                    sn TEXT NOT NULL DEFAULT '',
                    designation TEXT NOT NULL DEFAULT '',
                    nb_inverseurs INTEGER NOT NULL DEFAULT 2,
                    operateur TEXT NOT NULL DEFAULT '',
                    date TEXT NOT NULL DEFAULT '',
                    heure TEXT NOT NULL DEFAULT '',
                    scenario TEXT NOT NULL DEFAULT '',
                    resultat TEXT NOT NULL DEFAULT '',
                    details TEXT NOT NULL DEFAULT '',
                    timestamp TEXT NOT NULL DEFAULT ''
                )
                """
            )
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS operators (
                    name TEXT PRIMARY KEY,
                    active INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL DEFAULT ''
                )
                """
            )
            con.execute("CREATE INDEX IF NOT EXISTS idx_essais_lot_sn ON essais(lot, sn)")
            con.execute("CREATE INDEX IF NOT EXISTS idx_essais_lot_timestamp ON essais(lot, timestamp)")
            con.execute("CREATE INDEX IF NOT EXISTS idx_essais_timestamp ON essais(timestamp)")
            columns = {row["name"] for row in con.execute("PRAGMA table_info(essais)").fetchall()}
            if "nb_inverseurs" not in columns:
                con.execute("ALTER TABLE essais ADD COLUMN nb_inverseurs INTEGER NOT NULL DEFAULT 2")
        self.database_seed_operators_from_combo()

    def chrono_connect_db(self):
        con = sqlite3.connect(str(self.chrono_db_file), timeout=5.0, factory=ClosingSQLiteConnection)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        return con

    def chrono_init_db(self):
        with self.chrono_connect_db() as con:
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS mesures_chrono_contacts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    lot TEXT NOT NULL DEFAULT '',
                    date_test TEXT NOT NULL DEFAULT '',
                    relais TEXT NOT NULL DEFAULT '',
                    ambiance_c TEXT NOT NULL DEFAULT '',
                    nom_test TEXT NOT NULL DEFAULT '',
                    sn TEXT NOT NULL DEFAULT '',
                    relay_type TEXT NOT NULL DEFAULT '',
                    action TEXT NOT NULL DEFAULT '',
                    nb_inverseurs INTEGER NOT NULL DEFAULT 0,
                    capture_ms INTEGER NOT NULL DEFAULT 0,
                    pulse_ms INTEGER NOT NULL DEFAULT 0,
                    limite_temps_ms REAL NOT NULL DEFAULT 0,
                    limite_rebond_ms REAL NOT NULL DEFAULT 0,
                    resultat TEXT NOT NULL DEFAULT '',
                    overflow INTEGER NOT NULL DEFAULT 0,
                    details_json TEXT NOT NULL DEFAULT '',
                    events_json TEXT NOT NULL DEFAULT '',
                    timestamp TEXT NOT NULL DEFAULT ''
                )
                """
            )
            con.execute("CREATE INDEX IF NOT EXISTS idx_chrono_lot_sn ON mesures_chrono_contacts(lot, sn)")
            con.execute("CREATE INDEX IF NOT EXISTS idx_chrono_timestamp ON mesures_chrono_contacts(timestamp)")
            columns = {row["name"] for row in con.execute("PRAGMA table_info(mesures_chrono_contacts)").fetchall()}
            if "relay_type" not in columns:
                con.execute("ALTER TABLE mesures_chrono_contacts ADD COLUMN relay_type TEXT NOT NULL DEFAULT ''")

    def production_get_setting(self, key, default=""):
        with self.production_connect_db() as con:
            row = con.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return default if row is None else row["value"]

    def production_set_setting(self, key, value):
        with self.production_connect_db() as con:
            con.execute(
                "INSERT INTO settings(key, value) VALUES(?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (key, str(value)),
            )

    def database_default_operator_names(self):
        names = []
        if hasattr(self, "comboBox_prod_operateur"):
            for index in range(self.comboBox_prod_operateur.count()):
                name = self.comboBox_prod_operateur.itemText(index).strip()
                if name and name not in names:
                    names.append(name)
        if not names:
            names = ["O.MARECHAL", "Opérateur 1", "Opérateur 2", "Opérateur 3"]
        return names

    def database_seed_operators_from_combo(self):
        if getattr(self, "_database_seed_running", False):
            return
        self._database_seed_running = True
        try:
            names = self.database_default_operator_names()
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            with self.production_connect_db() as con:
                seeded_row = con.execute(
                    "SELECT value FROM settings WHERE key = ?",
                    ("operators_seeded",),
                ).fetchone()
                already_seeded = seeded_row is not None and str(seeded_row["value"] or "") == "1"
                current_count = con.execute("SELECT COUNT(*) AS n FROM operators").fetchone()["n"]
                if already_seeded:
                    return
                if current_count > 0:
                    con.execute(
                        "INSERT INTO settings(key, value) VALUES(?, ?) "
                        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                        ("operators_seeded", "1"),
                    )
                    return
                for name in names:
                    con.execute(
                        "INSERT OR IGNORE INTO operators(name, active, created_at) VALUES(?, 1, ?)",
                        (name, timestamp),
                    )
                con.execute(
                    "INSERT INTO settings(key, value) VALUES(?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    ("operators_seeded", "1"),
                )
        except Exception:
            pass
        finally:
            self._database_seed_running = False

    def database_load_operators_to_combo(self):
        try:
            with self.production_connect_db() as con:
                rows = con.execute(
                    "SELECT name FROM operators ORDER BY name COLLATE NOCASE ASC"
                ).fetchall()
        except Exception:
            return
        current = self.comboBox_prod_operateur.currentText().strip()
        self.comboBox_prod_operateur.blockSignals(True)
        try:
            self.comboBox_prod_operateur.clear()
            for row in rows:
                self.comboBox_prod_operateur.addItem(row["name"])
            if current and self.comboBox_prod_operateur.findText(current) >= 0:
                self.comboBox_prod_operateur.setCurrentText(current)
            else:
                self.comboBox_prod_operateur.setCurrentIndex(-1)
        finally:
            self.comboBox_prod_operateur.blockSignals(False)

    def database_admin_target(self):
        text = self.comboBox_db_target.currentText().strip().lower()
        return "chrono" if "chrono" in text else "production"

    def database_admin_is_chrono(self):
        return self.database_admin_target() == "chrono"

    def database_admin_active_db_file(self):
        return self.chrono_db_file if self.database_admin_is_chrono() else self.production_db_file

    def database_admin_target_changed(self):
        is_chrono = self.database_admin_is_chrono()
        if self.groupBox_db_operators is not None:
            self.groupBox_db_operators.setEnabled(not is_chrono)
            self.groupBox_db_operators.setTitle(
                "Opérateurs (base production uniquement)" if is_chrono else "Opérateurs"
            )
        if self.groupBox_db_lots is not None:
            self.groupBox_db_lots.setTitle(
                "Mesures chronométrie par lot / relais / test" if is_chrono else "Essais conservés par lot"
            )
        self.pushButton_db_lot_pdf.setEnabled(True)
        self.pushButton_db_lot_pdf.setToolTip(
            "Crée le PDF du lot sélectionné depuis la base Chronométrie."
            if is_chrono else
            "Crée le PDF du lot sélectionné depuis la base Production."
        )
        self.pushButton_db_lot_xlsx.setToolTip(
            "Crée le XLSX du lot sélectionné depuis la base Chronométrie."
            if is_chrono else
            "Crée le XLSX du lot sélectionné depuis la base Production."
        )
        self.pushButton_db_lot_delete.setText("SUPPRIMER MESURES LOT" if is_chrono else "SUPPRIMER LOT")
        self.lineEdit_db_lot_filter.setPlaceholderText(
            "Lot, relais, SN ou nom du test"
            if is_chrono else
            "Taper une partie du numéro de lot ou de la désignation"
        )
        self.database_admin_refresh()

    def database_admin_refresh(self):
        if not hasattr(self, "tableWidget_db_operators"):
            return
        if self.database_admin_is_chrono():
            self.chrono_database_admin_refresh()
            return
        search = self.lineEdit_db_lot_filter.text().strip() if hasattr(self, "lineEdit_db_lot_filter") else ""
        try:
            self.production_init_db()
            db_size = self.production_db_file.stat().st_size if self.production_db_file.exists() else 0
            with self.production_connect_db() as con:
                operator_rows = con.execute(
                    "SELECT name, created_at FROM operators ORDER BY name COLLATE NOCASE ASC"
                ).fetchall()
                params = []
                where = ""
                if search:
                    where = "WHERE lot LIKE ? OR designation LIKE ?"
                    params = [f"%{search}%", f"%{search}%"]
                lot_rows = con.execute(
                    f"""
                    SELECT
                        lot,
                        COUNT(*) AS nb_essais,
                        COUNT(DISTINCT sn) AS nb_sn_distincts,
                        SUM(CASE WHEN resultat LIKE 'ACCEPT%' THEN 1 ELSE 0 END) AS nb_acceptes,
                        SUM(CASE WHEN resultat LIKE 'REFUS%' OR resultat LIKE 'REJET%' THEN 1 ELSE 0 END) AS nb_refuses,
                        MIN(timestamp) AS premier_essai,
                        MAX(designation) AS designation,
                        MAX(nb_inverseurs) AS nb_inverseurs
                    FROM essais
                    {where}
                    GROUP BY lot
                    ORDER BY premier_essai DESC, lot COLLATE NOCASE ASC
                    """,
                    params,
                ).fetchall()
                nb_essais = con.execute("SELECT COUNT(*) AS n FROM essais").fetchone()["n"]
            self.label_db_admin_file.setText(str(self.production_db_file))
            suffix = f" - filtre : {search}" if search else ""
            self.label_db_admin_status.setText(
                f"Base OK - {nb_essais} essai(s) enregistré(s) - {len(lot_rows)} lot(s) affiché(s){suffix} - {db_size / 1024:.1f} Ko"
            )
            self.label_db_admin_status.setStyleSheet(
                "background-color: rgb(220,255,220); border: 2px solid rgb(50,150,50); padding: 8px; font-size: 11pt; font-weight: bold;"
            )
            self.tableWidget_db_operators.setSortingEnabled(False)
            self.tableWidget_db_operators.setRowCount(len(operator_rows))
            for row_index, row in enumerate(operator_rows):
                self.tableWidget_db_operators.setItem(row_index, 0, self.table_item_sort(row["name"], str(row["name"] or "").lower()))
                self.tableWidget_db_operators.setItem(row_index, 1, self.table_item_sort(self.format_datetime_fr(row["created_at"]), str(row["created_at"] or "")))
            self.tableWidget_db_operators.setSortingEnabled(True)
            self.tableWidget_db_lots.setSortingEnabled(False)
            self.tableWidget_db_lots.setColumnCount(7)
            self.tableWidget_db_lots.setHorizontalHeaderLabels(["Lot", "Désignation", "Inv.", "Date création essai", "Bon / mauvais", "Nb essais", "SN distincts"])
            self.tableWidget_db_lots.setRowCount(len(lot_rows))
            for row_index, record in enumerate(lot_rows):
                nb_acceptes = int(record["nb_acceptes"] or 0)
                nb_refuses = int(record["nb_refuses"] or 0)
                nb_essais = int(record["nb_essais"] or 0)
                nb_sn_distincts = int(record["nb_sn_distincts"] or 0)
                bon_mauvais = f"{nb_acceptes} bon / {nb_refuses} mauvais"
                lot_raw = str(record["lot"] or "")
                lot_display = lot_raw if lot_raw.strip() else "(lot vide)"
                vals = [
                    lot_display,
                    record["designation"],
                    self.clamp_nb_inverseurs(record["nb_inverseurs"]),
                    self.format_datetime_fr(record["premier_essai"]),
                    bon_mauvais,
                    nb_essais,
                    nb_sn_distincts,
                ]
                sorts = [
                    lot_raw.lower(),
                    str(record["designation"] or "").lower(),
                    self.clamp_nb_inverseurs(record["nb_inverseurs"]),
                    str(record["premier_essai"] or ""),
                    nb_refuses,
                    nb_essais,
                    nb_sn_distincts,
                ]
                for col, (val, sort_value) in enumerate(zip(vals, sorts)):
                    item = self.table_item_sort(val, sort_value)
                    if col == 0:
                        item.setData(Qt.UserRole + 1, lot_raw)
                    self.tableWidget_db_lots.setItem(row_index, col, item)
            self.tableWidget_db_lots.setSortingEnabled(True)
            for col, width in enumerate([80, 125, 45, 105, 100, 75, 85]):
                self.tableWidget_db_lots.setColumnWidth(col, width)
        except Exception as exc:
            self.label_db_admin_status.setText(f"Base non lisible : {exc}")
            self.label_db_admin_status.setStyleSheet(
                "background-color: rgb(255,225,225); border: 2px solid rgb(180,40,40); padding: 8px; font-size: 11pt; font-weight: bold;"
            )

    def chrono_database_admin_refresh(self):
        search = self.lineEdit_db_lot_filter.text().strip() if hasattr(self, "lineEdit_db_lot_filter") else ""
        try:
            self.voltage_init_db()
            db_size = self.chrono_db_file.stat().st_size if self.chrono_db_file.exists() else 0
            params = []
            where = ""
            if search:
                where = "WHERE lot LIKE ? OR relais LIKE ? OR nom_test LIKE ? OR sn LIKE ?"
                params = [f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"]
            with self.chrono_connect_db() as con:
                rows = con.execute(f"""
                    WITH toutes AS (
                        SELECT 'CHRONO' AS source, lot, relais, nom_test, sn, resultat, timestamp, relay_type
                        FROM mesures_chrono_contacts
                        UNION ALL
                        SELECT 'TENSION' AS source, lot, relais, nom_test, sn, resultat, timestamp, relay_type
                        FROM mesures_tension_fonctionnement
                    )
                    SELECT lot, relais, nom_test,
                           COUNT(DISTINCT sn) AS nb_sn,
                           SUM(CASE WHEN source='CHRONO' THEN 1 ELSE 0 END) AS nb_chrono,
                           SUM(CASE WHEN source='TENSION' THEN 1 ELSE 0 END) AS nb_tensions,
                           COUNT(*) AS nb_total,
                           SUM(CASE WHEN resultat='OK' THEN 1 ELSE 0 END) AS nb_ok,
                           SUM(CASE WHEN resultat<>'OK' THEN 1 ELSE 0 END) AS nb_defauts,
                           MAX(timestamp) AS derniere_mesure,
                           MAX(relay_type) AS relay_type
                    FROM toutes {where}
                    GROUP BY lot, relais, nom_test
                    ORDER BY derniere_mesure DESC, lot COLLATE NOCASE ASC
                """, params).fetchall()
                nb_chrono = con.execute("SELECT COUNT(*) AS n FROM mesures_chrono_contacts").fetchone()["n"]
                nb_tension = con.execute("SELECT COUNT(*) AS n FROM mesures_tension_fonctionnement").fetchone()["n"]
            self.label_db_admin_file.setText(str(self.chrono_db_file))
            suffix = f" - filtre : {search}" if search else ""
            self.label_db_admin_status.setText(
                f"Base chronométrie + tensions OK - {nb_chrono} chrono - {nb_tension} tension(s) - {len(rows)} groupe(s){suffix} - {db_size / 1024:.1f} Ko"
            )
            self.label_db_admin_status.setStyleSheet("background-color: rgb(220,245,255); border: 2px solid rgb(70,130,180); padding: 8px; font-size: 11pt; font-weight: bold;")
            table = self.tableWidget_db_lots
            table.setSortingEnabled(False)
            table.setColumnCount(11)
            table.setHorizontalHeaderLabels(["Lot", "Relais", "Nom test", "SN", "Chrono", "Tensions", "Total", "OK", "Défauts", "Dernière mesure", "Type"])
            table.setRowCount(len(rows))
            for row_index, record in enumerate(rows):
                lot_raw = str(record["lot"] or "")
                vals = [lot_raw, str(record["relais"] or ""), str(record["nom_test"] or ""), int(record["nb_sn"] or 0), int(record["nb_chrono"] or 0), int(record["nb_tensions"] or 0), int(record["nb_total"] or 0), int(record["nb_ok"] or 0), int(record["nb_defauts"] or 0), self.format_datetime_fr(record["derniere_mesure"]), str(record["relay_type"] or "")]
                for col, val in enumerate(vals):
                    item = self.table_item_sort(val, val)
                    if col == 0:
                        item.setData(Qt.UserRole + 1, lot_raw)
                        item.setData(Qt.UserRole + 2, str(record["relais"] or ""))
                        item.setData(Qt.UserRole + 3, str(record["nom_test"] or ""))
                    table.setItem(row_index, col, item)
            table.setSortingEnabled(True)
            for col, width in enumerate([80, 110, 135, 45, 60, 65, 55, 50, 60, 125, 85]):
                table.setColumnWidth(col, width)
        except Exception as exc:
            self.label_db_admin_status.setText(f"Base non lisible : {exc}")
            self.label_db_admin_status.setStyleSheet("background-color: rgb(255,225,225); border: 2px solid rgb(180,40,40); padding: 8px; font-size: 11pt; font-weight: bold;")

    def database_add_operator(self):
        name = self.lineEdit_db_operator.text().strip()
        if not name:
            QMessageBox.warning(self.window, "Opérateur", "Saisir un nom d'opérateur.")
            return
        try:
            self.production_init_db()
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            with self.production_connect_db() as con:
                con.execute(
                    "INSERT INTO operators(name, active, created_at) VALUES(?, 1, ?) "
                    "ON CONFLICT(name) DO UPDATE SET active = 1",
                    (name, timestamp),
                )
            self.lineEdit_db_operator.clear()
            self.database_load_operators_to_combo()
            self.database_admin_refresh()
            self.label_prod_status.setText(f"Opérateur ajouté : {name}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Opérateur", f"Ajout impossible : {exc}")

    def database_selected_operator(self):
        row = self.tableWidget_db_operators.currentRow()
        if row < 0:
            return ""
        item = self.tableWidget_db_operators.item(row, 0)
        return item.text().strip() if item is not None else ""

    def database_delete_operator(self):
        name = self.database_selected_operator()
        if not name:
            QMessageBox.warning(self.window, "Opérateur", "Sélectionner un opérateur.")
            return
        if QMessageBox.question(
            self.window,
            "Supprimer opérateur",
            f"Supprimer cet opérateur de la liste actuelle ?\n\n{name}\n\nLes anciens essais réalisés avec ce nom resteront conservés.",
        ) != QMessageBox.Yes:
            return
        try:
            with self.production_connect_db() as con:
                con.execute("DELETE FROM operators WHERE name = ?", (name,))
            self.database_load_operators_to_combo()
            self.database_admin_refresh()
            self.label_prod_status.setText(f"Opérateur supprimé de la liste : {name}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Opérateur", f"Suppression impossible : {exc}")

    def database_backup(self):
        db_file = self.database_admin_active_db_file()
        prefix = "chronometrie_contacts" if self.database_admin_is_chrono() else "production_essais"
        default_name = f"{prefix}_backup_{time.strftime('%Y%m%d_%H%M')}.sqlite3"
        path = self.ask_export_path(
            "Sauvegarder la base SQLite",
            default_name,
            "SQLite (*.sqlite3);;Tous les fichiers (*)",
            ".sqlite3",
        )
        if not path:
            return
        try:
            if self.database_admin_is_chrono():
                self.chrono_init_db()
                source_con = self.chrono_connect_db()
            else:
                self.production_init_db()
                source_con = self.production_connect_db()
            with source_con as source, sqlite3.connect(path, factory=ClosingSQLiteConnection) as destination:
                source.execute("PRAGMA wal_checkpoint(FULL)")
                source.backup(destination)
            self.database_admin_refresh()
            QMessageBox.information(self.window, "Sauvegarde", f"Sauvegarde créée :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Sauvegarde", f"Sauvegarde impossible : {exc}")

    def database_restore(self):
        db_file = self.database_admin_active_db_file()
        target_label = "chronométrie contacts" if self.database_admin_is_chrono() else "production Neutral Screen"
        path, _ = QFileDialog.getOpenFileName(
            self.window,
            "Restaurer une base SQLite",
            str(db_file.parent),
            "SQLite (*.sqlite3);;Tous les fichiers (*)",
        )
        if not path:
            return
        try:
            self.database_validate_restore_source(path)
        except Exception as exc:
            QMessageBox.warning(self.window, "Restaurer la base", f"Base source non utilisable pour {target_label} : {exc}")
            return
        if QMessageBox.question(
            self.window,
            "Restaurer la base",
            f"Cette opération remplace la base {target_label} active.\nUne sauvegarde de sécurité sera créée avant restauration.\n\nContinuer ?",
        ) != QMessageBox.Yes:
            return
        try:
            prefix = "chronometrie_contacts" if self.database_admin_is_chrono() else "production_essais"
            safety_name = db_file.with_name(f"{prefix}_avant_restauration_{time.strftime('%Y%m%d_%H%M%S')}.sqlite3")
            source_con = self.chrono_connect_db() if self.database_admin_is_chrono() else self.production_connect_db()
            with source_con as source, sqlite3.connect(str(safety_name)) as destination:
                source.execute("PRAGMA wal_checkpoint(FULL)")
                source.backup(destination)
            with sqlite3.connect(path, factory=ClosingSQLiteConnection) as source, sqlite3.connect(str(db_file), factory=ClosingSQLiteConnection) as destination:
                source.backup(destination)
            if self.database_admin_is_chrono():
                self.chrono_init_db()
            else:
                self.production_load_db()
                self.production_refresh_table()
                self.database_load_operators_to_combo()
            self.database_admin_refresh()
            QMessageBox.information(self.window, "Restaurer la base", f"Base restaurée.\nSauvegarde de sécurité :\n{safety_name}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Restaurer la base", f"Restauration impossible : {exc}")

    def database_export_csv(self):
        if self.database_admin_is_chrono():
            self.chrono_database_export_csv()
            return
        path = self.ask_export_path(
            "Exporter les essais en CSV",
            f"production_essais_{time.strftime('%Y%m%d_%H%M')}.csv",
            "CSV (*.csv);;Tous les fichiers (*)",
            ".csv",
        )
        if not path:
            return
        headers = ["lot", "sn", "designation", "nb_inverseurs", "operateur", "date", "heure", "scenario", "resultat", "details", "timestamp"]
        try:
            self.production_init_db()
            with self.production_connect_db() as con:
                rows = con.execute(
                    "SELECT lot, sn, designation, nb_inverseurs, operateur, date, heure, scenario, resultat, details, timestamp "
                    "FROM essais ORDER BY lot COLLATE NOCASE ASC, timestamp ASC, id ASC"
                ).fetchall()
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(headers)
                for row in rows:
                    writer.writerow([row[h] for h in headers])
            QMessageBox.information(self.window, "Export CSV", f"Export créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export CSV", f"Export impossible : {exc}")

    def database_maintenance(self):
        try:
            if self.database_admin_is_chrono():
                self.chrono_init_db()
                con_factory = self.chrono_connect_db
            else:
                self.production_init_db()
                con_factory = self.production_connect_db
            with con_factory() as con:
                con.execute("PRAGMA wal_checkpoint(TRUNCATE)")
                con.execute("VACUUM")
                con.execute("ANALYZE")
            if not self.database_admin_is_chrono():
                self.production_refresh_table()
            self.database_admin_refresh()
            QMessageBox.information(
                self.window,
                "Optimiser la base",
                "Base optimisée : journal WAL nettoyé, fichier compacté et index analysés.",
            )
        except Exception as exc:
            QMessageBox.warning(self.window, "Optimiser la base", f"Optimisation impossible : {exc}")

    def database_validate_restore_source(self, path):
        expected_table = "mesures_chrono_contacts" if self.database_admin_is_chrono() else "essais"
        with self.database_connect_file(path) as con:
            row = con.execute("PRAGMA quick_check").fetchone()
            result = row[0] if row is not None else "réponse vide"
            if str(result).lower() != "ok":
                raise RuntimeError(f"PRAGMA quick_check = {result}")
            if not self.database_external_has_table(con, expected_table):
                raise RuntimeError(f"table {expected_table} absente")

    def chrono_database_combined_rows(self):
        self.voltage_init_db()
        rows = []
        with self.chrono_connect_db() as con:
            chrono = con.execute("""
                SELECT lot, date_test, relais, ambiance_c, nom_test, sn, relay_type,
                       action, nb_inverseurs, capture_ms, pulse_ms, limite_temps_ms,
                       limite_rebond_ms, resultat, overflow, details_json, events_json, timestamp
                FROM mesures_chrono_contacts
            """).fetchall()
            voltages = con.execute("""
                SELECT lot, date_test, relais, ambiance_c, nom_test, sn, relay_type,
                       nb_inverseurs, pickup_global_v, dropout_global_v, pickup_json,
                       dropout_json, pickup_plausibility_status, dropout_plausibility_status,
                       ea_stop_confirmed, ea_final_voltage_v, resultat, timestamp
                FROM mesures_tension_fonctionnement
            """).fetchall()
        for r in chrono:
            rows.append({
                "source":"CHRONO", "lot":r["lot"], "date_test":r["date_test"], "relais":r["relais"],
                "ambiance_c":r["ambiance_c"], "nom_test":r["nom_test"], "sn":r["sn"], "relay_type":r["relay_type"],
                "action":r["action"], "nb_inverseurs":r["nb_inverseurs"], "capture_ms":r["capture_ms"], "pulse_ms":r["pulse_ms"],
                "limite_temps_ms":r["limite_temps_ms"], "limite_rebond_ms":r["limite_rebond_ms"],
                "tension_be_v":None, "tension_br_v":None,
                "tension_be_i1_v":None, "tension_be_i2_v":None, "tension_be_i3_v":None, "tension_be_i4_v":None,
                "tension_br_i1_v":None, "tension_br_i2_v":None, "tension_br_i3_v":None, "tension_br_i4_v":None,
                "plausibilite_be":"", "plausibilite_br":"",
                "arret_ea_confirme":"", "tension_ea_finale_v":None, "resultat":r["resultat"], "overflow":r["overflow"],
                "details_json":r["details_json"], "events_json":r["events_json"], "timestamp":r["timestamp"],
            })
        for r in voltages:
            pickup = self.chrono_json_dict(r["pickup_json"])
            dropout = self.chrono_json_dict(r["dropout_json"])
            details=json.dumps({"pickup":pickup,"dropout":dropout}, ensure_ascii=False)
            stop=int(r["ea_stop_confirmed"] if r["ea_stop_confirmed"] is not None else -1)
            rows.append({
                "source":"TENSION", "lot":r["lot"], "date_test":r["date_test"], "relais":r["relais"],
                "ambiance_c":r["ambiance_c"], "nom_test":r["nom_test"], "sn":r["sn"], "relay_type":r["relay_type"],
                "action":"BE/BR", "nb_inverseurs":r["nb_inverseurs"], "capture_ms":None, "pulse_ms":None,
                "limite_temps_ms":None, "limite_rebond_ms":None,
                "tension_be_v":r["pickup_global_v"], "tension_br_v":r["dropout_global_v"],
                "tension_be_i1_v":pickup.get("1"), "tension_be_i2_v":pickup.get("2"), "tension_be_i3_v":pickup.get("3"), "tension_be_i4_v":pickup.get("4"),
                "tension_br_i1_v":dropout.get("1"), "tension_br_i2_v":dropout.get("2"), "tension_br_i3_v":dropout.get("3"), "tension_br_i4_v":dropout.get("4"),
                "plausibilite_be":r["pickup_plausibility_status"], "plausibilite_br":r["dropout_plausibility_status"],
                "arret_ea_confirme":"OUI" if stop==1 else "NON" if stop==0 else "NON VÉRIFIÉ",
                "tension_ea_finale_v":r["ea_final_voltage_v"], "resultat":r["resultat"], "overflow":None,
                "details_json":details, "events_json":"", "timestamp":r["timestamp"],
            })
        rows.sort(key=lambda r:(str(r["lot"] or "").lower(), str(r["timestamp"] or ""), str(r["sn"] or "")))
        return rows

    def chrono_database_export_csv(self):
        path = self.ask_export_path("Exporter chronométrie et tensions en CSV", f"chronometrie_et_tensions_{time.strftime('%Y%m%d_%H%M')}.csv", "CSV (*.csv);;Tous les fichiers (*)", ".csv")
        if not path:
            return
        headers=["source","lot","date_test","relais","ambiance_c","nom_test","sn","relay_type","action","nb_inverseurs","capture_ms","pulse_ms","limite_temps_ms","limite_rebond_ms","tension_be_v","tension_br_v","tension_be_i1_v","tension_be_i2_v","tension_be_i3_v","tension_be_i4_v","tension_br_i1_v","tension_br_i2_v","tension_br_i3_v","tension_br_i4_v","plausibilite_be","plausibilite_br","arret_ea_confirme","tension_ea_finale_v","resultat","overflow","details_json","events_json","timestamp"]
        try:
            rows=self.chrono_database_combined_rows()
            with open(path,"w",encoding="utf-8-sig",newline="") as f:
                writer=csv.writer(f,delimiter=";"); writer.writerow(headers)
                for row in rows: writer.writerow([row[h] for h in headers])
            QMessageBox.information(self.window,"Export CSV",f"Export chronométrie + tensions créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window,"Export CSV",f"Export impossible : {exc}")

    def database_admin_export_dataset(self):
        if self.database_admin_is_chrono():
            headers = [
                ("source","Source"),("lot","Lot"),("date_test","Date test"),("relais","Relais"),
                ("ambiance_c","Ambiance °C"),("nom_test","Nom test"),("sn","SN"),("relay_type","Type"),
                ("action","Action"),("nb_inverseurs","Inv."),("capture_ms","Capture ms"),("pulse_ms","Pulse ms"),
                ("limite_temps_ms","Limite temps"),("limite_rebond_ms","Limite rebond"),
                ("tension_be_v","Tension collage/BE globale V"),("tension_br_v","Tension décollage/BR globale V"),
                ("tension_be_i1_v","BE I1 V"),("tension_be_i2_v","BE I2 V"),("tension_be_i3_v","BE I3 V"),("tension_be_i4_v","BE I4 V"),
                ("tension_br_i1_v","BR I1 V"),("tension_br_i2_v","BR I2 V"),("tension_br_i3_v","BR I3 V"),("tension_br_i4_v","BR I4 V"),
                ("plausibilite_be","Plausibilité BE"),("plausibilite_br","Plausibilité BR"),
                ("arret_ea_confirme","Arrêt EA confirmé"),("tension_ea_finale_v","Tension EA finale V"),
                ("resultat","Résultat"),("overflow","Overflow"),("details_json","Détails JSON"),
                ("events_json","Événements JSON"),("timestamp","Horodatage"),
            ]
            pdf_headers = [
                ("source","Source"),("lot","Lot"),("date_test","Date"),("relais","Relais"),
                ("nom_test","Test"),("sn","SN"),("relay_type","Type"),("action","Action"),
                ("tension_be_v","BE V"),("tension_br_v","BR V"),("resultat","Résultat"),
            ]
            return "chronometrie_et_tensions", "Base Chronométrie contacts et tensions", headers, pdf_headers, self.chrono_database_combined_rows()

        headers = [
            ("lot", "Lot"), ("sn", "SN"), ("designation", "Désignation"),
            ("nb_inverseurs", "Inv."), ("operateur", "Opérateur"), ("date", "Date"),
            ("heure", "Heure"), ("scenario", "Scénario"), ("resultat", "Résultat"),
            ("details", "Détails"), ("timestamp", "Horodatage"),
        ]
        pdf_headers = [
            ("lot", "Lot"), ("sn", "SN"), ("designation", "Désignation"),
            ("nb_inverseurs", "Inv."), ("date", "Date"), ("scenario", "Scénario"),
            ("resultat", "Résultat"),
        ]
        self.production_init_db()
        with self.production_connect_db() as con:
            rows = con.execute(
                """
                SELECT lot, sn, designation, nb_inverseurs, operateur, date, heure,
                       scenario, resultat, details, timestamp
                FROM essais
                ORDER BY lot COLLATE NOCASE ASC, timestamp ASC, id ASC
                """
            ).fetchall()
        return "production_essais", "Base Production Neutral Screen", headers, pdf_headers, rows

    def database_export_xlsx(self):
        try:
            prefix, title, headers, _pdf_headers, rows = self.database_admin_export_dataset()
        except Exception as exc:
            QMessageBox.warning(self.window, "Export XLSX", f"Lecture base impossible : {exc}")
            return
        path = self.ask_export_path(
            "Exporter la base en XLSX",
            f"{prefix}_{time.strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel (*.xlsx)",
            ".xlsx",
        )
        if not path:
            return
        try:
            self.write_table_xlsx(path, title, headers, rows)
            QMessageBox.information(self.window, "Export XLSX", f"Export XLSX créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export XLSX", f"Export impossible : {exc}")

    def database_export_pdf(self):
        try:
            prefix, title, _headers, pdf_headers, rows = self.database_admin_export_dataset()
        except Exception as exc:
            QMessageBox.warning(self.window, "Export PDF", f"Lecture base impossible : {exc}")
            return
        path = self.ask_export_path(
            "Exporter la base en PDF",
            f"{prefix}_{time.strftime('%Y%m%d_%H%M')}.pdf",
            "PDF (*.pdf)",
            ".pdf",
        )
        if not path:
            return
        try:
            self.write_table_pdf(path, title, pdf_headers, rows)
            QMessageBox.information(self.window, "Export PDF", f"Export PDF créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export PDF", f"Export impossible : {exc}")

    def write_table_xlsx(self, path, title, headers, rows):
        table_rows = [[label for _key, label in headers]]
        for row in rows:
            table_rows.append([row[key] for key, _label in headers])
        sheet_rows = []
        for r_idx, row in enumerate(table_rows, start=1):
            cells = []
            for c_idx, value in enumerate(row):
                cell_ref = f"{self.xlsx_col_name(c_idx)}{r_idx}"
                style_id = 2 if r_idx == 1 else 1
                if r_idx > 1 and isinstance(value, (int, float)) and value is not None:
                    cells.append(f'<c r="{cell_ref}" s="{style_id}"><v>{value}</v></c>')
                else:
                    cells.append(
                        f'<c r="{cell_ref}" t="inlineStr" s="{style_id}">'
                        f'<is><t>{self.xlsx_xml_escape(value)}</t></is></c>'
                    )
            sheet_rows.append(f'<row r="{r_idx}">{"".join(cells)}</row>')
        last_col = self.xlsx_col_name(max(0, len(headers) - 1))
        cols_xml = "".join(
            f'<col min="{idx}" max="{idx}" width="{22 if idx <= 6 else 32}" customWidth="1"/>'
            for idx in range(1, len(headers) + 1)
        )
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<dimension ref="A1:{last_col}{len(table_rows)}"/>'
            '<sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>'
            f'<cols>{cols_xml}</cols>'
            f'<sheetData>{"".join(sheet_rows)}</sheetData>'
            f'<autoFilter ref="A1:{last_col}{len(table_rows)}"/>'
            '</worksheet>'
        )
        self.write_xlsx_package(path, title, sheet_xml)

    def write_xlsx_package(self, path, title, sheet_xml):
        now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        files = {
            "[Content_Types].xml": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
                '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
                '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
                '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
                '</Types>'
            ),
            "_rels/.rels": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
                '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
                '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
                '</Relationships>'
            ),
            "docProps/app.xml": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
                'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
                '<Application>Neutral Screen RP2040</Application></Properties>'
            ),
            "docProps/core.xml": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                'xmlns:dc="http://purl.org/dc/elements/1.1/" '
                'xmlns:dcterms="http://purl.org/dc/terms/" '
                'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
                'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
                f'<dc:title>{self.xlsx_xml_escape(title)}</dc:title>'
                '<dc:creator>Neutral Screen RP2040</dc:creator>'
                f'<dcterms:created xsi:type="dcterms:W3CDTF">{now}</dcterms:created>'
                f'<dcterms:modified xsi:type="dcterms:W3CDTF">{now}</dcterms:modified>'
                '</cp:coreProperties>'
            ),
            "xl/workbook.xml": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                '<sheets><sheet name="Export" sheetId="1" r:id="rId1"/></sheets></workbook>'
            ),
            "xl/_rels/workbook.xml.rels": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
                '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
                '</Relationships>'
            ),
            "xl/styles.xml": (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                '<fonts count="2"><font><sz val="10"/><name val="Arial"/></font><font><b/><sz val="10"/><name val="Arial"/></font></fonts>'
                '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
                '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
                '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
                '<cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
                '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0"/></cellXfs>'
                '</styleSheet>'
            ),
            "xl/worksheets/sheet1.xml": sheet_xml,
        }
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, content in files.items():
                archive.writestr(name, content)

    def write_table_pdf(self, path, title, headers, rows):
        writer = QPdfWriter(path)
        writer.setPageSize(QPageSize(QPageSize.A4))
        if len(headers) > 9:
            writer.setPageOrientation(QPageLayout.Landscape)
        writer.setResolution(96)
        painter = QPainter(writer)
        if not painter.isActive():
            raise RuntimeError("QPainter PDF non actif.")
        try:
            page_w = writer.width()
            page_h = writer.height()
            margin = 28 if len(headers) > 9 else 34
            y = margin
            dense_table = len(headers) > 9
            line_h = 13 if dense_table else 16
            normal = QFont("Arial", 5 if dense_table else 7)
            bold = QFont("Arial", 6 if dense_table else 8)
            bold.setBold(True)
            title_font = QFont("Arial", 12 if dense_table else 14)
            title_font.setBold(True)

            def new_page():
                nonlocal y
                writer.newPage()
                y = margin
                draw_header()

            def draw_text(x, yy, text, font=None, max_chars=None):
                painter.setFont(font or normal)
                text = str(text or "")
                if max_chars and len(text) > max_chars:
                    text = text[: max_chars - 3] + "..."
                painter.drawText(int(x), int(yy), text)

            available = page_w - margin * 2
            # Une largeur est calculée pour chaque colonne. L'ancienne liste fixe
            # de neuf largeurs tronquait silencieusement les colonnes suivantes.
            weights = [max(5, min(22, len(str(label)) + 2)) for _key, label in headers]
            total = sum(weights) or 1
            widths = [max(20, int(weight * available / total)) for weight in weights]
            if widths:
                widths[-1] += available - sum(widths)
            x_positions = [margin]
            for width in widths[:-1]:
                x_positions.append(x_positions[-1] + width)

            def draw_header():
                nonlocal y
                painter.setFont(bold)
                for x, (_key, label), width in zip(x_positions, headers, widths):
                    painter.drawText(int(x), int(y), str(label)[:max(4, width // 7)])
                y += line_h
                painter.drawLine(margin, y - 11, page_w - margin, y - 11)

            painter.setFont(title_font)
            painter.drawText(margin, y, title)
            y += line_h * 2
            draw_text(margin, y, f"Lignes : {len(rows)}    Généré le : {time.strftime('%d/%m/%Y %H:%M')}", bold)
            y += line_h * 2
            draw_header()
            for row in rows:
                if y > page_h - margin:
                    new_page()
                for x, (key, _label), width in zip(x_positions, headers, widths):
                    draw_text(x, y, row[key], normal, max(4, width // 6))
                y += line_h
        finally:
            painter.end()

    def chrono_current_export_lot(self):
        lot = self.lineEdit_chrono_lot.text().strip()
        if not lot:
            QMessageBox.information(self.window, "Export lot mesures", "Renseigner le champ Lot dans l'onglet Mesures.")
            return ""
        return lot

    def chrono_records_for_lot(self, lot):
        self.chrono_init_db()
        with self.chrono_connect_db() as con:
            return con.execute(
                """
                SELECT lot, date_test, relais, ambiance_c, nom_test, sn,
                       relay_type, action, nb_inverseurs, capture_ms, pulse_ms,
                       limite_temps_ms, limite_rebond_ms, resultat, overflow,
                       details_json, events_json, timestamp
                FROM mesures_chrono_contacts
                WHERE lot = ?
                ORDER BY sn COLLATE NOCASE ASC, timestamp ASC, id ASC
                """,
                (lot,),
            ).fetchall()

    def chrono_json_dict(self, value):
        try:
            data = json.loads(value or "{}")
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def chrono_export_flat_rows(self, records):
        flat_rows = []
        for record in records:
            details = self.chrono_json_dict(record["details_json"])
            quality = details.get("quality") if isinstance(details.get("quality"), dict) else {}
            measure_rows = details.get("lignes_action")
            if not isinstance(measure_rows, list):
                measure_rows = details.get("inverseurs")
            if not isinstance(measure_rows, list):
                measure_rows = []
            for item in measure_rows:
                if not isinstance(item, dict):
                    continue
                if "temps_ms" in item and "debut" in item:
                    flat_rows.append({
                        "lot": record["lot"],
                        "date_test": record["date_test"],
                        "relais": record["relais"],
                        "ambiance_c": record["ambiance_c"],
                        "nom_test": record["nom_test"],
                        "sn": record["sn"],
                        "relay_type": record["relay_type"],
                        "action": record["action"],
                        "nb_inverseurs": record["nb_inverseurs"],
                        "resultat": record["resultat"],
                        "inverseur": item.get("inverseur", ""),
                        "mesure": item.get("mesure", ""),
                        "debut": item.get("debut", ""),
                        "fin": item.get("fin", ""),
                        "temps_ms": item.get("temps_ms"),
                        "sanction": self.chrono_export_sanction_label(item.get("ok")),
                        "spread_us": quality.get("spread_us"),
                        "loop_max_us": quality.get("loop_max_us"),
                        "events": quality.get("events"),
                        "overflow": "Oui" if int(record["overflow"] or 0) else "Non",
                        "timestamp": self.format_datetime_fr(record["timestamp"]) or record["timestamp"],
                    })
                else:
                    self.chrono_append_legacy_export_rows(flat_rows, record, item, quality)
        return flat_rows

    def chrono_append_legacy_export_rows(self, flat_rows, record, item, quality):
        inverseur = item.get("inverseur", "")
        mesure = str(item.get("mesure", "") or "")
        contact = str(item.get("contact", "") or "")
        base = {
            "lot": record["lot"],
            "date_test": record["date_test"],
            "relais": record["relais"],
            "ambiance_c": record["ambiance_c"],
            "nom_test": record["nom_test"],
            "sn": record["sn"],
            "relay_type": record["relay_type"],
            "action": record["action"],
            "nb_inverseurs": record["nb_inverseurs"],
            "resultat": record["resultat"],
            "inverseur": inverseur,
            "spread_us": quality.get("spread_us"),
            "loop_max_us": quality.get("loop_max_us"),
            "events": quality.get("events"),
            "overflow": "Oui" if int(record["overflow"] or 0) else "Non",
            "timestamp": self.format_datetime_fr(record["timestamp"]) or record["timestamp"],
        }
        flat_rows.append({
            **base,
            "mesure": f"{mesure} {inverseur}".strip(),
            "debut": "Commande",
            "fin": f"{contact} fermé" if contact else "Contact fermé",
            "temps_ms": item.get("temps_ms"),
            "sanction": self.chrono_export_sanction_label(item.get("temps_ok")),
        })
        flat_rows.append({
            **base,
            "mesure": f"Rebond fermeture {contact}".strip(),
            "debut": f"{contact} 1ère fermeture" if contact else "1ère fermeture",
            "fin": f"{contact} dernière fermeture" if contact else "dernière fermeture",
            "temps_ms": item.get("rebond_fermeture_ms"),
            "sanction": self.chrono_export_sanction_label(item.get("rebond_ok")),
        })

    def chrono_export_sanction_label(self, value):
        if value is None:
            return "EN ATTENTE"
        return "OK" if bool(value) else "DEFAUT"

    def chrono_export_limit_label(self, value_ms):
        try:
            value = float(value_ms)
        except Exception:
            value = 0.0
        if value <= 0:
            return ""
        if abs(value - round(value)) < 0.0005:
            text = str(int(round(value)))
        else:
            text = f"{value:.3f}".rstrip("0").rstrip(".")
        return f"<{text.replace('.', ',')}ms"

    def chrono_export_value_ms(self, value):
        if value is None:
            return "--"
        try:
            return f"{float(value):.3f}".replace(".", ",")
        except Exception:
            return str(value)

    def chrono_export_metric_order(self):
        """Ordre officiel des temps dans les rapports opérateur.

        Les rebonds d'ouverture restent enregistrés en base et disponibles dans
        l'oscillogramme, mais ne figurent pas dans les deux feuilles de synthèse
        demandées par l'utilisateur.
        """
        return [
            "enclenchement",
            "transfert_travail",
            "rebond_travail",
            "declenchement",
            "transfert_repos",
            "rebond_repos",
        ]

    def chrono_export_metric_label(self, metric, inv):
        labels = {
            "enclenchement": "Temps d'Enclenchement {inv} (ms)",
            "transfert_travail": "Temps de transfère {inv} (ms)",
            "rebond_travail": "Temps Rebond Travail Fermeture {inv} (ms)",
            "declenchement": "Temps de Déclenchement {inv} (ms)",
            "transfert_repos": "Temps de transfère {inv} retour (ms)",
            "rebond_repos": "Temps Rebond Repos Fermeture {inv} (ms)",
        }
        return labels.get(metric, str(metric)).format(inv=inv)

    def chrono_export_sanction_for_metric(self, metric, record):
        if metric in ("transfert_travail", "transfert_repos"):
            return self.chrono_export_limit_label(CHRONO_TRANSFER_LIMIT_MS)
        if metric in ("rebond_travail", "rebond_repos", "rebond_travail_ouverture", "rebond_repos_ouverture"):
            return self.chrono_export_limit_label(record["limite_rebond_ms"] or 2.0)
        return self.chrono_export_limit_label(record["limite_temps_ms"] or 1.5)

    def chrono_export_group_key(self, record):
        return (
            str(record["lot"] or ""),
            str(record["date_test"] or ""),
            str(record["relais"] or ""),
            str(record["ambiance_c"] or ""),
            str(record["nom_test"] or ""),
            str(record["sn"] or ""),
            str(record["relay_type"] or ""),
            int(record["nb_inverseurs"] or 0),
        )

    def chrono_export_group_data(self, records):
        """Regroupe BE/BR ou enclenchement/déclenchement par lot/relais/SN."""
        groups = {}
        for record in records:
            key = self.chrono_export_group_key(record)
            if key not in groups:
                groups[key] = {
                    "record": record,
                    "metrics": {},
                    "resultats": [],
                    "nb_inverseurs": int(record["nb_inverseurs"] or 1),
                }
            group = groups[key]
            group["resultats"].append(str(record["resultat"] or "").upper())
            group["nb_inverseurs"] = max(group["nb_inverseurs"], int(record["nb_inverseurs"] or 1))
            details = self.chrono_json_dict(record["details_json"])
            measure_rows = details.get("lignes_action")
            if not isinstance(measure_rows, list):
                measure_rows = details.get("inverseurs")
            if not isinstance(measure_rows, list):
                measure_rows = []
            for item in measure_rows:
                if not isinstance(item, dict):
                    continue
                metric = str(item.get("metric", "") or "")
                if metric not in self.chrono_export_metric_order():
                    continue
                try:
                    inv = int(item.get("inverseur") or 0)
                except Exception:
                    inv = 0
                if inv < 1:
                    continue
                group["metrics"][(inv, metric)] = {
                    "value": item.get("temps_ms"),
                    "sanction": self.chrono_export_sanction_for_metric(metric, record),
                }
        return list(groups.values())

    def chrono_export_find_voltage(self, chrono_record, voltage_records, used_voltage_ids):
        """Choisit la mesure tension la plus récente correspondant au même essai."""
        candidates = []
        for voltage in voltage_records:
            vid = int(voltage["id"] or 0)
            if vid in used_voltage_ids:
                continue
            if str(voltage["lot"] or "") != str(chrono_record["lot"] or ""):
                continue
            if str(voltage["sn"] or "") != str(chrono_record["sn"] or ""):
                continue
            vr = str(voltage["relais"] or "").strip()
            cr = str(chrono_record["relais"] or "").strip()
            if vr and cr and vr != cr:
                continue
            exact_test = int(
                bool(str(voltage["nom_test"] or "").strip())
                and str(voltage["nom_test"] or "").strip() == str(chrono_record["nom_test"] or "").strip()
            )
            candidates.append((exact_test, str(voltage["timestamp"] or ""), vid, voltage))
        if not candidates:
            return None
        candidates.sort(key=lambda item: (item[0], item[1], item[2]))
        selected = candidates[-1][3]
        used_voltage_ids.add(int(selected["id"] or 0))
        return selected

    def chrono_export_base_rows(self, record, chrono_result, voltage_result):
        return [
            ["Lot", "", record["lot"]],
            ["Date du test", "", record["date_test"]],
            ["Relais", "", record["relais"]],
            ["Ambiance", "", record["ambiance_c"]],
            ["Nom du Test", "", record["nom_test"]],
            ["Numéro de Relais", "", record["sn"]],
            ["Résultat chronométrie", "", chrono_result],
            ["Résultat tensions", "", voltage_result],
        ]

    def chrono_export_time_rows(self, metrics, nb_inverseurs):
        rows = []
        for inv in range(1, max(1, min(4, int(nb_inverseurs or 1))) + 1):
            for metric in self.chrono_export_metric_order():
                data = metrics.get((inv, metric), {})
                rows.append([
                    self.chrono_export_metric_label(metric, inv),
                    data.get("sanction", ""),
                    self.chrono_export_value_ms(data.get("value")),
                ])
        return rows

    def chrono_export_global_voltage_rows(self, voltage, detailed=False):
        if detailed:
            pickup_label = "Tension d'Enclenchement (globale)"
            dropout_label = "Tension de Rappel (globale)"
        else:
            pickup_label = "Tension d'Enclenchement"
            dropout_label = "Tension de Rappel"
        if voltage is None:
            return [
                [pickup_label, "", "--"],
                [dropout_label, "", "--"],
            ]
        return [
            [
                pickup_label,
                str(voltage["pickup_plausibility_status"] or ""),
                self.chrono_export_voltage_value(voltage["pickup_global_v"]),
            ],
            [
                dropout_label,
                str(voltage["dropout_plausibility_status"] or ""),
                self.chrono_export_voltage_value(voltage["dropout_global_v"]),
            ],
        ]

    def chrono_export_individual_voltage_rows(self, voltage, nb_inverseurs):
        pickup = self.chrono_json_dict(voltage["pickup_json"]) if voltage is not None else {}
        dropout = self.chrono_json_dict(voltage["dropout_json"]) if voltage is not None else {}
        rows = []
        for inv in range(1, max(1, min(4, int(nb_inverseurs or 1))) + 1):
            p = pickup.get(str(inv), pickup.get(inv))
            d = dropout.get(str(inv), dropout.get(inv))
            rows.append([
                f"Tension d'Enclenchement inverseur {inv}",
                "",
                self.chrono_export_voltage_value(p),
            ])
            rows.append([
                f"Tension de Rappel inverseur {inv}",
                "",
                self.chrono_export_voltage_value(d),
            ])
        return rows

    def chrono_export_measure_sheets(self, records, voltage_records=None):
        """Construit les deux feuilles officielles demandées pour chaque SN."""
        voltage_records = list(voltage_records or [])
        used_voltage_ids = set()
        summary_cards = []
        detail_cards = []

        for group in self.chrono_export_group_data(records):
            record = group["record"]
            chrono_result = "DEFAUT" if any(v == "DEFAUT" for v in group["resultats"]) else "OK"
            voltage = self.chrono_export_find_voltage(record, voltage_records, used_voltage_ids)
            voltage_result = str(voltage["resultat"] or "") if voltage is not None else "Aucune mesure tension associée"
            nb = max(group["nb_inverseurs"], int(voltage["nb_inverseurs"] or 1) if voltage is not None else 1)
            base_rows = self.chrono_export_base_rows(record, chrono_result, voltage_result)
            time_rows = self.chrono_export_time_rows(group["metrics"], nb)
            summary_cards.append(
                base_rows
                + self.chrono_export_global_voltage_rows(voltage, detailed=False)
                + time_rows
            )
            detail_cards.append(
                base_rows
                + self.chrono_export_global_voltage_rows(voltage, detailed=True)
                + self.chrono_export_individual_voltage_rows(voltage, nb)
                + time_rows
            )

        # Une tension peut exister sans chronométrie : elle reste visible et les
        # temps sont explicitement notés "--" au lieu de disparaître du rapport.
        for voltage in voltage_records:
            vid = int(voltage["id"] or 0)
            if vid in used_voltage_ids:
                continue
            pseudo = voltage
            nb = max(1, min(4, int(voltage["nb_inverseurs"] or 1)))
            base_rows = self.chrono_export_base_rows(
                pseudo,
                "Aucune mesure chronométrie associée",
                str(voltage["resultat"] or ""),
            )
            time_rows = self.chrono_export_time_rows({}, nb)
            summary_cards.append(
                base_rows
                + self.chrono_export_global_voltage_rows(voltage, detailed=False)
                + time_rows
            )
            detail_cards.append(
                base_rows
                + self.chrono_export_global_voltage_rows(voltage, detailed=True)
                + self.chrono_export_individual_voltage_rows(voltage, nb)
                + time_rows
            )
        return summary_cards, detail_cards

    def chrono_export_measure_cards(self, records, voltage_records=None):
        """Compatibilité historique : renvoie la première feuille uniquement."""
        summary_cards, _detail_cards = self.chrono_export_measure_sheets(records, voltage_records)
        return summary_cards

    def chrono_export_voltage_value(self, value):
        if value is None or value == "":
            return "--"
        try:
            return f"{float(value):.3f}"
        except Exception:
            return str(value)

    def chrono_export_voltage_rows(self, record):
        pickup = self.chrono_json_dict(record["pickup_json"])
        dropout = self.chrono_json_dict(record["dropout_json"])
        rows = [
            ["MESURES DE TENSION", "", ""],
            ["Horodatage tension", "", self.format_datetime_fr(record["timestamp"]) or record["timestamp"]],
            ["Résultat tensions", "", record["resultat"]],
            ["Tension collage / BE globale (V)", record["pickup_plausibility_status"], self.chrono_export_voltage_value(record["pickup_global_v"])],
            ["Tension décollage / BR globale (V)", record["dropout_plausibility_status"], self.chrono_export_voltage_value(record["dropout_global_v"])],
        ]
        nb = max(1, min(4, int(record["nb_inverseurs"] or 1)))
        for inv in range(1, nb + 1):
            p = pickup.get(str(inv), pickup.get(inv))
            d = dropout.get(str(inv), dropout.get(inv))
            rows.append([f"Tension collage / BE inverseur {inv} (V)", "", self.chrono_export_voltage_value(p)])
            rows.append([f"Tension décollage / BR inverseur {inv} (V)", "", self.chrono_export_voltage_value(d)])
        stop = int(record["ea_stop_confirmed"] if record["ea_stop_confirmed"] is not None else -1)
        stop_label = "OUI" if stop == 1 else "NON" if stop == 0 else "NON VÉRIFIÉ"
        rows.extend([
            ["Étalonnage utilisé", "", record["calibration_id"] if record["calibration_id"] is not None else "--"],
            ["Erreur contrôle étalonnage (V)", "", self.chrono_export_voltage_value(record["calibration_error_v"])],
            ["Arrêt EA confirmé", "", stop_label],
            ["Tension EA finale (V)", "", self.chrono_export_voltage_value(record["ea_final_voltage_v"])],
        ])
        return rows

    def chrono_export_headers(self):
        return [
            ("lot", "Lot"),
            ("date_test", "Date du test"),
            ("relais", "Relais"),
            ("ambiance_c", "Ambiance °C"),
            ("nom_test", "Nom du test"),
            ("sn", "SN"),
            ("relay_type", "Type relais"),
            ("action", "Action"),
            ("nb_inverseurs", "Nb inverseurs"),
            ("resultat", "Résultat mesure"),
            ("inverseur", "Inverseur"),
            ("mesure", "Mesure"),
            ("debut", "Début"),
            ("fin", "Fin"),
            ("temps_ms", "Temps (ms)"),
            ("sanction", "Sanction"),
            ("spread_us", "Écart indicatif 1er/dernier (µs)"),
            ("loop_max_us", "Loop max (µs)"),
            ("events", "Événements"),
            ("overflow", "Overflow"),
            ("timestamp", "Horodatage"),
        ]

    def chrono_export_current_lot_xlsx(self):
        lot = self.chrono_current_export_lot()
        if not lot:
            return
        try:
            records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "Export XLSX lot", f"Lecture base impossible : {exc}")
            return
        if not records and not voltage_records:
            QMessageBox.information(self.window, "Export XLSX lot", f"Aucune mesure chronométrie ou tension enregistrée pour le lot {lot}.")
            return
        default_name = f"chronometrie_et_tensions_lot_{self.filename_safe(lot)}.xlsx"
        path = self.ask_export_path("Exporter chronométrie et tensions en XLSX", default_name, "Excel (*.xlsx)", ".xlsx")
        if not path:
            return
        try:
            summary_cards, detail_cards = self.chrono_export_measure_sheets(records, voltage_records)
            self.write_chrono_lot_xlsx(path, lot, summary_cards, detail_cards)
            self.label_chrono_status.setText(f"Export XLSX chronométrie + tensions créé : {Path(path).name}")
            QMessageBox.information(self.window, "Export XLSX lot", f"Export XLSX créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export XLSX lot", f"Export impossible : {exc}")

    def chrono_export_current_lot_pdf(self):
        lot = self.chrono_current_export_lot()
        if not lot:
            return
        try:
            records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "Export PDF lot", f"Lecture base impossible : {exc}")
            return
        if not records and not voltage_records:
            QMessageBox.information(self.window, "Export PDF lot", f"Aucune mesure chronométrie ou tension enregistrée pour le lot {lot}.")
            return
        default_name = f"chronometrie_et_tensions_lot_{self.filename_safe(lot)}.pdf"
        path = self.ask_export_path("Exporter chronométrie et tensions en PDF", default_name, "PDF (*.pdf)", ".pdf")
        if not path:
            return
        try:
            summary_cards, detail_cards = self.chrono_export_measure_sheets(records, voltage_records)
            self.write_chrono_lot_pdf(path, lot, records, summary_cards, detail_cards, voltage_records)
            self.label_chrono_status.setText(f"Export PDF chronométrie + tensions créé : {Path(path).name}")
            QMessageBox.information(self.window, "Export PDF lot", f"Rapport PDF créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export PDF lot", f"Création PDF impossible : {exc}")

    def xlsx_xml_escape(self, value):
        text = "" if value is None else str(value)
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    def xlsx_col_name(self, col_index):
        name = ""
        col_index += 1
        while col_index:
            col_index, rem = divmod(col_index - 1, 26)
            name = chr(65 + rem) + name
        return name

    def write_chrono_lot_xlsx(self, path, lot, summary_cards, detail_cards):
        """Crée les deux feuilles opérateur demandées dans un même classeur."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        workbook = Workbook()
        first = workbook.active
        first.title = "Synthèse globale"
        second = workbook.create_sheet("Détail tensions")

        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill("solid", fgColor="17365D")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        result_fill = PatternFill("solid", fgColor="E2F0D9")

        def fill_sheet(sheet, sheet_title, cards):
            sheet.sheet_view.showGridLines = False
            sheet.column_dimensions["A"].width = 48
            sheet.column_dimensions["B"].width = 22
            sheet.column_dimensions["C"].width = 22
            sheet.merge_cells("A1:C1")
            cell = sheet["A1"]
            cell.value = f"{sheet_title} - Lot {lot}"
            cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
            cell.fill = title_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 24
            row_index = 3
            for card_index, card in enumerate(cards, start=1):
                if card_index > 1:
                    row_index += 1
                for label, sanction, value in card:
                    sheet.cell(row_index, 1, label)
                    sheet.cell(row_index, 2, sanction)
                    sheet.cell(row_index, 3, value)
                    for col in range(1, 4):
                        c = sheet.cell(row_index, col)
                        c.font = Font(name="Arial", size=10, bold=(label in ("Lot", "Numéro de Relais")))
                        c.border = border
                        c.alignment = Alignment(vertical="center", wrap_text=True)
                    if label in ("Résultat chronométrie", "Résultat tensions"):
                        for col in range(1, 4):
                            sheet.cell(row_index, col).fill = result_fill
                    if label.startswith("Tension d'"):
                        for col in range(1, 4):
                            sheet.cell(row_index, col).fill = section_fill
                    if isinstance(value, (int, float)):
                        sheet.cell(row_index, 3).number_format = "0.000"
                    row_index += 1
            sheet.freeze_panes = "A3"
            sheet.print_title_rows = "1:2"
            sheet.page_setup.orientation = "portrait"
            sheet.page_setup.fitToWidth = 1
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.print_options.horizontalCentered = True

        fill_sheet(first, "Synthèse globale", summary_cards)
        fill_sheet(second, "Détail tensions", detail_cards)
        workbook.save(path)

    def write_chrono_lot_pdf(self, path, lot, records, summary_cards, detail_cards, voltage_records=None):
        writer = QPdfWriter(path)
        writer.setPageSize(QPageSize(QPageSize.A4))
        writer.setResolution(96)
        painter = QPainter(writer)
        if not painter.isActive():
            raise RuntimeError("QPainter PDF non actif.")
        try:
            page_w = writer.width()
            page_h = writer.height()
            margin = 34
            y = margin
            line_h = 17
            normal = QFont("Arial", 9)
            bold = QFont("Arial", 9)
            bold.setBold(True)
            title_font = QFont("Arial", 15)
            title_font.setBold(True)

            def draw_text(x, yy, value, font=None, max_chars=None):
                painter.setFont(font or normal)
                value = str(value if value is not None else "")
                if max_chars and len(value) > max_chars:
                    value = value[: max_chars - 3] + "..."
                painter.drawText(int(x), int(yy), value)

            def new_page():
                nonlocal y
                writer.newPage()
                y = margin

            def draw_section(section_title, cards, force_new_page=False):
                nonlocal y
                if force_new_page:
                    new_page()
                painter.setFont(title_font)
                painter.drawText(margin, y, f"{section_title} - Lot {lot}")
                y += line_h * 2
                draw_text(margin, y, f"Généré le : {time.strftime('%d/%m/%Y %H:%M')}", normal)
                y += line_h * 2
                x_label = margin
                x_sanction = margin + 330
                x_value = margin + 455
                for card in cards:
                    if y + line_h * (len(card) + 2) > page_h - margin:
                        new_page()
                        draw_text(margin, y, f"{section_title} - Lot {lot} (suite)", bold)
                        y += line_h * 2
                    for label, sanction, value in card:
                        row_font = bold if label in ("Lot", "Numéro de Relais") else normal
                        draw_text(x_label, y, label, row_font, 48)
                        draw_text(x_sanction, y, sanction, normal, 16)
                        draw_text(x_value, y, value, normal, 18)
                        y += line_h
                    painter.drawLine(margin, y - 9, page_w - margin, y - 9)
                    y += line_h

            draw_section("FEUILLE 1 - SYNTHÈSE GLOBALE", summary_cards)
            draw_section("FEUILLE 2 - DÉTAIL TENSIONS", detail_cards, force_new_page=True)
        finally:
            painter.end()

    def database_backup_raw_files(self, reason_slug="avant_recreation"):
        """Copie brute de la base et de ses fichiers WAL/SHM si présents."""
        return self.database_backup_raw_files_for(self.production_db_file, "production_essais", reason_slug)

    def database_backup_raw_files_for(self, db_file, prefix, reason_slug="avant_recreation"):
        """Copie brute d'une base et de ses fichiers WAL/SHM si présents."""
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        backup_main = db_file.with_name(f"{prefix}_{reason_slug}_{timestamp}.sqlite3")
        copied = []
        sources = [
            (db_file, backup_main),
            (Path(str(db_file) + "-wal"), Path(str(backup_main) + "-wal")),
            (Path(str(db_file) + "-shm"), Path(str(backup_main) + "-shm")),
        ]
        for src, dst in sources:
            if src.exists() and src.stat().st_size > 0:
                shutil.copy2(src, dst)
                copied.append(dst)
        return copied

    def database_remove_active_files(self):
        self.database_remove_files_for(self.production_db_file)

    def database_remove_files_for(self, db_file):
        gc.collect()
        for path in (
            db_file,
            Path(str(db_file) + "-wal"),
            Path(str(db_file) + "-shm"),
        ):
            try:
                if path.exists():
                    path.unlink()
            except PermissionError as exc:
                raise RuntimeError(
                    f"Impossible de supprimer {path.name} : fichier verrouillé par Windows. "
                    "Fermer toute autre instance du logiciel, tout outil SQLite ou tout explorateur qui utilise cette base, puis réessayer. "
                    f"Détail : {exc}"
                )
            except Exception as exc:
                raise RuntimeError(f"Impossible de supprimer {path.name} : {exc}")

    def database_verify_integrity(self):
        with self.production_connect_db() as con:
            row = con.execute("PRAGMA quick_check").fetchone()
            result = row[0] if row is not None else "réponse vide"
            if str(result).lower() != "ok":
                raise RuntimeError(f"PRAGMA quick_check : {result}")

    def database_external_has_table(self, con, table_name):
        row = con.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        ).fetchone()
        return row is not None

    def database_validate_merge_source(self, path):
        path = Path(path)
        if not path.exists():
            raise RuntimeError("Fichier source introuvable.")
        active_db = self.database_admin_active_db_file()
        if path.resolve() == active_db.resolve():
            raise RuntimeError("La base source est la base active. Choisir une autre base à fusionner.")
        if self.database_admin_is_chrono():
            return self.chrono_database_validate_merge_source(path)
        required = ["lot", "sn", "designation", "operateur", "date", "heure", "scenario", "resultat", "details", "timestamp"]
        with self.database_connect_file(path) as con:
            row = con.execute("PRAGMA quick_check").fetchone()
            result = row[0] if row is not None else "réponse vide"
            if str(result).lower() != "ok":
                raise RuntimeError(f"Base source non saine : PRAGMA quick_check = {result}")
            if not self.database_external_has_table(con, "essais"):
                raise RuntimeError("La base source ne contient pas la table essais.")
            cols = {str(r["name"]) for r in con.execute("PRAGMA table_info(essais)").fetchall()}
            missing = [name for name in required if name not in cols]
            if missing:
                raise RuntimeError("Colonnes manquantes dans essais : " + ", ".join(missing))
            nb_essais = con.execute("SELECT COUNT(*) AS n FROM essais").fetchone()["n"]
            nb_ops = 0
            if self.database_external_has_table(con, "operators"):
                nb_ops = con.execute("SELECT COUNT(*) AS n FROM operators").fetchone()["n"]
        return nb_essais, nb_ops

    def chrono_database_validate_merge_source(self, path):
        with self.database_connect_file(path) as con:
            row = con.execute("PRAGMA quick_check").fetchone()
            result = row[0] if row is not None else "réponse vide"
            if str(result).lower() != "ok":
                raise RuntimeError(f"Base source non saine : PRAGMA quick_check = {result}")
            if not self.database_external_has_table(con, "mesures_chrono_contacts"):
                raise RuntimeError("La base source ne contient pas la table mesures_chrono_contacts.")
            required = ["lot", "date_test", "relais", "ambiance_c", "nom_test", "sn", "action", "resultat", "timestamp"]
            cols = {str(r["name"]) for r in con.execute("PRAGMA table_info(mesures_chrono_contacts)").fetchall()}
            missing = [name for name in required if name not in cols]
            if missing:
                raise RuntimeError("Colonnes manquantes dans mesures_chrono_contacts : " + ", ".join(missing))
            nb_chrono = con.execute("SELECT COUNT(*) AS n FROM mesures_chrono_contacts").fetchone()["n"]
            nb_tensions = 0
            if self.database_external_has_table(con, "mesures_tension_fonctionnement"):
                voltage_cols = {str(r["name"]) for r in con.execute("PRAGMA table_info(mesures_tension_fonctionnement)").fetchall()}
                voltage_required = {"lot", "date_test", "relais", "nom_test", "sn", "relay_type", "resultat", "timestamp"}
                if not voltage_required.issubset(voltage_cols):
                    raise RuntimeError("Table mesures_tension_fonctionnement incomplète dans la base source.")
                nb_tensions = con.execute("SELECT COUNT(*) AS n FROM mesures_tension_fonctionnement").fetchone()["n"]
        return nb_chrono, nb_tensions

    def database_merge_clicked(self):
        if getattr(self, "auto_neutral_running", False):
            self.big_message_box(
                "Fusion base impossible",
                "ESSAI EN COURS",
                "La fusion d'une base est interdite pendant un essai Neutral Screen Automatique.\n\nArrêtez ou terminez l'essai avant cette opération.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return
        path, _ = QFileDialog.getOpenFileName(
            self.window,
            "Importer / fusionner une base SQLite",
            str(self.database_admin_active_db_file().parent),
            "SQLite (*.sqlite3);;Tous les fichiers (*)",
        )
        if not path:
            return
        try:
            nb_essais, nb_ops = self.database_validate_merge_source(path)
        except Exception as exc:
            QMessageBox.warning(self.window, "Fusion base", f"Base source non utilisable : {exc}")
            return
        if not self.big_message_box(
            "Importer / fusionner une base",
            "FUSIONNER CETTE BASE ?",
            (
                f"Base source :\n{path}\n\n"
                f"Contenu détecté :\n- {nb_essais} mesure(s) chronométrie\n- {nb_ops} mesure(s) de tension\n\n"
                "La base chronométrie active sera sauvegardée avant fusion.\n"
                "Les mesures déjà présentes ne seront pas dupliquées."
            )
            if self.database_admin_is_chrono() else
            (
                f"Base source :\n{path}\n\nContenu détecté :\n- {nb_essais} essai(s) relais\n- {nb_ops} opérateur(s) dans la table operators\n\n"
                "La base active sera sauvegardée avant fusion.\nLes essais déjà présents ne seront pas dupliqués.\n"
                "Le mot de passe et les réglages de la base active ne seront pas remplacés."
            ),
            ok_text="OUI, FUSIONNER",
            cancel_text="ANNULER",
            icon=QMessageBox.Question,
        ):
            return
        if self.database_admin_is_chrono():
            self.chrono_database_merge_from_file(Path(path), nb_mesures_source=nb_essais, nb_tensions_source=nb_ops)
            return
        self.database_merge_from_file(Path(path), nb_essais_source=nb_essais)

    def database_merge_from_file(self, source_path, nb_essais_source=None):
        required = ["lot", "sn", "designation", "operateur", "date", "heure", "scenario", "resultat", "details", "timestamp"]
        try:
            backup_files = self.database_backup_raw_files("avant_fusion")
            self.production_init_db()
            inserted = 0
            skipped = 0
            operators_added = 0
            with self.database_connect_file(source_path) as source, self.production_connect_db() as dest:
                source_cols = {str(r["name"]) for r in source.execute("PRAGMA table_info(essais)").fetchall()}
                nb_inverseurs_expr = "nb_inverseurs" if "nb_inverseurs" in source_cols else "2 AS nb_inverseurs"
                rows = source.execute(
                    f"SELECT lot, sn, designation, {nb_inverseurs_expr}, operateur, date, heure, scenario, resultat, details, timestamp "
                    "FROM essais ORDER BY lot COLLATE NOCASE ASC, timestamp ASC, sn ASC"
                ).fetchall()
                source_operator_names = set()
                if self.database_external_has_table(source, "operators"):
                    for op in source.execute("SELECT name FROM operators ORDER BY name COLLATE NOCASE ASC").fetchall():
                        name = str(op["name"] or "").strip()
                        if name:
                            source_operator_names.add(name)
                for row in rows:
                    op_name = str(row["operateur"] or "").strip()
                    if op_name:
                        source_operator_names.add(op_name)
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
                for name in sorted(source_operator_names, key=str.lower):
                    before = dest.total_changes
                    dest.execute(
                        "INSERT OR IGNORE INTO operators(name, active, created_at) VALUES(?, 1, ?)",
                        (name, timestamp),
                    )
                    if dest.total_changes > before:
                        operators_added += 1
                duplicate_query = """
                    SELECT id FROM essais
                    WHERE lot = ? AND sn = ? AND designation = ? AND operateur = ?
                      AND date = ? AND heure = ? AND scenario = ? AND resultat = ?
                      AND details = ? AND timestamp = ?
                    LIMIT 1
                """
                insert_query = """
                    INSERT INTO essais(lot, sn, designation, nb_inverseurs, operateur, date, heure, scenario, resultat, details, timestamp)
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                for row in rows:
                    values = tuple(str(row[col] or "") for col in required)
                    exists = dest.execute(duplicate_query, values).fetchone()
                    if exists is not None:
                        skipped += 1
                        continue
                    insert_values = (
                        str(row["lot"] or ""),
                        str(row["sn"] or ""),
                        str(row["designation"] or ""),
                        self.clamp_nb_inverseurs(row["nb_inverseurs"]),
                        str(row["operateur"] or ""),
                        str(row["date"] or ""),
                        str(row["heure"] or ""),
                        str(row["scenario"] or ""),
                        str(row["resultat"] or ""),
                        str(row["details"] or ""),
                        str(row["timestamp"] or ""),
                    )
                    dest.execute(insert_query, insert_values)
                    inserted += 1
            self.database_load_operators_to_combo()
            self.production_refresh_table()
            self.database_admin_refresh()
            self.label_prod_status.setText(f"Fusion base OK : {inserted} essai(s) importé(s), {skipped} doublon(s) ignoré(s).")
            details = (
                f"Fusion terminée.\n\n"
                f"Essais source : {nb_essais_source if nb_essais_source is not None else inserted + skipped}\n"
                f"Essais importés : {inserted}\n"
                f"Doublons ignorés : {skipped}\n"
                f"Opérateurs ajoutés : {operators_added}"
            )
            if backup_files:
                details += "\n\nSauvegarde avant fusion :\n" + "\n".join(str(p) for p in backup_files)
            QMessageBox.information(self.window, "Fusion base", details)
        except Exception as exc:
            QMessageBox.warning(self.window, "Fusion base", f"Fusion impossible : {exc}")
            self.label_prod_status.setText(f"Fusion base impossible : {exc}")

    def chrono_database_merge_from_file(self, source_path, nb_mesures_source=None, nb_tensions_source=None):
        chrono_headers = [
            "lot", "date_test", "relais", "ambiance_c", "nom_test", "sn",
            "relay_type", "action", "nb_inverseurs", "capture_ms", "pulse_ms",
            "limite_temps_ms", "limite_rebond_ms", "resultat", "overflow",
            "details_json", "events_json", "timestamp",
        ]
        try:
            backup_files = self.database_backup_raw_files_for(self.chrono_db_file, "chronometrie_contacts", "avant_fusion")
            self.voltage_init_db()
            inserted = skipped = inserted_voltage = skipped_voltage = 0
            with self.database_connect_file(source_path) as source, self.chrono_connect_db() as dest:
                source_cols = {str(r["name"]) for r in source.execute("PRAGMA table_info(mesures_chrono_contacts)").fetchall()}
                select_parts = []
                for name in chrono_headers:
                    if name in source_cols:
                        select_parts.append(name)
                    elif name == "relay_type":
                        select_parts.append("'' AS relay_type")
                    elif name in ("nb_inverseurs", "capture_ms", "pulse_ms", "overflow"):
                        select_parts.append(f"0 AS {name}")
                    elif name in ("limite_temps_ms", "limite_rebond_ms"):
                        select_parts.append(f"0.0 AS {name}")
                    else:
                        select_parts.append(f"'' AS {name}")
                rows = source.execute("SELECT " + ", ".join(select_parts) + " FROM mesures_chrono_contacts ORDER BY timestamp ASC, id ASC").fetchall()
                duplicate_query = """
                    SELECT id FROM mesures_chrono_contacts
                    WHERE lot=? AND date_test=? AND relais=? AND nom_test=?
                      AND sn=? AND relay_type=? AND action=? AND timestamp=? LIMIT 1
                """
                insert_query = "INSERT INTO mesures_chrono_contacts(" + ",".join(chrono_headers) + ") VALUES(" + ",".join("?" for _ in chrono_headers) + ")"
                for row in rows:
                    duplicate_values=(str(row["lot"] or ""),str(row["date_test"] or ""),str(row["relais"] or ""),str(row["nom_test"] or ""),str(row["sn"] or ""),str(row["relay_type"] or ""),str(row["action"] or ""),str(row["timestamp"] or ""))
                    if dest.execute(duplicate_query,duplicate_values).fetchone() is not None:
                        skipped += 1; continue
                    dest.execute(insert_query,tuple(row[h] for h in chrono_headers)); inserted += 1

                if self.database_external_has_table(source, "mesures_tension_fonctionnement"):
                    source_voltage_cols = [str(r["name"]) for r in source.execute("PRAGMA table_info(mesures_tension_fonctionnement)").fetchall() if str(r["name"]) != "id"]
                    dest_voltage_cols = {str(r["name"]) for r in dest.execute("PRAGMA table_info(mesures_tension_fonctionnement)").fetchall()}
                    common = [name for name in source_voltage_cols if name in dest_voltage_cols]
                    if common:
                        voltage_rows = source.execute("SELECT " + ",".join(common) + " FROM mesures_tension_fonctionnement ORDER BY timestamp ASC, id ASC").fetchall()
                        voltage_insert = "INSERT INTO mesures_tension_fonctionnement(" + ",".join(common) + ") VALUES(" + ",".join("?" for _ in common) + ")"
                        for row in voltage_rows:
                            exists = dest.execute("""
                                SELECT id FROM mesures_tension_fonctionnement
                                WHERE lot=? AND sn=? AND relais=? AND nom_test=? AND timestamp=?
                                LIMIT 1
                            """, (str(row["lot"] or ""),str(row["sn"] or ""),str(row["relais"] or ""),str(row["nom_test"] or ""),str(row["timestamp"] or ""))).fetchone()
                            if exists is not None:
                                skipped_voltage += 1; continue
                            dest.execute(voltage_insert, tuple(row[name] for name in common)); inserted_voltage += 1
            self.database_admin_refresh()
            self.label_prod_status.setText(f"Fusion OK : {inserted} chrono + {inserted_voltage} tension(s), doublons {skipped + skipped_voltage}.")
            details = (
                "Fusion chronométrie + tensions terminée.\n\n"
                f"Chronométrie source : {nb_mesures_source if nb_mesures_source is not None else inserted + skipped}\n"
                f"Chronométrie importée : {inserted}\n"
                f"Tensions source : {nb_tensions_source if nb_tensions_source is not None else inserted_voltage + skipped_voltage}\n"
                f"Tensions importées : {inserted_voltage}\n"
                f"Doublons ignorés : {skipped + skipped_voltage}"
            )
            if backup_files:
                details += "\n\nSauvegarde avant fusion :\n" + "\n".join(str(p) for p in backup_files)
            QMessageBox.information(self.window, "Fusion base", details)
        except Exception as exc:
            QMessageBox.warning(self.window, "Fusion base", f"Fusion chronométrie/tensions impossible : {exc}")
            self.label_prod_status.setText(f"Fusion impossible : {exc}")

    def database_recreate_default_clicked(self):
        if getattr(self, "auto_neutral_running", False):
            self.big_message_box(
                "Base par défaut impossible",
                "ESSAI EN COURS",
                "La base ne peut pas être recréée pendant un essai Neutral Screen Automatique.\n\nArrêtez ou terminez l'essai avant cette opération.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return
        is_chrono = self.database_admin_is_chrono()
        base_label = "chronométrie contacts" if is_chrono else "production Neutral Screen"
        if not self.big_message_box(
            "Recréer base par défaut",
            "RECRÉER UNE BASE VIDE ?",
            f"Attention : cette opération crée une base {base_label} neuve vide.\n\n"
            "Les anciens enregistrements ne seront pas récupérés dans la nouvelle base.\n"
            "Une sauvegarde de l'ancien fichier sera faite si possible avant remplacement.\n\nContinuer ?",
            ok_text="OUI, RECRÉER",
            cancel_text="ANNULER",
            icon=QMessageBox.Warning,
        ):
            return
        if is_chrono:
            self.chrono_database_recreate_default(reason_slug="avant_recreation_manuelle", show_success=True)
            return
        self.database_recreate_default(reason_slug="avant_recreation_manuelle", show_success=True)

    def database_recreate_default(self, reason_slug="avant_recreation", show_success=False):
        backup_files = []
        access_code = str(getattr(self, "current_access_code", LOCK_ACCESS_CODE) or LOCK_ACCESS_CODE)
        try:
            backup_files = self.database_backup_raw_files(reason_slug)
            self.database_remove_active_files()
            self.production_data = {"version": "2.12.3", "last_context": {}, "records": [], "access_code": access_code}
            self.current_access_code = access_code
            self.production_init_db()
            self.production_set_setting("access_code", access_code)
            self.production_set_setting("last_context", "{}")
            self.production_set_setting("legacy_json_migrated", "1")
            self.database_load_operators_to_combo()
            self.production_clear_entry_fields()
            self.production_refresh_table()
            self.database_admin_refresh()
            self.label_prod_status.setText("Base par défaut recréée : base vide prête.")
            if show_success:
                details = "Base neuve vide créée."
                if backup_files:
                    details += "\n\nSauvegarde créée :\n" + "\n".join(str(p) for p in backup_files)
                else:
                    details += "\n\nAucun ancien fichier à sauvegarder."
                QMessageBox.information(self.window, "Base recréée", details)
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, "Base par défaut", f"Recréation impossible : {exc}")
            self.label_prod_status.setText(f"Base par défaut non recréée : {exc}")
            return False

    def chrono_database_recreate_default(self, reason_slug="avant_recreation", show_success=False):
        backup_files = []
        try:
            backup_files = self.database_backup_raw_files_for(self.chrono_db_file, "chronometrie_contacts", reason_slug)
            self.database_remove_files_for(self.chrono_db_file)
            self.voltage_init_db()
            self.database_admin_refresh()
            self.label_prod_status.setText("Base chronométrie recréée : base vide prête.")
            if show_success:
                details = "Base chronométrie neuve vide créée."
                if backup_files:
                    details += "\n\nSauvegarde créée :\n" + "\n".join(str(p) for p in backup_files)
                else:
                    details += "\n\nAucun ancien fichier à sauvegarder."
                QMessageBox.information(self.window, "Base recréée", details)
            return True
        except Exception as exc:
            QMessageBox.critical(self.window, "Base chronométrie", f"Recréation impossible : {exc}")
            self.label_prod_status.setText(f"Base chronométrie non recréée : {exc}")
            return False

    def database_handle_unreadable_startup(self, exc):
        if getattr(self, "_database_recovery_prompt_shown", False):
            self.label_prod_status.setText(f"Base production SQLite non lue : {exc}")
            return
        self._database_recovery_prompt_shown = True
        if self.big_message_box(
            "Base production illisible",
            "BASE PRODUCTION ILLISIBLE",
            f"Le fichier de base existe mais il ne peut pas être lu correctement.\n\nDétail : {exc}\n\nLe logiciel peut sauvegarder le fichier défectueux si possible, puis recréer une base par défaut vide.",
            ok_text="SAUVEGARDER ET RECRÉER",
            cancel_text="NE PAS TOUCHER",
            icon=QMessageBox.Critical,
        ):
            self.database_recreate_default(reason_slug="base_illisible", show_success=True)
        else:
            self.label_prod_status.setText(f"Base illisible non remplacée : {exc}")

    def database_selected_lot(self):
        row = self.tableWidget_db_lots.currentRow()
        if row < 0:
            return None
        item = self.tableWidget_db_lots.item(row, 0)
        if item is None:
            return None
        raw = item.data(Qt.UserRole + 1)
        if raw is not None:
            return str(raw)
        text = item.text().strip()
        return "" if text == "(lot vide)" else text

    def chrono_database_selected_group(self):
        row = self.tableWidget_db_lots.currentRow()
        if row < 0:
            return None
        item = self.tableWidget_db_lots.item(row, 0)
        if item is None:
            return None
        return {
            "lot": str(item.data(Qt.UserRole + 1) or item.text().strip()),
            "relais": str(item.data(Qt.UserRole + 2) or ""),
            "nom_test": str(item.data(Qt.UserRole + 3) or ""),
        }

    def database_lot_display(self, lot):
        return str(lot) if str(lot or "").strip() else "(lot vide)"

    def database_open_selected_lot_details(self):
        if self.database_admin_is_chrono():
            self.chrono_database_open_selected_details()
            return
        lot = self.database_selected_lot()
        if lot is None:
            QMessageBox.information(self.window, "Détail lot", "Sélectionner un lot.")
            return
        records = self.production_records_for_lot(lot)
        if not records:
            QMessageBox.information(self.window, "Détail lot", f"Aucun relais trouvé pour le lot {self.database_lot_display(lot)}.")
            return
        self.production_show_lot_details(lot, records)

    def database_export_selected_lot_pdf(self):
        if self.database_admin_is_chrono():
            group = self.chrono_database_selected_group()
            if group is None:
                QMessageBox.information(self.window, "PDF lot", "Sélectionner un lot ou groupe chronométrie.")
                return
            self.chrono_export_lot_pdf_from_database(group["lot"])
            return
        lot = self.database_selected_lot()
        if lot is None:
            QMessageBox.information(self.window, "PDF lot", "Sélectionner un lot.")
            return
        self.production_export_pdf_for_lot(lot)

    def database_export_selected_lot_xlsx(self):
        if self.database_admin_is_chrono():
            group = self.chrono_database_selected_group()
            if group is None:
                QMessageBox.information(self.window, "XLSX lot", "Sélectionner un lot ou groupe chronométrie.")
                return
            self.chrono_export_lot_xlsx_from_database(group["lot"])
            return
        lot = self.database_selected_lot()
        if lot is None:
            QMessageBox.information(self.window, "XLSX lot", "Sélectionner un lot.")
            return
        try:
            records = self.production_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "XLSX lot", f"Lecture base impossible : {exc}")
            return
        if not records:
            QMessageBox.information(self.window, "XLSX lot", f"Aucun essai trouvé pour le lot {self.database_lot_display(lot)}.")
            return
        headers = [
            ("lot", "Lot"), ("sn", "SN"), ("designation", "Désignation"),
            ("nb_inverseurs", "Inv."), ("operateur", "Opérateur"), ("date", "Date"),
            ("heure", "Heure"), ("scenario", "Scénario"), ("resultat", "Résultat"),
            ("details", "Détails"), ("timestamp", "Horodatage"),
        ]
        path = self.ask_export_path(
            "Exporter le lot en XLSX",
            f"production_lot_{self.filename_safe(lot)}.xlsx",
            "Excel (*.xlsx)",
            ".xlsx",
        )
        if not path:
            return
        try:
            self.write_table_xlsx(path, f"Production lot {lot}", headers, records)
            QMessageBox.information(self.window, "XLSX lot", f"Export XLSX créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "XLSX lot", f"Export impossible : {exc}")

    def chrono_export_lot_xlsx_from_database(self, lot):
        try:
            records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "XLSX lot", f"Lecture base impossible : {exc}")
            return
        if not records and not voltage_records:
            QMessageBox.information(self.window, "XLSX lot", f"Aucune mesure chronométrie ou tension trouvée pour le lot {self.database_lot_display(lot)}.")
            return
        path = self.ask_export_path("Exporter chronométrie et tensions en XLSX", f"chronometrie_et_tensions_lot_{self.filename_safe(lot)}.xlsx", "Excel (*.xlsx)", ".xlsx")
        if not path:
            return
        try:
            self.write_chrono_lot_xlsx(path, lot, self.chrono_export_measure_cards(records, voltage_records))
            QMessageBox.information(self.window, "XLSX lot", f"Export XLSX créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "XLSX lot", f"Export impossible : {exc}")

    def chrono_export_lot_pdf_from_database(self, lot):
        try:
            records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "PDF lot", f"Lecture base impossible : {exc}")
            return
        if not records and not voltage_records:
            QMessageBox.information(self.window, "PDF lot", f"Aucune mesure chronométrie ou tension trouvée pour le lot {self.database_lot_display(lot)}.")
            return
        path = self.ask_export_path("Exporter chronométrie et tensions en PDF", f"chronometrie_et_tensions_lot_{self.filename_safe(lot)}.pdf", "PDF (*.pdf)", ".pdf")
        if not path:
            return
        try:
            summary_cards, detail_cards = self.chrono_export_measure_sheets(records, voltage_records)
            self.write_chrono_lot_pdf(path, lot, records, summary_cards, detail_cards, voltage_records)
            QMessageBox.information(self.window, "PDF lot", f"Export PDF créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "PDF lot", f"Création PDF impossible : {exc}")

    def database_delete_selected_lot(self):
        if getattr(self, "auto_neutral_running", False):
            self.big_message_box(
                "Suppression lot impossible",
                "ESSAI EN COURS",
                "La suppression d'un lot est interdite pendant un essai Neutral Screen Automatique.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return
        if self.database_admin_is_chrono():
            self.chrono_database_delete_selected_group()
            return
        lot = self.database_selected_lot()
        if lot is None:
            QMessageBox.information(self.window, "Supprimer lot", "Sélectionner un lot à supprimer.")
            return
        try:
            with self.production_connect_db() as con:
                count = con.execute("SELECT COUNT(*) AS n FROM essais WHERE lot = ?", (lot,)).fetchone()["n"]
        except Exception as exc:
            QMessageBox.warning(self.window, "Supprimer lot", f"Lecture impossible : {exc}")
            return
        lot_label = self.database_lot_display(lot)
        if count <= 0:
            QMessageBox.information(self.window, "Supprimer lot", f"Aucun essai trouvé pour le lot {lot_label}.")
            return
        if not self.big_message_box(
            "Supprimer lot",
            "SUPPRIMER CE LOT ?",
            f"Lot : {lot_label}\n\n"
            f"Cette opération supprimera définitivement {count} relais/essai(s) de la base active.\n"
            "Une sauvegarde de sécurité sera faite avant suppression.\n\n"
            "Continuer ?",
            ok_text="OUI, SUPPRIMER",
            cancel_text="ANNULER",
            icon=QMessageBox.Warning,
        ):
            return
        try:
            backup_files = self.database_backup_raw_files("avant_suppression_lot")
            with self.production_connect_db() as con:
                con.execute("DELETE FROM essais WHERE lot = ?", (lot,))
            self.production_refresh_table()
            self.database_admin_refresh()
            self.label_prod_status.setText(f"Lot supprimé : {lot_label} ({count} essai(s)).")
            details = f"Lot supprimé : {lot_label}\nEssais supprimés : {count}"
            if backup_files:
                details += "\n\nSauvegarde avant suppression :\n" + "\n".join(str(p) for p in backup_files)
            QMessageBox.information(self.window, "Supprimer lot", details)
        except Exception as exc:
            QMessageBox.warning(self.window, "Supprimer lot", f"Suppression impossible : {exc}")

    def chrono_database_open_selected_details(self):
        group = self.chrono_database_selected_group()
        if group is None:
            QMessageBox.information(self.window, "Détail chronométrie", "Sélectionner un groupe de mesures.")
            return
        try:
            self.voltage_init_db()
            with self.chrono_connect_db() as con:
                chrono = con.execute("""
                    SELECT timestamp, date_test, sn, relay_type, action, nb_inverseurs, capture_ms, pulse_ms,
                           resultat, overflow, ambiance_c, details_json
                    FROM mesures_chrono_contacts
                    WHERE lot=? AND relais=? AND nom_test=?
                """, (group["lot"], group["relais"], group["nom_test"])).fetchall()
                volts = con.execute("""
                    SELECT timestamp, date_test, sn, relay_type, nb_inverseurs, pickup_global_v, dropout_global_v,
                           resultat, ambiance_c, pickup_json, dropout_json
                    FROM mesures_tension_fonctionnement
                    WHERE lot=? AND relais=? AND nom_test=?
                """, (group["lot"], group["relais"], group["nom_test"])).fetchall()
        except Exception as exc:
            QMessageBox.warning(self.window, "Détail chronométrie", f"Lecture impossible : {exc}")
            return
        combined=[]
        for r in chrono:
            combined.append((str(r["timestamp"] or ""), [self.format_datetime_fr(r["timestamp"]) or r["date_test"], "CHRONO", r["sn"], r["relay_type"], r["action"], r["nb_inverseurs"], r["capture_ms"], r["pulse_ms"], "", "", r["resultat"], "Oui" if int(r["overflow"] or 0) else "Non", r["ambiance_c"], r["details_json"]]))
        for r in volts:
            detail=json.dumps({"pickup":self.chrono_json_dict(r["pickup_json"]),"dropout":self.chrono_json_dict(r["dropout_json"])}, ensure_ascii=False)
            combined.append((str(r["timestamp"] or ""), [self.format_datetime_fr(r["timestamp"]) or r["date_test"], "TENSION", r["sn"], r["relay_type"], "BE/BR", r["nb_inverseurs"], "", "", self.chrono_export_voltage_value(r["pickup_global_v"]), self.chrono_export_voltage_value(r["dropout_global_v"]), r["resultat"], "", r["ambiance_c"], detail]))
        combined.sort(key=lambda x:x[0])
        if not combined:
            QMessageBox.information(self.window, "Détail chronométrie", "Aucune mesure trouvée pour cette sélection.")
            return
        dialog=QDialog(self.window); dialog.setWindowTitle(f"Chronométrie et tensions - Lot {self.database_lot_display(group['lot'])}"); dialog.resize(1350,560)
        layout=QVBoxLayout(dialog); title=QLabel(f"Lot {self.database_lot_display(group['lot'])} - Relais {group['relais'] or '-'} - Test {group['nom_test'] or '-'}"); title.setStyleSheet("font-size: 14pt; font-weight: bold; color: black;"); layout.addWidget(title)
        table=QTableWidget(dialog); headers=["Date","Source","SN","Type","Action","Inv.","Capture","Pulse","BE V","BR V","Résultat","Overflow","Ambiance","Détails"]; table.setColumnCount(len(headers)); table.setHorizontalHeaderLabels(headers); table.setRowCount(len(combined)); table.setEditTriggers(QTableWidget.NoEditTriggers); table.setSelectionBehavior(QTableWidget.SelectRows)
        for row,(_ts,vals) in enumerate(combined):
            for col,val in enumerate(vals): table.setItem(row,col,self.table_item(val))
        table.horizontalHeader().setStretchLastSection(True); layout.addWidget(table); close_button=QPushButton("Fermer"); close_button.clicked.connect(dialog.accept); layout.addWidget(close_button); dialog.exec()

    def chrono_database_delete_selected_group(self):
        group = self.chrono_database_selected_group()
        if group is None:
            QMessageBox.information(self.window, "Supprimer mesures", "Sélectionner un groupe de mesures à supprimer.")
            return
        try:
            self.voltage_init_db()
            with self.chrono_connect_db() as con:
                chrono_count=con.execute("SELECT COUNT(*) AS n FROM mesures_chrono_contacts WHERE lot=? AND relais=? AND nom_test=?", (group["lot"],group["relais"],group["nom_test"])).fetchone()["n"]
                voltage_count=con.execute("SELECT COUNT(*) AS n FROM mesures_tension_fonctionnement WHERE lot=? AND relais=? AND nom_test=?", (group["lot"],group["relais"],group["nom_test"])).fetchone()["n"]
        except Exception as exc:
            QMessageBox.warning(self.window, "Supprimer mesures", f"Lecture impossible : {exc}"); return
        total=int(chrono_count or 0)+int(voltage_count or 0)
        if total<=0:
            QMessageBox.information(self.window, "Supprimer mesures", "Aucune mesure trouvée pour cette sélection."); return
        label=f"Lot {self.database_lot_display(group['lot'])} / Relais {group['relais'] or '-'} / Test {group['nom_test'] or '-'}"
        if not self.big_message_box("Supprimer mesures", "SUPPRIMER CES MESURES ?", f"{label}\n\nChronométrie : {chrono_count}\nTensions : {voltage_count}\nTotal : {total}\n\nUne sauvegarde sera faite avant suppression.", ok_text="OUI, SUPPRIMER", cancel_text="ANNULER", icon=QMessageBox.Warning): return
        try:
            backup_files=self.database_backup_raw_files_for(self.chrono_db_file,"chronometrie_contacts","avant_suppression_mesures")
            with self.chrono_connect_db() as con:
                args=(group["lot"],group["relais"],group["nom_test"])
                con.execute("DELETE FROM mesures_chrono_contacts WHERE lot=? AND relais=? AND nom_test=?",args)
                con.execute("DELETE FROM mesures_tension_fonctionnement WHERE lot=? AND relais=? AND nom_test=?",args)
            self.database_admin_refresh(); self.label_prod_status.setText(f"Mesures supprimées : {total}.")
            details=f"Chronométrie supprimée : {chrono_count}\nTensions supprimées : {voltage_count}"
            if backup_files: details += "\n\nSauvegarde :\n"+"\n".join(str(p) for p in backup_files)
            QMessageBox.information(self.window,"Supprimer mesures",details)
        except Exception as exc:
            QMessageBox.warning(self.window,"Supprimer mesures",f"Suppression impossible : {exc}")

    def production_migrate_legacy_json_if_needed(self):
        if self.production_get_setting("legacy_json_migrated", "0") == "1":
            return
        if not self.production_legacy_json_file.exists():
            self.production_set_setting("legacy_json_migrated", "1")
            return
        try:
            with open(self.production_legacy_json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                self.production_set_setting("legacy_json_migrated", "1")
                return
            records = data.get("records", [])
            if not isinstance(records, list):
                records = []
            with self.production_connect_db() as con:
                count = con.execute("SELECT COUNT(*) AS n FROM essais").fetchone()["n"]
                if count == 0:
                    for record in records:
                        if not isinstance(record, dict):
                            continue
                        con.execute(
                            """
                            INSERT INTO essais(lot, sn, designation, nb_inverseurs, operateur, date, heure, scenario, resultat, details, timestamp)
                            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                str(record.get("lot", "") or ""),
                                str(record.get("sn", "") or ""),
                                str(record.get("designation", "") or ""),
                                self.clamp_nb_inverseurs(record.get("nb_inverseurs", 2)),
                                str(record.get("operateur", "") or ""),
                                str(record.get("date", "") or ""),
                                str(record.get("heure", "") or ""),
                                str(record.get("scenario", "") or ""),
                                str(record.get("resultat", "") or ""),
                                str(record.get("details", "") or ""),
                                str(record.get("timestamp", "") or ""),
                            ),
                        )
            if isinstance(data.get("last_context", {}), dict):
                self.production_set_setting("last_context", json.dumps(data.get("last_context", {}), ensure_ascii=False))
            if data.get("access_code"):
                self.production_set_setting("access_code", str(data.get("access_code")))
            self.production_set_setting("legacy_json_migrated", "1")
        except Exception as exc:
            self.label_prod_status.setText(f"Migration JSON production impossible : {exc}")

    def production_load_db(self):
        try:
            self.production_init_db()
            self.database_verify_integrity()
            self.production_migrate_legacy_json_if_needed()
            last_context_raw = self.production_get_setting("last_context", "{}")
            try:
                last_context = json.loads(last_context_raw)
                if not isinstance(last_context, dict):
                    last_context = {}
            except Exception:
                last_context = {}
            self.current_access_code = str(self.production_get_setting("access_code", LOCK_ACCESS_CODE) or LOCK_ACCESS_CODE)
            self.production_data = {
                "version": "2.12.3",
                "last_context": last_context,
                "records": [],
                "access_code": self.current_access_code,
            }
            self.database_load_operators_to_combo()
            self.production_clear_entry_fields()
        except Exception as exc:
            self.database_handle_unreadable_startup(exc)

    def production_save_db(self):
        self.production_init_db()
        self.production_data["version"] = "2.12.3"
        self.production_data["access_code"] = self.current_access_code
        self.production_set_setting("access_code", self.current_access_code)
        self.production_set_setting(
            "last_context",
            json.dumps(self.production_data.get("last_context", {}), ensure_ascii=False),
        )

    def production_save_context(self):
        self.production_data["last_context"] = self.production_context()
        self.production_save_db()
        self._auto_start_prompts_done = False
        self.set_auto_finish_validation_state(False)
        self.label_prod_status.setText(f"Nouvel essai production prêt : {self.production_db_file.name}")

    def production_missing_required_fields(self, include_sn=False):
        missing = []
        if not self.comboBox_prod_scenario.currentText().strip():
            missing.append("Scénario pour l'essai")
        if not self.lineEdit_prod_lot.text().strip():
            missing.append("Numéro de lot")
        if not self.lineEdit_prod_designation.text().strip():
            missing.append("Désignation")
        try:
            self.production_nb_inverseurs()
        except Exception:
            missing.append("Nombre d'inverseurs")
        if not self.comboBox_prod_operateur.currentText().strip():
            missing.append("Opérateur")
        if not self.dateEdit_prod_date.date().isValid():
            missing.append("Date")
        if include_sn and not self.lineEdit_SN.text().strip():
            missing.append("Numéro de SN")
        return missing

    def production_show_missing_fields_message(self, missing):
        detail = "Le test ne démarrera pas car les champs suivants ne sont pas remplis :\n\n"
        detail += "\n".join(f"- {name}" for name in missing)
        detail += "\n\nCompléter ces informations dans l'onglet Production avant de passer au test."
        self.big_message_box(
            "Informations production incomplètes",
            "ESSAI PRODUCTION INCOMPLET",
            detail,
            ok_text="COMPRIS",
            icon=QMessageBox.Warning,
        )

    def production_prepare_and_open_auto_test(self):
        missing = self.production_missing_required_fields()
        if missing:
            self.production_show_missing_fields_message(missing)
            self.label_prod_status.setText("Passage au test refusé : informations production incomplètes.")
            return
        self.production_sync_nb_inverseurs_to_auto()
        lot = self.lineEdit_prod_lot.text().strip()
        if getattr(self, "_lot_session_active", False) and not getattr(self, "_active_lot_finished", True) and lot != getattr(self, "_active_lot", ""):
            self.big_message_box(
                "Lot en cours",
                "LOT NON CLOTURE",
                f"Le lot {self._active_lot} est encore en cours.\n\n"
                f"Appuyez sur LOT FINI avant de préparer le lot {lot}.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return
        self.production_autofill_from_lot()
        self.production_save_context()
        self._lot_session_active = True
        self._active_lot_finished = False
        self._active_lot = lot
        tab = self.window.findChild(QWidget, "tab_neutral_auto")
        if tab is not None:
            self.set_tab_internal(tab)
        self.label_auto_status.setText(f"Production prête - lot {lot} en cours - lancer MARCHE AUTO")
        self.label_auto_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")

    def production_reload_base(self):
        self.production_load_db()
        self.production_refresh_table()
        self.database_admin_refresh()
        self.label_prod_status.setText("Base production rechargée.")

    def production_refresh_scenarios(self):
        if not hasattr(self, "comboBox_prod_scenario"):
            return
        current = self.comboBox_prod_scenario.currentText().strip()
        names = self.scenarios_names()
        self.comboBox_prod_scenario.clear()
        self.comboBox_prod_scenario.addItems(names)
        if current in names:
            self.comboBox_prod_scenario.setCurrentText(current)
        elif names:
            self.comboBox_prod_scenario.setCurrentIndex(0)

    def on_production_scenario_changed(self, scenario_name):
        scenario_name = str(scenario_name or "").strip()
        if not scenario_name or getattr(self, "_refreshing_scenario_combos", False):
            return
        if hasattr(self, "comboBox_auto_scenario") and self.comboBox_auto_scenario.findText(scenario_name) >= 0:
            self.comboBox_auto_scenario.setCurrentText(scenario_name)

    def production_refresh_table(self):
        search = self.lineEdit_prod_search_lot.text().strip()
        try:
            self.production_init_db()
            with self.production_connect_db() as con:
                params = []
                where = ""
                if search:
                    where = "WHERE lot LIKE ?"
                    params.append(f"%{search}%")
                rows = con.execute(
                    """
                    SELECT lot, sn, date, heure, operateur, designation, nb_inverseurs, scenario, resultat, details, timestamp
                    FROM essais
                    {where}
                    ORDER BY lot COLLATE NOCASE ASC, timestamp ASC, id ASC
                    """.format(where=where),
                    params,
                ).fetchall()
            by_lot = {}
            for record in rows:
                lot = str(record["lot"] or "")
                by_lot.setdefault(lot, []).append(record)
            records = []
            for lot, lot_records in sorted(by_lot.items(), key=lambda item: item[0].lower()):
                summary = self.production_lot_summary(lot_records)
                ordered_by_time = sorted(lot_records, key=lambda r: str(r["timestamp"] or f"{r['date']} {r['heure']}"))
                last_record = ordered_by_time[-1] if ordered_by_time else lot_records[-1]
                records.append({
                    "lot": lot,
                    "nb_essais": summary["total"],
                    "sn_distincts": summary["sn_distincts"],
                    "nb_acceptes": summary["acceptes"],
                    "nb_refuses": summary["refuses"],
                    "premier_sn": summary["premier_sn"],
                    "dernier_sn": summary["dernier_sn"],
                    "dernier_essai": str(last_record["timestamp"] or ""),
                    "nb_inverseurs": self.clamp_nb_inverseurs(last_record["nb_inverseurs"]),
                    "designation": str(last_record["designation"] or ""),
                })
        except Exception as exc:
            records = []
            self.label_prod_status.setText(f"Base production non lue : {exc}")
        table = self.tableWidget_prod_records
        table.setRowCount(len(records))
        for row, record in enumerate(records):
            vals = [
                record["lot"],
                record["nb_essais"],
                record["sn_distincts"],
                record["nb_acceptes"] or 0,
                record["nb_refuses"] or 0,
                record["premier_sn"] or "",
                record["dernier_sn"] or "",
                self.format_datetime_fr(record["dernier_essai"]),
                record["nb_inverseurs"],
                record["designation"],
            ]
            for col, val in enumerate(vals):
                table.setItem(row, col, self.table_item(val))
        widths = [110, 75, 85, 75, 75, 90, 90, 145, 45, 165]
        for col, width in enumerate(widths):
            table.setColumnWidth(col, width)
        suffix = f" - filtre : {search}" if search else ""
        self.label_prod_status.setText(f"Historique lots : {len(records)} lot(s) affiché(s){suffix}. Double-clic = détail relais.")

    def production_selected_lot(self):
        selected = self.tableWidget_prod_records.selectedItems()
        if selected:
            row = selected[0].row()
            item = self.tableWidget_prod_records.item(row, 0)
            if item is not None and item.text().strip():
                return item.text().strip()
        return self.lineEdit_prod_lot.text().strip()

    def filename_safe(self, value):
        value = str(value or "").strip()
        value = re.sub(r"[^A-Za-z0-9._-]+", "_", value)
        return value.strip("_") or "lot"

    def export_directory(self):
        """Répertoire unique des exports : même dossier que l'EXE en production."""
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        return EXPORT_DIR

    def export_default_path(self, default_name):
        return str(self.export_directory() / str(default_name))

    def normalize_export_path(self, selected_path, default_name, extension):
        """Force le fichier exporté dans le dossier de l'EXE.

        Le sélecteur de fichier sert uniquement à choisir le nom du fichier.
        Même si l'opérateur navigue ailleurs, le dossier final reste celui de l'EXE.
        """
        if not selected_path:
            return ""
        selected = Path(str(selected_path))
        name = selected.name.strip() or str(default_name)
        if extension and not name.lower().endswith(extension.lower()):
            name += extension
        final_path = self.export_directory() / name
        try:
            selected_is_final = selected.resolve() == final_path.resolve()
        except Exception:
            selected_is_final = False
        if final_path.exists() and not selected_is_final:
            if QMessageBox.question(
                self.window,
                "Remplacer fichier",
                f"Le fichier existe déjà dans le dossier de l'EXE :\n\n{final_path}\n\nLe remplacer ?",
            ) != QMessageBox.Yes:
                return ""
        return str(final_path)

    def ask_export_path(self, title, default_name, file_filter, extension):
        path, _ = QFileDialog.getSaveFileName(
            self.window,
            title,
            self.export_default_path(default_name),
            file_filter,
        )
        return self.normalize_export_path(path, default_name, extension)

    def format_datetime_fr(self, value):
        value = str(value or "").strip()
        if not value:
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
            try:
                return time.strftime("%d/%m/%Y %H:%M", time.strptime(value, fmt))
            except ValueError:
                pass
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return time.strftime("%d/%m/%Y", time.strptime(value, fmt))
            except ValueError:
                pass
        return value

    def production_record_datetime_fr(self, record):
        timestamp = str(record["timestamp"] or "").strip()
        if timestamp:
            return self.format_datetime_fr(timestamp)
        date = str(record["date"] or "").strip()
        heure = str(record["heure"] or "").strip()
        return self.format_datetime_fr(f"{date} {heure}".strip())

    def production_lot_date_label(self, records):
        if not records:
            return "Date essai : -"
        ordered = sorted(records, key=lambda r: str(r["timestamp"] or f"{r['date']} {r['heure']}"))
        first = self.production_record_datetime_fr(ordered[0])
        last = self.production_record_datetime_fr(ordered[-1])
        if not first:
            return "Date essai : -"
        if first == last:
            return f"Date essai : {first}"
        return f"Période essais : {first} au {last}"

    def production_sn_sort_key(self, sn):
        """Tri naturel des SN : 1, 2, 3... 10 au lieu de 1, 10, 11... 2."""
        text = str(sn or "").strip()
        if not text:
            return ((2, ""),)
        parts = re.split(r"(\d+)", text)
        key = []
        for part in parts:
            if not part:
                continue
            if part.isdigit():
                key.append((0, int(part), len(part), part))
            else:
                key.append((1, part.lower(), part))
        return tuple(key)

    def production_sort_records_by_sn(self, records):
        return sorted(
            list(records or []),
            key=lambda r: (
                self.production_sn_sort_key(r["sn"]),
                str(r["timestamp"] or ""),
                str(r["date"] or ""),
                str(r["heure"] or ""),
            ),
        )

    def production_distinct_sn_values(self, records):
        values = []
        seen = set()
        for record in records or []:
            sn = str(record["sn"] or "").strip()
            key = sn.lower()
            if not sn or key in seen:
                continue
            seen.add(key)
            values.append(sn)
        return sorted(values, key=self.production_sn_sort_key)

    def production_records_for_lot(self, lot):
        with self.production_connect_db() as con:
            records = con.execute(
                """
                SELECT lot, sn, date, heure, operateur, designation, nb_inverseurs, scenario, resultat, details, timestamp
                FROM essais
                WHERE lot = ?
                ORDER BY timestamp ASC, id ASC
                """,
                (lot,),
            ).fetchall()
        return self.production_sort_records_by_sn(records)

    def production_lot_summary(self, records):
        total = len(records)
        distinct_sn = self.production_distinct_sn_values(records)
        acceptes = sum(1 for r in records if "ACCEPT" in str(r["resultat"]).upper())
        refuses = sum(1 for r in records if "REFUS" in str(r["resultat"]).upper() or "REJET" in str(r["resultat"]).upper())
        pct_acceptes = (acceptes * 100.0 / total) if total else 0.0
        pct_refuses = (refuses * 100.0 / total) if total else 0.0
        return {
            "total": total,
            "sn_distincts": len(distinct_sn),
            "premier_sn": distinct_sn[0] if distinct_sn else "",
            "dernier_sn": distinct_sn[-1] if distinct_sn else "",
            "acceptes": acceptes,
            "refuses": refuses,
            "pct_acceptes": pct_acceptes,
            "pct_refuses": pct_refuses,
        }

    def production_open_selected_lot_details(self):
        lot = self.production_selected_lot()
        if not lot:
            return
        try:
            records = self.production_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "Détail lot", f"Lecture base impossible : {exc}")
            return
        if not records:
            QMessageBox.warning(self.window, "Détail lot", f"Aucun relais enregistré pour le lot : {lot}")
            return
        self.production_show_lot_details(lot, records)

    def production_show_lot_details(self, lot, records):
        summary = self.production_lot_summary(records)
        dialog = QDialog(self.window)
        dialog.setWindowTitle(f"Détail lot {lot}")
        dialog.resize(980, 520)
        layout = QVBoxLayout(dialog)

        header_row = QHBoxLayout()
        title = QLabel(f"Lot {lot} - {summary['total']} essai(s) - {summary['sn_distincts']} SN distinct(s)")
        title.setStyleSheet("font-size: 16pt; font-weight: bold; color: black;")
        title.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        summary_label = QLabel(
            f"Acceptés : {summary['acceptes']} ({summary['pct_acceptes']:.1f} %)   "
            f"Refusés : {summary['refuses']} ({summary['pct_refuses']:.1f} %)"
        )
        summary_label.setStyleSheet(
            "background-color: rgb(245,245,220); color: black; "
            "font-size: 12pt; font-weight: bold; border: 2px solid rgb(160,140,70); padding: 6px;"
        )
        summary_label.setAlignment(Qt.AlignCenter)
        header_row.addWidget(title, 1)
        header_row.addWidget(summary_label, 0)
        layout.addLayout(header_row)

        table = QTableWidget(dialog)
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["SN", "Date essai", "Résultat", "Opérateur", "Inv.", "Scénario", "Désignation", "Détails"])
        table.setRowCount(len(records))
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        for row, record in enumerate(records):
            vals = [
                record["sn"],
                self.production_record_datetime_fr(record),
                record["resultat"],
                record["operateur"],
                self.clamp_nb_inverseurs(record["nb_inverseurs"]),
                record["scenario"],
                record["designation"],
                record["details"],
            ]
            for col, val in enumerate(vals):
                table.setItem(row, col, self.table_item(val))
        widths = [90, 130, 90, 110, 45, 170, 150, 260]
        for col, width in enumerate(widths):
            table.setColumnWidth(col, width)
        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table)

        row_buttons = QHBoxLayout()
        row_buttons.addStretch(1)
        pdf_button = QPushButton("Créer PDF du lot")
        close_button = QPushButton("Fermer")
        pdf_button.clicked.connect(lambda: self.production_export_pdf_for_lot(lot))
        close_button.clicked.connect(dialog.accept)
        row_buttons.addWidget(pdf_button)
        row_buttons.addWidget(close_button)
        layout.addLayout(row_buttons)

        dialog.exec()

    def production_export_pdf_lot(self):
        lot = self.production_selected_lot()
        if not lot:
            QMessageBox.warning(self.window, "PDF lot", "Sélectionnez une ligne du tableau ou renseignez un numéro de lot.")
            return
        self.production_export_pdf_for_lot(lot)

    def production_export_pdf_for_lot(self, lot):
        try:
            records = self.production_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "PDF lot", f"Lecture base impossible : {exc}")
            return
        if not records:
            QMessageBox.warning(self.window, "PDF lot", f"Aucun essai enregistré pour le lot : {lot}")
            return

        default_name = f"rapport_lot_{self.filename_safe(lot)}.pdf"
        path = self.ask_export_path(
            "Exporter le rapport PDF du lot",
            default_name,
            "PDF (*.pdf)",
            ".pdf",
        )
        if not path:
            return
        try:
            self.write_lot_pdf(path, lot, records)
            self.label_prod_status.setText(f"PDF lot créé : {Path(path).name}")
            QMessageBox.information(self.window, "PDF lot", f"Rapport PDF créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "PDF lot", f"Création PDF impossible : {exc}")

    def write_lot_pdf(self, path, lot, records):
        writer = QPdfWriter(path)
        writer.setPageSize(QPageSize(QPageSize.A4))
        writer.setResolution(96)
        painter = QPainter(writer)
        if not painter.isActive():
            raise RuntimeError("QPainter PDF non actif.")
        try:
            page_w = writer.width()
            page_h = writer.height()
            margin = 38
            y = margin
            line_h = 18
            normal = QFont("Arial", 9)
            bold = QFont("Arial", 9)
            bold.setBold(True)
            title_font = QFont("Arial", 16)
            title_font.setBold(True)

            def new_page():
                nonlocal y
                writer.newPage()
                y = margin

            def draw_text(x, yy, text, font=None):
                painter.setFont(font or normal)
                painter.drawText(int(x), int(yy), str(text or ""))

            summary = self.production_lot_summary(records)
            designation = str(records[0]["designation"] or "")
            nb_inverseurs_values = sorted({self.clamp_nb_inverseurs(r["nb_inverseurs"]) for r in records})
            nb_inverseurs_label = ", ".join(str(v) for v in nb_inverseurs_values) if nb_inverseurs_values else "-"
            operateurs = sorted({str(r["operateur"] or "") for r in records if str(r["operateur"] or "")})
            date_label = self.production_lot_date_label(records)

            painter.setFont(title_font)
            painter.drawText(margin, y, f"Rapport Neutral Screen - Lot {lot}")
            painter.setFont(bold)
            summary_x = page_w - margin - 220
            painter.drawText(summary_x, y, f"Acceptés : {summary['acceptes']} ({summary['pct_acceptes']:.1f} %)")
            painter.drawText(summary_x, y + line_h, f"Refusés : {summary['refuses']} ({summary['pct_refuses']:.1f} %)")
            painter.drawText(summary_x, y + line_h * 2, f"Total essais : {summary['total']}")
            painter.drawText(summary_x, y + line_h * 3, f"SN distincts : {summary['sn_distincts']}")
            y += line_h * 4 + 8
            draw_text(margin, y, date_label, normal)
            y += line_h
            draw_text(margin, y, f"Désignation : {designation}", normal)
            y += line_h
            draw_text(margin, y, f"Nombre d'inverseurs : {nb_inverseurs_label}", normal)
            y += line_h
            draw_text(margin, y, f"Opérateur(s) : {', '.join(operateurs) if operateurs else '-'}", normal)
            y += line_h
            draw_text(
                margin,
                y,
                f"Essais : {summary['total']}    SN distincts : {summary['sn_distincts']}    Acceptés : {summary['acceptes']}    Refusés : {summary['refuses']}",
                bold,
            )
            y += 30

            headers = ["SN", "Date essai", "Résultat", "Inv.", "Scénario", "Détails"]
            widths = [75, 115, 72, 35, 135, page_w - margin * 2 - 432]
            x_positions = [margin]
            for width in widths[:-1]:
                x_positions.append(x_positions[-1] + width)

            def draw_header():
                nonlocal y
                painter.setFont(bold)
                for x, header in zip(x_positions, headers):
                    painter.drawText(int(x), int(y), header)
                y += line_h
                painter.drawLine(margin, y - 12, page_w - margin, y - 12)

            draw_header()
            painter.setFont(normal)
            for record in records:
                if y > page_h - margin:
                    new_page()
                    draw_header()
                vals = [
                    record["sn"],
                    self.production_record_datetime_fr(record),
                    record["resultat"],
                    self.clamp_nb_inverseurs(record["nb_inverseurs"]),
                    record["scenario"],
                    record["details"],
                ]
                for x, val in zip(x_positions, vals):
                    text = str(val or "")
                    if len(text) > 42:
                        text = text[:39] + "..."
                    painter.drawText(int(x), int(y), text)
                y += line_h
        finally:
            painter.end()

    def production_record_result(self, resultat, details=""):
        missing = self.production_missing_required_fields(include_sn=True)
        if missing:
            self.production_show_missing_fields_message(missing)
            self.label_prod_status.setText("Enregistrement refusé : informations production incomplètes.")
            self.auto_neutral_running = False
            self.set_auto_finish_validation_state(False)
            return False
        context = self.production_context()
        current_sn = context.get("sn", "")
        try:
            self.production_init_db()
            heure = time.strftime("%H:%M:%S")
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            scenario = self.current_runtime_scenario_name or context.get("scenario", "")
            with self.production_connect_db() as con:
                con.execute(
                    """
                    INSERT INTO essais(lot, sn, designation, nb_inverseurs, operateur, date, heure, scenario, resultat, details, timestamp)
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        context.get("lot", ""),
                        current_sn,
                        context.get("designation", ""),
                        context.get("nb_inverseurs", 2),
                        context.get("operateur", ""),
                        context.get("date", ""),
                        heure,
                        scenario,
                        resultat,
                        str(details or ""),
                        timestamp,
                    ),
                )
            self.last_finished_sn = str(current_sn or "")
            self.last_finished_result = str(resultat or "")
            self.increment_sn_after_result(current_sn)
            self.production_data["last_context"] = self.production_context()
            self.production_save_db()
            self.production_refresh_table()
            self.database_admin_refresh()
            return True
        except Exception as exc:
            self.label_prod_status.setText(f"Erreur enregistrement production : {exc}")
            return False

    def increment_sn_after_result(self, current_sn):
        current_sn = str(current_sn or "").strip()
        next_sn = self.next_sn_value(current_sn)
        if next_sn:
            self.lineEdit_SN.setText(next_sn)

    def next_sn_value(self, current_sn):
        current_sn = str(current_sn or "").strip()
        if not current_sn:
            return ""
        match = re.search(r"(\d+)$", current_sn)
        if not match:
            return current_sn
        number = match.group(1)
        prefix = current_sn[:match.start(1)]
        return f"{prefix}{int(number) + 1:0{len(number)}d}"

    def initialiser_unites(self):
        index_baud_rapide = self.comboBox_baudrate.findText("921600")
        if index_baud_rapide >= 0:
            self.comboBox_baudrate.setCurrentIndex(index_baud_rapide)
        index_baud_rapide_production = self.comboBox_baudrate_production.findText("921600")
        if index_baud_rapide_production >= 0:
            self.comboBox_baudrate_production.setCurrentIndex(index_baud_rapide_production)
        self.appliquer_mode_temps_global("ms")
        self.comboBox_unite_neutral_be.setCurrentText("ms")
        self.comboBox_unite_neutral_br.setCurrentText("ms")
        self.comboBox_unite_neutral_bebr.setCurrentText("ms")

    def connect_signals(self):
        self.pushButton_rafraichir_ports.clicked.connect(self.refresh_ports)
        self.pushButton_connecter.clicked.connect(self.connect_serial)
        self.pushButton_deconnecter.clicked.connect(self.disconnect_serial)
        self.pushButton_rafraichir_ports_production.clicked.connect(self.refresh_ports)
        self.pushButton_connecter_production.clicked.connect(self.connect_serial)
        self.pushButton_deconnecter_production.clicked.connect(self.disconnect_serial)
        self.comboBox_ports.currentTextChanged.connect(
            lambda text: self.sync_combo_text(self.comboBox_ports_production, text)
        )
        self.comboBox_ports_production.currentTextChanged.connect(
            lambda text: self.sync_combo_text(self.comboBox_ports, text)
        )
        self.comboBox_baudrate.currentTextChanged.connect(
            lambda text: self.sync_combo_text(self.comboBox_baudrate_production, text)
        )
        self.comboBox_baudrate_production.currentTextChanged.connect(
            lambda text: self.sync_combo_text(self.comboBox_baudrate, text)
        )
        self.pushButton_prod_save_context.clicked.connect(self.production_prepare_and_open_auto_test)
        self.pushButton_prod_reload_base.clicked.connect(self.production_reload_base)
        self.pushButton_prod_export_pdf_lot.clicked.connect(self.production_export_pdf_lot)
        self.lineEdit_prod_search_lot.textChanged.connect(self.production_refresh_table)
        self.pushButton_prod_search_clear.clicked.connect(self.lineEdit_prod_search_lot.clear)
        self.lineEdit_prod_lot.editingFinished.connect(self.production_autofill_from_lot)
        self.lineEdit_prod_lot.textEdited.connect(self.production_on_lot_edited)
        self.lineEdit_prod_lot.textChanged.connect(self.sync_auto_production_labels)
        self.lineEdit_prod_designation.textChanged.connect(self.sync_auto_production_labels)
        self.spinBox_prod_nb_inverseurs.valueChanged.connect(lambda _value: self.production_sync_nb_inverseurs_to_auto())
        self.tableWidget_prod_records.itemDoubleClicked.connect(self.production_open_selected_lot_details)
        self.comboBox_prod_scenario.currentTextChanged.connect(self.on_production_scenario_changed)
        self.pushButton_db_refresh.clicked.connect(self.database_admin_refresh)
        self.pushButton_db_backup.clicked.connect(self.database_backup)
        self.pushButton_db_restore.clicked.connect(self.database_restore)
        self.pushButton_db_export_csv.clicked.connect(self.database_export_csv)
        self.pushButton_db_export_xlsx.clicked.connect(self.database_export_xlsx)
        self.pushButton_db_export_pdf.clicked.connect(self.database_export_pdf)
        self.pushButton_db_vacuum.clicked.connect(self.database_maintenance)
        self.pushButton_db_recreate_default.clicked.connect(self.database_recreate_default_clicked)
        self.pushButton_db_merge.clicked.connect(self.database_merge_clicked)
        self.comboBox_db_target.currentTextChanged.connect(self.database_admin_target_changed)
        self.pushButton_db_operator_add.clicked.connect(self.database_add_operator)
        self.pushButton_db_operator_delete.clicked.connect(self.database_delete_operator)
        self.lineEdit_db_lot_filter.textChanged.connect(self.database_admin_refresh)
        self.pushButton_db_lot_open.clicked.connect(self.database_open_selected_lot_details)
        self.pushButton_db_lot_pdf.clicked.connect(self.database_export_selected_lot_pdf)
        self.pushButton_db_lot_xlsx.clicked.connect(self.database_export_selected_lot_xlsx)
        self.pushButton_db_lot_delete.clicked.connect(self.database_delete_selected_lot)
        self.tableWidget_db_lots.itemDoubleClicked.connect(lambda _item: self.database_open_selected_lot_details())

        self.pushButton_mode_us.clicked.connect(lambda: self.appliquer_mode_temps_global("µs"))
        self.pushButton_mode_ms.clicked.connect(lambda: self.appliquer_mode_temps_global("ms"))
        self.pushButton_mode_s.clicked.connect(lambda: self.appliquer_mode_temps_global("s"))
        self.pushButton_mode_min.clicked.connect(lambda: self.appliquer_mode_temps_global("min"))
        self.pushButton_mode_h.clicked.connect(lambda: self.appliquer_mode_temps_global("h"))

        self.pushButton_demarrer.clicked.connect(self.start_cycle)
        self.pushButton_pause.clicked.connect(lambda: self.send_command("PAUSE"))
        self.pushButton_reprendre.clicked.connect(lambda: self.send_command("RESUME"))
        self.pushButton_arret.clicked.connect(self.stop_general)
        self.pushButton_status.clicked.connect(lambda: self.send_command("STATUS?"))

        self.pushButton_neutral_be.clicked.connect(self.neutral_pulse_be)
        self.pushButton_neutral_br.clicked.connect(self.neutral_pulse_br)
        self.pushButton_neutral_bebr.clicked.connect(self.neutral_pulse_bebr)
        self.pushButton_neutral_stop.clicked.connect(self.neutral_stop)

        self.pushButton_auto_neutral_marche.clicked.connect(self.auto_neutral_start)
        self.pushButton_auto_neutral_arret.clicked.connect(self.auto_neutral_stop)
        self.pushButton_auto_lot_fini.clicked.connect(self.auto_finish_lot)
        self.comboBox_auto_scenario.currentIndexChanged.connect(self.on_auto_scenario_changed)
        self.pushButton_auto_scenario_recharger.clicked.connect(self.on_recharger_scenarios)
        self.pushButton_auto_scenario_editer.clicked.connect(self.aller_onglet_editeur_scenarios)
        self.lineEdit_auto_nb_inverseurs.textChanged.connect(self.on_auto_nb_inverseurs_changed)
        self.checkBox_auto_pulses_particuliers.toggled.connect(self.on_auto_pulses_particuliers_changed)
        self.lineEdit_chrono_lot.editingFinished.connect(self.chrono_autofill_from_lot)
        self.lineEdit_chrono_lot.textEdited.connect(self.chrono_on_lot_edited)
        self.pushButton_chrono_mesure_be.clicked.connect(lambda: self.chrono_start_measure("BE"))
        self.pushButton_chrono_mesure_br.clicked.connect(lambda: self.chrono_start_measure("BR"))
        self.pushButton_chrono_mesure_be_br.clicked.connect(self.chrono_start_measure_be_br)
        self.pushButton_chrono_export_xlsx_lot.clicked.connect(self.chrono_export_current_lot_xlsx)
        self.pushButton_chrono_export_pdf_lot.clicked.connect(self.chrono_export_current_lot_pdf)
        self.pushButton_chrono_effacer.clicked.connect(self.chrono_clear_results)
        self.pushButton_voltage_copy_chrono.clicked.connect(self.voltage_copy_from_chrono)
        self.pushButton_voltage_ea_refresh.clicked.connect(self.voltage_refresh_ea_ports)
        self.pushButton_voltage_ea_connect.clicked.connect(self.voltage_connect_ea)
        self.pushButton_voltage_ea_disconnect.clicked.connect(self.voltage_disconnect_ea)
        self.pushButton_voltage_pickup.clicked.connect(lambda: self.voltage_start_test("PICKUP"))
        self.pushButton_voltage_dropout.clicked.connect(lambda: self.voltage_start_test("DROPOUT"))
        self.pushButton_voltage_cycle.clicked.connect(lambda: self.voltage_start_test("CYCLE"))
        self.pushButton_voltage_measure_all.clicked.connect(self.voltage_start_measure_all)
        self.pushButton_voltage_stop.clicked.connect(self.voltage_stop_clicked)
        self.pushButton_voltage_export_xlsx.clicked.connect(self.voltage_export_lot_xlsx)
        self.pushButton_voltage_export_pdf.clicked.connect(self.voltage_export_lot_pdf)
        self.spinBox_voltage_nb_inverseurs.valueChanged.connect(lambda _value: self.voltage_on_nb_inverseurs_changed())
        self.doubleSpinBox_voltage_vmax.valueChanged.connect(self.voltage_update_ramp_limits)
        for spin in (
            self.doubleSpinBox_voltage_vmax,
            self.doubleSpinBox_voltage_ramp_up_s,
            self.doubleSpinBox_voltage_ramp_down_s,
            self.doubleSpinBox_voltage_current_limit,
            self.doubleSpinBox_voltage_chrono_v,
            self.doubleSpinBox_voltage_interphase_s,
        ):
            spin.valueChanged.connect(self.voltage_save_measure_settings)
            spin.editingFinished.connect(self.voltage_commit_measure_settings)
        self.comboBox_voltage_relay_type.currentTextChanged.connect(self.voltage_update_relay_type_ui)
        self.pushButton_voltage_open_calibration.clicked.connect(lambda: self.set_tab_internal(self.tab_voltage_calibration))
        self.pushButton_calibration_request_ads.clicked.connect(self.voltage_calibration_request_ads)
        self.pushButton_calibration_capture_low.clicked.connect(lambda: self.voltage_calibration_request_capture("LOW"))
        self.pushButton_calibration_capture_high.clicked.connect(lambda: self.voltage_calibration_request_capture("HIGH"))
        self.pushButton_calibration_capture_check.clicked.connect(lambda: self.voltage_calibration_request_capture("CHECK"))
        self.pushButton_calibration_calculate.clicked.connect(self.voltage_calibration_calculate)
        self.pushButton_calibration_save_activate.clicked.connect(self.voltage_calibration_save_activate)
        self.pushButton_calibration_invalidate.clicked.connect(self.voltage_calibration_invalidate_active)
        self.pushButton_calibration_clear.clicked.connect(self.voltage_calibration_clear_captures)
        self.comboBox_chrono_type_relais.currentTextChanged.connect(self.chrono_update_relay_type_ui)
        self.pushButton_oscillo_export_xlsx.clicked.connect(self.oscillo_export_current_xlsx)
        self.pushButton_oscillo_export_pdf.clicked.connect(self.oscillo_export_current_pdf)
        self.pushButton_oscillo_vue_complete.clicked.connect(lambda _checked=False: self.oscillo_set_full_view(update=True))
        self.pushButton_oscillo_zoom_fronts.clicked.connect(self.oscillo_zoom_fronts)
        self.comboBox_oscillo_capture.currentIndexChanged.connect(self.oscillo_on_capture_changed)
        self.comboBox_oscillo_display_mode.currentIndexChanged.connect(lambda _index=0: self.oscillo_update_view())
        self.pushButton_oscillo_load_saved.clicked.connect(self.oscillo_load_saved_capture_dialog)
        self.pushButton_oscillo_zoom_contact.clicked.connect(self.oscillo_zoom_contact)
        self.pushButton_oscillo_zoom_rebonds.clicked.connect(self.oscillo_zoom_rebonds)
        self.pushButton_oscillo_zoom_in.clicked.connect(self.oscillo_zoom_in_factor)
        self.pushButton_oscillo_zoom_out.clicked.connect(self.oscillo_zoom_out_factor)
        self.spinBox_oscillo_zoom_factor.valueChanged.connect(lambda _value: self.oscillo_update_zoom_button_labels())
        self.spinBox_oscillo_zoom_start_us.valueChanged.connect(lambda _value: self.oscillo_update_view())
        self.spinBox_oscillo_zoom_end_us.valueChanged.connect(lambda _value: self.oscillo_update_view())
        self.spinBox_oscillo_cursor_a_us.valueChanged.connect(lambda _value: self.oscillo_update_view())
        self.spinBox_oscillo_cursor_b_us.valueChanged.connect(lambda _value: self.oscillo_update_view())

        self.comboBox_editor_scenarios.currentIndexChanged.connect(self.on_editor_scenario_changed)
        self.pushButton_editor_nouveau.clicked.connect(self.editor_nouveau_scenario)
        self.pushButton_editor_dupliquer.clicked.connect(self.editor_dupliquer_scenario)
        self.pushButton_editor_supprimer.clicked.connect(self.editor_supprimer_scenario)
        self.pushButton_editor_sauvegarder.clicked.connect(self.editor_sauvegarder_scenario)
        self.pushButton_editor_recharger.clicked.connect(self.on_recharger_scenarios)
        self.pushButton_editor_ajouter_etape.clicked.connect(self.editor_ajouter_etape)
        self.pushButton_editor_supprimer_etape.clicked.connect(self.editor_supprimer_etape)
        self.pushButton_editor_monter_etape.clicked.connect(lambda: self.editor_deplacer_etape(-1))
        self.pushButton_editor_descendre_etape.clicked.connect(lambda: self.editor_deplacer_etape(1))
        self.pushButton_editor_importer.clicked.connect(self.editor_importer_json)
        self.pushButton_editor_exporter.clicked.connect(self.editor_exporter_json)
        self.pushButton_editor_mot_de_passe.clicked.connect(self.changer_mot_de_passe_acces)

        self.radioButton_monostable.toggled.connect(self.update_mode_fields)
        self.radioButton_bistable.toggled.connect(self.update_mode_fields)

    def appliquer_mode_temps_global(self, unite):
        for combo in (self.comboBox_unite_on, self.comboBox_unite_off,
                      self.comboBox_unite_set, self.comboBox_unite_reset):
            index = combo.findText(unite)
            if index >= 0:
                combo.setCurrentIndex(index)
        self.label_mode_temps_actif.setText(f"Actif : {unite}")

        if unite == "µs":
            vals = ("100", "100", "100", "100")
        elif unite == "ms":
            vals = ("1000", "1000", "30", "30")
        else:
            vals = ("1", "1", "1", "1")
        self.lineEdit_temps_on.setText(vals[0])
        self.lineEdit_temps_off.setText(vals[1])
        self.lineEdit_impulsion_set.setText(vals[2])
        self.lineEdit_impulsion_reset.setText(vals[3])
        self.update_mode_fields()

    def sync_combo_text(self, combo, text):
        if combo.currentText() == text:
            return
        combo.blockSignals(True)
        index = combo.findText(text)
        if index >= 0:
            combo.setCurrentIndex(index)
        combo.blockSignals(False)

    def refresh_ports_keep_selection(self):
        current = self.comboBox_ports.currentText() or self.comboBox_ports_production.currentText()
        self.refresh_ports()
        if current:
            for combo in (self.comboBox_ports, self.comboBox_ports_production):
                index = combo.findText(current)
                if index >= 0:
                    combo.blockSignals(True)
                    combo.setCurrentIndex(index)
                    combo.blockSignals(False)

    def refresh_ports(self):
        current = self.comboBox_ports.currentText() or self.comboBox_ports_production.currentText()
        ports = [port.device for port in serial.tools.list_ports.comports()]
        for combo in (self.comboBox_ports, self.comboBox_ports_production):
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(ports)
            if current in ports:
                combo.setCurrentText(current)
            combo.blockSignals(False)

    def candidate_rp2040_ports(self):
        ports = list(serial.tools.list_ports.comports())
        scored = []
        for port in ports:
            text = " ".join(
                str(v or "") for v in (
                    port.device,
                    port.description,
                    port.manufacturer,
                    port.product,
                    port.hwid,
                )
            ).lower()
            score = 0
            for word in ("rp2040", "pico", "waveshare", "usb serial", "usb-sérial", "cdc", "arduino"):
                if word in text:
                    score += 10
            if "bluetooth" in text:
                score -= 20
            if "com" in str(port.device).lower() or "/dev/tty" in str(port.device).lower():
                score += 1
            scored.append((score, port.device, port))
        scored.sort(key=lambda item: item[0], reverse=True)
        return [device for score, device, _port in scored if score > 0]

    def auto_connect_rp2040(self, force=False):
        if self._auto_connect_in_progress or self.is_connected() or self._manual_disconnect:
            return
        if self._auto_connect_done and not force:
            return
        self._auto_connect_done = True
        self._auto_connect_in_progress = True
        try:
            self.refresh_ports()
            candidates = self.candidate_rp2040_ports()
            if not candidates:
                self._last_connection_error = "RP2040 non détecté"
                self.update_connection_status_visual()
                return
            for port in candidates:
                if self.connect_serial(port_override=port, silent=True):
                    self.log(f"Connexion automatique RP2040 : {port}")
                    return
            self._last_connection_error = "Connexion auto impossible"
            self.update_connection_status_visual()
        finally:
            self._auto_connect_in_progress = False

    def is_connected(self):
        try:
            return self.ser is not None and self.ser.is_open
        except Exception:
            return False

    def current_serial_port_still_present(self):
        if self.ser is None:
            return False
        try:
            port = str(self.ser.port or "")
        except Exception:
            return False
        if not port:
            return False
        try:
            return port in {p.device for p in serial.tools.list_ports.comports()}
        except Exception:
            return True

    def mark_serial_disconnected(self, reason):
        reason = str(reason or "Déconnexion USB détectée")
        if self.ser is None and self.reader is None:
            self._last_connection_error = reason
            self.update_button_states()
            self.update_connection_status_visual()
            return
        old_reader = self.reader
        old_ser = self.ser
        self.reader = None
        self.ser = None
        self.rp2040_ea_chrono_capable = False
        self._auto_connect_done = False
        self._manual_disconnect = False
        try:
            if old_reader is not None:
                old_reader.stop()
        except Exception:
            pass
        try:
            if old_ser is not None and old_ser.is_open:
                old_ser.close()
        except Exception:
            pass
        if getattr(self, "auto_neutral_running", False):
            self.remember_auto_usb_interrupt(reason)
            self.auto_neutral_running = False
            self.auto_next_action = None
            try:
                self.auto_neutral_timer.stop()
            except Exception:
                pass
            self.label_auto_status.setText("Automatique : arrêté - déconnexion USB")
            self.label_auto_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            self.label_auto_resultat.setText("Résultat : interrompu - USB déconnecté")
            self.label_auto_resultat.setStyleSheet("background-color: rgb(255,80,80); color: black; font-size: 14pt; font-weight: bold; border: 3px solid rgb(160,0,0);")
        current_tab = self.tabWidget_principal.currentWidget()
        if current_tab is not None and current_tab.objectName() == "tab_neutral_auto":
            self.set_tab_internal(self.tab_production_accueil)
            self.label_prod_status.setText("Déconnexion USB détectée : retour Production.")
        self._last_connection_error = reason
        self.log(f"Déconnexion série : {reason}")
        self.update_button_states()
        self.update_connection_status_visual()

    def remember_auto_usb_interrupt(self, reason):
        sn = self.lineEdit_SN.text().strip()
        self.interrupted_auto_sn = sn
        self.interrupted_auto_scenario = str(self.current_runtime_scenario_name or self.comboBox_auto_scenario.currentText() or "").strip()
        self.interrupted_auto_reason = str(reason or "Déconnexion USB")
        self.set_auto_finish_validation_state(False)
        self._auto_start_prompts_done = True
        if sn:
            self.label_prod_status.setText(
                f"Essai interrompu par coupure USB : SN {sn} non enregistré. Rebranchez puis relancez pour refaire ce SN."
            )

    def connect_serial(self, port_override=None, silent=False):
        self.rp2040_ea_chrono_capable = False
        if self.is_connected():
            return True
        self._manual_disconnect = False
        port = (port_override or self.comboBox_ports.currentText()).strip()
        baudrate_text = self.comboBox_baudrate.currentText().strip()
        if not port:
            self._last_connection_error = "Aucun port COM sélectionné"
            self.update_connection_status_visual()
            if not silent:
                QMessageBox.warning(self.window, "Port COM", "Aucun port COM sélectionné.")
            return False
        try:
            baudrate = int(baudrate_text)
        except ValueError:
            self._last_connection_error = "Baudrate invalide"
            self.update_connection_status_visual()
            if not silent:
                QMessageBox.warning(self.window, "Baudrate", "Baudrate invalide.")
            return False
        try:
            for combo in (self.comboBox_ports, self.comboBox_ports_production):
                if combo.findText(port) >= 0:
                    combo.setCurrentText(port)
            self.ser = serial.Serial(
                port=port, baudrate=baudrate, timeout=0.005,
                write_timeout=0.5, dsrdtr=False, rtscts=False,
            )
            time.sleep(0.2)
            self.reader = SerialReader(self.ser)
            self.reader.line_received.connect(self.on_line_received)
            self.reader.error_received.connect(self.on_serial_error)
            self.reader.start()
            self._last_connected_port = port
            self.set_connection_status_text(f"État connexion : connecté sur {port} à {baudrate} bauds")
            self.log(f"Connexion ouverte : {port} / {baudrate}")
            # On demande immédiatement un STATUS pour ne pas attendre l'AUTO 1 s.
            # V2.12.3 : repartir d'un état LED neutre et forcer un refresh complet
            # dès la première trame, puis demander un STATUS immédiat.
            self.contacts_known_values = [None] * 8
            self.contacts_last_values = (None,) * 8
            self.contacts_force_refresh = True
            self.initialiser_leds_contacts()
            self.initialiser_auto_neutral()
            QTimer.singleShot(150, lambda: self.send_led_command("CONNECTED"))
            QTimer.singleShot(300, lambda: self.send_command("STATUS?"))
            self._last_connection_error = ""
            self.update_connection_status_visual()
            self.update_button_states()
            return True
        except Exception as exc:
            self.ser = None
            self._last_connection_error = str(exc)
            self.update_connection_status_visual()
            if not silent:
                QMessageBox.critical(self.window, "Erreur connexion série", str(exc))
        self.update_button_states()
        return False

    def disconnect_serial(self):
        self._manual_disconnect = True
        if getattr(self, "auto_neutral_running", False):
            self.auto_neutral_stop(send_stop=False)
        try:
            self.send_led_command("BOOT")
            if self.reader is not None:
                self.reader.stop()
                self.reader.wait(1000)
                self.reader = None
            if self.ser is not None:
                if self.ser.is_open:
                    self.ser.close()
                self.ser = None
            self._last_connected_port = ""
            self.set_connection_status_text("État connexion : déconnecté")
            self.log("Connexion fermée.")
        except Exception as exc:
            self.log(f"Erreur fermeture série : {exc}")
            self._last_connection_error = str(exc)
        self.update_button_states()
        self.update_connection_status_visual()

    def set_connection_status_text(self, text):
        self.label_etat_connexion.setText(text)
        self.label_etat_connexion_production.setText(text)

    def update_connection_status_visual(self):
        if not hasattr(self, "label_etat_connexion_production"):
            return
        if self.ser is not None and not self.current_serial_port_still_present():
            self.mark_serial_disconnected("USB déconnecté")
            return
        if self.is_connected():
            port = ""
            try:
                port = self.ser.port if self.ser is not None else ""
            except Exception:
                port = ""
            text = f"✓ CONNECTÉ RP2040 {port}".strip()
            style = (
                "background-color: rgb(0,150,70); color: white; "
                "font-size: 13pt; font-weight: bold; "
                "border: 3px solid rgb(0,80,35); border-radius: 5px; padding: 4px;"
            )
            self.label_etat_connexion.setStyleSheet(style)
            self.label_etat_connexion_production.setStyleSheet(style)
            self.set_connection_status_text(text)
            return

        self._connection_alert_phase = (self._connection_alert_phase + 1) % 2
        self._auto_reconnect_counter += 1
        if not self._manual_disconnect and self._auto_reconnect_counter >= 5:
            self._auto_reconnect_counter = 0
            QTimer.singleShot(0, lambda: self.auto_connect_rp2040(force=True))
        if self._last_connection_error:
            message = f"✗ NON CONNECTÉ - {self._last_connection_error} - APPUYER SUR CONNECTER"
        else:
            message = "✗ NON CONNECTÉ - APPUYER SUR CONNECTER"
        if self._connection_alert_phase:
            bg = "rgb(190,0,0)"
            fg = "white"
        else:
            bg = "rgb(230,45,45)"
            fg = "white"
        style = (
            f"background-color: {bg}; color: {fg}; "
            "font-size: 12pt; font-weight: bold; "
            "border: 3px solid rgb(150,0,0); border-radius: 5px; padding: 4px;"
        )
        self.label_etat_connexion.setStyleSheet(style)
        self.label_etat_connexion_production.setStyleSheet(style)
        self.set_connection_status_text(message)

    def update_button_states(self):
        connected = self.is_connected()
        voltage_busy = bool(
            getattr(self, "voltage_test_running", False)
            or getattr(self, "measure_all_active", False)
        )
        self.pushButton_connecter.setEnabled(not connected)
        self.pushButton_deconnecter.setEnabled(connected)
        self.pushButton_connecter_production.setEnabled(not connected)
        self.pushButton_deconnecter_production.setEnabled(connected)
        self.pushButton_demarrer.setEnabled(connected and not voltage_busy)
        self.pushButton_pause.setEnabled(connected and not voltage_busy)
        self.pushButton_reprendre.setEnabled(connected and not voltage_busy)
        self.pushButton_arret.setEnabled(connected)
        self.pushButton_status.setEnabled(connected)
        self.pushButton_neutral_be.setEnabled(connected and not voltage_busy)
        self.pushButton_neutral_br.setEnabled(connected and not voltage_busy)
        self.pushButton_neutral_bebr.setEnabled(connected and not voltage_busy)
        self.pushButton_neutral_stop.setEnabled(connected)
        if hasattr(self, "pushButton_auto_neutral_marche"):
            self.pushButton_auto_neutral_marche.setEnabled(connected and not self.auto_neutral_running and not self.auto_end_validation_pending and not voltage_busy)
            self.pushButton_auto_neutral_arret.setEnabled(connected and self.auto_neutral_running)
            self.pushButton_auto_lot_fini.setEnabled(not self.auto_neutral_running)
            self.lineEdit_auto_nb_inverseurs.setEnabled(not self.auto_neutral_running)
            self.update_auto_pulse_fields_state()
            self.comboBox_auto_scenario.setEnabled(not self.auto_neutral_running)
            self.pushButton_auto_scenario_recharger.setEnabled(not self.auto_neutral_running)
            self.pushButton_auto_scenario_editer.setEnabled(not self.auto_neutral_running)
        if hasattr(self, "pushButton_chrono_mesure_be"):
            chrono_enabled = connected and not getattr(self, "auto_neutral_running", False) and not getattr(self, "chrono_measure_running", False) and not voltage_busy
            self.pushButton_chrono_mesure_be.setEnabled(chrono_enabled)
            self.pushButton_chrono_mesure_br.setEnabled(chrono_enabled)
            self.pushButton_chrono_mesure_be_br.setEnabled(chrono_enabled)
            self.pushButton_chrono_export_xlsx_lot.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_chrono_export_pdf_lot.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_export_xlsx.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_export_pdf.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_vue_complete.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_zoom_fronts.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_zoom_contact.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_zoom_rebonds.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_zoom_in.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_zoom_out.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.comboBox_oscillo_capture.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.comboBox_oscillo_contact.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.comboBox_oscillo_display_mode.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.pushButton_oscillo_load_saved.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.spinBox_oscillo_zoom_factor.setEnabled(not getattr(self, "chrono_measure_running", False))
            self.comboBox_chrono_type_relais.setEnabled(not getattr(self, "chrono_measure_running", False))
        if hasattr(self, "pushButton_voltage_pickup"):
            self.voltage_update_button_states()

    def update_mode_fields(self):
        bistable = self.radioButton_bistable.isChecked()
        self.lineEdit_impulsion_set.setEnabled(bistable)
        self.lineEdit_impulsion_reset.setEnabled(bistable)
        self.comboBox_unite_set.setEnabled(bistable)
        self.comboBox_unite_reset.setEnabled(bistable)

    def log(self, text):
        horodatage = time.strftime("%d/%m/%Y %H:%M")
        self.textEdit_log.append(f"[{horodatage}] {text}")

    def normaliser_unite(self, unite):
        unite = unite.strip()
        if unite == "us":
            return "µs"
        return unite

    def lire_valeur_us(self, line_edit, combo_unite, nom_champ):
        texte = line_edit.text().strip().replace(" ", "")
        unite = self.normaliser_unite(combo_unite.currentText())
        if "," in texte or "." in texte:
            raise ValueError(
                f"{nom_champ} doit être une valeur entière dans son unité.\n"
                f"Exemple : pour 0,1 ms, utiliser 100 µs."
            )
        if not texte.isdigit():
            raise ValueError(f"{nom_champ} doit être un entier positif.")
        valeur = int(texte)
        if unite not in UNIT_FACTORS_US:
            raise ValueError(f"Unité invalide pour {nom_champ} : {unite}")
        mini, maxi = UNIT_LIMITS[unite]
        if valeur < mini or valeur > maxi:
            raise ValueError(
                f"{nom_champ} hors plage pour l'unité {unite}.\n"
                f"Plage autorisée : {mini} à {maxi} {unite}."
            )
        valeur_us = valeur * UNIT_FACTORS_US[unite]
        if valeur_us < 1 or valeur_us > MAX_US_TOTAL:
            raise ValueError(f"{nom_champ} dépasse la limite interne autorisée.")
        return valeur_us

    def lire_cycles(self):
        texte = self.lineEdit_nombre_cycles.text().strip().replace(" ", "")
        if not texte.isdigit():
            raise ValueError("Le nombre de cycles doit être un entier. 0 = infini.")
        valeur = int(texte)
        if valeur < 0:
            raise ValueError("Le nombre de cycles doit être positif.")
        if valeur > 4_294_967_295:
            raise ValueError("Le nombre de cycles dépasse la limite autorisée.")
        return valeur

    def start_cycle(self):
        try:
            mode = "BISTABLE" if self.radioButton_bistable.isChecked() else "MONO"
            on_us = self.lire_valeur_us(self.lineEdit_temps_on, self.comboBox_unite_on, "Temps ON / état ON")
            off_us = self.lire_valeur_us(self.lineEdit_temps_off, self.comboBox_unite_off, "Temps OFF / état OFF")
            if mode == "BISTABLE":
                set_us = self.lire_valeur_us(self.lineEdit_impulsion_set, self.comboBox_unite_set, "Impulsion SET")
                reset_us = self.lire_valeur_us(self.lineEdit_impulsion_reset, self.comboBox_unite_reset, "Impulsion RESET")
            else:
                set_us = 1
                reset_us = 1
            cycles = self.lire_cycles()
            command = f"START_US;{mode};{on_us};{off_us};{set_us};{reset_us};{cycles}"
            self.label_derniere_commande_us.setText(
                f"Dernière commande : ON={on_us}µs OFF={off_us}µs SET={set_us}µs RESET={reset_us}µs"
            )
            self.send_led_command("CONNECTED")
            self.send_command(command)
        except Exception as exc:
            QMessageBox.warning(self.window, "Paramètres invalides", str(exc))

    def valeur_tension_basse_info(self):
        if self.lineEdit_tension_basse_info is None:
            return "basse"
        texte = self.lineEdit_tension_basse_info.text().strip().replace(",", ".")
        return texte if texte else "basse"

    def valeur_tension_haute_info(self):
        if self.lineEdit_tension_haute_info is None:
            return "haute"
        texte = self.lineEdit_tension_haute_info.text().strip().replace(",", ".")
        return texte if texte else "haute"

    def texte_tension_basse_info(self):
        val = self.valeur_tension_basse_info()
        return f"{val} V" if val.replace(".", "", 1).isdigit() else f"tension {val}"

    def texte_tension_haute_info(self):
        val = self.valeur_tension_haute_info()
        return f"{val} V" if val.replace(".", "", 1).isdigit() else f"tension {val}"

    def neutral_pulse_be(self):
        try:
            d = self.lire_valeur_us(self.lineEdit_neutral_be, self.comboBox_unite_neutral_be, "Neutral screen pulse BE")
            self.label_neutral_derniere_commande.setText(f"Dernière commande neutral screen : BE {d} µs - sélection {self.texte_tension_haute_info()}")
            self.send_led_command("BE")
            self.send_command(f"PULSE_US;BE;{d}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Paramètre neutral screen invalide", str(exc))

    def neutral_pulse_br(self):
        try:
            d = self.lire_valeur_us(self.lineEdit_neutral_br, self.comboBox_unite_neutral_br, "Neutral screen pulse BR")
            self.label_neutral_derniere_commande.setText(f"Dernière commande neutral screen : BR {d} µs - sélection {self.texte_tension_haute_info()}")
            self.send_led_command("BR")
            self.send_command(f"PULSE_US;BR;{d}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Paramètre neutral screen invalide", str(exc))

    def neutral_pulse_bebr(self):
        try:
            d = self.lire_valeur_us(self.lineEdit_neutral_bebr, self.comboBox_unite_neutral_bebr, "Neutral screen pulse BE/BR")
            self.label_neutral_derniere_commande.setText(f"Dernière commande neutral screen : BE/BR {d} µs - sélection {self.texte_tension_basse_info()}")
            self.send_led_command("BEBR")
            self.send_command(f"PULSE_US;BEBR;{d}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Paramètre neutral screen invalide", str(exc))

    def neutral_stop(self):
        self.label_neutral_derniere_commande.setText("Dernière commande neutral screen : arrêt immédiat sorties")
        self.send_command("STOP")
        self.send_led_command("CONNECTED")

    def stop_general(self):
        self.send_command("STOP")
        self.send_led_command("CONNECTED")

    def send_command(self, command):
        if not self.is_connected():
            QMessageBox.warning(self.window, "Connexion", "RP2040 non connecté.")
            return
        try:
            line = command.strip() + "\n"
            self.ser.write(line.encode("utf-8"))
            self.ser.flush()
            self.log(f"> {command}")
        except Exception as exc:
            self.mark_serial_disconnected(str(exc))
            self.label_prod_status.setText(f"Erreur série : {exc}")

    def send_led_command(self, mode):
        if not self.is_connected():
            return
        try:
            line = f"LED;{mode}\n"
            self.ser.write(line.encode("utf-8"))
            self.ser.flush()
            self.log(f"> LED;{mode}")
        except Exception as exc:
            self.log(f"Erreur LED série : {exc}")
            self.mark_serial_disconnected(str(exc))

    def on_serial_error(self, message):
        self.log(f"Erreur série : {message}")
        self.mark_serial_disconnected(message)

    def on_line_received(self, line):
        if line.startswith("CONTACT;"):
            self.parse_contact_frame(line)
            return
        if line.startswith("OUT;"):
            self.parse_output_frame(line)
            return
        if line.startswith("VSEL;"):
            self.parse_vsel_frame(line)
            return
        if line.startswith("MEASURE;") or line.startswith("MEASURE_EVT;"):
            self.parse_measure_frame(line)
            return
        if line.startswith("VSCAN;") or line.startswith("ADS;") or line.startswith("COIL;"):
            self.voltage_parse_frame(line)
            return
        self.log(f"< {line}")
        if line.startswith("STATUS;") and "PULSE_DONE" in line and getattr(self, "chrono_auto_prereset_pending", False):
            self.chrono_auto_prereset_pending = False
            self.label_chrono_status.setText("Pré-positionnement RESET terminé - lancement mesure BE.")
            QTimer.singleShot(150, self.chrono_start_next_auto_measure)
            self.parse_status(line)
            return
        if line.startswith("STATUS;"):
            self.parse_status(line)
        elif line.startswith("ERREUR;"):
            self.label_etat_essai.setText("État : erreur")
        elif line.startswith("RP2040_RELAIS_28VDC_PRET"):
            self.rp2040_ea_chrono_capable = "EA_CHRONO_NO_GP26" in line.upper()
            self.label_etat_essai.setText("État : prêt")

    def chrono_clear_results(self, clear_pair_cache=True, clear_oscillo_history=True):
        self.chrono_events = []
        self.chrono_current = {}
        self.lineEdit_chrono_resultat.clear()
        self.tableWidget_chrono_results.setRowCount(0)
        self.tableWidget_chrono_events.setRowCount(0)
        self.oscillo_clear_view(clear_history=clear_oscillo_history)
        if clear_pair_cache:
            self.chrono_result_display_key = None
            self.chrono_result_rows_by_action = {}
        self.label_chrono_status.setText("Prêt - lancer une mesure BE ou BR.")
        self.label_chrono_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")

    def chrono_relay_type(self):
        text = self.comboBox_chrono_type_relais.currentText().strip().lower()
        return "MONOSTABLE" if text.startswith("mono") else "BISTABLE"

    def chrono_update_relay_type_ui(self, *_args):
        if self.chrono_relay_type() == "MONOSTABLE":
            self.pushButton_chrono_mesure_be.setText("MESURER ENCLENCHEMENT : OFF vers ON")
            self.pushButton_chrono_mesure_br.setText("MESURER DÉCLENCHEMENT : ON vers OFF")
            self.pushButton_chrono_mesure_be_br.setText("MESURER ON / OFF AUTOMATIQUE")
            self.label_chrono_status.setText("Mode monostable : GP14 pilote la bobine, GP15 non utilisé.")
        else:
            self.pushButton_chrono_mesure_be.setText("MESURER BE : repos vers travail")
            self.pushButton_chrono_mesure_br.setText("MESURER BR : travail vers repos")
            self.pushButton_chrono_mesure_be_br.setText("MESURER BE / BR AUTOMATIQUE")
            self.label_chrono_status.setText("Mode bistable : BE sur GP14, BR sur GP15.")
        self.label_chrono_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")

    def chrono_float_ms(self, line_edit, nom):
        texte = line_edit.text().strip().replace(" ", "").replace(",", ".")
        try:
            valeur = float(texte)
        except Exception as exc:
            raise ValueError(f"{nom} doit être un nombre en ms.") from exc
        if valeur <= 0:
            raise ValueError(f"{nom} doit être supérieur à 0 ms.")
        return valeur

    def chrono_metadata(self):
        if not self.lineEdit_chrono_date.text().strip():
            self.lineEdit_chrono_date.setText(QDate.currentDate().toString("dd/MM/yyyy"))
        return {
            "lot": self.lineEdit_chrono_lot.text().strip(),
            "date_test": self.lineEdit_chrono_date.text().strip(),
            "relais": self.lineEdit_chrono_relais.text().strip(),
            "ambiance_c": self.lineEdit_chrono_ambiance.text().strip(),
            "nom_test": self.lineEdit_chrono_nom_test.text().strip(),
            "sn": self.lineEdit_chrono_sn.text().strip(),
        }

    def chrono_format_input_ms(self, value, default=""):
        try:
            numeric = float(value)
        except Exception:
            return str(default)
        if abs(numeric - round(numeric)) < 0.0005:
            return str(int(round(numeric)))
        return f"{numeric:.3f}".rstrip("0").rstrip(".").replace(".", ",")

    def chrono_clear_lot_entry_fields(self, keep_lot=True):
        lot = self.lineEdit_chrono_lot.text().strip()
        self._chrono_lot_autofill_running = True
        try:
            if keep_lot:
                self.lineEdit_chrono_lot.setText(lot)
            else:
                self.lineEdit_chrono_lot.clear()
            self.lineEdit_chrono_date.setText(QDate.currentDate().toString("dd/MM/yyyy"))
            self.lineEdit_chrono_relais.clear()
            self.lineEdit_chrono_ambiance.clear()
            self.lineEdit_chrono_nom_test.clear()
            self.lineEdit_chrono_sn.clear()
            self.lineEdit_chrono_resultat.clear()
        finally:
            self._chrono_lot_autofill_running = False
        self.chrono_result_display_key = None
        self.chrono_result_rows_by_action = {}
        self.tableWidget_chrono_results.setRowCount(0)
        self.tableWidget_chrono_events.setRowCount(0)
        self.oscillo_clear_view()

    def chrono_on_lot_edited(self, _text=""):
        if getattr(self, "_chrono_lot_autofill_running", False):
            return
        self.chrono_clear_lot_entry_fields(keep_lot=True)
        self.label_chrono_status.setText("Lot modifié - valider le champ Lot pour rechercher en base.")
        self.label_chrono_status.setStyleSheet("background-color: rgb(255,235,150); color: black; font-weight: bold; border: 2px solid rgb(160,100,0);")

    def chrono_autofill_from_lot(self):
        if getattr(self, "_chrono_lot_autofill_running", False):
            return
        lot = self.lineEdit_chrono_lot.text().strip()
        if not lot:
            self.chrono_clear_lot_entry_fields(keep_lot=False)
            self.label_chrono_status.setText("Lot vide - renseigner un numéro de lot avant mesure.")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            return
        try:
            self.chrono_init_db()
            with self.chrono_connect_db() as con:
                record = con.execute(
                    """
                    SELECT lot, date_test, relais, ambiance_c, nom_test, sn,
                           relay_type, nb_inverseurs, limite_temps_ms, limite_rebond_ms
                    FROM mesures_chrono_contacts
                    WHERE lot = ?
                    ORDER BY timestamp DESC, id DESC
                    LIMIT 1
                    """,
                    (lot,),
                ).fetchone()
        except Exception as exc:
            self.label_chrono_status.setText(f"Recherche lot chronométrie impossible : {exc}")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            return
        if record is None:
            self.chrono_clear_lot_entry_fields(keep_lot=True)
            self.label_chrono_status.setText(f"Nouveau lot chronométrie : {lot} - compléter tous les champs avant mesure.")
            self.label_chrono_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")
            return

        last_sn = str(record["sn"] or "").strip()
        next_sn = self.next_sn_value(last_sn)
        self._chrono_lot_autofill_running = True
        try:
            self.lineEdit_chrono_lot.setText(str(record["lot"] or lot))
            self.lineEdit_chrono_date.setText(str(record["date_test"] or QDate.currentDate().toString("dd/MM/yyyy")))
            self.lineEdit_chrono_relais.setText(str(record["relais"] or ""))
            self.lineEdit_chrono_ambiance.setText(str(record["ambiance_c"] or ""))
            self.lineEdit_chrono_nom_test.setText(str(record["nom_test"] or ""))
            relay_type = str(record["relay_type"] or "").strip().lower()
            if "mono" in relay_type:
                self.comboBox_chrono_type_relais.setCurrentText("Monostable")
            else:
                self.comboBox_chrono_type_relais.setCurrentText("Bistable")
            self.spinBox_chrono_nb_inverseurs.setValue(self.clamp_nb_inverseurs(record["nb_inverseurs"], 2))
            self.lineEdit_chrono_limite_temps_ms.setText(self.chrono_format_input_ms(record["limite_temps_ms"], "1,5"))
            self.lineEdit_chrono_limite_rebond_ms.setText(self.chrono_format_input_ms(record["limite_rebond_ms"], "2"))
            if last_sn and next_sn and next_sn != last_sn:
                self.lineEdit_chrono_sn.setText(next_sn)
                sn_status = f"prochain SN prêt : {next_sn}"
            elif last_sn:
                self.lineEdit_chrono_sn.clear()
                sn_status = f"dernier SN {last_sn} non incrémentable, saisir le SN suivant"
            else:
                self.lineEdit_chrono_sn.clear()
                sn_status = "aucun dernier SN trouvé, saisir le SN"
            self.lineEdit_chrono_resultat.clear()
        finally:
            self._chrono_lot_autofill_running = False
        self.chrono_result_display_key = None
        self.chrono_result_rows_by_action = {}
        self.tableWidget_chrono_results.setRowCount(0)
        self.tableWidget_chrono_events.setRowCount(0)
        self.oscillo_clear_view()
        self.chrono_update_relay_type_ui()
        self.label_chrono_status.setText(f"Lot chronométrie existant chargé : {lot} - {sn_status}.")
        self.label_chrono_status.setStyleSheet("background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);")

    def chrono_result_key(self, meta, relay_type, nb_inv):
        return (
            str(meta.get("lot", "")),
            str(meta.get("date_test", "")),
            str(meta.get("relais", "")),
            str(meta.get("nom_test", "")),
            str(meta.get("sn", "")),
            str(relay_type or ""),
            int(nb_inv or 0),
        )

    def chrono_validate_metadata(self):
        meta = self.chrono_metadata()
        labels = {
            "lot": "Lot",
            "date_test": "Date du test",
            "relais": "Relais",
            "ambiance_c": "Ambiance",
            "nom_test": "Nom du Test",
            "sn": "Numéro de Relais",
        }
        missing = [labels[key] for key, value in meta.items() if not str(value).strip()]
        if missing:
            raise ValueError(
                "Champs obligatoires manquants : "
                + ", ".join(missing)
                + ".\n\nAucune mesure ne peut être démarrée tant que tous les champs ne sont pas remplis."
            )
        return meta

    def chrono_start_measure_be_br(self):
        try:
            pulse_ms = int(self.spinBox_chrono_pulse_ms.value())
            if int(self.spinBox_chrono_capture_ms.value()) < pulse_ms:
                raise ValueError("La durée pulse / maintien doit être inférieure ou égale à la fenêtre de capture.")
            self.chrono_float_ms(self.lineEdit_chrono_limite_temps_ms, "Sanction temps max")
            self.chrono_float_ms(self.lineEdit_chrono_limite_rebond_ms, "Sanction rebond max")
            self.chrono_validate_metadata()
            self.chrono_clear_results()
            self.chrono_auto_sequence_active = True
            self.lineEdit_chrono_resultat.setText("EN COURS")
            if self.chrono_relay_type() == "MONOSTABLE":
                self.chrono_auto_sequence_queue = ["MONO_ON", "MONO_OFF"]
                self.chrono_auto_prereset_pending = False
                self.chrono_measure_running = True
                self.update_button_states()
                self.label_chrono_status.setText("Automatique monostable : mesure enclenchement puis déclenchement sur GP14.")
                self.label_chrono_status.setStyleSheet("background-color: rgb(255,235,150); color: black; font-weight: bold; border: 2px solid rgb(160,100,0);")
                QTimer.singleShot(0, self.chrono_start_next_auto_measure)
            else:
                self.chrono_auto_sequence_queue = ["BE", "BR"]
                self.chrono_auto_prereset_pending = True
                self.chrono_measure_running = True
                self.update_button_states()
                self.label_chrono_status.setText("Pré-positionnement RESET : pulse BR non mesuré avant chronométrie BE/BR.")
                self.label_chrono_status.setStyleSheet("background-color: rgb(255,235,150); color: black; font-weight: bold; border: 2px solid rgb(160,100,0);")
                source_suffix = ";EA" if getattr(self, "chrono_external_supply_mode", False) else ""
                self.send_command(f"PULSE_US;BR;{pulse_ms * 1000}{source_suffix}")
            return True
        except Exception as exc:
            self.chrono_measure_running = False
            self.chrono_auto_sequence_active = False
            self.chrono_auto_sequence_queue = []
            self.chrono_auto_prereset_pending = False
            self.update_button_states()
            self.label_chrono_status.setText(f"Mesure refusée : {exc}")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            QMessageBox.warning(self.window, "Chronométrie contacts", str(exc))
            return False

    def chrono_start_next_auto_measure(self):
        if not self.chrono_auto_sequence_active or not self.chrono_auto_sequence_queue:
            return
        next_action = self.chrono_auto_sequence_queue.pop(0)
        self.chrono_start_measure(next_action, from_auto_sequence=True)

    def chrono_start_measure(self, action, from_auto_sequence=False):
        try:
            if not from_auto_sequence:
                self.chrono_auto_sequence_active = False
                self.chrono_auto_sequence_queue = []
                self.chrono_auto_prereset_pending = False
            action = str(action or "").strip().upper()
            if self.chrono_relay_type() == "MONOSTABLE" and action == "BE":
                action = "MONO_ON"
            elif self.chrono_relay_type() == "MONOSTABLE" and action == "BR":
                action = "MONO_OFF"
            if action not in ("BE", "BR", "MONO_ON", "MONO_OFF"):
                raise ValueError("Action chronométrie invalide.")
            capture_ms = int(self.spinBox_chrono_capture_ms.value())
            pulse_ms = int(self.spinBox_chrono_pulse_ms.value())
            nb_inv = int(self.spinBox_chrono_nb_inverseurs.value())
            if pulse_ms > capture_ms:
                raise ValueError("La durée pulse doit être inférieure ou égale à la fenêtre de capture.")
            limite_temps = self.chrono_float_ms(self.lineEdit_chrono_limite_temps_ms, "Sanction temps max")
            limite_rebond = self.chrono_float_ms(self.lineEdit_chrono_limite_rebond_ms, "Sanction rebond max")
            meta = self.chrono_validate_metadata()
            relay_type = "MONOSTABLE" if action.startswith("MONO_") else "BISTABLE"
            display_key = self.chrono_result_key(meta, relay_type, nb_inv)
            previous_display_key = self.chrono_result_display_key
            if self.chrono_result_display_key != display_key:
                self.chrono_result_rows_by_action = {}
                self.chrono_result_display_key = display_key
            self.chrono_clear_results(
                clear_pair_cache=False,
                clear_oscillo_history=(previous_display_key != display_key),
            )
            self.chrono_current = {
                "action": action,
                "capture_us": capture_ms * 1000,
                "pulse_us": pulse_ms * 1000,
                "nb_inv": nb_inv,
                "capture_ms": capture_ms,
                "pulse_ms": pulse_ms,
                "limite_temps_ms": limite_temps,
                "limite_rebond_ms": limite_rebond,
                "meta": meta,
                "relay_type": relay_type,
                "start_bits": None,
                "end_bits": None,
                "overflow": False,
            }
            self.lineEdit_chrono_resultat.setText("EN COURS")
            self.chrono_measure_running = True
            self.update_button_states()
            self.label_chrono_status.setText(f"Mesure {action} en cours : capture {capture_ms} ms, pulse {pulse_ms} ms.")
            self.label_chrono_status.setStyleSheet("background-color: rgb(255,235,150); color: black; font-weight: bold; border: 2px solid rgb(160,100,0);")
            source_suffix = ";EA" if getattr(self, "chrono_external_supply_mode", False) else ""
            if action == "MONO_ON":
                self.send_command(f"MEASURE_MONO;ON;{capture_ms * 1000};{pulse_ms * 1000};{nb_inv}{source_suffix}")
            elif action == "MONO_OFF":
                self.send_command(f"MEASURE_MONO;OFF;{capture_ms * 1000};{pulse_ms * 1000};{nb_inv}{source_suffix}")
            else:
                self.send_command(f"MEASURE_CONTACTS;{action};{capture_ms * 1000};{pulse_ms * 1000};{nb_inv}{source_suffix}")
        except Exception as exc:
            self.chrono_measure_running = False
            if not from_auto_sequence:
                self.chrono_auto_sequence_active = False
                self.chrono_auto_sequence_queue = []
            self.update_button_states()
            self.label_chrono_status.setText(f"Mesure refusée : {exc}")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            QMessageBox.warning(self.window, "Chronométrie contacts", str(exc))
            if getattr(self, "measure_all_active", False) and self.measure_all_phase == "CHRONO":
                self.voltage_measure_all_fail(f"Chronométrie refusée : {exc}")

    def parse_measure_fields(self, line):
        fields = line.split(";")
        data = {"frame": fields[0], "kind": ""}
        start_index = 1
        if len(fields) > 1 and "=" not in fields[1]:
            data["kind"] = fields[1]
            start_index = 2
        for field in fields[start_index:]:
            if "=" in field:
                k, v = field.split("=", 1)
                data[k] = v
        return data

    def parse_measure_frame(self, line):
        data = self.parse_measure_fields(line)
        frame = data.get("frame", "")
        kind = data.get("kind", "")
        if frame == "MEASURE_EVT":
            self.chrono_events.append({
                "i": int(data.get("I", len(self.chrono_events))),
                "t_us": int(data.get("T_US", "0")),
                "contact": data.get("CONTACT", "?"),
                "state": data.get("STATE", "?"),
            })
            self.chrono_refresh_events_table()
            return
        if kind == "BEGIN":
            self.chrono_events = []
            self.tableWidget_chrono_events.setRowCount(0)
            self.tableWidget_chrono_results.setRowCount(0)
            self.oscillo_clear_view(clear_history=False)
            self.chrono_current.update(data)
            self.chrono_current["start_bits"] = self.chrono_parse_int(data.get("START_BITS"))
            self.label_chrono_status.setText(f"Capture RP2040 démarrée : {data.get('ACTION', '?')}")
            return
        if kind == "END":
            self.chrono_current.update(data)
            self.chrono_current["end_bits"] = self.chrono_parse_int(data.get("END_BITS"))
            self.chrono_current["overflow"] = data.get("OVERFLOW", "0") == "1"
            self.chrono_current["event_capacity"] = self.chrono_parse_int(data.get("EVENT_CAPACITY"), 0)
            self.chrono_current["dropped_events"] = self.chrono_parse_int(data.get("DROPPED_EVENTS"), 0)
            self.chrono_current["loop_max_us"] = self.chrono_parse_int(data.get("LOOP_MAX_US"), 0)
            self.chrono_measure_running = False
            self.update_button_states()
            self.oscillo_store_current_capture(select=True)
            self.chrono_compute_results()
            self.oscillo_set_full_view(update=False)
            self.oscillo_update_view()
            return
        if kind == "ERROR":
            reason = data.get("REASON", "erreur inconnue")
            self.chrono_measure_running = False
            self.chrono_auto_sequence_active = False
            self.chrono_auto_sequence_queue = []
            self.update_button_states()
            self.lineEdit_chrono_resultat.setText("DEFAUT")
            self.label_chrono_status.setText(f"Mesure refusée par le RP2040 : {reason}")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
            if getattr(self, "measure_all_active", False) and self.measure_all_phase == "CHRONO":
                self.voltage_measure_all_fail(f"Chronométrie refusée par le RP2040 : {reason}")

    def chrono_parse_int(self, value, default=None):
        try:
            return int(str(value))
        except Exception:
            return default

    def chrono_bit_state(self, bits, contact):
        idx = self.chrono_contact_index(contact)
        if bits is None or idx is None:
            return "--"
        return "1" if ((bits >> idx) & 0x01) else "0"

    def chrono_contact_index(self, contact):
        names = {"R1": 0, "R2": 1, "R3": 2, "R4": 3, "T1": 4, "T2": 5, "T3": 6, "T4": 7}
        return names.get(str(contact).upper())

    def chrono_refresh_events_table(self):
        self.tableWidget_chrono_events.setRowCount(len(self.chrono_events))
        for row, event in enumerate(self.chrono_events):
            values = [
                str(event["i"]),
                str(event["t_us"]),
                event["contact"],
                "fermé" if str(event["state"]) == "1" else "ouvert",
            ]
            for col, value in enumerate(values):
                self.tableWidget_chrono_events.setItem(row, col, QTableWidgetItem(value))

    def oscillo_capture_action_label(self, action):
        action = str(action or "").upper()
        labels = {
            "BE": "BE - repos vers travail",
            "BR": "BR - travail vers repos",
            "MONO_ON": "MONO ON - repos vers travail",
            "MONO_OFF": "MONO OFF - travail vers repos",
            "CYCLE_BE_BR": "Cycle complet BE → BR",
            "CYCLE_MONO": "Cycle complet MONO ON → MONO OFF",
        }
        return labels.get(action, action or "Mesure")

    def oscillo_capture_snapshot(self):
        current = copy.deepcopy(self.chrono_current or {})
        events = copy.deepcopy(self.chrono_events or [])
        action = str(current.get("ACTION") or current.get("action") or "").upper()
        capture_us = self.chrono_parse_int(current.get("CAPTURE_US") or current.get("capture_us"), None)
        if capture_us is None:
            try:
                capture_us = int(float(current.get("capture_ms") or 0) * 1000)
            except Exception:
                capture_us = 0
        if not capture_us:
            if events:
                capture_us = max([int(event.get("t_us", 0)) for event in events] + [1])
            else:
                try:
                    capture_us = int(self.spinBox_chrono_capture_ms.value()) * 1000
                except Exception:
                    capture_us = 1
        return {
            "action": action,
            "events": events,
            "current": current,
            "start_bits": current.get("start_bits"),
            "capture_us": int(capture_us or 0),
            "title": self.oscillo_title_from_current(current),
        }

    def oscillo_title_from_current(self, current):
        current = dict(current or {})
        meta = current.get("meta") if isinstance(current.get("meta"), dict) else {}
        action = str(current.get("ACTION") or current.get("action") or "?").upper()
        lot = meta.get("lot") or self.lineEdit_chrono_lot.text().strip()
        sn = meta.get("sn") or self.lineEdit_chrono_sn.text().strip()
        relais = meta.get("relais") or self.lineEdit_chrono_relais.text().strip()
        return f"Oscillogramme {action} - Lot {lot} - SN {sn} - {relais}".strip()

    def oscillo_store_current_capture(self, select=True):
        capture = self.oscillo_capture_snapshot()
        action = capture.get("action") or "MESURE"
        if not hasattr(self, "oscillo_captures"):
            self.oscillo_captures = {}
        self.oscillo_captures["LAST"] = capture
        self.oscillo_captures[action] = capture
        self.oscillo_rebuild_combined_captures()
        if select:
            selected = action
            if action == "BR" and self.oscillo_captures.get("CYCLE_BE_BR"):
                selected = "CYCLE_BE_BR"
            elif action == "MONO_OFF" and self.oscillo_captures.get("CYCLE_MONO"):
                selected = "CYCLE_MONO"
            self.oscillo_selected_capture_key = selected
        self.oscillo_refresh_capture_combo(select_key=self.oscillo_selected_capture_key if select else None)

    def oscillo_reconstructed_start_bits(self, action, nb_inv):
        nb_inv = max(1, min(4, int(nb_inv or 4)))
        action = str(action or "").upper()
        bits = 0
        if action in ("BE", "MONO_ON"):
            # Départ attendu : repos fermé R, travail ouvert T.
            for i in range(nb_inv):
                bits |= (1 << i)
        elif action in ("BR", "MONO_OFF"):
            # Départ attendu : travail fermé T, repos ouvert R.
            for i in range(nb_inv):
                bits |= (1 << (4 + i))
        return bits

    def oscillo_rebuild_combined_captures(self):
        captures = getattr(self, "oscillo_captures", {}) or {}
        for first_key, second_key, cycle_key, label_first, label_second in (
            ("BE", "BR", "CYCLE_BE_BR", "BE", "BR"),
            ("MONO_ON", "MONO_OFF", "CYCLE_MONO", "MONO ON", "MONO OFF"),
        ):
            first = captures.get(first_key)
            second = captures.get(second_key)
            if not first or not second:
                captures.pop(cycle_key, None)
                continue
            first_events = copy.deepcopy(first.get("events", []))
            second_events = copy.deepcopy(second.get("events", []))
            first_capture_us = int(first.get("capture_us") or 0)
            if first_events:
                first_capture_us = max(first_capture_us, max(int(e.get("t_us", 0)) for e in first_events) + 1)
            offset = max(1, first_capture_us) + OSCILLO_COMBINED_GAP_US
            events = []
            idx = 0
            for event in sorted(first_events, key=lambda e: (int(e.get("t_us", 0)), int(e.get("i", 0)))):
                e = copy.deepcopy(event)
                e["i"] = idx
                e["phase"] = label_first
                events.append(e)
                idx += 1
            for event in sorted(second_events, key=lambda e: (int(e.get("t_us", 0)), int(e.get("i", 0)))):
                e = copy.deepcopy(event)
                e["i"] = idx
                e["phase"] = label_second
                e["t_us"] = int(e.get("t_us", 0)) + offset
                events.append(e)
                idx += 1
            current = copy.deepcopy(first.get("current", {}))
            current["ACTION"] = cycle_key
            current["action"] = cycle_key
            current["combined_cycle"] = True
            current["phase_markers"] = [{"t_us": 0, "label": label_first}, {"t_us": offset, "label": label_second}]
            try:
                nb_inv = int(current.get("NB_INV") or current.get("nb_inv") or second.get("current", {}).get("NB_INV") or second.get("current", {}).get("nb_inv") or 4)
            except Exception:
                nb_inv = 4
            current["NB_INV"] = max(1, min(4, nb_inv))
            second_capture_us = int(second.get("capture_us") or 0)
            if second_events:
                second_capture_us = max(second_capture_us, max(int(e.get("t_us", 0)) for e in second_events) + 1)
            captures[cycle_key] = {
                "action": cycle_key,
                "events": events,
                "current": current,
                "start_bits": first.get("start_bits"),
                "capture_us": offset + max(1, second_capture_us),
                "title": f"Oscillogramme {label_first} → {label_second} - cycle complet",
            }
        self.oscillo_captures = captures

    def oscillo_refresh_capture_combo(self, select_key=None):
        if not hasattr(self, "comboBox_oscillo_capture"):
            return
        previous = self.comboBox_oscillo_capture.blockSignals(True)
        try:
            current_key = select_key or self.comboBox_oscillo_capture.currentData() or getattr(self, "oscillo_selected_capture_key", "LAST")
            self.comboBox_oscillo_capture.clear()
            if getattr(self, "oscillo_captures", None):
                self.comboBox_oscillo_capture.addItem("Dernière mesure", "LAST")
                order = ["CYCLE_BE_BR", "BE", "BR", "CYCLE_MONO", "MONO_ON", "MONO_OFF"]
                for key in order:
                    cap = self.oscillo_captures.get(key)
                    if cap:
                        self.comboBox_oscillo_capture.addItem(
                            f"{self.oscillo_capture_action_label(key)} ({len(cap.get('events', []))} fronts)",
                            key,
                        )
                if current_key not in self.oscillo_captures:
                    current_key = "LAST"
                for index in range(self.comboBox_oscillo_capture.count()):
                    if self.comboBox_oscillo_capture.itemData(index) == current_key:
                        self.comboBox_oscillo_capture.setCurrentIndex(index)
                        break
                self.oscillo_selected_capture_key = current_key
            else:
                self.comboBox_oscillo_capture.addItem("Aucune mesure", "LAST")
                self.oscillo_selected_capture_key = "LAST"
        finally:
            self.comboBox_oscillo_capture.blockSignals(previous)

    def oscillo_on_capture_changed(self, _index=0):
        key = self.comboBox_oscillo_capture.currentData() if hasattr(self, "comboBox_oscillo_capture") else "LAST"
        self.oscillo_selected_capture_key = key or "LAST"
        self.oscillo_set_full_view(update=False)
        self.oscillo_update_view()

    def oscillo_active_capture(self):
        captures = getattr(self, "oscillo_captures", {}) or {}
        key = getattr(self, "oscillo_selected_capture_key", None)
        if hasattr(self, "comboBox_oscillo_capture"):
            key = self.comboBox_oscillo_capture.currentData() or key
        if key in captures:
            return captures[key]
        if "LAST" in captures:
            return captures["LAST"]
        return self.oscillo_capture_snapshot()

    def oscillo_active_events(self):
        return list((self.oscillo_active_capture() or {}).get("events", []))

    def oscillo_active_current(self):
        return dict((self.oscillo_active_capture() or {}).get("current", {}))

    def oscillo_active_start_bits(self):
        return (self.oscillo_active_capture() or {}).get("start_bits")

    def oscillo_display_mode(self):
        if hasattr(self, "comboBox_oscillo_display_mode"):
            mode = self.comboBox_oscillo_display_mode.currentData()
            if mode in (OSCILLO_DISPLAY_ELECTRIC, OSCILLO_DISPLAY_LOGIC, OSCILLO_DISPLAY_SYNTHESIS):
                return mode
        return OSCILLO_DISPLAY_ELECTRIC

    def oscillo_display_mode_label(self):
        mode = self.oscillo_display_mode()
        if mode == OSCILLO_DISPLAY_LOGIC:
            return "Logique contact"
        if mode == OSCILLO_DISPLAY_SYNTHESIS:
            return "Synthèse transfert/rebonds"
        return "Électrique GPIO"

    def oscillo_load_saved_capture_dialog(self):
        """Recharge un ancien lot/SN depuis chronometrie_contacts.sqlite3.

        Sélection en deux étapes pour éviter une liste mélangeant tous les lots
        et tous les SN : d'abord le lot, puis uniquement les SN de ce lot.
        Les boîtes sont éditables : l'opérateur peut choisir dans la liste ou
        saisir directement le lot/SN s'il s'en souvient.
        """
        if getattr(self, "chrono_measure_running", False):
            QMessageBox.information(self.window, "Rappel oscillogramme", "Impossible pendant une capture en cours.")
            return

        try:
            self.chrono_init_db()
            with self.chrono_connect_db() as con:
                lot_rows = con.execute(
                    """
                    SELECT lot,
                           COUNT(*) AS nb_mesures,
                           COUNT(DISTINCT sn) AS nb_sn,
                           MAX(timestamp) AS ts
                    FROM mesures_chrono_contacts
                    GROUP BY lot
                    ORDER BY ts DESC, lot COLLATE NOCASE ASC
                    LIMIT 500
                    """
                ).fetchall()
        except Exception as exc:
            QMessageBox.warning(self.window, "Rappel oscillogramme", f"Lecture base impossible : {exc}")
            return

        if not lot_rows:
            QMessageBox.information(self.window, "Rappel oscillogramme", "Aucune mesure chronométrie enregistrée.")
            return

        current_lot = self.lineEdit_chrono_lot.text().strip() if hasattr(self, "lineEdit_chrono_lot") else ""
        lot_choices = []
        lot_map = {}
        lot_index = 0
        for index, row in enumerate(lot_rows):
            raw_lot = str(row["lot"] or "").strip()
            display_lot = raw_lot if raw_lot else "(lot vide)"
            label = (
                f"{display_lot} | {int(row['nb_sn'] or 0)} SN | "
                f"{int(row['nb_mesures'] or 0)} mesure(s) | dernière {self.format_datetime_fr(row['ts'])}"
            )
            lot_choices.append(label)
            lot_map[label] = raw_lot
            if current_lot and raw_lot.lower() == current_lot.lower():
                lot_index = index

        lot_choice, ok = QInputDialog.getItem(
            self.window,
            "Rappel oscillogramme - choix du lot",
            "Choisir un lot dans la liste ou saisir le numéro de lot :",
            lot_choices,
            lot_index,
            True,
        )
        if not ok:
            return
        selected_lot = lot_map.get(str(lot_choice), str(lot_choice or "").strip())
        if selected_lot == "(lot vide)":
            selected_lot = ""

        try:
            with self.chrono_connect_db() as con:
                sn_rows = con.execute(
                    """
                    SELECT lot, relais, nom_test, sn, MAX(timestamp) AS ts, COUNT(*) AS n
                    FROM mesures_chrono_contacts
                    WHERE lot = ?
                    GROUP BY lot, relais, nom_test, sn
                    ORDER BY sn COLLATE NOCASE ASC, ts DESC
                    """,
                    (selected_lot,),
                ).fetchall()
        except Exception as exc:
            QMessageBox.warning(self.window, "Rappel oscillogramme", f"Lecture SN impossible : {exc}")
            return

        if not sn_rows:
            QMessageBox.information(
                self.window,
                "Rappel oscillogramme",
                f"Aucun SN trouvé pour le lot {selected_lot or '(lot vide)'}.",
            )
            return

        sn_rows = [dict(row) for row in sn_rows]
        sn_rows.sort(
            key=lambda row: (
                self.production_sn_sort_key(row.get("sn", "")),
                str(row.get("relais", "") or "").lower(),
                str(row.get("nom_test", "") or "").lower(),
                str(row.get("ts", "") or ""),
            )
        )

        current_sn = self.lineEdit_chrono_sn.text().strip() if hasattr(self, "lineEdit_chrono_sn") else ""
        sn_choices = []
        sn_map = {}
        sn_index = 0
        for index, row in enumerate(sn_rows):
            raw_sn = str(row["sn"] or "").strip()
            display_sn = raw_sn if raw_sn else "(SN vide)"
            label = (
                f"{display_sn} | Relais {row['relais'] or '-'} | Test {row['nom_test'] or '-'} | "
                f"{int(row['n'] or 0)} mesure(s) | dernière {self.format_datetime_fr(row['ts'])}"
            )
            sn_choices.append(label)
            sn_map[label] = dict(row)
            if current_sn and raw_sn.lower() == current_sn.lower():
                sn_index = index

        sn_choice, ok = QInputDialog.getItem(
            self.window,
            "Rappel oscillogramme - choix du SN",
            f"Lot {selected_lot or '(lot vide)'} : choisir un SN ou saisir le SN :",
            sn_choices,
            sn_index,
            True,
        )
        if not ok:
            return

        selected = sn_map.get(str(sn_choice))
        if selected is None:
            typed_sn = str(sn_choice or "").strip()
            if typed_sn == "(SN vide)":
                typed_sn = ""
            matches = [dict(row) for row in sn_rows if str(row["sn"] or "").strip().lower() == typed_sn.lower()]
            if not matches:
                QMessageBox.information(
                    self.window,
                    "Rappel oscillogramme",
                    f"SN {typed_sn or '(SN vide)'} introuvable dans le lot {selected_lot or '(lot vide)'}.",
                )
                return
            if len(matches) == 1:
                selected = matches[0]
            else:
                detail_choices = []
                detail_map = {}
                for row in matches:
                    label = (
                        f"SN {row['sn'] or '(SN vide)'} | Relais {row['relais'] or '-'} | "
                        f"Test {row['nom_test'] or '-'} | {int(row['n'] or 0)} mesure(s) | "
                        f"dernière {self.format_datetime_fr(row['ts'])}"
                    )
                    detail_choices.append(label)
                    detail_map[label] = row
                detail_choice, ok = QInputDialog.getItem(
                    self.window,
                    "Rappel oscillogramme - précision mesure",
                    "Plusieurs groupes existent pour ce SN. Choisir le groupe :",
                    detail_choices,
                    0,
                    False,
                )
                if not ok or not detail_choice:
                    return
                selected = detail_map.get(detail_choice)

        if selected is None:
            return

        if hasattr(self, "lineEdit_chrono_lot"):
            self.lineEdit_chrono_lot.setText(str(selected.get("lot", "") or ""))
        if hasattr(self, "lineEdit_chrono_sn"):
            self.lineEdit_chrono_sn.setText(str(selected.get("sn", "") or ""))
        self.oscillo_load_saved_capture_group(selected)

    def oscillo_load_saved_capture_group(self, selected):
        try:
            with self.chrono_connect_db() as con:
                records = con.execute(
                    """
                    SELECT lot, date_test, relais, ambiance_c, nom_test, sn,
                           relay_type, action, nb_inverseurs, capture_ms, pulse_ms,
                           limite_temps_ms, limite_rebond_ms, resultat, overflow,
                           details_json, events_json, timestamp
                    FROM mesures_chrono_contacts
                    WHERE lot = ? AND relais = ? AND nom_test = ? AND sn = ?
                    ORDER BY timestamp ASC, id ASC
                    """,
                    (
                        str(selected.get("lot", "")),
                        str(selected.get("relais", "")),
                        str(selected.get("nom_test", "")),
                        str(selected.get("sn", "")),
                    ),
                ).fetchall()
        except Exception as exc:
            QMessageBox.warning(self.window, "Rappel oscillogramme", f"Chargement impossible : {exc}")
            return
        if not records:
            QMessageBox.information(self.window, "Rappel oscillogramme", "Aucune capture trouvée pour cette sélection.")
            return
        captures = {}
        for record in records:
            action = str(record["action"] or "").upper()
            try:
                events = json.loads(record["events_json"] or "[]")
                if not isinstance(events, list):
                    events = []
            except Exception:
                events = []
            nb_inv = max(1, min(4, int(record["nb_inverseurs"] or 4)))
            current = {
                "ACTION": action,
                "action": action,
                "NB_INV": nb_inv,
                "nb_inv": nb_inv,
                "capture_ms": int(record["capture_ms"] or 0),
                "pulse_ms": int(record["pulse_ms"] or 0),
                "limite_temps_ms": float(record["limite_temps_ms"] or 0),
                "limite_rebond_ms": float(record["limite_rebond_ms"] or 0),
                "relay_type": str(record["relay_type"] or ""),
                "meta": {
                    "lot": record["lot"],
                    "date_test": record["date_test"],
                    "relais": record["relais"],
                    "ambiance_c": record["ambiance_c"],
                    "nom_test": record["nom_test"],
                    "sn": record["sn"],
                },
            }
            start_bits = self.oscillo_reconstructed_start_bits(action, nb_inv)
            capture_us = int(record["capture_ms"] or 0) * 1000
            if events:
                capture_us = max(capture_us, max(int(e.get("t_us", 0)) for e in events) + 1)
            cap = {
                "action": action,
                "events": events,
                "current": current,
                "start_bits": start_bits,
                "capture_us": max(1, capture_us),
                "title": self.oscillo_title_from_current(current),
            }
            captures["LAST"] = cap
            captures[action] = cap
        self.oscillo_captures = captures
        self.oscillo_rebuild_combined_captures()
        selected_key = "CYCLE_BE_BR" if self.oscillo_captures.get("CYCLE_BE_BR") else "CYCLE_MONO" if self.oscillo_captures.get("CYCLE_MONO") else "LAST"
        self.oscillo_selected_capture_key = selected_key
        self.oscillo_refresh_capture_combo(select_key=selected_key)
        self.oscillo_set_full_view(update=False)
        self.oscillo_update_view()
        meta = selected
        self.label_oscillo_status.setText(f"Oscillogramme rappelé : lot {meta.get('lot')} / SN {meta.get('sn')}")
        self.label_oscillo_status.setStyleSheet("background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);")

    def oscillo_capture_us(self):
        capture = self.oscillo_active_capture()
        capture_us = self.chrono_parse_int(capture.get("capture_us"), None)
        if capture_us is not None and capture_us > 0:
            return capture_us
        current = capture.get("current", {}) if isinstance(capture.get("current", {}), dict) else {}
        for key in ("CAPTURE_US", "capture_us"):
            parsed = self.chrono_parse_int(current.get(key), None)
            if parsed is not None and parsed > 0:
                return parsed
        try:
            return int(float(current.get("capture_ms") or self.spinBox_chrono_capture_ms.value()) * 1000)
        except Exception:
            return max([int(event.get("t_us", 0)) for event in self.oscillo_active_events()] + [1])

    def oscillo_title(self):
        capture = self.oscillo_active_capture()
        return str(capture.get("title") or self.oscillo_title_from_current(capture.get("current", {})))

    def oscillo_capture_end_us(self):
        capture_us = self.oscillo_capture_us()
        events = self.oscillo_active_events()
        if events:
            capture_us = max(capture_us, max(int(event.get("t_us", 0)) for event in events) + 1)
        return max(1, int(capture_us))

    def oscillo_configure_time_controls(self):
        max_us = max(1, self.oscillo_capture_end_us())
        for spin in (
            self.spinBox_oscillo_zoom_start_us,
            self.spinBox_oscillo_zoom_end_us,
            self.spinBox_oscillo_cursor_a_us,
            self.spinBox_oscillo_cursor_b_us,
        ):
            previous = spin.blockSignals(True)
            spin.setRange(0, max_us)
            spin.setSingleStep(1)
            if spin.value() > max_us:
                spin.setValue(max_us)
            spin.blockSignals(previous)

    def oscillo_set_spin_value(self, spin, value):
        previous = spin.blockSignals(True)
        spin.setValue(max(0, int(value)))
        spin.blockSignals(previous)

    def oscillo_set_time_window(self, start_us, end_us, cursor_a_us=None, cursor_b_us=None, update=True):
        """Applique une fenêtre temps en bloquant les signaux des spinbox.

        Cette fonction remplace les mises à jour successives qui pouvaient donner
        l'impression que le dézoom ne fonctionnait pas lorsque début/fin étaient
        temporairement incohérents pendant la modification des champs.
        """
        capture_us = self.oscillo_capture_end_us()
        start_us = max(0, min(capture_us - 1, int(start_us)))
        end_us = max(1, min(capture_us, int(end_us)))
        if end_us <= start_us:
            if start_us >= capture_us - 1:
                start_us = max(0, capture_us - 2)
                end_us = capture_us
            else:
                end_us = min(capture_us, start_us + 1)
        if cursor_a_us is None:
            cursor_a_us = start_us
        if cursor_b_us is None:
            cursor_b_us = end_us
        cursor_a_us = max(0, min(capture_us, int(cursor_a_us)))
        cursor_b_us = max(0, min(capture_us, int(cursor_b_us)))

        self.oscillo_configure_time_controls()
        controls = (
            (self.spinBox_oscillo_zoom_start_us, start_us),
            (self.spinBox_oscillo_zoom_end_us, end_us),
            (self.spinBox_oscillo_cursor_a_us, cursor_a_us),
            (self.spinBox_oscillo_cursor_b_us, cursor_b_us),
        )
        previous_states = []
        try:
            for spin, _value in controls:
                previous_states.append((spin, spin.blockSignals(True)))
            for spin, value in controls:
                spin.setValue(int(value))
        finally:
            for spin, previous in previous_states:
                spin.blockSignals(previous)
        if update:
            self.oscillo_update_view()

    def oscillo_view_range(self):
        capture_us = self.oscillo_capture_end_us()
        start_us = max(0, min(capture_us - 1, int(self.spinBox_oscillo_zoom_start_us.value())))
        end_us = max(1, min(capture_us, int(self.spinBox_oscillo_zoom_end_us.value())))
        if end_us <= start_us:
            if start_us >= capture_us - 1:
                start_us = max(0, capture_us - 2)
                end_us = capture_us
            else:
                end_us = min(capture_us, start_us + 1)
            # Remet les champs en cohérence sans déclencher une cascade de signaux.
            self.oscillo_set_spin_value(self.spinBox_oscillo_zoom_start_us, start_us)
            self.oscillo_set_spin_value(self.spinBox_oscillo_zoom_end_us, end_us)
        return start_us, end_us

    def oscillo_cursor_values(self):
        capture_us = self.oscillo_capture_end_us()
        cursor_a = max(0, min(capture_us, int(self.spinBox_oscillo_cursor_a_us.value())))
        cursor_b = max(0, min(capture_us, int(self.spinBox_oscillo_cursor_b_us.value())))
        return cursor_a, cursor_b

    def oscillo_update_delta_label(self):
        cursor_a, cursor_b = self.oscillo_cursor_values()
        delta = cursor_b - cursor_a
        self.label_oscillo_delta.setText(f"Δ A/B : {delta} µs = {delta / 1000.0:.3f} ms")

    def oscillo_set_full_view(self, update=True):
        capture_us = self.oscillo_capture_end_us()
        self.oscillo_set_time_window(0, capture_us, 0, capture_us, update=update)

    def oscillo_set_zoom_window(self, first_t, last_t, margin_us=50, min_span_us=20):
        capture_us = self.oscillo_capture_end_us()
        first_t = int(first_t)
        last_t = int(last_t)
        if last_t < first_t:
            first_t, last_t = last_t, first_t
        raw_span = max(0, last_t - first_t)
        target_span = max(int(min_span_us), raw_span + 2 * max(0, int(margin_us)))
        center = (first_t + last_t) // 2
        start_us = center - target_span // 2
        end_us = start_us + target_span
        if start_us < 0:
            end_us -= start_us
            start_us = 0
        if end_us > capture_us:
            start_us = max(0, start_us - (end_us - capture_us))
            end_us = capture_us
        if end_us <= start_us:
            end_us = min(capture_us, start_us + 1)
        self.oscillo_set_time_window(start_us, end_us, first_t, last_t, update=True)

    def oscillo_zoom_fronts(self):
        events = self.oscillo_active_events()
        if not events:
            self.oscillo_set_full_view()
            return
        times = sorted(int(event.get("t_us", 0)) for event in events)
        self.oscillo_set_zoom_window(times[0], times[-1], margin_us=100, min_span_us=200)

    def oscillo_zoom_contact(self):
        events = self.oscillo_active_events()
        if not events:
            self.oscillo_set_full_view()
            return
        contact = self.comboBox_oscillo_contact.currentText().strip().upper() if hasattr(self, "comboBox_oscillo_contact") else "AUTO"
        if contact == "AUTO" or contact not in CHRONO_CONTACT_NAMES:
            self.oscillo_zoom_fronts()
            return
        times = sorted(int(event.get("t_us", 0)) for event in events if str(event.get("contact", "")).upper() == contact)
        if not times:
            self.label_oscillo_status.setText(f"Aucun front capturé sur {contact}.")
            self.label_oscillo_status.setStyleSheet("background-color: rgb(255,170,60); color: black; font-weight: bold; border: 2px solid rgb(180,90,0);")
            return
        span = max(1, times[-1] - times[0])
        self.oscillo_set_zoom_window(times[0], times[-1], margin_us=max(10, int(span * 0.05)), min_span_us=80)

    def oscillo_expected_bounce_contacts(self):
        current = self.oscillo_active_current()
        action = str(current.get("ACTION") or current.get("action") or "").upper()
        nb_inv = self.chrono_parse_int(current.get("NB_INV") or current.get("nb_inv"), 4)
        nb_inv = max(1, min(4, nb_inv or 4))
        if action in ("BE", "MONO_ON"):
            return [f"R{i}" for i in range(1, nb_inv + 1)] + [f"T{i}" for i in range(1, nb_inv + 1)]
        if action in ("BR", "MONO_OFF"):
            return [f"T{i}" for i in range(1, nb_inv + 1)] + [f"R{i}" for i in range(1, nb_inv + 1)]
        return CHRONO_CONTACT_NAMES

    def oscillo_selected_or_expected_contacts(self):
        selected = self.comboBox_oscillo_contact.currentText().strip().upper() if hasattr(self, "comboBox_oscillo_contact") else "AUTO"
        if selected in CHRONO_CONTACT_NAMES:
            return [selected]
        return self.oscillo_expected_bounce_contacts()

    def oscillo_rebound_zone_from_events(self, events, contact, state, phase=None):
        expected = str(int(state))
        phase_norm = str(phase or "").upper().strip()
        times = []
        for event in events:
            if str(event.get("contact", "")).upper() != contact:
                continue
            if str(event.get("state", "")) != expected:
                continue
            if phase_norm and str(event.get("phase", "")).upper().strip() != phase_norm:
                continue
            try:
                times.append(int(event.get("t_us", 0)))
            except Exception:
                pass
        times.sort()
        if not times:
            return None
        return times[0], times[-1], max(0, times[-1] - times[0])

    def oscillo_rebound_candidates_for_contact(self, events, contact):
        current = self.oscillo_active_current()
        action = str(current.get("ACTION") or current.get("action") or "").upper()
        combined = bool(current.get("combined_cycle"))
        candidates = []

        def add_candidate(state, label, phase=None):
            zone = self.oscillo_rebound_zone_from_events(events, contact, state, phase=phase)
            if zone is not None:
                first_t, last_t, span = zone
                candidates.append({
                    "contact": contact,
                    "state": int(state),
                    "type": label,
                    "phase": str(phase or "").strip(),
                    "first_t": first_t,
                    "last_t": last_t,
                    "span": span,
                })

        if combined:
            phase_directions = [
                ("BE", "forward"),
                ("MONO ON", "forward"),
                ("BR", "reverse"),
                ("MONO OFF", "reverse"),
            ]
            for phase, direction in phase_directions:
                if direction == "forward":
                    if contact.startswith("R"):
                        add_candidate(0, "ouverture", phase=phase)
                    if contact.startswith("T"):
                        add_candidate(1, "fermeture", phase=phase)
                else:
                    if contact.startswith("T"):
                        add_candidate(0, "ouverture", phase=phase)
                    if contact.startswith("R"):
                        add_candidate(1, "fermeture", phase=phase)
            return candidates

        if action in ("BE", "MONO_ON"):
            if contact.startswith("R"):
                add_candidate(0, "ouverture")
            if contact.startswith("T"):
                add_candidate(1, "fermeture")
        elif action in ("BR", "MONO_OFF"):
            if contact.startswith("T"):
                add_candidate(0, "ouverture")
            if contact.startswith("R"):
                add_candidate(1, "fermeture")
        else:
            add_candidate(0, "ouverture")
            add_candidate(1, "fermeture")
        return candidates

    def oscillo_find_rebound_window_metier(self):
        """Cherche une vraie zone métier de rebond fermeture/ouverture.

        Le bouton ZOOM REBONDS ne zoome plus sur le plus petit intervalle brut.
        Il sélectionne la zone de rebond complète la plus longue parmi les
        contacts attendus ou le contact choisi, ce qui correspond mieux aux
        mesures sanctionnées dans l'onglet Mesures.
        """
        events = self.oscillo_active_events()
        if not events:
            return None
        contacts = self.oscillo_selected_or_expected_contacts()
        candidates = []
        for contact in contacts:
            candidates.extend(self.oscillo_rebound_candidates_for_contact(events, contact))
        candidates = [c for c in candidates if c["first_t"] is not None and c["last_t"] is not None]
        if not candidates:
            return None
        # Priorité métier : durée de rebond la plus longue, puis contact dans l'ordre R/T naturel.
        def contact_rank(contact):
            try:
                return CHRONO_CONTACT_NAMES.index(contact)
            except ValueError:
                return 99
        candidates.sort(key=lambda c: (-int(c["span"]), contact_rank(c["contact"]), int(c["first_t"])))
        return candidates[0]

    def oscillo_zoom_rebonds(self):
        events = self.oscillo_active_events()
        if not events:
            self.oscillo_set_full_view()
            return
        candidate = self.oscillo_find_rebound_window_metier()
        if candidate is not None:
            first_t = int(candidate["first_t"])
            last_t = int(candidate["last_t"])
            span = max(0, int(candidate["span"]))
            margin = max(10, int(max(20, span) * 0.20))
            self.oscillo_set_zoom_window(first_t, last_t, margin_us=margin, min_span_us=80)
            phase_txt = f" {candidate.get('phase')}" if candidate.get('phase') else ""
            self.label_oscillo_status.setText(
                f"Zoom rebond {candidate['type']}{phase_txt} {candidate['contact']} : "
                f"{first_t} à {last_t} µs - Δ {span} µs = {span / 1000.0:.3f} ms"
            )
            self.label_oscillo_status.setStyleSheet("background-color: rgb(255,170,60); color: black; font-weight: bold; border: 2px solid rgb(180,90,0);")
            return
        self.oscillo_zoom_fronts()

    def oscillo_zoom_factor(self):
        try:
            return max(2, min(100, int(self.spinBox_oscillo_zoom_factor.value())))
        except Exception:
            return 5

    def oscillo_update_zoom_button_labels(self):
        if not hasattr(self, "pushButton_oscillo_zoom_in"):
            return
        factor = self.oscillo_zoom_factor() if hasattr(self, "spinBox_oscillo_zoom_factor") else 5
        self.pushButton_oscillo_zoom_in.setText(f"ZOOM + x{factor}")
        self.pushButton_oscillo_zoom_out.setText(f"DÉZOOM x{factor}")

    def oscillo_zoom_centered(self, new_span):
        capture_us = self.oscillo_capture_end_us()
        view_start_us, view_end_us = self.oscillo_view_range()
        cursor_a_us, cursor_b_us = self.oscillo_cursor_values()
        new_span = max(1, min(capture_us, int(new_span)))
        center = (view_start_us + view_end_us) // 2
        start_us = center - new_span // 2
        end_us = start_us + new_span
        if start_us < 0:
            end_us -= start_us
            start_us = 0
        if end_us > capture_us:
            start_us = max(0, start_us - (end_us - capture_us))
            end_us = capture_us
        if end_us <= start_us:
            end_us = min(capture_us, start_us + 1)
        self.oscillo_set_time_window(start_us, end_us, cursor_a_us, cursor_b_us, update=True)

    def oscillo_zoom_in_factor(self):
        view_start_us, view_end_us = self.oscillo_view_range()
        span = max(1, view_end_us - view_start_us)
        if span <= 1:
            self.oscillo_update_view()
            return
        factor = self.oscillo_zoom_factor()
        # Arrondi supérieur : un span de quelques microsecondes reste zoomable
        # sans tomber brutalement à 0.
        new_span = max(1, (span + factor - 1) // factor)
        self.oscillo_zoom_centered(new_span)

    def oscillo_zoom_out_factor(self):
        capture_us = self.oscillo_capture_end_us()
        view_start_us, view_end_us = self.oscillo_view_range()
        span = max(1, view_end_us - view_start_us)
        if view_start_us <= 0 and view_end_us >= capture_us:
            self.oscillo_set_full_view()
            return
        factor = self.oscillo_zoom_factor()
        new_span = min(capture_us, max(span + 1, span * factor))
        self.oscillo_zoom_centered(new_span)

    def oscillo_update_view(self):
        events = self.oscillo_active_events()
        start_bits = self.oscillo_active_start_bits()
        capture_us = self.oscillo_capture_end_us()
        self.oscillo_configure_time_controls()
        view_start_us, view_end_us = self.oscillo_view_range()
        cursor_a_us, cursor_b_us = self.oscillo_cursor_values()
        self.oscillo_update_delta_label()
        self.oscillo_update_zoom_button_labels()
        current = self.oscillo_active_current()
        display_mode = self.oscillo_display_mode()
        self.oscillo_canvas.set_data(
            events, start_bits, capture_us, self.oscillo_title(),
            view_start_us, view_end_us, cursor_a_us, cursor_b_us,
            display_mode, current,
        )
        self.tableWidget_oscillo_points.setRowCount(len(events))
        for row, event in enumerate(events):
            logic_state = str(event.get("state", "0"))
            electric_state = "0" if logic_state == "1" else "1"
            values = [
                str(event.get("i", row)),
                str(event.get("t_us", "")),
                str(event.get("phase", "")),
                str(event.get("contact", "")),
                logic_state,
                electric_state,
                "fermé" if logic_state == "1" else "ouvert",
            ]
            for col, value in enumerate(values):
                self.tableWidget_oscillo_points.setItem(row, col, QTableWidgetItem(value))
        if events:
            pixels = max(1, self.oscillo_canvas.width() - 80)
            span_us = max(1, view_end_us - view_start_us)
            us_per_px = span_us / pixels
            visible_events = sum(1 for event in events if view_start_us <= int(event.get("t_us", 0)) <= view_end_us)
            action_label = self.oscillo_capture_action_label((self.oscillo_active_capture() or {}).get("action"))
            if us_per_px < 0.1:
                resolution_txt = f"{us_per_px:.3f} µs/px"
            elif us_per_px < 10:
                resolution_txt = f"{us_per_px:.2f} µs/px"
            else:
                resolution_txt = f"{us_per_px:.1f} µs/px"
            mode_txt = "loupe vectorielle" if span_us <= 200 else "vue échelle réelle"
            factor = self.oscillo_zoom_factor() if hasattr(self, "spinBox_oscillo_zoom_factor") else 5
            interaction_txt = "molette zoom, clic droit déplacement, clic gauche sélection, M+clic droit mesure"
            self.label_oscillo_status.setText(
                f"{action_label} - {self.oscillo_display_mode_label()} - {visible_events}/{len(events)} fronts visibles - {span_us} µs - {resolution_txt} - {mode_txt} - zoom x{factor} - {interaction_txt}"
            )
            self.label_oscillo_status.setStyleSheet("background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);")
        else:
            self.label_oscillo_status.setText("Aucun front capturé pour la capture sélectionnée.")
            self.label_oscillo_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")
        self.tableWidget_oscillo_points.resizeColumnsToContents()

    def oscillo_clear_view(self, clear_history=True):
        self.tableWidget_oscillo_points.setRowCount(0)
        if clear_history:
            self.oscillo_captures = {}
            self.oscillo_selected_capture_key = "LAST"
            self.oscillo_refresh_capture_combo(select_key="LAST")
        self.oscillo_canvas.set_data([], None, 0, "Oscillogramme contacts", 0, None, None, None, self.oscillo_display_mode(), {})
        for spin in (
            self.spinBox_oscillo_zoom_start_us,
            self.spinBox_oscillo_zoom_end_us,
            self.spinBox_oscillo_cursor_a_us,
            self.spinBox_oscillo_cursor_b_us,
        ):
            self.oscillo_set_spin_value(spin, 0)
        self.label_oscillo_delta.setText("Δ A/B : -- µs")
        self.label_oscillo_status.setText("Aucune mesure affichée.")
        self.label_oscillo_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")

    def oscillo_curve_rows(self):
        events = self.oscillo_active_events()
        capture_us = self.oscillo_capture_us()
        view_start_us, view_end_us = self.oscillo_view_range()
        cursor_a_us, cursor_b_us = self.oscillo_cursor_values()
        delta_us = cursor_b_us - cursor_a_us
        start_bits = self.oscillo_active_start_bits()
        states = {}
        for idx, contact in enumerate(CHRONO_CONTACT_NAMES):
            states[contact] = 1 if start_bits is not None and ((int(start_bits) >> idx) & 0x01) else 0
        rows = []

        def append_row(t_us, event_label, phase=""):
            row = {
                "temps_us": int(t_us),
                "temps_ms": f"{int(t_us) / 1000.0:.6f}",
                "phase": phase,
                "mode_affichage": self.oscillo_display_mode_label(),
                "view_start_us": view_start_us,
                "view_end_us": view_end_us,
                "cursor_a_us": cursor_a_us,
                "cursor_b_us": cursor_b_us,
                "delta_us": delta_us,
                "evenement": event_label,
            }
            for contact in CHRONO_CONTACT_NAMES:
                logic_value = int(states[contact])
                row[f"{contact}_logique"] = logic_value
                row[f"{contact}_electrique"] = 0 if logic_value else 1
            rows.append(row)

        append_row(0, "Début capture", "")
        sorted_events = sorted(
            events,
            key=lambda event: (int(event.get("t_us", 0)), int(event.get("i", 0))),
        )
        for event in sorted_events:
            contact = str(event.get("contact", "")).upper()
            if contact not in states:
                continue
            t_us = int(event.get("t_us", 0))
            phase = str(event.get("phase", ""))
            append_row(t_us, f"Avant front {contact}", phase)
            states[contact] = 1 if str(event.get("state", "0")) == "1" else 0
            append_row(t_us, f"Après front {contact}", phase)
        append_row(capture_us, "Fin capture", "")
        return rows

    def oscillo_export_current_xlsx(self):
        events = self.oscillo_active_events()
        if not events:
            QMessageBox.information(self.window, "Export XLSX courbe", "Aucun oscillogramme à exporter.")
            return
        current = self.oscillo_active_current()
        meta = current.get("meta") if isinstance(current.get("meta"), dict) else {}
        lot = meta.get("lot") or self.lineEdit_chrono_lot.text().strip() or "sans_lot"
        sn = meta.get("sn") or self.lineEdit_chrono_sn.text().strip() or "sans_sn"
        action = str(current.get("ACTION") or current.get("action") or "mesure").lower()
        default_name = f"oscillogramme_{self.filename_safe(lot)}_{self.filename_safe(sn)}_{self.filename_safe(action)}.xlsx"
        path = self.ask_export_path(
            "Exporter oscillogramme en XLSX",
            default_name,
            "Excel (*.xlsx)",
            ".xlsx",
        )
        if not path:
            return
        headers = [
            ("temps_us", "Temps (µs)"),
            ("temps_ms", "Temps (ms)"),
            ("phase", "Phase"),
            ("mode_affichage", "Mode affichage"),
            ("view_start_us", "Zoom début (µs)"),
            ("view_end_us", "Zoom fin (µs)"),
            ("cursor_a_us", "Curseur A (µs)"),
            ("cursor_b_us", "Curseur B (µs)"),
            ("delta_us", "Delta B-A (µs)"),
        ]
        for contact in CHRONO_CONTACT_NAMES:
            headers.append((f"{contact}_logique", f"{contact} logique"))
            headers.append((f"{contact}_electrique", f"{contact} électrique"))
        headers.append(("evenement", "Événement"))
        try:
            self.write_table_xlsx(path, self.oscillo_title(), headers, self.oscillo_curve_rows())
            self.label_oscillo_status.setText(f"Export XLSX courbe créé : {Path(path).name}")
            QMessageBox.information(self.window, "Export XLSX courbe", f"Export XLSX créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export XLSX courbe", f"Export impossible : {exc}")

    def oscillo_export_current_pdf(self):
        events = self.oscillo_active_events()
        if not events:
            QMessageBox.information(self.window, "Export PDF courbe", "Aucun oscillogramme à exporter.")
            return
        current = self.oscillo_active_current()
        meta = current.get("meta") if isinstance(current.get("meta"), dict) else {}
        lot = meta.get("lot") or self.lineEdit_chrono_lot.text().strip() or "sans_lot"
        sn = meta.get("sn") or self.lineEdit_chrono_sn.text().strip() or "sans_sn"
        action = str(current.get("ACTION") or current.get("action") or "mesure").lower()
        default_name = f"oscillogramme_{self.filename_safe(lot)}_{self.filename_safe(sn)}_{self.filename_safe(action)}.pdf"
        path = self.ask_export_path(
            "Exporter oscillogramme en PDF",
            default_name,
            "PDF (*.pdf)",
            ".pdf",
        )
        if not path:
            return
        try:
            self.write_oscillo_pdf(path)
            self.label_oscillo_status.setText(f"Export PDF courbe créé : {Path(path).name}")
            QMessageBox.information(self.window, "Export PDF courbe", f"PDF créé :\n{path}")
        except Exception as exc:
            QMessageBox.warning(self.window, "Export PDF courbe", f"Création PDF impossible : {exc}")

    def write_oscillo_pdf(self, path):
        writer = QPdfWriter(path)
        writer.setPageSize(QPageSize(QPageSize.A4))
        writer.setResolution(96)
        painter = QPainter(writer)
        try:
            page_w = writer.width()
            page_h = writer.height()
            margin = 35
            title_font = QFont("Arial", 14, QFont.Bold)
            normal = QFont("Arial", 9)
            painter.setFont(title_font)
            painter.drawText(margin, margin + 10, self.oscillo_title())
            painter.setFont(normal)
            current = self.oscillo_active_current()
            meta = current.get("meta") if isinstance(current.get("meta"), dict) else {}
            action = str(current.get("ACTION") or current.get("action") or "?").upper()
            view_start_us, view_end_us = self.oscillo_view_range()
            cursor_a_us, cursor_b_us = self.oscillo_cursor_values()
            events = self.oscillo_active_events()
            info = (
                f"Action : {action}    Lot : {meta.get('lot', self.lineEdit_chrono_lot.text().strip())}    "
                f"SN : {meta.get('sn', self.lineEdit_chrono_sn.text().strip())}    "
                f"Événements : {len(events)}"
            )
            painter.drawText(margin, margin + 34, info)
            painter.drawText(
                margin,
                margin + 50,
                f"Vue : {view_start_us} à {view_end_us} µs    Curseurs : A={cursor_a_us} µs, B={cursor_b_us} µs, Δ={cursor_b_us - cursor_a_us} µs",
            )
            display_mode = self.oscillo_display_mode()
            if display_mode == OSCILLO_DISPLAY_SYNTHESIS:
                draw_synthesis_oscillogram(
                    painter,
                    margin,
                    margin + 70,
                    page_w - margin * 2,
                    page_h - margin * 2 - 70,
                    events,
                    self.oscillo_active_start_bits(),
                    self.oscillo_capture_end_us(),
                    current,
                    self.oscillo_title(),
                    view_start_us,
                    view_end_us,
                    cursor_a_us,
                    cursor_b_us,
                )
            else:
                draw_logic_oscillogram(
                    painter,
                    margin,
                    margin + 70,
                    page_w - margin * 2,
                    page_h - margin * 2 - 70,
                    events,
                    self.oscillo_active_start_bits(),
                    self.oscillo_capture_end_us(),
                    self.oscillo_title(),
                    view_start_us,
                    view_end_us,
                    cursor_a_us,
                    cursor_b_us,
                    display_mode,
                    current.get("phase_markers") if isinstance(current, dict) else None,
                    current,
                )
        finally:
            painter.end()

    def chrono_format_ms(self, value):
        if value is None:
            return "--"
        return f"{value:.3f}"

    def chrono_status_item(self, ok):
        text = "OK" if ok else "DEFAUT"
        item = QTableWidgetItem(text)
        if ok:
            item.setBackground(QColor(200, 255, 210))
            item.setForeground(QColor(0, 90, 35))
        else:
            item.setBackground(QColor(255, 200, 200))
            item.setForeground(QColor(130, 0, 0))
        item.setData(Qt.UserRole, 0 if ok else 1)
        return item

    def chrono_value_item(self, value, sort_value=None):
        item = QTableWidgetItem(str(value))
        if sort_value is not None:
            item.setData(Qt.UserRole, sort_value)
        return item

    def chrono_events_for_contact(self, contact):
        return [e for e in self.chrono_events if e["contact"] == contact]

    def chrono_first_state_us(self, contact, state):
        expected = str(int(state))
        for event in self.chrono_events:
            if event["contact"] == contact and str(event["state"]) == expected:
                return int(event["t_us"])
        return None

    def chrono_first_state_and_bounce(self, contact, state):
        """Retourne première occurrence d'un état et durée de rebond sur cet état.

        state = 1 : rebond de fermeture, contact qui revient plusieurs fois fermé.
        state = 0 : rebond d'ouverture, contact qui revient plusieurs fois ouvert.

        La logique reste volontairement symétrique avec l'ancien calcul de rebond
        fermeture : première occurrence -> dernière occurrence dans la capture.
        """
        expected = str(int(state))
        events = [e for e in self.chrono_events if e["contact"] == contact]
        times = [int(e["t_us"]) for e in events if str(e["state"]) == expected]
        if not times:
            return None, None
        first_time = times[0]
        last_time = times[-1]
        return first_time / 1000.0, max(0.0, (last_time - first_time) / 1000.0)

    def chrono_first_close_and_bounce(self, contact):
        return self.chrono_first_state_and_bounce(contact, 1)

    def chrono_first_open_and_bounce(self, contact):
        return self.chrono_first_state_and_bounce(contact, 0)

    def chrono_first_close_us(self, contact):
        return self.chrono_first_state_us(contact, 1)

    def chrono_first_open_us(self, contact):
        return self.chrono_first_state_us(contact, 0)

    def chrono_delta_ms(self, start_us, end_us):
        if start_us is None or end_us is None or end_us < start_us:
            return None
        return (end_us - start_us) / 1000.0

    def chrono_result_row(self, inv, metric_key, label, start_label, end_label, value_ms, ok):
        return {
            "inverseur": inv,
            "metric": metric_key,
            "mesure": label,
            "debut": start_label,
            "fin": end_label,
            "temps_ms": value_ms,
            "ok": ok,
        }

    def chrono_pending_result_row(self, inv, metric_key, label):
        return self.chrono_result_row(inv, metric_key, label, "--", "--", None, None)

    def chrono_pending_item(self):
        item = QTableWidgetItem("EN ATTENTE")
        item.setTextAlignment(Qt.AlignCenter)
        item.setBackground(QColor(235, 235, 235))
        item.setForeground(QColor(80, 80, 80))
        item.setData(Qt.UserRole, 2)
        return item

    def chrono_format_us(self, value):
        if value is None:
            return "--"
        return f"{int(value)}"

    def chrono_save_measure(self, result, details):
        self.chrono_init_db()
        meta = dict(self.chrono_current.get("meta") or self.chrono_metadata())
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        with self.chrono_connect_db() as con:
            con.execute(
                """
                INSERT INTO mesures_chrono_contacts(
                    lot, date_test, relais, ambiance_c, nom_test, sn, relay_type, action,
                    nb_inverseurs, capture_ms, pulse_ms, limite_temps_ms,
                    limite_rebond_ms, resultat, overflow, details_json,
                    events_json, timestamp
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    meta.get("lot", ""),
                    meta.get("date_test", ""),
                    meta.get("relais", ""),
                    meta.get("ambiance_c", ""),
                    meta.get("nom_test", ""),
                    meta.get("sn", ""),
                    str(self.chrono_current.get("relay_type") or self.chrono_relay_type()),
                    str(self.chrono_current.get("ACTION") or self.chrono_current.get("action") or ""),
                    int(self.chrono_current.get("NB_INV") or self.chrono_current.get("nb_inv") or 0),
                    int(self.chrono_current.get("capture_ms") or 0),
                    int(self.chrono_current.get("pulse_ms") or 0),
                    float(self.chrono_current.get("limite_temps_ms") or 0),
                    float(self.chrono_current.get("limite_rebond_ms") or 0),
                    str(result or ""),
                    1 if self.chrono_current.get("overflow") else 0,
                    json.dumps(details, ensure_ascii=False),
                    json.dumps(self.chrono_events, ensure_ascii=False),
                    timestamp,
                ),
            )

    def chrono_increment_sn_after_save(self):
        current_sn = self.lineEdit_chrono_sn.text().strip()
        next_sn = self.next_sn_value(current_sn)
        if next_sn:
            self.lineEdit_chrono_sn.setText(next_sn)

    def chrono_actions_done_for_current_sn(self):
        meta = dict(self.chrono_current.get("meta") or self.chrono_metadata())
        with self.chrono_connect_db() as con:
            rows = con.execute(
                """
                SELECT DISTINCT action
                FROM mesures_chrono_contacts
                WHERE lot = ? AND relais = ? AND nom_test = ? AND sn = ? AND relay_type = ?
                """,
                (
                    meta.get("lot", ""),
                    meta.get("relais", ""),
                    meta.get("nom_test", ""),
                    meta.get("sn", ""),
                    str(self.chrono_current.get("relay_type") or self.chrono_relay_type()),
                ),
            ).fetchall()
        return {str(row["action"] or "").upper() for row in rows}

    def chrono_pair_complete_for_current_sn(self):
        actions = self.chrono_actions_done_for_current_sn()
        if str(self.chrono_current.get("relay_type") or self.chrono_relay_type()) == "MONOSTABLE":
            return "MONO_ON" in actions and "MONO_OFF" in actions
        return "BE" in actions and "BR" in actions

    def chrono_build_action_rows(self, action, nb_inv, limite_temps, limite_rebond):
        action_rows = []
        first_close_us_by_contact = {}
        target_contacts = []
        action_ok = True
        forward = action in ("BE", "MONO_ON")

        for inv in range(1, nb_inv + 1):
            contact_r = f"R{inv}"
            contact_t = f"T{inv}"
            if forward:
                command_label = "BE ON" if action == "BE" else "GP14 ON"
                open_us = self.chrono_first_open_us(contact_r)
                close_us = self.chrono_first_close_us(contact_t)
                temps_ms, rebond_fermeture_ms = self.chrono_first_close_and_bounce(contact_t)
                _open_time_ms, rebond_ouverture_ms = self.chrono_first_open_and_bounce(contact_r)
                transfert_ms = self.chrono_delta_ms(open_us, close_us)
                rows = [
                    self.chrono_result_row(
                        inv, "enclenchement",
                        f"Temps d'Enclenchement {inv} (ms)",
                        command_label, f"{contact_t} fermé", temps_ms,
                        temps_ms is not None and temps_ms < limite_temps,
                    ),
                    self.chrono_result_row(
                        inv, "transfert_travail",
                        f"Temps de transfert {inv} (ms)",
                        f"{contact_r} ouvert", f"{contact_t} fermé", transfert_ms,
                        transfert_ms is not None and transfert_ms < CHRONO_TRANSFER_LIMIT_MS,
                    ),
                    self.chrono_result_row(
                        inv, "rebond_repos_ouverture",
                        f"Temps Rebond Repos Ouverture {inv} (ms)",
                        f"{contact_r} 1ère ouverture", f"{contact_r} dernière ouverture", rebond_ouverture_ms,
                        rebond_ouverture_ms is not None and rebond_ouverture_ms < limite_rebond,
                    ),
                    self.chrono_result_row(
                        inv, "rebond_travail",
                        f"Temps Rebond Travail Fermeture {inv} (ms)",
                        f"{contact_t} 1ère fermeture", f"{contact_t} dernière fermeture", rebond_fermeture_ms,
                        rebond_fermeture_ms is not None and rebond_fermeture_ms < limite_rebond,
                    ),
                ]
                target_contacts.append(contact_t)
                first_close_us_by_contact[contact_t] = close_us
            else:
                command_label = "BR ON" if action == "BR" else "GP14 OFF"
                open_us = self.chrono_first_open_us(contact_t)
                close_us = self.chrono_first_close_us(contact_r)
                temps_ms, rebond_fermeture_ms = self.chrono_first_close_and_bounce(contact_r)
                _open_time_ms, rebond_ouverture_ms = self.chrono_first_open_and_bounce(contact_t)
                transfert_ms = self.chrono_delta_ms(open_us, close_us)
                rows = [
                    self.chrono_result_row(
                        inv, "declenchement",
                        f"Temps de Déclenchement {inv} (ms)",
                        command_label, f"{contact_r} fermé", temps_ms,
                        temps_ms is not None and temps_ms < limite_temps,
                    ),
                    self.chrono_result_row(
                        inv, "transfert_repos",
                        f"Temps de transfert {inv} retour (ms)",
                        f"{contact_t} ouvert", f"{contact_r} fermé", transfert_ms,
                        transfert_ms is not None and transfert_ms < CHRONO_TRANSFER_LIMIT_MS,
                    ),
                    self.chrono_result_row(
                        inv, "rebond_travail_ouverture",
                        f"Temps Rebond Travail Ouverture {inv} (ms)",
                        f"{contact_t} 1ère ouverture", f"{contact_t} dernière ouverture", rebond_ouverture_ms,
                        rebond_ouverture_ms is not None and rebond_ouverture_ms < limite_rebond,
                    ),
                    self.chrono_result_row(
                        inv, "rebond_repos",
                        f"Temps Rebond Repos Fermeture {inv} (ms)",
                        f"{contact_r} 1ère fermeture", f"{contact_r} dernière fermeture", rebond_fermeture_ms,
                        rebond_fermeture_ms is not None and rebond_fermeture_ms < limite_rebond,
                    ),
                ]
                target_contacts.append(contact_r)
                first_close_us_by_contact[contact_r] = close_us

            for row in rows:
                action_ok = action_ok and bool(row["ok"])
            action_rows.extend(rows)

        return action_rows, action_ok, first_close_us_by_contact, target_contacts

    def chrono_combined_result_rows(self, nb_inv):
        rows_by_metric = {}
        for rows in getattr(self, "chrono_result_rows_by_action", {}).values():
            for row in rows:
                rows_by_metric[(int(row["inverseur"]), row["metric"])] = row

        order = [
            ("enclenchement", "Temps d'Enclenchement {inv} (ms)"),
            ("transfert_travail", "Temps de transfert {inv} (ms)"),
            ("rebond_repos_ouverture", "Temps Rebond Repos Ouverture {inv} (ms)"),
            ("rebond_travail", "Temps Rebond Travail Fermeture {inv} (ms)"),
            ("declenchement", "Temps de Déclenchement {inv} (ms)"),
            ("transfert_repos", "Temps de transfert {inv} retour (ms)"),
            ("rebond_travail_ouverture", "Temps Rebond Travail Ouverture {inv} (ms)"),
            ("rebond_repos", "Temps Rebond Repos Fermeture {inv} (ms)"),
        ]
        display_rows = []
        for inv in range(1, nb_inv + 1):
            for metric, label_template in order:
                display_rows.append(
                    rows_by_metric.get(
                        (inv, metric),
                        self.chrono_pending_result_row(inv, metric, label_template.format(inv=inv)),
                    )
                )
        return display_rows

    def chrono_fill_results_table(self, display_rows):
        self.tableWidget_chrono_results.setRowCount(len(display_rows))
        for row_index, row in enumerate(display_rows):
            values = [
                str(row["inverseur"]),
                row["mesure"],
                row["debut"],
                row["fin"],
                self.chrono_format_ms(row["temps_ms"]),
            ]
            for col, value in enumerate(values):
                sort_value = row["temps_ms"] if col == 4 and row["temps_ms"] is not None else None
                self.tableWidget_chrono_results.setItem(row_index, col, self.chrono_value_item(value, sort_value))
            if row["ok"] is None:
                self.tableWidget_chrono_results.setItem(row_index, 5, self.chrono_pending_item())
            else:
                self.tableWidget_chrono_results.setItem(row_index, 5, self.chrono_status_item(bool(row["ok"])))

    def chrono_compute_results(self):
        action = str(self.chrono_current.get("ACTION") or self.chrono_current.get("action") or "").upper()
        nb_inv = self.chrono_parse_int(self.chrono_current.get("NB_INV") or self.chrono_current.get("nb_inv"), 1)
        nb_inv = max(1, min(4, nb_inv))
        limite_temps = self.chrono_float_ms(self.lineEdit_chrono_limite_temps_ms, "Sanction temps max")
        limite_rebond = self.chrono_float_ms(self.lineEdit_chrono_limite_rebond_ms, "Sanction rebond max")

        action_rows, global_ok, first_close_us_by_contact, target_contacts = self.chrono_build_action_rows(
            action, nb_inv, limite_temps, limite_rebond
        )
        self.chrono_result_rows_by_action[action] = action_rows
        display_rows = self.chrono_combined_result_rows(nb_inv)
        self.chrono_fill_results_table(display_rows)

        close_values = [first_close_us_by_contact.get(contact) for contact in target_contacts]
        close_values_ok = all(value is not None for value in close_values)
        spread_us = (max(close_values) - min(close_values)) if close_values_ok and close_values else None
        spread_warning = spread_us is not None and spread_us > CHRONO_SPREAD_INFO_US
        loop_max_us = int(self.chrono_current.get("loop_max_us") or 0)
        dropped_events = int(self.chrono_current.get("dropped_events") or 0)
        event_capacity = int(self.chrono_current.get("event_capacity") or 0)
        event_count = self.chrono_parse_int(self.chrono_current.get("EVENTS"), len(self.chrono_events))
        loop_warning = loop_max_us > CHRONO_LOOP_WARN_US if loop_max_us else False
        # L’écart premier/dernier contact est une information de synchronisme mécanique.
        # Il ne participe pas au verdict OK/DEFAUT du relais.

        if self.chrono_current.get("overflow"):
            global_ok = False
        quality = {
            "spread_reference_us": CHRONO_SPREAD_INFO_US,
            "spread_us": spread_us,
            "spread_warning": spread_warning,
            "spread_informative_only": True,
            "first_close_us_by_contact": first_close_us_by_contact,
            "loop_max_us": loop_max_us,
            "loop_warn_us": CHRONO_LOOP_WARN_US,
            "loop_warning": loop_warning,
            "events": event_count,
            "event_capacity": event_capacity,
            "overflow": bool(self.chrono_current.get("overflow")),
            "dropped_events": dropped_events,
        }
        result = "OK" if global_ok else "DEFAUT"
        if getattr(self, "measure_all_active", False) and self.measure_all_phase == "CHRONO":
            self.measure_all_chrono_results[action] = result
        self.lineEdit_chrono_resultat.setText(result)
        try:
            self.chrono_save_measure(result, {
                "inverseurs": action_rows,
                "lignes_action": action_rows,
                "lignes_affichage": display_rows,
                "quality": quality,
            })
            saved = True
        except Exception as exc:
            saved = False
            self.label_chrono_status.setText(f"Mesure {action} terminée mais sauvegarde impossible : {exc}")
            self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
        pair_complete = False
        if not saved and getattr(self, "measure_all_active", False) and self.measure_all_phase == "CHRONO":
            self.voltage_measure_all_fail("Sauvegarde de la chronométrie impossible.")
        if saved:
            if self.chrono_auto_sequence_active and self.chrono_auto_sequence_queue:
                next_action = self.chrono_auto_sequence_queue[0]
                self.chrono_measure_running = True
                self.update_button_states()
                self.label_chrono_status.setText(f"Mesure {action} sauvegardée - lancement mesure {next_action}.")
                self.label_chrono_status.setStyleSheet("background-color: rgb(255,235,150); color: black; font-weight: bold; border: 2px solid rgb(160,100,0);")
                self.tableWidget_chrono_results.resizeColumnsToContents()
                self.tableWidget_chrono_events.resizeColumnsToContents()
                QTimer.singleShot(150, self.chrono_start_next_auto_measure)
                return

            pair_complete = self.chrono_pair_complete_for_current_sn()
            if pair_complete:
                self.chrono_increment_sn_after_save()
                self.chrono_auto_sequence_active = False
                self.chrono_auto_sequence_queue = []
                if str(self.chrono_current.get("relay_type") or self.chrono_relay_type()) == "MONOSTABLE":
                    suffix_sn = " - enclenchement et déclenchement sauvegardés - SN suivant prêt"
                else:
                    suffix_sn = " - BE et BR sauvegardés - SN suivant prêt"
            else:
                if str(self.chrono_current.get("relay_type") or self.chrono_relay_type()) == "MONOSTABLE":
                    missing = "DÉCLENCHEMENT" if action == "MONO_ON" else "ENCLENCHEMENT"
                else:
                    missing = "BR" if action == "BE" else "BE"
                suffix_sn = f" - sauvegardée - faire aussi {missing} avant SN suivant"
            if spread_us is None:
                spread_info = "info écart inter-inverseurs non calculable"
            elif spread_warning:
                spread_info = f"info écart inter-inverseurs {self.chrono_format_us(spread_us)} µs > {CHRONO_SPREAD_INFO_US} µs (indicatif)"
            else:
                spread_info = f"info écart inter-inverseurs {self.chrono_format_us(spread_us)} µs / {CHRONO_SPREAD_INFO_US} µs"

            if global_ok:
                loop_info = f" - attention loop max {loop_max_us} µs" if loop_warning else f" - loop max {loop_max_us} µs"
                self.label_chrono_status.setText(
                    f"Mesure {action} terminée : OK - {spread_info}{loop_info}{suffix_sn}"
                )
                if loop_warning or spread_warning:
                    self.label_chrono_status.setStyleSheet("background-color: rgb(255,170,60); color: black; font-weight: bold; border: 2px solid rgb(180,90,0);")
                else:
                    self.label_chrono_status.setStyleSheet("background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);")
            else:
                suffix_parts = []
                if self.chrono_current.get("overflow"):
                    suffix_parts.append(f"buffer saturé {event_count}/{event_capacity}, perdus {dropped_events}")
                if loop_warning:
                    suffix_parts.append(f"attention scrutation max {loop_max_us} µs")
                if spread_warning:
                    suffix_parts.append(spread_info)
                suffix = " - " + " - ".join(suffix_parts) if suffix_parts else ""
                self.label_chrono_status.setText(f"Mesure {action} terminée : DEFAUT{suffix}{suffix_sn}")
                self.label_chrono_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
        self.tableWidget_chrono_results.resizeColumnsToContents()
        self.tableWidget_chrono_events.resizeColumnsToContents()
        if (
            saved and pair_complete
            and getattr(self, "measure_all_active", False)
            and self.measure_all_phase == "CHRONO"
        ):
            QTimer.singleShot(0, self.voltage_measure_all_finish)

    def initialiser_leds_contacts(self):
        # Onglet Cyclage : R1/R2/R3/R4 en vert, T1/T2/T3/T4 en rouge
        self.set_contact_led(self.label_led_reset_contact_1, "R1", None, "green")
        self.set_contact_led(self.label_led_reset_contact_2, "R2", None, "green")
        self.set_contact_led(self.label_led_reset_contact_5, "R3", None, "green")
        self.set_contact_led(self.label_led_reset_contact_6, "R4", None, "green")
        self.set_contact_led(self.label_led_latch_contact_1, "T1", None, "red")
        self.set_contact_led(self.label_led_latch_contact_2, "T2", None, "red")
        self.set_contact_led(self.label_led_latch_contact_5, "T3", None, "red")
        self.set_contact_led(self.label_led_latch_contact_6, "T4", None, "red")

        # Onglet Neutral screen : mêmes 8 contacts
        self.set_contact_led(self.label_led_reset_contact_3, "R1", None, "green")
        self.set_contact_led(self.label_led_reset_contact_4, "R2", None, "green")
        self.set_contact_led(self.label_led_reset_contact_7, "R3", None, "green")
        self.set_contact_led(self.label_led_reset_contact_8, "R4", None, "green")
        self.set_contact_led(self.label_led_latch_contact_3, "T1", None, "red")
        self.set_contact_led(self.label_led_latch_contact_4, "T2", None, "red")
        self.set_contact_led(self.label_led_latch_contact_7, "T3", None, "red")
        self.set_contact_led(self.label_led_latch_contact_8, "T4", None, "red")

        self.label_neutral_contact_summary.setText(
            "R1: -- | R2: -- | R3: -- | R4: -- | T1: -- | T2: -- | T3: -- | T4: --"
        )
        self.label_neutral_contact_summary.setStyleSheet(
            "background-color: rgb(80,80,80); color: white; font-weight: bold; border: 1px solid black;"
        )

        if hasattr(self, "label_auto_led_r1"):
            self.set_contact_led(self.label_auto_led_r1, "R1", None, "green")
            self.set_contact_led(self.label_auto_led_r2, "R2", None, "green")
            self.set_contact_led(self.label_auto_led_r3, "R3", None, "green")
            self.set_contact_led(self.label_auto_led_r4, "R4", None, "green")
            self.set_contact_led(self.label_auto_led_t1, "T1", None, "red")
            self.set_contact_led(self.label_auto_led_t2, "T2", None, "red")
            self.set_contact_led(self.label_auto_led_t3, "T3", None, "red")
            self.set_contact_led(self.label_auto_led_t4, "T4", None, "red")
            self.auto_apply_inverseur_filter()

        if hasattr(self, "label_chrono_led_r1"):
            self.set_contact_led(self.label_chrono_led_r1, "R1", None, "green")
            self.set_contact_led(self.label_chrono_led_r2, "R2", None, "green")
            self.set_contact_led(self.label_chrono_led_r3, "R3", None, "green")
            self.set_contact_led(self.label_chrono_led_r4, "R4", None, "green")
            self.set_contact_led(self.label_chrono_led_t1, "T1", None, "red")
            self.set_contact_led(self.label_chrono_led_t2, "T2", None, "red")
            self.set_contact_led(self.label_chrono_led_t3, "T3", None, "red")
            self.set_contact_led(self.label_chrono_led_t4, "T4", None, "red")

        if hasattr(self, "label_voltage_led_r1"):
            self.voltage_refresh_contact_leds()

    def set_contact_led(self, label, titre, valeur, couleur):
        label.setText("")
        if valeur is None:
            label.setToolTip(f"{titre} : état inconnu")
            label.setStyleSheet(
                "background-color: rgb(45,45,45); border: 2px solid rgb(15,15,15); border-radius: 14px;"
            )
            return
        actif = str(valeur) == "1"
        if actif:
            if couleur == "green":
                label.setStyleSheet(
                    "background-color: qradialgradient(cx:0.35, cy:0.35, radius:0.9, "
                    "fx:0.35, fy:0.35, stop:0 rgb(230,255,230), stop:0.30 rgb(90,255,125), "
                    "stop:0.65 rgb(0,215,70), stop:1 rgb(0,110,35));"
                    "border: 2px solid rgb(190,255,190); border-radius: 14px;"
                )
            else:
                label.setStyleSheet(
                    "background-color: qradialgradient(cx:0.35, cy:0.35, radius:0.9, "
                    "fx:0.35, fy:0.35, stop:0 rgb(255,230,230), stop:0.30 rgb(255,95,95), "
                    "stop:0.65 rgb(235,0,0), stop:1 rgb(120,0,0));"
                    "border: 2px solid rgb(255,190,190); border-radius: 14px;"
                )
            label.setToolTip(f"{titre} : contact fermé détecté")
        else:
            if couleur == "green":
                label.setStyleSheet(
                    "background-color: qradialgradient(cx:0.35, cy:0.35, radius:0.9, "
                    "fx:0.35, fy:0.35, stop:0 rgb(80,90,80), stop:0.45 rgb(45,55,45), "
                    "stop:1 rgb(22,28,22)); border: 2px solid rgb(12,18,12); border-radius: 14px;"
                )
            else:
                label.setStyleSheet(
                    "background-color: qradialgradient(cx:0.35, cy:0.35, radius:0.9, "
                    "fx:0.35, fy:0.35, stop:0 rgb(90,80,80), stop:0.45 rgb(55,45,45), "
                    "stop:1 rgb(28,22,22)); border: 2px solid rgb(18,12,12); border-radius: 14px;"
                )
            label.setToolTip(f"{titre} : contact ouvert / non détecté")

    def update_contacts_feedback(self, reset1, reset2, reset3, reset4, latch1, latch2, latch3, latch4):
        # V2.12.3 : si un champ est absent (None) dans une trame partielle, on
        # conserve la dernière valeur connue au lieu d'éteindre/griser la LED.
        recus = [reset1, reset2, reset3, reset4, latch1, latch2, latch3, latch4]
        fusion = []
        for i, v in enumerate(recus):
            if v is None:
                ancienne = self.contacts_known_values[i]
                fusion.append(ancienne)  # garde l'état connu (peut être None au tout début)
            else:
                fusion.append(str(v))
        nouvel_etat = tuple(fusion)

        if (not self.contacts_force_refresh) and nouvel_etat == self.contacts_last_values:
            return
        self.contacts_force_refresh = False
        self.contacts_last_values = nouvel_etat
        self.contacts_known_values = list(nouvel_etat)

        r1, r2, r3, r4, t1, t2, t3, t4 = nouvel_etat

        # Onglet Cyclage
        self.set_contact_led(self.label_led_reset_contact_1, "R1", r1, "green")
        self.set_contact_led(self.label_led_reset_contact_2, "R2", r2, "green")
        self.set_contact_led(self.label_led_reset_contact_5, "R3", r3, "green")
        self.set_contact_led(self.label_led_reset_contact_6, "R4", r4, "green")
        self.set_contact_led(self.label_led_latch_contact_1, "T1", t1, "red")
        self.set_contact_led(self.label_led_latch_contact_2, "T2", t2, "red")
        self.set_contact_led(self.label_led_latch_contact_5, "T3", t3, "red")
        self.set_contact_led(self.label_led_latch_contact_6, "T4", t4, "red")

        # Onglet Neutral screen
        self.set_contact_led(self.label_led_reset_contact_3, "R1", r1, "green")
        self.set_contact_led(self.label_led_reset_contact_4, "R2", r2, "green")
        self.set_contact_led(self.label_led_reset_contact_7, "R3", r3, "green")
        self.set_contact_led(self.label_led_reset_contact_8, "R4", r4, "green")
        self.set_contact_led(self.label_led_latch_contact_3, "T1", t1, "red")
        self.set_contact_led(self.label_led_latch_contact_4, "T2", t2, "red")
        self.set_contact_led(self.label_led_latch_contact_7, "T3", t3, "red")
        self.set_contact_led(self.label_led_latch_contact_8, "T4", t4, "red")

        if hasattr(self, "label_auto_led_r1"):
            self.set_contact_led(self.label_auto_led_r1, "R1", r1, "green")
            self.set_contact_led(self.label_auto_led_r2, "R2", r2, "green")
            self.set_contact_led(self.label_auto_led_r3, "R3", r3, "green")
            self.set_contact_led(self.label_auto_led_r4, "R4", r4, "green")
            self.set_contact_led(self.label_auto_led_t1, "T1", t1, "red")
            self.set_contact_led(self.label_auto_led_t2, "T2", t2, "red")
            self.set_contact_led(self.label_auto_led_t3, "T3", t3, "red")
            self.set_contact_led(self.label_auto_led_t4, "T4", t4, "red")
            self.auto_apply_inverseur_filter()

        if hasattr(self, "label_chrono_led_r1"):
            self.set_contact_led(self.label_chrono_led_r1, "R1", r1, "green")
            self.set_contact_led(self.label_chrono_led_r2, "R2", r2, "green")
            self.set_contact_led(self.label_chrono_led_r3, "R3", r3, "green")
            self.set_contact_led(self.label_chrono_led_r4, "R4", r4, "green")
            self.set_contact_led(self.label_chrono_led_t1, "T1", t1, "red")
            self.set_contact_led(self.label_chrono_led_t2, "T2", t2, "red")
            self.set_contact_led(self.label_chrono_led_t3, "T3", t3, "red")
            self.set_contact_led(self.label_chrono_led_t4, "T4", t4, "red")

        if hasattr(self, "label_voltage_led_r1"):
            self.voltage_refresh_contact_leds()

        def txt(v):
            if v is None:
                return "--"
            return "1" if str(v) == "1" else "0"

        self.label_neutral_contact_summary.setText(
            f"R1:{txt(r1)} | R2:{txt(r2)} | R3:{txt(r3)} | R4:{txt(r4)} | "
            f"T1:{txt(t1)} | T2:{txt(t2)} | T3:{txt(t3)} | T4:{txt(t4)}"
        )
        if hasattr(self, "label_chrono_contact_summary"):
            self.label_chrono_contact_summary.setText(
                f"R1:{txt(r1)} | R2:{txt(r2)} | R3:{txt(r3)} | R4:{txt(r4)} | "
                f"T1:{txt(t1)} | T2:{txt(t2)} | T3:{txt(t3)} | T4:{txt(t4)}"
            )

        latch_actif = any(str(v) == "1" for v in (t1, t2, t3, t4))
        reset_actif = any(str(v) == "1" for v in (r1, r2, r3, r4))

        if latch_actif:
            summary_style = "background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;"
            self.label_neutral_contact_summary.setStyleSheet(summary_style)
        elif reset_actif:
            summary_style = "background-color: rgb(0,176,80); color: white; font-weight: bold; border: 2px solid black;"
            self.label_neutral_contact_summary.setStyleSheet(summary_style)
        else:
            summary_style = "background-color: rgb(70,70,70); color: white; font-weight: bold; border: 1px solid black;"
            self.label_neutral_contact_summary.setStyleSheet(summary_style)
        if hasattr(self, "label_chrono_contact_summary"):
            self.label_chrono_contact_summary.setStyleSheet(summary_style)

    def refresh_auto_leds_from_known_values(self):
        """Rafraîchit immédiatement les LED de l'onglet auto depuis le dernier état connu.

        Objectif V2.12.3 :
        - si N passe de 2 à 4, R3/R4/T3/T4 ne doivent pas rester grisées ;
        - si N passe de 4 à 2, R3/R4/T3/T4 doivent être grisées immédiatement ;
        - ne pas attendre une nouvelle trame STATUS ou CONTACT.
        """
        if not hasattr(self, "label_auto_led_r1"):
            return
        vals = list(getattr(self, "contacts_known_values", [None] * 8))
        while len(vals) < 8:
            vals.append(None)
        r1, r2, r3, r4, t1, t2, t3, t4 = vals[:8]

        # Remet d'abord les 8 LED à leur vrai état connu.
        self.set_contact_led(self.label_auto_led_r1, "R1", r1, "green")
        self.set_contact_led(self.label_auto_led_r2, "R2", r2, "green")
        self.set_contact_led(self.label_auto_led_r3, "R3", r3, "green")
        self.set_contact_led(self.label_auto_led_r4, "R4", r4, "green")
        self.set_contact_led(self.label_auto_led_t1, "T1", t1, "red")
        self.set_contact_led(self.label_auto_led_t2, "T2", t2, "red")
        self.set_contact_led(self.label_auto_led_t3, "T3", t3, "red")
        self.set_contact_led(self.label_auto_led_t4, "T4", t4, "red")

        # Puis applique le grisage des inverseurs hors N.
        self.auto_apply_inverseur_filter()

    def on_auto_nb_inverseurs_changed(self):
        """Réagit immédiatement au changement de N inverseurs dans l'IHM."""
        if not hasattr(self, "lineEdit_auto_nb_inverseurs"):
            return
        try:
            self.lire_auto_nb_inverseurs()
            self.auto_update_tension_labels()
            self.refresh_auto_leds_from_known_values()
            if hasattr(self, "render_auto_logigramme"):
                self.render_auto_logigramme()
            if hasattr(self, "label_auto_status") and not getattr(self, "auto_neutral_running", False):
                self.label_auto_status.setText(
                    f"Automatique : prêt — évaluation sur {self.lire_auto_nb_inverseurs()} inverseur(s)"
                )
        except Exception:
            # Saisie temporairement invalide pendant que l'utilisateur tape.
            self.auto_apply_inverseur_filter()
            if hasattr(self, "label_auto_status") and not getattr(self, "auto_neutral_running", False):
                self.label_auto_status.setText("Automatique : nombre d'inverseurs invalide, saisir 1 à 4")

    # --- Affichage sorties : état réel toujours à jour, flash purement visuel ---
    def render_output_labels(self):
        """Affiche l'état réel courant. Appelé hors flash."""
        s1 = "ON" if self.current_out1 == "1" else "OFF" if self.current_out1 == "0" else self.current_out1
        s2 = "ON" if self.current_out2 == "1" else "OFF" if self.current_out2 == "0" else self.current_out2
        self.label_etat_sortie1.setText(f"Sortie 1 : {s1}")
        self.label_etat_sortie2.setText(f"Sortie 2 : {s2}")
        self.label_neutral_sortie1.setText(f"BE / sortie 1 : GP14 / Pico pin 19 → MOSFET 1 | État : {s1}")
        self.label_neutral_sortie2.setText(f"BR / sortie 2 : GP15 / Pico pin 20 → MOSFET 2 | État : {s2}")

    def set_output_state(self, out1, out2):
        """Met à jour l'état réel. N'écrase pas le texte si un flash est en cours."""
        self.current_out1 = str(out1)
        self.current_out2 = str(out2)
        if not self.flash_active:
            self.render_output_labels()

    def end_flash(self):
        self.flash_active = False
        self.render_output_labels()

    def flash_output_pulse(self, out1, out2, pulse_name, duree_us):
        # Le flash est purement visuel et ne bloque jamais la mise à jour de
        # current_out*. Si un flash est déjà actif, on le réarme proprement.
        self.flash_active = True
        if str(out1) == "1":
            self.label_etat_sortie1.setText(f"Sortie 1 : PULSE {pulse_name} {duree_us} µs")
            self.label_neutral_sortie1.setText(f"BE / sortie 1 : PULSE {pulse_name} {duree_us} µs")
        if str(out2) == "1":
            self.label_etat_sortie2.setText(f"Sortie 2 : PULSE {pulse_name} {duree_us} µs")
            self.label_neutral_sortie2.setText(f"BR / sortie 2 : PULSE {pulse_name} {duree_us} µs")
        self.flash_timer.start(180)

    def update_voltage_selection_label(self, voltage_text, sel32=None):
        if self.label_neutral_tension_selection is None:
            return

        haute = self.texte_tension_haute_info()
        basse = self.texte_tension_basse_info()

        v = str(voltage_text).strip().upper()
        high_selected = (sel32 == "1") or v in ("HIGH", "HAUTE", "NO", "32V", "32 V", "32")

        if high_selected:
            txt = f"Tension sélectionnée : {haute} - voie haute / NO"
            style = "background-color: rgb(255,235,210); color: rgb(120,55,0); font-weight: bold; border: 2px solid rgb(210,120,20);"
        else:
            txt = f"Tension sélectionnée : {basse} - voie basse / NC"
            style = "background-color: rgb(220,245,220); color: rgb(0,90,35); font-weight: bold; border: 2px solid rgb(0,150,60);"

        self.label_neutral_tension_selection.setText(txt)
        self.label_neutral_tension_selection.setStyleSheet(style)

    def parse_vsel_frame(self, line):
        fields = line.split(";")
        sel32 = None
        voltage = None
        raison = fields[1] if len(fields) > 1 else ""
        for field in fields:
            if field.startswith("SEL32="):
                sel32 = field.split("=", 1)[1]
            elif field.startswith("VSEL="):
                voltage = field.split("=", 1)[1]
        if voltage is None:
            voltage = "HIGH" if sel32 == "1" else "LOW"
        self.update_voltage_selection_label(voltage, sel32)

    def parse_output_frame(self, line):
        fields = line.split(";")
        raison = fields[1] if len(fields) > 1 else ""
        out1 = out2 = None
        pulse_name = ""
        duree_us = ""
        for field in fields:
            if field.startswith("OUT1="):
                out1 = field.split("=", 1)[1]
            elif field.startswith("OUT2="):
                out2 = field.split("=", 1)[1]
            elif field.startswith("PULSE="):
                pulse_name = field.split("=", 1)[1]
            elif field.startswith("DUREE_US="):
                duree_us = field.split("=", 1)[1]

        if raison == "PULSE":
            # Flash visuel uniquement ; l'état réel sera confirmé par CHANGE/STATUS.
            f1 = out1 if out1 is not None else "0"
            f2 = out2 if out2 is not None else "0"
            self.flash_output_pulse(f1, f2, pulse_name, duree_us)
        else:
            if out1 is None:
                out1 = self.current_out1
            if out2 is None:
                out2 = self.current_out2
            self.set_output_state(out1, out2)

    def parse_contact_frame(self, line):
        fields = line.split(";")
        reset1 = reset2 = reset3 = reset4 = None
        latch1 = latch2 = latch3 = latch4 = None
        for field in fields:
            if field.startswith("IN_RESET1="):
                reset1 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET2="):
                reset2 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET3="):
                reset3 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET4="):
                reset4 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH1="):
                latch1 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH2="):
                latch2 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH3="):
                latch3 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH4="):
                latch4 = field.split("=", 1)[1]
        self.update_contacts_feedback(reset1, reset2, reset3, reset4, latch1, latch2, latch3, latch4)

    def parse_status(self, line):
        fields = line.split(";")
        raison = fields[1] if len(fields) > 1 else "?"
        self.label_etat_essai.setText(f"État : {raison}")
        out1 = out2 = "?"
        cycle = "?"
        vsel = None
        sel32 = None
        reset1 = reset2 = reset3 = reset4 = None
        latch1 = latch2 = latch3 = latch4 = None
        for field in fields:
            if field.startswith("OUT1="):
                out1 = field.split("=", 1)[1]
            elif field.startswith("OUT2="):
                out2 = field.split("=", 1)[1]
            elif field.startswith("CYCLE="):
                cycle = field.split("=", 1)[1]
            elif field.startswith("VSEL="):
                vsel = field.split("=", 1)[1]
            elif field.startswith("SEL32="):
                sel32 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET1="):
                reset1 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET2="):
                reset2 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET3="):
                reset3 = field.split("=", 1)[1]
            elif field.startswith("IN_RESET4="):
                reset4 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH1="):
                latch1 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH2="):
                latch2 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH3="):
                latch3 = field.split("=", 1)[1]
            elif field.startswith("IN_LATCH4="):
                latch4 = field.split("=", 1)[1]
        self.set_output_state(out1, out2)
        if vsel is not None or sel32 is not None:
            self.update_voltage_selection_label(vsel if vsel is not None else "", sel32)
        self.label_cycle_actuel.setText(f"Cycle : {cycle}")
        self.update_contacts_feedback(reset1, reset2, reset3, reset4, latch1, latch2, latch3, latch4)


    # ------------------------------------------------------------------
    # Neutral Screen Automatique V2.12.3
    # ------------------------------------------------------------------
    def auto_update_tension_labels(self):
        tension_basse = self.texte_tension_basse_info()
        tension_haute = self.texte_tension_haute_info()
        try:
            n = self.lire_auto_nb_inverseurs()
            suffixe_court = f" | N={n}"
            suffixe_complet = f" | N={n} inverseur(s)"
        except Exception:
            suffixe_court = " | N=?"
            suffixe_complet = " | N invalide"

        texte_basse = f"Basse / NC : {tension_basse}"
        texte_basse_complet = f"Voie basse / NC : {tension_basse}"
        texte_haute_court = f"Haute / NO : {tension_haute}{suffixe_court}"
        texte_haute_complet = f"Voie haute / NO : {tension_haute}{suffixe_complet}"
        self.label_auto_tension_basse.setText(texte_basse)
        self.label_auto_tension_haute.setText(texte_haute_court)
        self.label_auto_tension_basse.setToolTip(texte_basse_complet)
        self.label_auto_tension_haute.setToolTip(texte_haute_complet)
        self.auto_apply_inverseur_filter()

    def lire_auto_delai_ms(self):
        texte = self.lineEdit_auto_delai_ms.text().strip().replace(" ", "")
        if not texte.isdigit():
            raise ValueError("Le temps entre étapes doit être un entier en ms.")
        valeur = int(texte)
        if valeur < 100 or valeur > 2000:
            raise ValueError("Le temps entre étapes doit être compris entre 100 ms et 2000 ms.")
        return valeur

    def lire_auto_nb_inverseurs(self):
        texte = self.lineEdit_auto_nb_inverseurs.text().strip().replace(" ", "")
        if not texte.isdigit():
            raise ValueError("Le nombre d'inverseurs doit être un entier entre 1 et 4.")
        valeur = int(texte)
        if valeur < 1 or valeur > 4:
            raise ValueError("Le nombre d'inverseurs doit être compris entre 1 et 4.")
        return valeur

    def lire_auto_pulse_ms(self, line_edit, nom_champ):
        texte = line_edit.text().strip().replace(" ", "")
        if not texte.isdigit():
            raise ValueError(f"{nom_champ} doit être un entier en ms.")
        valeur = int(texte)
        # La norme demande 10 ±1 ms pour BE et BR ; le champ reste volontairement modifiable.
        # Limites larges mais raisonnables pour éviter les erreurs de saisie.
        if valeur < 1 or valeur > 2000:
            raise ValueError(f"{nom_champ} doit être compris entre 1 ms et 2000 ms.")
        return valeur

    def lire_auto_pulse_us(self, pulse_name):
        pulse_name = pulse_name.upper()
        if pulse_name == "BEBR":
            return self.lire_auto_pulse_ms(self.lineEdit_auto_pulse_bebr_ms, "Pulse automatique BE/BR") * 1000
        if pulse_name == "BE":
            return self.lire_auto_pulse_ms(self.lineEdit_auto_pulse_be_ms, "Pulse automatique BE") * 1000
        if pulse_name == "BR":
            return self.lire_auto_pulse_ms(self.lineEdit_auto_pulse_br_ms, "Pulse automatique BR") * 1000
        raise ValueError(f"Pulse automatique inconnu : {pulse_name}")

    def texte_auto_pulse_ms(self, pulse_name):
        us = self.lire_auto_pulse_us(pulse_name)
        return f"{us // 1000} ms"

    def auto_pulses_particuliers_actifs(self):
        return hasattr(self, "checkBox_auto_pulses_particuliers") and self.checkBox_auto_pulses_particuliers.isChecked()

    def update_auto_pulse_fields_state(self):
        if not hasattr(self, "lineEdit_auto_pulse_bebr_ms"):
            return
        override_actif = self.auto_pulses_particuliers_actifs()
        enabled = (not self.auto_neutral_running) and override_actif
        for edit in (
            self.lineEdit_auto_delai_ms,
            self.lineEdit_auto_pulse_bebr_ms,
            self.lineEdit_auto_pulse_be_ms,
            self.lineEdit_auto_pulse_br_ms,
        ):
            edit.setEnabled(enabled)
            edit.setToolTip(
                "Mode particulier actif : cette durée IHM est modifiable."
                if override_actif else
                "Mode normal : durée verrouillée pour éviter une modification opérateur accidentelle."
            )
        self.checkBox_auto_pulses_particuliers.setEnabled(not self.auto_neutral_running)
        if hasattr(self, "label_auto_pulse_info"):
            self.label_auto_pulse_info.setText(
                "Durées IHM actives" if override_actif else "Durées scénario actives"
            )

    def on_auto_pulses_particuliers_changed(self, _checked=False):
        self.update_auto_pulse_fields_state()
        self.render_auto_logigramme()

    def auto_contact_values(self):
        return tuple(self.contacts_known_values)

    def auto_active_contacts(self):
        """Retourne uniquement R1..RN et T1..TN selon le nombre d'inverseurs actifs."""
        n = self.lire_auto_nb_inverseurs()
        vals = self.auto_contact_values()
        r = vals[0:n]
        t = vals[4:4+n]
        return n, r, t

    def auto_contacts_known(self):
        try:
            _n, r, t = self.auto_active_contacts()
        except Exception:
            return False
        return all(v is not None for v in r) and all(v is not None for v in t)

    def auto_count_green_red(self):
        _n, r, t = self.auto_active_contacts()
        green_count = sum(1 for v in r if str(v) == "1")
        red_count = sum(1 for v in t if str(v) == "1")
        return green_count, red_count

    def auto_is_neutral_position(self):
        if not self.auto_contacts_known():
            return None
        n, _r, _t = self.auto_active_contacts()
        green_count, red_count = self.auto_count_green_red()
        # Définition opérationnelle Neutral screen (labo) :
        #   NON neutre uniquement pour 2 états propres sur les N pôles actifs :
        #     - tous les Repos R fermés ET aucun Travail T  (franchement au repos)
        #     - tous les Travail T fermés ET aucun Repos R   (franchement au travail)
        #   Tout autre état = neutral screen : tout ouvert, état mixte, ou
        #   incomplet (au moins un contact manquant).
        # (Critère plus large que MIL-PRF-39016 §3.5.4.2 strict — qui ne retient
        #  comme neutre que "tous contacts ouverts" — donc plus sévère/conservateur.)
        franchement_repos = (green_count == n and red_count == 0)
        franchement_travail = (red_count == n and green_count == 0)
        return not (franchement_repos or franchement_travail)

    def auto_is_latched_red(self):
        if not self.auto_contacts_known():
            return None
        _n, r, t = self.auto_active_contacts()
        return all(str(v) == "0" for v in r) and all(str(v) == "1" for v in t)

    def auto_is_reset_green(self):
        if not self.auto_contacts_known():
            return None
        _n, r, t = self.auto_active_contacts()
        return all(str(v) == "1" for v in r) and all(str(v) == "0" for v in t)

    def auto_apply_inverseur_filter(self):
        """Grise les LED hors nombre d'inverseurs actif dans l'onglet automatique."""
        if not hasattr(self, "label_auto_led_r1"):
            return
        try:
            n = self.lire_auto_nb_inverseurs()
        except Exception:
            n = 4
        leds_r = [self.label_auto_led_r1, self.label_auto_led_r2, self.label_auto_led_r3, self.label_auto_led_r4]
        leds_t = [self.label_auto_led_t1, self.label_auto_led_t2, self.label_auto_led_t3, self.label_auto_led_t4]
        for i, lab in enumerate(leds_r + leds_t):
            idx = (i % 4) + 1
            if idx > n:
                lab.setStyleSheet(
                    "background-color: rgb(120,120,120); border: 2px solid rgb(80,80,80); "
                    "border-radius: 14px;"
                )
                lab.setToolTip(f"Contact ignoré : inverseur {idx} hors nombre actif ({n})")

    def auto_timer_timeout(self):
        if not self.auto_neutral_running:
            return
        callback = self.auto_next_action
        self.auto_next_action = None
        if callback is not None:
            callback()

    def auto_schedule(self, callback, delay_ms=None):
        if not self.auto_neutral_running:
            return
        if delay_ms is None:
            delay_ms = self.lire_auto_delai_ms()
        self.auto_next_action = callback
        self.auto_neutral_timer.start(int(delay_ms))

    def auto_wait_status_then(self, callback):
        if not self.auto_neutral_running:
            return
        self.send_command("STATUS?")
        self.auto_schedule(callback, 120)

    def default_scenarios_data(self):
        return {
            "version": "2.12.3",
            "scenarios": [
                {"name":"Neutral screen norme","description":"Scénario strict MIL-PRF-39016H 4.8.7.7 : BE/BR max 3 avec vérification neutre, si pas neutre après 3 essais accepté, sinon BE puis vérification latch, répéter BE/BR avec vérification neutre, puis BR avec vérification reset.","steps":[
                    {"action":"BEBR","pulse_ms":10,"check":"NEUTRAL","max_attempts":3,"on_fail":"ACCEPT","description":"Recherche position neutre"},
                    {"action":"BE","pulse_ms":10,"check":"LATCH_RED","max_attempts":1,"on_fail":"REJECT","description":"Vérifier latch / rouges T"},
                    {"action":"BEBR","pulse_ms":10,"check":"NEUTRAL","max_attempts":1,"on_fail":"REJECT","description":"4.8.7.7b(2) : BE/BR, vérifier neutre"},
                    {"action":"BR","pulse_ms":10,"check":"RESET_GREEN","max_attempts":1,"on_fail":"REJECT","description":"Vérifier reset / verts R"}]},
                {"name":"BR puis BE","description":"Essai simple reset puis latch.","steps":[
                    {"action":"BR","pulse_ms":10,"check":"RESET_GREEN","max_attempts":1,"on_fail":"REJECT","description":"Reset d'abord"},
                    {"action":"BE","pulse_ms":10,"check":"LATCH_RED","max_attempts":1,"on_fail":"REJECT","description":"Latch ensuite"}]},
                {"name":"BE puis BR","description":"Essai simple latch puis reset.","steps":[
                    {"action":"BE","pulse_ms":10,"check":"LATCH_RED","max_attempts":1,"on_fail":"REJECT","description":"Latch d'abord"},
                    {"action":"BR","pulse_ms":10,"check":"RESET_GREEN","max_attempts":1,"on_fail":"REJECT","description":"Reset ensuite"}]},
                {"name":"BE/BR seulement x3","description":"Recherche de neutre uniquement, maximum 3 essais.","steps":[
                    {"action":"BEBR","pulse_ms":10,"check":"NEUTRAL","max_attempts":3,"on_fail":"ACCEPT","description":"Neutral screen seul"}]}
            ]
        }

    def resolve_scenario_file(self):
        if getattr(sys, "frozen", False):
            base = Path(sys.executable).resolve().parent
        else:
            base = APP_DIR
        candidate = base / "neutral_scenarios.json"
        try:
            base.mkdir(parents=True, exist_ok=True)
            with open(candidate, "a", encoding="utf-8"):
                pass
            return candidate
        except Exception:
            appdata = Path.home() / "AppData" / "Roaming" / "OutilsLabo" / "RelaisRP2040"
            appdata.mkdir(parents=True, exist_ok=True)
            return appdata / "neutral_scenarios.json"

    def resolve_production_db_file(self):
        if getattr(sys, "frozen", False):
            base = Path(sys.executable).resolve().parent
        else:
            base = APP_DIR
        candidate = base / "production_essais.sqlite3"
        try:
            base.mkdir(parents=True, exist_ok=True)
            with open(candidate, "ab"):
                pass
            return candidate
        except Exception:
            appdata = Path.home() / "AppData" / "Roaming" / "OutilsLabo" / "RelaisRP2040"
            appdata.mkdir(parents=True, exist_ok=True)
            return appdata / "production_essais.sqlite3"

    def resolve_chrono_db_file(self):
        if getattr(sys, "frozen", False):
            base = Path(sys.executable).resolve().parent
        else:
            base = APP_DIR
        candidate = base / "chronometrie_contacts.sqlite3"
        try:
            base.mkdir(parents=True, exist_ok=True)
            with open(candidate, "ab"):
                pass
            return candidate
        except Exception:
            appdata = Path.home() / "AppData" / "Roaming" / "OutilsLabo" / "RelaisRP2040"
            appdata.mkdir(parents=True, exist_ok=True)
            return appdata / "chronometrie_contacts.sqlite3"

    def resolve_legacy_production_json_file(self):
        if getattr(sys, "frozen", False):
            base = Path(sys.executable).resolve().parent
        else:
            base = APP_DIR
        return base / "production_essais.json"

    def scenarios_load_or_create(self):
        try:
            if not self.scenario_file.exists() or self.scenario_file.stat().st_size == 0:
                self.scenarios_data = self.default_scenarios_data()
                self.scenarios_save_file()
            else:
                with open(self.scenario_file, "r", encoding="utf-8") as f:
                    self.scenarios_data = json.load(f)
                if "scenarios" not in self.scenarios_data or not isinstance(self.scenarios_data["scenarios"], list):
                    raise ValueError("Structure JSON invalide : clé scenarios absente.")
        except Exception as exc:
            QMessageBox.warning(self.window, "Scénarios", f"Impossible de charger neutral_scenarios.json.\nUn fichier par défaut va être recréé.\n\nDétail : {exc}")
            self.scenarios_data = self.default_scenarios_data()
            self.scenarios_save_file()
        self.label_editor_fichier.setText(f"Fichier : {self.scenario_file}")

    def scenarios_save_file(self):
        with open(self.scenario_file, "w", encoding="utf-8") as f:
            json.dump(self.scenarios_data, f, ensure_ascii=False, indent=2)

    def scenarios_names(self):
        return [str(s.get("name", "")).strip() for s in self.scenarios_data.get("scenarios", []) if str(s.get("name", "")).strip()]

    def get_scenario_by_name(self, name):
        for scenario in self.scenarios_data.get("scenarios", []):
            if scenario.get("name") == name:
                return scenario
        return None

    def scenarios_refresh_all(self, prefer_name=None):
        self._refreshing_scenario_combos = True
        names = self.scenarios_names()
        if not names:
            self.scenarios_data = self.default_scenarios_data()
            self.scenarios_save_file()
            names = self.scenarios_names()
        current_auto = prefer_name or self.comboBox_auto_scenario.currentText()
        current_editor = prefer_name or self.comboBox_editor_scenarios.currentText()
        self.comboBox_auto_scenario.clear(); self.comboBox_auto_scenario.addItems(names)
        self.comboBox_editor_scenarios.clear(); self.comboBox_editor_scenarios.addItems(names)
        for combo, wanted in ((self.comboBox_auto_scenario, current_auto), (self.comboBox_editor_scenarios, current_editor)):
            if wanted in names: combo.setCurrentText(wanted)
            elif names: combo.setCurrentIndex(0)
        self.production_refresh_scenarios()
        self._refreshing_scenario_combos = False
        self.on_production_scenario_changed(self.comboBox_prod_scenario.currentText())
        self.on_auto_scenario_changed(); self.on_editor_scenario_changed()
        self.label_editor_status.setText("Éditeur : scénarios chargés")

    def on_recharger_scenarios(self):
        if self.auto_neutral_running:
            QMessageBox.warning(self.window, "Scénarios", "Impossible de recharger pendant un essai automatique.")
            return
        self.scenarios_load_or_create(); self.scenarios_refresh_all(); self.log("Scénarios rechargés.")

    def on_auto_scenario_changed(self):
        if getattr(self, "_refreshing_scenario_combos", False): return
        self.render_auto_logigramme()

    def on_editor_scenario_changed(self):
        if getattr(self, "_refreshing_scenario_combos", False): return
        self.editor_load_selected_scenario()

    def aller_onglet_editeur_scenarios(self):
        tab = self.window.findChild(QWidget, "tab_neutral_scenario_editor")
        tabw = self.window.findChild(QWidget, "tabWidget_principal")
        if tab is not None and tabw is not None:
            try: tabw.setCurrentWidget(tab)
            except Exception: pass

    def table_item(self, text):
        item = QTableWidgetItem("" if text is None else str(text))
        item.setToolTip("" if text is None else str(text))
        item.setTextAlignment(Qt.AlignCenter)
        return item

    def table_item_sort(self, text, sort_value=None):
        item = SortableTableWidgetItem("" if text is None else str(text))
        item.setToolTip("" if text is None else str(text))
        item.setTextAlignment(Qt.AlignCenter)
        if sort_value is not None:
            item.setData(Qt.UserRole, sort_value)
        return item


    def configure_table_readability(self, table, widths, row_height=28, stretch_last=False, word_wrap=False):
        """Configure les tableaux pour limiter les mots coupés."""
        try:
            table.setWordWrap(bool(word_wrap))
            table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            table.setAlternatingRowColors(True)
            table.verticalHeader().setDefaultSectionSize(row_height)
            table.verticalHeader().setMinimumSectionSize(row_height)
            table.horizontalHeader().setStretchLastSection(stretch_last)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
            for col, width in enumerate(widths):
                table.setColumnWidth(col, int(width))
        except Exception:
            pass

    def init_auto_table(self):
        table = self.tableWidget_auto_logigramme
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["État", "N°", "Action", "Durée", "Vérification", "Essai", "Si échec", "Résultat"])
        try:
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.setSelectionMode(QTableWidget.SingleSelection)
        except Exception:
            pass
        self.configure_table_readability(
            table,
            widths=[70, 42, 82, 82, 130, 72, 98, 205],
            row_height=28,
            stretch_last=False
        )


    def init_editor_table(self):
        table = self.tableWidget_editor_steps
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels(["Action", "Durée ms", "Vérification", "Max essais", "Si échec", "Description", "Note"])
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked | QTableWidget.EditKeyPressed)
        self.configure_table_readability(
            table,
            widths=EDITOR_TABLE_WIDTHS,
            row_height=42,
            stretch_last=False,
            word_wrap=True
        )
        try:
            table.setColumnHidden(6, True)  # colonne Note inutile à l'écran
        except Exception:
            pass


    def editor_allowed_actions(self):
        return ["BE", "BR", "BEBR", "STATUS", "PAUSE", "STOP"]

    def editor_allowed_checks(self):
        return ["NONE", "NEUTRAL", "LATCH_RED", "RESET_GREEN"]

    def editor_allowed_on_fail(self):
        return ["REJECT", "ACCEPT", "CONTINUE", "STOP"]

    def editor_make_combo(self, values, current):
        combo = QComboBox()
        combo.addItems(values)
        current = str(current or "").strip().upper()
        if current in values:
            combo.setCurrentText(current)
        else:
            combo.setCurrentIndex(0)
        combo.setMinimumWidth(82)
        combo.setStyleSheet("background-color: white; color: black;")
        return combo

    def editor_make_spin(self, minimum, maximum, current, suffix="", min_width=82):
        spin = QSpinBox()
        spin.setMinimum(minimum)
        spin.setMaximum(maximum)
        try:
            spin.setValue(int(current))
        except Exception:
            spin.setValue(minimum)
        if suffix:
            spin.setSuffix(suffix)
        spin.setMinimumWidth(int(min_width))
        spin.setStyleSheet("background-color: white; color: black;")
        return spin

    def editor_set_step_row(self, row, step):
        """Remplit une ligne avec des widgets guidés au lieu de champs libres."""
        step = dict(step or {})
        action = step.get("action", "BE")
        check = step.get("check", "NONE")
        on_fail = step.get("on_fail", "REJECT")
        pulse_ms = step.get("pulse_ms", 10)
        max_attempts = step.get("max_attempts", 1)
        description = step.get("description", "")

        self.tableWidget_editor_steps.setCellWidget(
            row, 0, self.editor_make_combo(self.editor_allowed_actions(), self.normalize_action(action))
        )
        self.tableWidget_editor_steps.setCellWidget(
            row, 1, self.editor_make_spin(0, 2000, pulse_ms, "", 104)
        )
        self.tableWidget_editor_steps.setCellWidget(
            row, 2, self.editor_make_combo(self.editor_allowed_checks(), self.normalize_check(check))
        )
        self.tableWidget_editor_steps.setCellWidget(
            row, 3, self.editor_make_spin(1, 99, max_attempts, "", 92)
        )
        self.tableWidget_editor_steps.setCellWidget(
            row, 4, self.editor_make_combo(self.editor_allowed_on_fail(), self.normalize_on_fail(on_fail))
        )

        desc_item = self.table_item(description)
        desc_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        desc_item.setToolTip(str(description))
        self.tableWidget_editor_steps.setItem(row, 5, desc_item)
        note = "Guidé"
        self.tableWidget_editor_steps.setItem(row, 6, self.table_item(note))

    def editor_cell_widget_value(self, row, col):
        widget = self.tableWidget_editor_steps.cellWidget(row, col)
        if isinstance(widget, QComboBox):
            return widget.currentText().strip()
        if isinstance(widget, QSpinBox):
            return widget.value()
        item = self.tableWidget_editor_steps.item(row, col)
        return "" if item is None else item.text().strip()

    def editor_step_from_row(self, row):
        return {
            "action": self.editor_cell_widget_value(row, 0),
            "pulse_ms": self.editor_cell_widget_value(row, 1),
            "check": self.editor_cell_widget_value(row, 2),
            "max_attempts": self.editor_cell_widget_value(row, 3),
            "on_fail": self.editor_cell_widget_value(row, 4),
            "description": self.editor_cell_widget_value(row, 5),
        }

    def editor_all_steps_from_table(self):
        steps = []
        for row in range(self.tableWidget_editor_steps.rowCount()):
            step = self.editor_step_from_row(row)
            if not any(str(step.get(k, "")).strip() for k in ("action", "description")):
                continue
            steps.append(step)
        return steps

    def editor_warn_guided_fields(self):
        self.label_editor_status.setText(
            "Éditeur : champs guidés — Action/Vérification/Si échec par listes ; Durée/Max essais bornés."
        )

    def normalize_action(self, value):
        raw = str(value or "").strip().upper()
        action = raw.replace("/", "").replace("+", "").replace("-", "").replace("_", "")
        if action == "BEBR": return "BEBR"
        if action in ("BE", "BR", "STATUS", "PAUSE", "STOP"): return action
        raise ValueError(f"Action invalide : {value}. Actions : BE, BR, BEBR, STATUS, PAUSE, STOP.")

    def normalize_check(self, value):
        check = str(value or "NONE").strip().upper()
        aliases = {"AUCUNE":"NONE", "RIEN":"NONE", "NEUTRE":"NEUTRAL", "ROUGE":"LATCH_RED", "LATCH":"LATCH_RED", "VERT":"RESET_GREEN", "RESET":"RESET_GREEN"}
        check = aliases.get(check, check)
        if check not in ("NONE", "NEUTRAL", "LATCH_RED", "RESET_GREEN"):
            raise ValueError(f"Vérification invalide : {value}. Vérifications : NONE, NEUTRAL, LATCH_RED, RESET_GREEN.")
        return check

    def normalize_on_fail(self, value):
        on_fail = str(value or "REJECT").strip().upper()
        aliases = {"REJETER":"REJECT", "ACCEPTER":"ACCEPT", "CONTINUER":"CONTINUE", "ARRET":"STOP", "ARRÊT":"STOP"}
        on_fail = aliases.get(on_fail, on_fail)
        if on_fail not in ("REJECT", "ACCEPT", "CONTINUE", "STOP"):
            raise ValueError(f"Si échec invalide : {value}. Valeurs : REJECT, ACCEPT, CONTINUE, STOP.")
        return on_fail

    def validate_step(self, step):
        action = self.normalize_action(step.get("action", ""))
        check = self.normalize_check(step.get("check", "NONE"))
        on_fail = self.normalize_on_fail(step.get("on_fail", "REJECT"))
        pulse_ms = int(step.get("pulse_ms", 10))
        if pulse_ms < 0 or pulse_ms > 2000: raise ValueError("Durée ms doit être comprise entre 0 et 2000.")
        if action in ("BE", "BR", "BEBR") and pulse_ms < 1: raise ValueError("Un pulse BE/BR/BE/BR doit durer au moins 1 ms.")
        max_attempts = int(step.get("max_attempts", 1))
        if max_attempts < 1 or max_attempts > 99: raise ValueError("Max essais doit être compris entre 1 et 99.")
        return {"action":action,"pulse_ms":pulse_ms,"check":check,"max_attempts":max_attempts,"on_fail":on_fail,"description":str(step.get("description", "") or "")}

    def validate_scenario(self, scenario):
        name = str(scenario.get("name", "")).strip()
        if not name: raise ValueError("Le scénario doit avoir un nom.")
        steps_in = scenario.get("steps", [])
        if not isinstance(steps_in, list) or not steps_in: raise ValueError("Le scénario doit contenir au moins une étape.")
        steps = [self.validate_step(step) for step in steps_in]
        return {"name":name, "description":str(scenario.get("description", "") or ""), "steps":steps}

    def render_auto_logigramme(self):
        if not hasattr(self, "tableWidget_auto_logigramme"): return
        self.init_auto_table()
        scenario = self.get_scenario_by_name(self.comboBox_auto_scenario.currentText())
        steps = scenario.get("steps", []) if scenario else []
        table = self.tableWidget_auto_logigramme; table.setRowCount(len(steps))
        for i, raw_step in enumerate(steps):
            try:
                step = self.validate_step(raw_step)
                duree = f'{step["pulse_ms"]} ms'
                if self.auto_pulses_particuliers_actifs() and step["action"] in ("BE", "BR", "BEBR"):
                    duree = f'{self.scenario_pulse_us_for_step(step) // 1000} ms IHM'
                vals = ["à faire", i+1, step["action"], duree, step["check"], f'0/{step["max_attempts"]}', step["on_fail"], ""]
            except Exception as exc:
                vals = ["erreur", i+1, raw_step.get("action","?"), raw_step.get("pulse_ms","?"), f"ERREUR : {exc}", "?", "?", ""]
            for col, val in enumerate(vals): table.setItem(i, col, self.table_item(val))
            self.set_auto_row_color(i, "idle")
        self.configure_table_readability(table, widths=[70, 42, 82, 82, 130, 72, 98, 205], row_height=28)

    def set_auto_row_color(self, row, state, result_text=None, attempt_text=None):
        if row < 0 or row >= self.tableWidget_auto_logigramme.rowCount():
            return
        colors = {
            "idle": QColor(235, 235, 235),
            "active": QColor(255, 235, 120),
            "done": QColor(210, 230, 255),
            "ok": QColor(120, 255, 120),
            "fail": QColor(255, 120, 120),
            "stop": QColor(255, 180, 80),
        }
        labels = {
            "idle": "à faire",
            "active": "en cours",
            "done": "fait",
            "ok": "OK",
            "fail": "KO",
            "stop": "arrêt",
        }
        color = colors.get(state, colors["idle"])
        for col in range(self.tableWidget_auto_logigramme.columnCount()):
            item = self.tableWidget_auto_logigramme.item(row, col)
            if item is None:
                item = self.table_item("")
                self.tableWidget_auto_logigramme.setItem(row, col, item)
            item.setBackground(color)
            item.setToolTip(item.text())

        item_state = self.tableWidget_auto_logigramme.item(row, 0)
        item_state.setText(labels.get(state, state))
        item_state.setToolTip(labels.get(state, state))

        if attempt_text is not None:
            item = self.tableWidget_auto_logigramme.item(row, 5)
            item.setText(str(attempt_text))
            item.setToolTip(str(attempt_text))

        if result_text is not None:
            item = self.tableWidget_auto_logigramme.item(row, 7)
            item.setText(str(result_text))
            item.setToolTip(str(result_text))


    def editor_load_selected_scenario(self):
        if not hasattr(self, "tableWidget_editor_steps"):
            return
        self.init_editor_table()
        scenario = self.get_scenario_by_name(self.comboBox_editor_scenarios.currentText())
        if scenario is None:
            return
        self.lineEdit_editor_nom.setText(str(scenario.get("name", "")))
        self.textEdit_editor_description.setPlainText(str(scenario.get("description", "")))
        steps = scenario.get("steps", [])
        self.tableWidget_editor_steps.setRowCount(len(steps))
        for row, step in enumerate(steps):
            try:
                clean = self.validate_step(step)
            except Exception:
                clean = {"action": "BE", "pulse_ms": 10, "check": "NONE", "max_attempts": 1, "on_fail": "REJECT", "description": f"Étape corrigée : {step}"}
            self.editor_set_step_row(row, clean)
        self.configure_table_readability(self.tableWidget_editor_steps, widths=EDITOR_TABLE_WIDTHS, row_height=42, word_wrap=True)
        try:
            self.tableWidget_editor_steps.setColumnHidden(6, True)
        except Exception:
            pass
        self.editor_warn_guided_fields()

    def editor_read_scenario_from_ui(self):
        name = self.lineEdit_editor_nom.text().strip()
        desc = self.textEdit_editor_description.toPlainText().strip()
        steps = self.editor_all_steps_from_table()
        return self.validate_scenario({"name": name, "description": desc, "steps": steps})

    def editor_nouveau_scenario(self):
        base="Nouveau scénario"; names=set(self.scenarios_names()); name=base; idx=1
        while name in names: idx+=1; name=f"{base} {idx}"
        scenario={"name":name,"description":"Nouveau scénario à compléter.","steps":[{"action":"BE","pulse_ms":10,"check":"LATCH_RED","max_attempts":1,"on_fail":"REJECT","description":"Étape exemple"}]}
        self.scenarios_data.setdefault("scenarios",[]).append(scenario); self.scenarios_save_file(); self.scenarios_refresh_all(name); self.label_editor_status.setText(f"Éditeur : scénario créé : {name}")

    def editor_dupliquer_scenario(self):
        src=self.comboBox_editor_scenarios.currentText(); sc=self.get_scenario_by_name(src)
        if sc is None: return
        new=copy.deepcopy(sc); base=f"{src} copie"; names=set(self.scenarios_names()); name=base; idx=1
        while name in names: idx+=1; name=f"{base} {idx}"
        new["name"]=name; self.scenarios_data.setdefault("scenarios",[]).append(new); self.scenarios_save_file(); self.scenarios_refresh_all(name); self.label_editor_status.setText(f"Éditeur : scénario dupliqué : {name}")

    def editor_supprimer_scenario(self):
        name=self.comboBox_editor_scenarios.currentText()
        if len(self.scenarios_data.get("scenarios",[]))<=1: QMessageBox.warning(self.window,"Scénarios","Impossible de supprimer le dernier scénario."); return
        if QMessageBox.question(self.window,"Supprimer scénario",f"Supprimer le scénario : {name} ?") != QMessageBox.Yes: return
        self.scenarios_data["scenarios"]=[s for s in self.scenarios_data.get("scenarios",[]) if s.get("name")!=name]
        self.scenarios_save_file(); self.scenarios_refresh_all(); self.label_editor_status.setText(f"Éditeur : scénario supprimé : {name}")

    def editor_sauvegarder_scenario(self):
        try: scenario=self.editor_read_scenario_from_ui()
        except Exception as exc: QMessageBox.warning(self.window,"Scénario invalide",str(exc)); return
        old=self.comboBox_editor_scenarios.currentText(); names=[s.get("name") for s in self.scenarios_data.get("scenarios",[])]
        if scenario["name"] != old and scenario["name"] in names: QMessageBox.warning(self.window,"Scénario invalide","Un scénario avec ce nom existe déjà."); return
        for i,s in enumerate(self.scenarios_data.get("scenarios",[])):
            if s.get("name")==old: self.scenarios_data["scenarios"][i]=scenario; break
        else: self.scenarios_data.setdefault("scenarios",[]).append(scenario)
        self.scenarios_save_file(); self.scenarios_refresh_all(scenario["name"]); self.label_editor_status.setText(f"Éditeur : scénario sauvegardé : {scenario['name']}")

    def editor_ajouter_etape(self):
        row = self.tableWidget_editor_steps.rowCount()
        self.tableWidget_editor_steps.insertRow(row)
        self.editor_set_step_row(
            row,
            {"action": "BE", "pulse_ms": 10, "check": "NONE", "max_attempts": 1, "on_fail": "REJECT", "description": "Nouvelle étape"}
        )
        self.tableWidget_editor_steps.selectRow(row)
        self.configure_table_readability(self.tableWidget_editor_steps, widths=EDITOR_TABLE_WIDTHS, row_height=42, word_wrap=True)
        self.editor_warn_guided_fields()

    def editor_supprimer_etape(self):
        row = self.tableWidget_editor_steps.currentRow()
        if row >= 0:
            self.tableWidget_editor_steps.removeRow(row)
            self.editor_warn_guided_fields()

    def editor_deplacer_etape(self, sens):
        row = self.tableWidget_editor_steps.currentRow()
        new = row + sens
        if row < 0 or new < 0 or new >= self.tableWidget_editor_steps.rowCount():
            return

        steps = self.editor_all_steps_from_table()
        if row >= len(steps) or new >= len(steps):
            return

        steps[row], steps[new] = steps[new], steps[row]
        self.tableWidget_editor_steps.setRowCount(len(steps))
        for r, step in enumerate(steps):
            self.editor_set_step_row(r, step)

        self.tableWidget_editor_steps.selectRow(new)
        self.configure_table_readability(self.tableWidget_editor_steps, widths=EDITOR_TABLE_WIDTHS, row_height=42, word_wrap=True)
        self.editor_warn_guided_fields()

    def editor_importer_json(self):
        path,_=QFileDialog.getOpenFileName(self.window,"Importer scénarios JSON",str(self.scenario_file.parent),"JSON (*.json)")
        if not path: return
        try:
            with open(path,"r",encoding="utf-8") as f: data=json.load(f)
            if "scenarios" not in data or not isinstance(data["scenarios"],list): raise ValueError("Le JSON ne contient pas de liste scenarios.")
            for sc in data["scenarios"]: self.validate_scenario(sc)
            self.scenarios_data=data; self.scenarios_save_file(); self.scenarios_refresh_all(); self.label_editor_status.setText(f"Éditeur : import OK depuis {path}")
        except Exception as exc: QMessageBox.warning(self.window,"Import JSON impossible",str(exc))

    def editor_exporter_json(self):
        path = self.ask_export_path("Exporter scénarios JSON", "neutral_scenarios_export.json", "JSON (*.json)", ".json")
        if not path: return
        try:
            with open(path,"w",encoding="utf-8") as f: json.dump(self.scenarios_data,f,ensure_ascii=False,indent=2)
            self.label_editor_status.setText(f"Éditeur : export OK vers {path}")
        except Exception as exc: QMessageBox.warning(self.window,"Export JSON impossible",str(exc))

    # Surcharges de l'automate fixe : exécution depuis scénario
    def initialiser_auto_neutral(self):
        if hasattr(self,"tableWidget_auto_logigramme"): self.init_auto_table()
        old_flow=self.window.findChild(QWidget,"groupBox_auto_logigramme")
        if old_flow is not None: old_flow.setVisible(False)
        self.auto_neutral_running=False; self.auto_neutral_attempt=0; self.auto_next_action=None; self.runtime_step_index=0; self.runtime_attempt=0; self.current_runtime_steps=[]
        self.set_auto_finish_validation_state(False)
        if hasattr(self,"auto_neutral_timer"): self.auto_neutral_timer.stop()
        if hasattr(self,"label_auto_status"):
            self.label_auto_status.setText("Automatique : prêt"); self.label_auto_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")
        if hasattr(self,"label_auto_resultat"):
            self.label_auto_resultat.setText("Résultat : --"); self.label_auto_resultat.setStyleSheet("background-color: rgb(90,90,90); color: white; font-size: 14pt; font-weight: bold; border: 2px solid black;")
        if hasattr(self, "refresh_auto_leds_from_known_values"):
            self.refresh_auto_leds_from_known_values()
        self.update_button_states()

    def scenario_selected_for_run(self):
        name=self.comboBox_auto_scenario.currentText().strip(); sc=self.get_scenario_by_name(name)
        if sc is None: raise ValueError("Aucun scénario sélectionné.")
        return self.validate_scenario(sc)

    def scenario_pulse_us_for_step(self, step):
        action = self.normalize_action(step.get("action"))
        if self.auto_pulses_particuliers_actifs() and action in ("BE", "BR", "BEBR"):
            return self.lire_auto_pulse_us(action)
        return int(step.get("pulse_ms", 0)) * 1000

    def big_message_box(self, title, main_text, detail_text="", ok_text="VALIDER", cancel_text=None, icon=QMessageBox.Warning):
        dialog = QDialog(self.window)
        dialog.setWindowTitle(title)
        dialog.setModal(True)
        dialog.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dialog.setFixedSize(760, 360)
        dialog.setStyleSheet("""
            QDialog {
                background-color: rgb(255,255,245);
                border: 5px solid rgb(40,40,40);
            }
            QLabel#mainTitle {
                color: black;
                font-size: 24pt;
                font-weight: bold;
            }
            QLabel#detailText {
                color: black;
                font-size: 16pt;
                font-weight: bold;
            }
            QPushButton {
                color: black;
                background-color: white;
                font-size: 18pt;
                font-weight: bold;
                min-width: 220px;
                min-height: 58px;
                border: 2px solid rgb(90,90,90);
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:focus {
                border: 5px solid rgb(0,120,215);
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(34, 28, 34, 28)
        layout.setSpacing(16)

        title_label = QLabel(main_text)
        title_label.setObjectName("mainTitle")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setWordWrap(True)
        title_label.setMinimumHeight(68)
        layout.addWidget(title_label)

        if detail_text:
            detail_label = QLabel(detail_text)
            detail_label.setObjectName("detailText")
            detail_label.setAlignment(Qt.AlignCenter)
            detail_label.setWordWrap(True)
            detail_label.setMinimumHeight(92)
            layout.addWidget(detail_label)

        layout.addStretch(1)

        button_row = QHBoxLayout()
        button_row.setSpacing(24)
        button_row.addStretch(1)
        ok_button = QPushButton(ok_text)
        ok_button.clicked.connect(dialog.accept)
        button_row.addWidget(ok_button)
        if cancel_text is not None:
            cancel_button = QPushButton(cancel_text)
            cancel_button.clicked.connect(dialog.reject)
            button_row.addWidget(cancel_button)
            cancel_button.setDefault(True)
            cancel_button.setFocus()
        else:
            ok_button.setDefault(True)
            ok_button.setFocus()
        button_row.addStretch(1)
        layout.addLayout(button_row)

        return dialog.exec() == QDialog.Accepted

    def big_sn_input_dialog(self, initial_sn):
        dialog = QDialog(self.window)
        dialog.setWindowTitle("NUMERO DE SERIE")
        dialog.setModal(True)
        dialog.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dialog.setFixedSize(760, 360)
        dialog.setStyleSheet("""
            QDialog {
                background-color: rgb(245,250,255);
                border: 5px solid rgb(40,40,40);
            }
            QLabel#mainTitle {
                color: black;
                font-size: 24pt;
                font-weight: bold;
            }
            QLineEdit {
                color: black;
                background-color: white;
                font-size: 32pt;
                font-weight: bold;
                min-height: 62px;
                padding: 6px;
            }
            QPushButton {
                color: black;
                background-color: white;
                font-size: 18pt;
                font-weight: bold;
                min-width: 210px;
                min-height: 58px;
                border: 2px solid rgb(90,90,90);
                border-radius: 6px;
                padding: 6px;
            }
            QPushButton:focus {
                border: 5px solid rgb(0,120,215);
            }
        """)
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(34, 28, 34, 28)
        layout.setSpacing(16)

        title_label = QLabel("SCANNER OU SAISIR LE SN\nDU PREMIER RELAIS")
        title_label.setObjectName("mainTitle")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setWordWrap(True)
        title_label.setMinimumHeight(86)
        layout.addWidget(title_label)

        edit = QLineEdit(str(initial_sn or ""))
        edit.setAlignment(Qt.AlignCenter)
        edit.selectAll()
        layout.addWidget(edit)
        layout.addStretch(1)

        button_row = QHBoxLayout()
        button_row.setSpacing(24)
        button_row.addStretch(1)
        ok_button = QPushButton("VALIDER SN")
        cancel_button = QPushButton("ANNULER")
        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        button_row.addWidget(ok_button)
        button_row.addWidget(cancel_button)
        button_row.addStretch(1)
        layout.addLayout(button_row)

        ok_button.setDefault(True)
        edit.setFocus()
        ok = dialog.exec() == QDialog.Accepted
        return edit.text().strip(), ok

    def ensure_auto_production_ready(self):
        missing = self.production_missing_required_fields(include_sn=False)
        if missing:
            self.production_show_missing_fields_message(missing)
            self.label_prod_status.setText("MARCHE AUTO refusé : informations production incomplètes.")
            tab = self.window.findChild(QWidget, "tab_production_accueil")
            if tab is not None:
                self.set_tab_internal(tab)
            return False
        if self.auto_end_validation_pending:
            self.big_message_box(
                "Test fini non validé",
                "TEST FINI NON VALIDÉ",
                "Validez le relais précédent avec le bouton TEST FINI ou la touche ENTRÉE.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return False
        if self.interrupted_auto_sn:
            sn = self.interrupted_auto_sn
            scenario = self.interrupted_auto_scenario or self.comboBox_auto_scenario.currentText().strip()
            if not self.big_message_box(
                "Reprise essai interrompu",
                "ESSAI INTERROMPU PAR USB",
                f"Le relais SN {sn} n'a pas été enregistré.\n\nVoulez-vous refaire ce même relais ?\nScénario : {scenario or '-'}",
                ok_text=f"REFAIRE SN {sn}",
                cancel_text="ANNULER",
                icon=QMessageBox.Warning,
            ):
                return False
            self.lineEdit_SN.setText(sn)
            if scenario and self.comboBox_auto_scenario.findText(scenario) >= 0:
                self.comboBox_auto_scenario.setCurrentText(scenario)
            self.interrupted_auto_sn = ""
            self.interrupted_auto_scenario = ""
            self.interrupted_auto_reason = ""
            self._auto_start_prompts_done = True
            self.production_data["last_context"] = self.production_context()
            try:
                self.production_save_db()
            except Exception as exc:
                self.label_prod_status.setText(f"Reprise SN non sauvegardée dans la base : {exc}")
            return True
        if self._auto_start_prompts_done:
            missing = self.production_missing_required_fields(include_sn=True)
            if missing:
                self.production_show_missing_fields_message(missing)
                self.label_prod_status.setText("MARCHE AUTO refusé : informations production/SN incomplètes.")
                tab = self.window.findChild(QWidget, "tab_production_accueil")
                if tab is not None:
                    self.tabWidget_principal.setCurrentWidget(tab)
                return False
            return True

        if not self.big_message_box(
            "Contrôle alimentations",
            "VÉRIFIER LES 2 ALIMENTATIONS",
            "Les deux tensions doivent être réglées correctement sur les alimentations avant de lancer l'essai.",
            ok_text="TENSIONS OK",
            cancel_text="ANNULER",
            icon=QMessageBox.Warning,
        ):
            return False

        sn_initial = self.lineEdit_SN.text().strip()
        sn, ok = self.big_sn_input_dialog(sn_initial)
        if not ok:
            return False
        sn = str(sn or "").strip()
        if not sn:
            self.big_message_box(
                "SN obligatoire",
                "SN OBLIGATOIRE",
                "Le numéro de série du premier relais est obligatoire.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
        )
            return False

        self.lineEdit_SN.setText(sn)
        self._auto_start_prompts_done = True
        self.production_data["last_context"] = self.production_context()
        try:
            self.production_save_db()
        except Exception as exc:
            self.label_prod_status.setText(f"SN non sauvegardé dans la base : {exc}")
        return True

    def set_auto_finish_validation_state(self, pending):
        self.auto_end_validation_pending = bool(pending)

    def show_auto_finish_validation_dialog(self, accepted):
        if not getattr(self, "auto_end_validation_pending", False):
            return

        sn = str(getattr(self, "last_finished_sn", "") or "").strip() or "-"
        if accepted:
            result_text = "RELAIS ACCEPTÉ"
            result_color = "rgb(0,190,80)"
            result_color_dim = "rgb(120,255,170)"
            border_color = "rgb(0,90,35)"
        else:
            result_text = "RELAIS REJETÉ"
            result_color = "rgb(230,0,0)"
            result_color_dim = "rgb(255,120,120)"
            border_color = "rgb(130,0,0)"

        dialog = QDialog(self.window)
        dialog.setWindowTitle("TEST FINI")
        dialog.setModal(True)
        dialog.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dialog.setFixedSize(820, 420)
        dialog.setStyleSheet("""
            QDialog {
                background-color: rgb(245,245,245);
                border: 6px solid rgb(35,35,35);
            }
            QLabel#title {
                color: black;
                font-size: 28pt;
                font-weight: bold;
            }
            QLabel#sn {
                color: black;
                font-size: 22pt;
                font-weight: bold;
            }
        """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(36, 30, 36, 30)
        layout.setSpacing(20)

        title = QLabel(result_text)
        title.setObjectName("title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        sn_label = QLabel(f"SN : {sn}")
        sn_label.setObjectName("sn")
        sn_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(sn_label)

        button = QPushButton("TEST FINI")
        button.setMinimumHeight(150)
        button.setDefault(True)
        button.setFocus()
        button.clicked.connect(dialog.accept)
        layout.addWidget(button)

        def set_button_style(color):
            button.setStyleSheet(
                f"background-color: {color}; color: black; "
                "font-size: 38pt; font-weight: bold; "
                f"border: 8px solid {border_color}; border-radius: 10px;"
            )

        blink_state = {"on": True}
        set_button_style(result_color)
        blink_timer = QTimer(dialog)
        blink_timer.setInterval(450)

        def blink():
            blink_state["on"] = not blink_state["on"]
            set_button_style(result_color if blink_state["on"] else result_color_dim)

        blink_timer.timeout.connect(blink)
        blink_timer.start()

        QShortcut(QKeySequence("Return"), dialog).activated.connect(dialog.accept)
        QShortcut(QKeySequence("Enter"), dialog).activated.connect(dialog.accept)

        accepted_dialog = dialog.exec() == QDialog.Accepted
        blink_timer.stop()
        if accepted_dialog and getattr(self, "auto_end_validation_pending", False):
            self.auto_validate_end_of_test()

    def auto_validate_end_of_test(self):
        if not getattr(self, "auto_end_validation_pending", False):
            return
        self.set_auto_finish_validation_state(False)
        self.label_auto_status.setText(f"Test fini validé - SN suivant prêt : {self.lineEdit_SN.text().strip()}")
        self.label_auto_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")
        self.label_auto_resultat.setText("Résultat : --")
        self.label_auto_resultat.setStyleSheet(
            "background-color: rgb(90,90,90); color: white; "
            "font-size: 14pt; font-weight: bold; border: 2px solid black;"
        )
        self.send_led_command("CONNECTED")
        self.update_button_states()

    def auto_finish_lot(self):
        if getattr(self, "auto_neutral_running", False):
            self.big_message_box(
                "Lot fini impossible",
                "ESSAI EN COURS",
                "Arrêtez ou terminez l'essai en cours avant de clôturer le lot.",
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )
            return
        lot = self.lineEdit_prod_lot.text().strip() or "(lot non renseigné)"
        if not self.big_message_box(
            "Confirmation lot fini",
            "LOT FINI ?",
            f"Confirmer la fin du test sur le lot : {lot}",
            ok_text="OUI LOT FINI",
            cancel_text="NON",
            icon=QMessageBox.Question,
        ):
            return
        self.set_auto_finish_validation_state(False)
        self._auto_start_prompts_done = False
        self._lot_session_active = False
        self._active_lot_finished = True
        self._active_lot = ""
        self.interrupted_auto_sn = ""
        self.interrupted_auto_scenario = ""
        self.interrupted_auto_reason = ""
        self.lineEdit_SN.clear()
        self.production_data["last_context"] = self.production_context()
        try:
            self.production_save_db()
        except Exception as exc:
            self.label_prod_status.setText(f"Lot fini : contexte non sauvegardé : {exc}")
        self.label_auto_status.setText("Lot fini - retour Production")
        self.label_auto_status.setStyleSheet("background-color: rgb(70,70,70); color: white; font-weight: bold; border: 2px solid black;")
        self.label_auto_resultat.setText("Résultat : --")
        self.label_auto_resultat.setStyleSheet(
            "background-color: rgb(90,90,90); color: white; "
            "font-size: 14pt; font-weight: bold; border: 2px solid black;"
        )
        self.label_prod_status.setText(f"Lot fini : {lot}")
        self.send_led_command("CONNECTED")
        self.set_tab_internal(self.tab_production_accueil)
        self.update_button_states()

    def auto_neutral_start(self):
        if not self.is_connected(): QMessageBox.warning(self.window,"Connexion","RP2040 non connecté."); return
        try:
            self.production_sync_nb_inverseurs_to_auto()
            self.lire_auto_delai_ms(); self.lire_auto_nb_inverseurs()
            if self.auto_pulses_particuliers_actifs():
                self.lire_auto_pulse_us("BEBR"); self.lire_auto_pulse_us("BE"); self.lire_auto_pulse_us("BR")
            scenario=self.scenario_selected_for_run()
        except Exception as exc: QMessageBox.warning(self.window,"Paramètre automatique invalide",str(exc)); return
        if not self.ensure_auto_production_ready():
            return
        self.set_auto_finish_validation_state(False)
        self.button_help_filter.hide()
        self.current_runtime_scenario_name=scenario["name"]; self.current_runtime_steps=copy.deepcopy(scenario["steps"]); self.runtime_step_index=0; self.runtime_attempt=0; self.auto_neutral_running=True; self.contacts_force_refresh=True
        self.render_auto_logigramme(); self.auto_update_tension_labels(); n=self.lire_auto_nb_inverseurs()
        mode_durees = "durées IHM particulières" if self.auto_pulses_particuliers_actifs() else "durées scénario"
        self.label_auto_status.setText(f"Automatique : scénario '{scenario['name']}' avec {n} inverseur(s) - {mode_durees}"); self.label_auto_status.setStyleSheet("background-color: rgb(255,235,120); color: black; font-weight: bold; border: 2px solid rgb(180,120,0);")
        self.label_auto_resultat.setText("Résultat : essai en cours"); self.label_auto_resultat.setStyleSheet("background-color: rgb(255,235,120); color: black; font-size: 14pt; font-weight: bold; border: 2px solid rgb(180,120,0);")
        self.update_button_states(); self.send_led_command("CONNECTED"); self.send_command("STATUS?"); self.send_command("STOP"); self.auto_schedule(self.scenario_execute_current_step,250)

    def auto_neutral_stop(self, send_stop=True):
        self.auto_neutral_running=False; self.auto_next_action=None
        if hasattr(self,"auto_neutral_timer"): self.auto_neutral_timer.stop()
        if send_stop and self.is_connected(): self.send_command("STOP")
        if 0 <= getattr(self,"runtime_step_index",-1) < self.tableWidget_auto_logigramme.rowCount(): self.set_auto_row_color(self.runtime_step_index,"stop","Arrêt opérateur")
        self.label_auto_status.setText("Automatique : arrêt demandé"); self.label_auto_status.setStyleSheet("background-color: rgb(255,180,80); color: black; font-weight: bold; border: 2px solid rgb(160,80,0);")
        self.label_auto_resultat.setText("Résultat : arrêté par opérateur"); self.label_auto_resultat.setStyleSheet("background-color: rgb(255,180,80); color: black; font-size: 14pt; font-weight: bold; border: 2px solid rgb(160,80,0);")
        self.send_led_command("CONNECTED")
        self.update_button_states()

    def scenario_execute_current_step(self):
        if not self.auto_neutral_running: return
        if self.runtime_step_index >= len(self.current_runtime_steps): self.auto_finish_accept_ok(); return
        step=self.current_runtime_steps[self.runtime_step_index]; self.runtime_attempt += 1
        action=self.normalize_action(step.get("action")); pulse_ms=int(step.get("pulse_ms",0)); check=self.normalize_check(step.get("check","NONE")); max_attempts=int(step.get("max_attempts",1))
        self.set_auto_row_color(self.runtime_step_index,"active","Action en cours",f"{self.runtime_attempt}/{max_attempts}")
        if action in ("BE","BR","BEBR"):
            pulse_us=self.scenario_pulse_us_for_step(step)
            if pulse_us < 1000: self.auto_finish_reject("durée pulse invalide"); return
            pulse_ms_effectif = max(1, (pulse_us + 999) // 1000)
            duree_source = "IHM" if self.auto_pulses_particuliers_actifs() else "scénario"
            self.auto_update_tension_labels(); self.label_auto_status.setText(f"Automatique : étape {self.runtime_step_index+1}/{len(self.current_runtime_steps)} - {action} {pulse_ms_effectif} ms ({duree_source}) - vérif {check}")
            self.send_led_command(action)
            self.send_command(f"PULSE_US;{action};{pulse_us}")
            wait_ms=max(120,pulse_ms_effectif+self.lire_auto_delai_ms())
            self.auto_schedule(lambda: self.auto_wait_status_then(self.scenario_check_current_step), wait_ms); return
        if action == "STATUS": self.send_command("STATUS?"); self.auto_schedule(self.scenario_check_current_step,self.lire_auto_delai_ms()); return
        if action == "PAUSE": self.label_auto_status.setText(f"Automatique : pause {pulse_ms} ms"); self.auto_schedule(self.scenario_step_success,max(1,pulse_ms)); return
        if action == "STOP": self.send_command("STOP"); self.auto_schedule(self.scenario_step_success,self.lire_auto_delai_ms()); return
        self.auto_finish_reject(f"action non supportée : {action}")

    def scenario_check_current_step(self):
        if not self.auto_neutral_running: return
        if self.runtime_step_index >= len(self.current_runtime_steps): self.auto_finish_accept_ok(); return
        step=self.current_runtime_steps[self.runtime_step_index]; check=self.normalize_check(step.get("check","NONE")); max_attempts=int(step.get("max_attempts",1)); on_fail=self.normalize_on_fail(step.get("on_fail","REJECT"))
        ok=True; detail="Aucune vérification"
        # V2.12.3 : garde anti faux-ACCEPT. Si la vérif dépend des contacts
        # mais que le retour RP2040 n'est pas encore connu, on rejette
        # explicitement au lieu de consommer un essai vers ACCEPT.
        if check in ("NEUTRAL","LATCH_RED","RESET_GREEN") and not self.auto_contacts_known():
            self.set_auto_row_color(self.runtime_step_index,"fail","Contacts inconnus (retour RP2040 absent)",f"{self.runtime_attempt}/{max_attempts}")
            self.auto_finish_reject("contacts inconnus : retour RP2040 absent"); return
        if check == "NEUTRAL":
            res=self.auto_is_neutral_position(); ok=False if res is None else bool(res); detail="NEUTRAL OK" if ok else "NEUTRAL non atteint"
        elif check == "LATCH_RED":
            res=self.auto_is_latched_red(); ok=False if res is None else bool(res); detail="LATCH_RED OK" if ok else "LATCH_RED KO"
        elif check == "RESET_GREEN":
            res=self.auto_is_reset_green(); ok=False if res is None else bool(res); detail="RESET_GREEN OK" if ok else "RESET_GREEN KO"
        if ok:
            self.set_auto_row_color(self.runtime_step_index,"ok",detail,f"{self.runtime_attempt}/{max_attempts}"); self.scenario_step_success(); return
        if self.runtime_attempt < max_attempts:
            self.set_auto_row_color(self.runtime_step_index,"done",f"{detail} - répétition",f"{self.runtime_attempt}/{max_attempts}")
            self.label_auto_status.setText(f"Automatique : étape {self.runtime_step_index+1} non conforme, répétition {self.runtime_attempt+1}/{max_attempts}")
            self.auto_schedule(self.scenario_execute_current_step); return
        self.set_auto_row_color(self.runtime_step_index,"fail",detail,f"{self.runtime_attempt}/{max_attempts}")
        if on_fail == "ACCEPT": self.auto_finish_accept_no_neutral()
        elif on_fail == "CONTINUE": self.scenario_step_success()
        elif on_fail == "STOP": self.auto_neutral_stop(send_stop=True)
        else: self.auto_finish_reject(detail)

    def scenario_step_success(self):
        if not self.auto_neutral_running: return
        self.runtime_step_index += 1; self.runtime_attempt=0
        if self.runtime_step_index >= len(self.current_runtime_steps): self.auto_finish_accept_ok()
        else: self.auto_schedule(self.scenario_execute_current_step,80)

    def auto_finish_accept_no_neutral(self):
        self.auto_neutral_running=False; self.auto_neutral_timer.stop(); self.label_auto_status.setText(f"Automatique : accepté par scénario '{self.current_runtime_scenario_name}'")
        self.label_auto_status.setStyleSheet("background-color: rgb(0,176,80); color: white; font-weight: bold; border: 2px solid black;")
        self.label_auto_resultat.setText("Résultat : RELAIS ACCEPTÉ"); self.label_auto_resultat.setStyleSheet("background-color: rgb(0,220,80); color: black; font-size: 14pt; font-weight: bold; border: 3px solid rgb(0,100,0);")
        if not self.production_record_result("ACCEPTÉ", "Accepté : pas neutre après essais scénario"):
            self.send_command("STOP"); self.update_button_states(); return
        self.set_auto_finish_validation_state(True)
        self.send_led_command("ACCEPT"); self.send_command("STOP"); self.update_button_states()
        QTimer.singleShot(0, lambda: self.show_auto_finish_validation_dialog(True))

    def auto_finish_accept_ok(self):
        self.auto_neutral_running=False; self.auto_neutral_timer.stop(); self.label_auto_status.setText(f"Automatique : scénario '{self.current_runtime_scenario_name}' terminé conforme")
        self.label_auto_status.setStyleSheet("background-color: rgb(0,176,80); color: white; font-weight: bold; border: 2px solid black;")
        self.label_auto_resultat.setText("Résultat : RELAIS ACCEPTÉ"); self.label_auto_resultat.setStyleSheet("background-color: rgb(0,220,80); color: black; font-size: 14pt; font-weight: bold; border: 3px solid rgb(0,100,0);")
        if not self.production_record_result("ACCEPTÉ", "Scénario terminé conforme"):
            self.send_command("STOP"); self.update_button_states(); return
        self.set_auto_finish_validation_state(True)
        self.send_led_command("ACCEPT"); self.send_command("STOP"); self.update_button_states()
        QTimer.singleShot(0, lambda: self.show_auto_finish_validation_dialog(True))

    def auto_finish_reject(self, raison):
        self.auto_neutral_running=False; self.auto_neutral_timer.stop()
        if 0 <= getattr(self,"runtime_step_index",-1) < self.tableWidget_auto_logigramme.rowCount(): self.set_auto_row_color(self.runtime_step_index,"fail",raison)
        self.label_auto_status.setText(f"Automatique : rejet - {raison}"); self.label_auto_status.setStyleSheet("background-color: rgb(220,0,0); color: white; font-weight: bold; border: 2px solid black;")
        self.label_auto_resultat.setText("Résultat : RELAIS REJETÉ"); self.label_auto_resultat.setStyleSheet("background-color: rgb(255,80,80); color: black; font-size: 14pt; font-weight: bold; border: 3px solid rgb(160,0,0);")
        raison_led = str(raison).lower()
        if any(mot in raison_led for mot in ("inconnu", "absent", "invalide", "non support")):
            self.send_led_command("ERROR")
        else:
            self.send_led_command("REJECT")
        if not self.production_record_result("REFUSÉ", raison):
            self.send_command("STOP"); self.update_button_states(); return
        self.set_auto_finish_validation_state(True)
        self.send_command("STOP"); self.update_button_states()
        QTimer.singleShot(0, lambda: self.show_auto_finish_validation_dialog(False))


    # ------------------------------------------------------------------
    # Étalonnage ADS1115 + pont diviseur V2.12.3
    # ------------------------------------------------------------------
    def voltage_calibration_init_db(self):
        self.chrono_init_db()
        with self.chrono_connect_db() as con:
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS calibrations_tension_ads1115 (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    calibration_date TEXT NOT NULL DEFAULT '',
                    operator TEXT NOT NULL DEFAULT '',
                    meter_reference TEXT NOT NULL DEFAULT '',
                    valid_days INTEGER NOT NULL DEFAULT 365,
                    low_actual_v REAL NOT NULL DEFAULT 0,
                    low_raw INTEGER NOT NULL DEFAULT 0,
                    high_actual_v REAL NOT NULL DEFAULT 0,
                    high_raw INTEGER NOT NULL DEFAULT 0,
                    check_actual_v REAL NOT NULL DEFAULT 0,
                    check_raw INTEGER NOT NULL DEFAULT 0,
                    divider_ratio REAL NOT NULL DEFAULT 0,
                    offset_mv INTEGER NOT NULL DEFAULT 0,
                    check_calculated_v REAL NOT NULL DEFAULT 0,
                    check_error_v REAL NOT NULL DEFAULT 0,
                    tolerance_v REAL NOT NULL DEFAULT 0.05,
                    is_valid INTEGER NOT NULL DEFAULT 0,
                    is_active INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL DEFAULT ''
                )
                """
            )
            con.execute("CREATE INDEX IF NOT EXISTS idx_calibration_active ON calibrations_tension_ads1115(is_active, is_valid)")

    def voltage_calibration_request_ads(self):
        if not self.is_connected():
            QMessageBox.warning(self.window, "Étalonnage tension", "RP2040 non connecté.")
            return
        self.send_command("ADS?")
        self.label_calibration_status.setText("Lecture ADS1115 demandée...")

    def voltage_calibration_request_capture(self, target):
        if not self.is_connected():
            QMessageBox.warning(self.window, "Étalonnage tension", "RP2040 non connecté.")
            return
        self.calibration_pending_capture = str(target).upper()
        self.send_command("ADS?")
        self.label_calibration_status.setText(f"Capture fraîche du point {self.calibration_pending_capture} demandée...")

    def voltage_calibration_accept_ads_frame(self, raw):
        try:
            raw = int(raw)
        except Exception:
            return
        self.voltage_last_adc_raw = raw
        self.voltage_calibration_update_live()
        target = str(getattr(self, "calibration_pending_capture", "") or "").upper()
        if target not in ("LOW", "HIGH", "CHECK"):
            return
        self.calibration_pending_capture = ""
        self.calibration_capture_raw[target] = raw
        mapping = {
            "LOW": self.lineEdit_calibration_low_raw,
            "HIGH": self.lineEdit_calibration_high_raw,
            "CHECK": self.lineEdit_calibration_check_raw,
        }
        mapping[target].setText(str(raw))
        self.calibration_calculated = None
        self.label_calibration_status.setText(f"RAW {raw} capturé pour le point {target}.")
        self.label_calibration_status.setStyleSheet("background-color: rgb(70,110,150); color: white; border: 2px solid black; font-weight: bold;")

    def voltage_calibration_update_live(self):
        raw = "--" if self.voltage_last_adc_raw is None else str(self.voltage_last_adc_raw)
        mv = "--" if self.voltage_last_adc_mv is None else f"{self.voltage_last_adc_mv / 1000.0:.3f} V"
        if hasattr(self, "label_calibration_live"):
            self.label_calibration_live.setText(f"RAW\n{raw}\n{mv}")

    def voltage_calibration_calculate(self):
        raw_low = self.calibration_capture_raw.get("LOW")
        raw_high = self.calibration_capture_raw.get("HIGH")
        raw_check = self.calibration_capture_raw.get("CHECK")
        if raw_low is None or raw_high is None:
            QMessageBox.warning(self.window, "Étalonnage tension", "Capturer obligatoirement le point bas et le point haut.")
            return None
        # Le firmware borne les valeurs négatives à zéro avant conversion.
        raw_low_calc = max(0, int(raw_low))
        raw_high_calc = max(0, int(raw_high))
        if raw_high_calc <= raw_low_calc:
            QMessageBox.warning(self.window, "Étalonnage tension", "Le RAW du point haut doit être supérieur au RAW du point bas.")
            return None
        low_v = float(self.doubleSpinBox_calibration_low_actual_v.value())
        high_v = float(self.doubleSpinBox_calibration_high_actual_v.value())
        if high_v <= low_v + 1.0:
            QMessageBox.warning(self.window, "Étalonnage tension", "Les deux valeurs multimètre sont trop proches ou inversées.")
            return None
        slope_mv_per_count = ((high_v - low_v) * 1000.0) / float(raw_high_calc - raw_low_calc)
        ratio = slope_mv_per_count / 0.125
        offset_mv_float = low_v * 1000.0 - raw_low_calc * 0.125 * ratio
        offset_mv = int(round(offset_mv_float))
        if not (1.0 <= ratio <= 100.0):
            QMessageBox.warning(self.window, "Étalonnage tension", f"Rapport calculé hors plage firmware : {ratio:.6f}.")
            return None
        if not (-500 <= offset_mv <= 500):
            QMessageBox.warning(self.window, "Étalonnage tension", f"Offset calculé hors plage firmware : {offset_mv} mV.")
            return None
        check_actual_v = float(self.doubleSpinBox_calibration_check_actual_v.value())
        check_calculated_v = None
        check_error_v = None
        if raw_check is not None:
            raw_check_calc = max(0, int(raw_check))
            check_calculated_v = (raw_check_calc * 0.125 * ratio + offset_mv) / 1000.0
            check_error_v = check_calculated_v - check_actual_v
        tolerance_v = float(self.doubleSpinBox_calibration_tolerance_v.value())
        valid = check_error_v is not None and abs(check_error_v) <= tolerance_v
        self.calibration_calculated = {
            "ratio": float(ratio), "offset_mv": int(offset_mv),
            "low_actual_v": low_v, "low_raw": int(raw_low),
            "high_actual_v": high_v, "high_raw": int(raw_high),
            "check_actual_v": check_actual_v, "check_raw": None if raw_check is None else int(raw_check),
            "check_calculated_v": check_calculated_v, "check_error_v": check_error_v,
            "tolerance_v": tolerance_v, "is_valid": bool(valid),
        }
        calc_txt = "--" if check_calculated_v is None else f"{check_calculated_v:.4f}"
        err_txt = "--" if check_error_v is None else f"{check_error_v:+.4f}"
        self.label_calibration_coefficients.setText(
            f"Rapport calculé : {ratio:.6f} | Offset : {offset_mv:+d} mV | Contrôle calculé : {calc_txt} V | Erreur : {err_txt} V"
        )
        if raw_check is None:
            msg = "Calibration 2 points calculée, mais le contrôle intermédiaire n'a pas été capturé : activation interdite."
            style = "background-color: rgb(190,110,0); color: white; border: 2px solid black; font-weight: bold;"
        elif valid:
            msg = f"CONFORME : erreur de contrôle {check_error_v:+.4f} V, tolérance ±{tolerance_v:.3f} V."
            style = "background-color: rgb(0,145,70); color: white; border: 2px solid rgb(0,70,30); font-weight: bold;"
        else:
            msg = f"NON CONFORME : erreur de contrôle {check_error_v:+.4f} V, tolérance ±{tolerance_v:.3f} V."
            style = "background-color: rgb(190,0,0); color: white; border: 2px solid black; font-weight: bold;"
        self.label_calibration_status.setText(msg)
        self.label_calibration_status.setStyleSheet(style)
        return self.calibration_calculated

    def voltage_calibration_save_activate(self):
        data = self.voltage_calibration_calculate()
        if not data or not data.get("is_valid"):
            QMessageBox.warning(self.window, "Étalonnage tension", "Activation refusée : le contrôle intermédiaire doit être conforme.")
            return
        operator = self.lineEdit_calibration_operator.text().strip()
        meter = self.lineEdit_calibration_meter.text().strip()
        if not operator or not meter:
            QMessageBox.warning(self.window, "Étalonnage tension", "Renseigner l'opérateur et la référence du multimètre.")
            return
        date_txt = self.dateEdit_calibration_date.date().toString("yyyy-MM-dd")
        valid_days = int(self.spinBox_calibration_valid_days.value())
        self.voltage_calibration_init_db()
        with self.chrono_connect_db() as con:
            con.execute("UPDATE calibrations_tension_ads1115 SET is_active = 0")
            cur = con.execute(
                """
                INSERT INTO calibrations_tension_ads1115 (
                    calibration_date, operator, meter_reference, valid_days,
                    low_actual_v, low_raw, high_actual_v, high_raw,
                    check_actual_v, check_raw, divider_ratio, offset_mv,
                    check_calculated_v, check_error_v, tolerance_v,
                    is_valid, is_active, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 1, ?)
                """,
                (
                    date_txt, operator, meter, valid_days,
                    data["low_actual_v"], data["low_raw"], data["high_actual_v"], data["high_raw"],
                    data["check_actual_v"], data["check_raw"], data["ratio"], data["offset_mv"],
                    data["check_calculated_v"], data["check_error_v"], data["tolerance_v"],
                    time.strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            calibration_id = int(cur.lastrowid)
        self.voltage_calibration_load_active()
        self.voltage_calibration_refresh_history()
        self.label_calibration_status.setText(f"Étalonnage #{calibration_id} enregistré et activé.")
        self.label_calibration_status.setStyleSheet("background-color: rgb(0,145,70); color: white; border: 2px solid rgb(0,70,30); font-weight: bold;")

    def voltage_calibration_clear_captures(self):
        self.calibration_capture_raw = {"LOW": None, "HIGH": None, "CHECK": None}
        self.calibration_pending_capture = ""
        self.calibration_calculated = None
        for edit in (self.lineEdit_calibration_low_raw, self.lineEdit_calibration_high_raw, self.lineEdit_calibration_check_raw):
            edit.setText("--")
        self.label_calibration_coefficients.setText("Rapport calculé : -- | Offset : -- mV | Contrôle calculé : -- V | Erreur : -- V")
        self.label_calibration_status.setText("Captures effacées. L'étalonnage actif enregistré n'est pas modifié.")
        self.label_calibration_status.setStyleSheet("background-color: rgb(100,100,100); color: white; border: 2px solid black; font-weight: bold;")

    def voltage_calibration_invalidate_active(self):
        if not self.active_voltage_calibration:
            QMessageBox.information(self.window, "Étalonnage tension", "Aucun étalonnage actif.")
            return
        if QMessageBox.question(self.window, "Étalonnage tension", "Invalider l'étalonnage actif ? Les mesures officielles de tension seront bloquées.") != QMessageBox.Yes:
            return
        with self.chrono_connect_db() as con:
            con.execute("UPDATE calibrations_tension_ads1115 SET is_active = 0, is_valid = 0 WHERE id = ?", (int(self.active_voltage_calibration["id"]),))
        self.active_voltage_calibration = None
        self.voltage_calibration_apply_to_operation_tab()
        self.voltage_calibration_refresh_history()
        self.label_calibration_status.setText("Étalonnage actif invalidé.")
        self.label_calibration_status.setStyleSheet("background-color: rgb(190,0,0); color: white; border: 2px solid black; font-weight: bold;")

    def voltage_calibration_row_is_current(self, row):
        if row is None or not int(row["is_valid"] or 0) or not int(row["is_active"] or 0):
            return False
        qdate = QDate.fromString(str(row["calibration_date"] or ""), "yyyy-MM-dd")
        if not qdate.isValid():
            return False
        age_days = qdate.daysTo(QDate.currentDate())
        return 0 <= age_days <= int(row["valid_days"] or 0)

    def voltage_calibration_load_active(self):
        self.voltage_calibration_init_db()
        with self.chrono_connect_db() as con:
            row = con.execute(
                "SELECT * FROM calibrations_tension_ads1115 WHERE is_active = 1 ORDER BY id DESC LIMIT 1"
            ).fetchone()
        self.active_voltage_calibration = dict(row) if row is not None and self.voltage_calibration_row_is_current(row) else None
        self.voltage_calibration_apply_to_operation_tab()

    def voltage_calibration_apply_to_operation_tab(self):
        cal = self.active_voltage_calibration
        if cal:
            self.doubleSpinBox_voltage_divider_ratio.setValue(float(cal["divider_ratio"]))
            self.spinBox_voltage_offset_mv.setValue(int(cal["offset_mv"]))
            date_fr = QDate.fromString(str(cal["calibration_date"]), "yyyy-MM-dd").toString("dd/MM/yyyy")
            text = f"Étalonnage #{cal['id']} valide - {date_fr} - erreur {float(cal['check_error_v']):+.4f} V"
            style = "background-color: rgb(205,245,215); border: 1px solid rgb(0,120,55); font-weight: bold;"
        else:
            self.doubleSpinBox_voltage_divider_ratio.setValue(12.818182)
            self.spinBox_voltage_offset_mv.setValue(0)
            text = "Étalonnage : absent, invalide ou expiré"
            style = "background-color: rgb(255,205,205); border: 1px solid rgb(170,0,0); font-weight: bold;"
        self.label_voltage_calibration_summary.setText(text)
        self.label_voltage_calibration_summary.setStyleSheet(style)
        self.voltage_update_button_states()

    def voltage_calibration_refresh_history(self):
        self.voltage_calibration_init_db()
        with self.chrono_connect_db() as con:
            rows = con.execute("SELECT * FROM calibrations_tension_ads1115 ORDER BY id DESC LIMIT 100").fetchall()
        table = self.tableWidget_calibration_history
        table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            current = self.voltage_calibration_row_is_current(row)
            if current:
                status = "ACTIVE VALIDE"
            elif int(row["is_active"] or 0) and int(row["is_valid"] or 0):
                status = "EXPIRÉE"
            elif int(row["is_valid"] or 0):
                status = "VALIDE NON ACTIVE"
            else:
                status = "INVALIDÉE"
            values = [
                row["id"], QDate.fromString(str(row["calibration_date"]), "yyyy-MM-dd").toString("dd/MM/yyyy"),
                row["operator"], row["meter_reference"], f"{float(row['divider_ratio']):.6f}",
                int(row["offset_mv"]), f"{float(row['check_error_v']):+.4f}", status,
            ]
            for c, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if status == "ACTIVE VALIDE": item.setBackground(QColor(205,245,215))
                elif status in ("EXPIRÉE", "INVALIDÉE"): item.setBackground(QColor(255,215,215))
                table.setItem(r, c, item)
        table.resizeColumnsToContents()

    def voltage_active_calibration_required(self):
        cal = self.active_voltage_calibration
        if not cal:
            return False, "Aucun étalonnage ADS1115 valide et non expiré n'est actif."
        return True, ""

    def voltage_relay_type(self):
        return "BISTABLE" if "bistable" in self.comboBox_voltage_relay_type.currentText().strip().lower() else "MONOSTABLE"

    def voltage_measure_settings_snapshot(self):
        """Retourne les réglages opérateur réellement validés dans les spinbox."""
        for spin in (
            self.doubleSpinBox_voltage_vmax,
            self.doubleSpinBox_voltage_ramp_up_s,
            self.doubleSpinBox_voltage_ramp_down_s,
            self.doubleSpinBox_voltage_current_limit,
            self.doubleSpinBox_voltage_chrono_v,
            self.doubleSpinBox_voltage_interphase_s,
        ):
            spin.interpretText()
        return {
            "vmax_v": float(self.doubleSpinBox_voltage_vmax.value()),
            "ramp_up_s": float(self.doubleSpinBox_voltage_ramp_up_s.value()),
            "ramp_down_s": float(self.doubleSpinBox_voltage_ramp_down_s.value()),
            "current_limit_a": float(self.doubleSpinBox_voltage_current_limit.value()),
            "chrono_supply_v": float(self.doubleSpinBox_voltage_chrono_v.value()),
            "interphase_s": float(self.doubleSpinBox_voltage_interphase_s.value()),
        }

    def voltage_load_measure_settings(self):
        """Recharge les derniers réglages de mesure depuis le dossier de l'EXE."""
        path = Path(self.voltage_measure_settings_file)
        if not path.exists():
            return
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return
            spins = (
                self.doubleSpinBox_voltage_vmax,
                self.doubleSpinBox_voltage_ramp_up_s,
                self.doubleSpinBox_voltage_ramp_down_s,
                self.doubleSpinBox_voltage_current_limit,
                self.doubleSpinBox_voltage_chrono_v,
                self.doubleSpinBox_voltage_interphase_s,
            )
            self._voltage_loading_measure_settings = True
            for spin in spins:
                spin.blockSignals(True)
            try:
                if "vmax_v" in data:
                    self.doubleSpinBox_voltage_vmax.setValue(float(data["vmax_v"]))
                self.voltage_update_ramp_limits()
                mapping = (
                    ("ramp_up_s", self.doubleSpinBox_voltage_ramp_up_s),
                    ("ramp_down_s", self.doubleSpinBox_voltage_ramp_down_s),
                    ("current_limit_a", self.doubleSpinBox_voltage_current_limit),
                    ("chrono_supply_v", self.doubleSpinBox_voltage_chrono_v),
                    ("interphase_s", self.doubleSpinBox_voltage_interphase_s),
                )
                for key, spin in mapping:
                    if key in data:
                        spin.setValue(float(data[key]))
            finally:
                for spin in spins:
                    spin.blockSignals(False)
                self._voltage_loading_measure_settings = False
        except Exception:
            self._voltage_loading_measure_settings = False

    def voltage_save_measure_settings(self, *_args):
        """Sauvegarde les réglages opérateur sans interrompre l'essai en cas d'erreur disque."""
        if getattr(self, "_voltage_loading_measure_settings", False):
            return
        try:
            data = self.voltage_measure_settings_snapshot()
            path = Path(self.voltage_measure_settings_file)
            path.parent.mkdir(parents=True, exist_ok=True)
            temporary = path.with_name(path.name + ".tmp")
            temporary.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            os.replace(temporary, path)
        except Exception:
            pass

    def voltage_commit_measure_settings(self):
        """Valide explicitement une saisie clavier et la rend persistante."""
        for spin in (
            self.doubleSpinBox_voltage_vmax,
            self.doubleSpinBox_voltage_ramp_up_s,
            self.doubleSpinBox_voltage_ramp_down_s,
            self.doubleSpinBox_voltage_current_limit,
            self.doubleSpinBox_voltage_chrono_v,
            self.doubleSpinBox_voltage_interphase_s,
        ):
            spin.interpretText()
        self.voltage_save_measure_settings()

    def voltage_capture_run_settings(self):
        """Fige les réglages au démarrage afin qu'ils restent identiques durant tout le cycle."""
        self.voltage_run_settings = self.voltage_measure_settings_snapshot()
        self.voltage_save_measure_settings()
        return dict(self.voltage_run_settings)

    def voltage_run_setting(self, key, fallback_widget):
        try:
            return float(self.voltage_run_settings.get(key, fallback_widget.value()))
        except Exception:
            return float(fallback_widget.value())

    def voltage_update_relay_type_ui(self, *_args):
        bistable = self.voltage_relay_type() == "BISTABLE"
        if bistable:
            self.pushButton_voltage_pickup.setText("MESURER BASCULEMENT BE")
            self.pushButton_voltage_dropout.setText("MESURER BASCULEMENT BR")
            self.pushButton_voltage_cycle.setText("CYCLE BE + BR")
            self.label_voltage_ramp_up.setText("Rampe BE 0→max")
            self.label_voltage_ramp_down.setText("Rampe BR 0→max")
            self.label_voltage_interphase.setText("Attente BE→BR")
            headers = ["Inverseur", "Tension BE (V)", "Tension BR (V)", "État"]
        else:
            self.pushButton_voltage_pickup.setText("MESURER COLLAGE")
            self.pushButton_voltage_dropout.setText("MESURER DÉCOLLAGE")
            self.pushButton_voltage_cycle.setText("CYCLE COLLAGE + DÉCOLLAGE")
            self.label_voltage_ramp_up.setText("Rampe BE 0→max")
            self.label_voltage_ramp_down.setText("Rampe BE max→0")
            self.label_voltage_interphase.setText("Maintien entre mesures")
            headers = ["Inverseur", "Tension collage (V)", "Tension décollage (V)", "État"]
        self.label_voltage_interphase_note.setText("Délai réel contrôlé")
        self.tableWidget_voltage_results.setHorizontalHeaderLabels(headers)
        self.voltage_update_ramp_limits()
        self.voltage_refresh_results_table()

    def voltage_update_ramp_limits(self, *_args):
        vmax = max(0.0, float(self.doubleSpinBox_voltage_vmax.value()))
        raw_max_duration_s = vmax / EAPSU.MIN_VOLTAGE_SLOPE_V_PER_S
        decimals = int(self.doubleSpinBox_voltage_ramp_up_s.decimals())
        scale = 10 ** max(0, decimals)
        rounded_down_s = math.floor(raw_max_duration_s * scale + 1e-12) / scale
        max_duration_s = max(float(self.doubleSpinBox_voltage_ramp_up_s.minimum()), rounded_down_s)
        for spin in (self.doubleSpinBox_voltage_ramp_up_s, self.doubleSpinBox_voltage_ramp_down_s):
            spin.setMaximum(max_duration_s)
            spin.setToolTip(
                f"Pente minimale EA : {EAPSU.MIN_VOLTAGE_SLOPE_V_PER_S:.3f} V/s. "
                f"Pour une rampe de {vmax:.3f} V, durée maximale autorisée : {max_duration_s:.3f} s."
            )
        self.label_voltage_accuracy.setText(
            f"Premier passage avant rebonds | pente EA mini {EAPSU.MIN_VOLTAGE_SLOPE_V_PER_S:.3f} V/s "
            f"| durée maxi actuelle {max_duration_s:.3f} s"
        )
        self.label_voltage_accuracy.setToolTip(
            "La durée maximale est recalculée automatiquement à partir de Vmax. "
            "Le contrôle de plausibilité ne remplace jamais la tension ADS1115 capturée."
        )
        return max_duration_s

    def voltage_refresh_ea_ports(self):
        current = self.comboBox_voltage_ea_port.currentText().strip() if hasattr(self, "comboBox_voltage_ea_port") else ""
        ports = [p.device for p in serial.tools.list_ports.comports()]
        rp_port = ""
        try:
            rp_port = str(self.ser.port) if self.ser is not None and self.ser.is_open else ""
        except Exception:
            pass
        self.comboBox_voltage_ea_port.blockSignals(True)
        try:
            self.comboBox_voltage_ea_port.clear()
            for port in ports:
                self.comboBox_voltage_ea_port.addItem(port)
            if current and current in ports:
                self.comboBox_voltage_ea_port.setCurrentText(current)
            elif rp_port and len(ports) > 1:
                for port in ports:
                    if port != rp_port:
                        self.comboBox_voltage_ea_port.setCurrentText(port)
                        break
            if not ports:
                self.comboBox_voltage_ea_port.addItem("Aucun port")
        finally:
            self.comboBox_voltage_ea_port.blockSignals(False)
        self.voltage_update_button_states()

    def voltage_connect_ea(self):
        if self.ea_psu.connected:
            return
        port = self.comboBox_voltage_ea_port.currentText().strip()
        if not port or port == "Aucun port":
            QMessageBox.warning(self.window, "Alimentation EA", "Aucun port série sélectionné.")
            return
        try:
            rp_port = str(self.ser.port) if self.ser is not None and self.ser.is_open else ""
        except Exception:
            rp_port = ""
        if rp_port and port == rp_port:
            QMessageBox.warning(self.window, "Alimentation EA", "Le port EA doit être différent du port RP2040.")
            return
        try:
            baud = int(self.comboBox_voltage_ea_baudrate.currentText())
        except Exception:
            baud = 9600
        try:
            self.ea_psu.connect(port, baud, timeout_s=1.0)
            identity = self.ea_psu.query("*IDN?")
            self.ea_psu.identity = str(identity or "").strip()
            self.ea_psu.set_remote()
            self.ea_psu.output(False)
            self.label_voltage_ea_status.setText(f"EA connectée {port} @ {baud} - {identity or 'identification non reçue'}")
            self.label_voltage_ea_status.setStyleSheet("background-color: rgb(0,150,70); color: white; font-weight: bold; border: 1px solid rgb(0,80,35);")
        except Exception as exc:
            self.ea_psu.disconnect()
            self.label_voltage_ea_status.setText(f"Erreur connexion EA : {exc}")
            self.label_voltage_ea_status.setStyleSheet("background-color: rgb(190,0,0); color: white; font-weight: bold; border: 1px solid black;")
            QMessageBox.warning(self.window, "Alimentation EA", f"Connexion impossible :\n{exc}")
        self.voltage_update_button_states()

    def voltage_disconnect_ea(self):
        if self.voltage_test_running:
            self.voltage_abort_test("Déconnexion EA demandée")
        stop_confirmed = True
        local_confirmed = True
        if self.ea_psu.connected:
            stop_confirmed = self.voltage_stop_ea_and_confirm("déconnexion EA", show_alert=True)
            if stop_confirmed:
                try:
                    self.ea_psu.set_local()
                except Exception:
                    local_confirmed = False
        self.ea_psu.disconnect()
        if stop_confirmed and local_confirmed:
            self.label_voltage_ea_status.setText("EA non connectée - arrêt confirmé")
            self.label_voltage_ea_status.setStyleSheet("background-color: rgb(95,95,95); color: white; font-weight: bold; border: 1px solid black;")
        elif stop_confirmed:
            self.label_voltage_ea_status.setText("EA non connectée - arrêt confirmé, retour local non confirmé")
            self.label_voltage_ea_status.setStyleSheet("background-color: rgb(210,145,0); color: black; font-weight: bold; border: 2px solid rgb(130,80,0);")
        else:
            self.label_voltage_ea_status.setText("EA déconnectée - ARRÊT NON CONFIRMÉ : couper manuellement")
            self.label_voltage_ea_status.setStyleSheet("background-color: rgb(190,0,0); color: white; font-weight: bold; border: 2px solid black;")
        self.voltage_update_button_states()

    def voltage_copy_from_chrono(self):
        self.lineEdit_voltage_lot.setText(self.lineEdit_chrono_lot.text())
        self.lineEdit_voltage_relais.setText(self.lineEdit_chrono_relais.text())
        self.lineEdit_voltage_sn.setText(self.lineEdit_chrono_sn.text())
        self.lineEdit_voltage_date.setText(self.lineEdit_chrono_date.text() or QDate.currentDate().toString("dd/MM/yyyy"))
        self.lineEdit_voltage_ambiance.setText(self.lineEdit_chrono_ambiance.text())
        self.lineEdit_voltage_test.setText((self.lineEdit_chrono_nom_test.text().strip() or "Chronométrie") + " - collage/décollage")
        self.spinBox_voltage_nb_inverseurs.setValue(self.spinBox_chrono_nb_inverseurs.value())
        self.comboBox_voltage_relay_type.setCurrentText(
            "Monostable" if self.chrono_relay_type() == "MONOSTABLE" else "Bistable 2 bobines"
        )
        self.label_voltage_status.setText("Informations et type de relais repris depuis l'onglet Chronométrie contacts.")

    def voltage_measure_all_context(self):
        """Fige l'identification, le type et la tension EA de chronométrie."""
        meta = self.voltage_metadata()
        settings = self.voltage_measure_settings_snapshot()
        return {
            **meta,
            "relay_type": self.voltage_relay_type(),
            "nb_inverseurs": int(self.spinBox_voltage_nb_inverseurs.value()),
            "chrono_supply_v": float(settings["chrono_supply_v"]),
        }

    def voltage_apply_measure_all_context_to_chrono(self, context=None):
        """Recopie le contexte tension vers l'onglet chronométrie sans auto-remplissage parasite."""
        context = dict(context or self.measure_all_context or self.voltage_measure_all_context())
        self._chrono_lot_autofill_running = True
        try:
            self.lineEdit_chrono_lot.setText(str(context.get("lot", "")))
            self.lineEdit_chrono_relais.setText(str(context.get("relais", "")))
            self.lineEdit_chrono_sn.setText(str(context.get("sn", "")))
            self.lineEdit_chrono_date.setText(str(context.get("date_test", "")) or QDate.currentDate().toString("dd/MM/yyyy"))
            self.lineEdit_chrono_ambiance.setText(str(context.get("ambiance_c", "")))
            self.lineEdit_chrono_nom_test.setText(str(context.get("nom_test", "")))
        finally:
            self._chrono_lot_autofill_running = False
        self.comboBox_chrono_type_relais.setCurrentText(
            "Monostable" if str(context.get("relay_type", "")).upper() == "MONOSTABLE" else "Bistable"
        )
        self.spinBox_chrono_nb_inverseurs.setValue(max(1, min(4, int(context.get("nb_inverseurs", 1)))))
        self.chrono_update_relay_type_ui()

    def voltage_validate_measure_all_chrono_settings(self):
        """Valide les paramètres chronométrie avant de commencer la phase tension."""
        pulse_ms = int(self.spinBox_chrono_pulse_ms.value())
        capture_ms = int(self.spinBox_chrono_capture_ms.value())
        if capture_ms < pulse_ms:
            raise ValueError(
                "Mesure totale impossible : dans l'onglet Chronométrie contacts, "
                "la fenêtre de capture doit être supérieure ou égale à la durée pulse / maintien."
            )
        self.chrono_float_ms(self.lineEdit_chrono_limite_temps_ms, "Sanction temps max")
        self.chrono_float_ms(self.lineEdit_chrono_limite_rebond_ms, "Sanction rebond max")
        self.chrono_validate_metadata()

    def voltage_start_measure_all(self):
        """Lance les tensions puis la chronométrie, alimentées uniquement par l'EA."""
        if getattr(self, "measure_all_active", False) or self.voltage_test_running or self.chrono_measure_running:
            QMessageBox.warning(self.window, "Mesurer tout", "Une mesure est déjà en cours.")
            return
        try:
            context = self.voltage_measure_all_context()
            self.voltage_apply_measure_all_context_to_chrono(context)
            self.voltage_validate_measure_all_chrono_settings()
            if not self.is_connected():
                raise RuntimeError("RP2040 non connecté.")
            if not self.ea_psu.connected:
                raise RuntimeError("Alimentation EA non connectée.")
            if not getattr(self, "rp2040_ea_chrono_capable", False):
                raise RuntimeError(
                    "Firmware RP2040 incompatible avec la chronométrie alimentée par l'EA. "
                    "Téléverser le firmware V2.12.3 R8 contenant EA_CHRONO_NO_GP26."
                )
        except Exception as exc:
            QMessageBox.warning(self.window, "Mesurer tout", str(exc))
            return

        self.measure_all_active = True
        self.measure_all_phase = "VOLTAGE"
        self.measure_all_context = dict(context)
        self.measure_all_chrono_results = {}
        self.measure_all_static_confirmation = {}
        self.chrono_external_supply_mode = False
        chrono_v = float(context.get("chrono_supply_v", 0.0))
        self.label_voltage_status.setText(
            f"MESURE TOTALE 1/2 — cycle tensions EA en cours. Ensuite l'EA sera réglée "
            f"automatiquement à {chrono_v:.3f} V pour la chronométrie."
        )
        self.label_voltage_status.setStyleSheet(
            "background-color: rgb(0,90,160); color: white; font-weight: bold; border: 2px solid rgb(0,45,90);"
        )
        self.update_button_states()
        self.voltage_start_test("CYCLE")
        if not self.voltage_test_running and self.measure_all_active:
            self.voltage_measure_all_fail("Le cycle de tensions n'a pas pu démarrer.")

    @staticmethod
    def voltage_format_static_confirmation(info):
        info = dict(info or {})
        measured = info.get("measured_voltage_v")
        measured_text = "non relue" if measured is None else f"{float(measured):.3f} V"
        errors = "; ".join(
            str(item) for item in info.get("errors", []) if str(item).strip()
        ) or "confirmation absente"
        return (
            f"Cible={float(info.get('target_voltage_v') or 0.0):.3f} V, "
            f"mesurée={measured_text}, sortie={info.get('output_state') or '--'}, "
            f"générateur={info.get('generator_selection') or '--'}, "
            f"SCPI={info.get('scpi_error') or '--'}. Détail : {errors}"
        )

    def voltage_measure_all_start_chrono(self):
        """Passe automatiquement l'EA en tension continue puis lance la chronométrie."""
        if not self.measure_all_active:
            return
        self.measure_all_phase = "EA_STATIC_PREP"
        target_v = float(self.measure_all_context.get(
            "chrono_supply_v", self.doubleSpinBox_voltage_chrono_v.value()
        ))
        current_limit_a = self.voltage_run_setting(
            "current_limit_a", self.doubleSpinBox_voltage_current_limit
        )
        self.label_voltage_status.setText(
            f"MESURE TOTALE — préparation automatique de l'EA à {target_v:.3f} V "
            "pour la chronométrie contacts..."
        )
        self.label_voltage_status.setStyleSheet(
            "background-color: rgb(0,90,160); color: white; font-weight: bold; border: 2px solid rgb(0,45,90);"
        )
        try:
            info = self.ea_psu.configure_static_output_and_confirm(target_v, current_limit_a)
        except Exception as exc:
            info = {"confirmed": False, "target_voltage_v": target_v, "errors": [str(exc)]}
        self.measure_all_static_confirmation = dict(info)
        if not bool(info.get("confirmed")):
            diagnostic = self.voltage_format_static_confirmation(info)
            self.voltage_measure_all_fail(
                "La tension continue de chronométrie n'a pas été confirmée. " + diagnostic
            )
            return

        self.chrono_external_supply_mode = True
        self.measure_all_phase = "CHRONO"
        self.voltage_apply_measure_all_context_to_chrono(self.measure_all_context)
        self.measure_all_chrono_results = {}
        measured_v = float(info.get("measured_voltage_v") or target_v)
        self.label_voltage_status.setText(
            f"MESURE TOTALE 2/2 — EA confirmée à {measured_v:.3f} V. "
            "Chronométrie contacts automatique en cours, source fixe neutral screen isolée."
        )
        self.label_voltage_status.setStyleSheet(
            "background-color: rgb(0,90,160); color: white; font-weight: bold; border: 2px solid rgb(0,45,90);"
        )
        started = self.chrono_start_measure_be_br()
        if started is False or not self.chrono_measure_running:
            self.voltage_measure_all_fail(
                "Les tensions ont été sauvegardées, mais la chronométrie EA n'a pas pu démarrer."
            )

    def voltage_measure_all_finish(self):
        if not self.measure_all_active or self.measure_all_phase != "CHRONO":
            return
        original_sn = str(self.measure_all_context.get("sn", ""))
        next_sn = self.lineEdit_chrono_sn.text().strip()
        if next_sn:
            self.lineEdit_voltage_sn.setText(next_sn)
        chrono_values = list(self.measure_all_chrono_results.values())
        chrono_result = "DEFAUT" if any(value != "OK" for value in chrono_values) else "OK"
        voltage_result = str(getattr(self, "voltage_last_saved_result", "") or "OK")
        stop_confirmed = self.voltage_stop_ea_and_confirm("fin de MESURER TOUT", show_alert=False)
        self.chrono_external_supply_mode = False
        self.measure_all_active = False
        self.measure_all_phase = ""
        self.measure_all_context = {}
        self.measure_all_chrono_results = {}
        self.update_button_states()
        message = (
            f"MESURE TOTALE TERMINÉE — SN {original_sn} : tensions={voltage_result}, "
            f"chronométrie={chrono_result}, arrêt EA={'confirmé' if stop_confirmed else 'NON CONFIRMÉ'}."
        )
        if next_sn and next_sn != original_sn:
            message += f" SN suivant prêt : {next_sn}."
        if not stop_confirmed:
            message += " COUPER MANUELLEMENT L'ALIMENTATION EA."
        self.label_voltage_status.setText(message)
        self.label_voltage_status.setStyleSheet(
            "background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);"
            if chrono_result == "OK" and voltage_result == "OK" and stop_confirmed else
            "background-color: rgb(220,145,0); color: black; font-weight: bold; border: 2px solid rgb(130,80,0);"
        )
        if not stop_confirmed:
            self.big_message_box(
                "Sécurité alimentation EA",
                "ARRÊT EA NON CONFIRMÉ",
                self.voltage_last_stop_diagnostic or message,
                ok_text="J'AI COUPÉ MANUELLEMENT",
                icon=QMessageBox.Critical,
            )

    def voltage_measure_all_fail(self, reason, preserve_status=False):
        if not getattr(self, "measure_all_active", False):
            return
        stop_needed = bool(self.chrono_external_supply_mode or self.measure_all_phase in ("EA_STATIC_PREP", "CHRONO"))
        stop_confirmed = True
        if stop_needed and getattr(self, "ea_psu", None) is not None and self.ea_psu.connected:
            stop_confirmed = self.voltage_stop_ea_and_confirm("interruption de MESURER TOUT", show_alert=False)
        self.chrono_external_supply_mode = False
        self.measure_all_active = False
        self.measure_all_phase = ""
        self.measure_all_context = {}
        self.measure_all_chrono_results = {}
        self.chrono_auto_sequence_active = False
        self.chrono_auto_sequence_queue = []
        self.chrono_auto_prereset_pending = False
        self.chrono_measure_running = False
        self.update_button_states()
        if not preserve_status:
            suffix = "" if stop_confirmed else " — ARRÊT EA NON CONFIRMÉ, COUPER MANUELLEMENT"
            self.label_voltage_status.setText(f"MESURE TOTALE INTERROMPUE — {reason}{suffix}")
            self.label_voltage_status.setStyleSheet(
                "background-color: rgb(190,0,0); color: white; font-weight: bold; border: 2px solid black;"
            )
        if not stop_confirmed:
            self.big_message_box(
                "Sécurité alimentation EA",
                "ARRÊT EA NON CONFIRMÉ",
                self.voltage_last_stop_diagnostic or str(reason),
                ok_text="J'AI COUPÉ MANUELLEMENT",
                icon=QMessageBox.Critical,
            )

    def voltage_stop_clicked(self):
        """Le bouton d'arrêt couvre la phase tension et la phase chronométrie de MESURER TOUT."""
        if getattr(self, "measure_all_active", False):
            if self.voltage_test_running:
                self.voltage_abort_test("Mesure totale arrêtée par l'opérateur.")
                return
            try:
                if self.is_connected():
                    self.send_command("STOP")
            except Exception:
                pass
            self.voltage_measure_all_fail("arrêt demandé par l'opérateur")
            return
        self.voltage_abort_test()

    def voltage_reset_assessment_state(self):
        """Réinitialise les verdicts propres à un essai de tension."""
        self.voltage_plausibility = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_result_override = ""
        self.voltage_ea_stop_confirmation = self.voltage_empty_stop_confirmation()
        self.voltage_last_stop_diagnostic = ""
        self.voltage_last_saved_result = ""

    @staticmethod
    def voltage_empty_stop_confirmation():
        return {
            "confirmed": None,
            "generator_selection_before": "",
            "generator_selection": "",
            "generator_state": "",
            "output_state": "",
            "measured_voltage_v": None,
            "scpi_error": "",
            "errors": [],
            "context": "",
            "poll_count": 0,
            "confirmation_elapsed_s": 0.0,
        }

    def voltage_stop_ea_and_confirm(self, context, show_alert=False):
        info = self.voltage_empty_stop_confirmation()
        info["context"] = str(context or "arrêt")
        try:
            if not self.ea_psu.connected:
                raise RuntimeError("Alimentation EA non connectée.")
            checked = self.ea_psu.safe_stop_and_confirm()
            info.update(checked)
        except Exception as exc:
            info["confirmed"] = False
            info["errors"] = list(info.get("errors", [])) + [str(exc)]
        self.voltage_ea_stop_confirmation = info
        if info.get("confirmed"):
            self.voltage_last_stop_diagnostic = ""
            return True

        self.voltage_result_override = "ARRET_EA_NON_CONFIRME"
        voltage = info.get("measured_voltage_v")
        voltage_text = "non relue" if voltage is None else f"{float(voltage):.3f} V"
        errors = "; ".join(str(item) for item in info.get("errors", []) if str(item).strip()) or "confirmation absente"
        elapsed = float(info.get("confirmation_elapsed_s") or 0.0)
        polls = int(info.get("poll_count") or 0)
        alert = (
            "ARRÊT EA NON CONFIRMÉ — COUPER MANUELLEMENT L'ALIMENTATION. "
            f"Contexte : {info['context']}. Sortie={info.get('output_state') or '--'}, "
            f"générateur initial={info.get('generator_selection_before') or '--'}, "
            f"générateur final={info.get('generator_selection') or info.get('generator_state') or '--'}, "
            f"tension={voltage_text}, contrôles={polls}, durée={elapsed:.2f} s, "
            f"SCPI={info.get('scpi_error') or '--'}. Détail : {errors}"
        )
        self.voltage_last_stop_diagnostic = alert
        if hasattr(self, "label_voltage_ea_status"):
            self.label_voltage_ea_status.setText(alert)
            self.label_voltage_ea_status.setStyleSheet(
                "background-color: rgb(190,0,0); color: white; font-weight: bold; border: 2px solid black;"
            )
        if show_alert:
            self.big_message_box(
                "Sécurité alimentation EA",
                "ARRÊT EA NON CONFIRMÉ",
                alert,
                ok_text="J'AI COUPÉ MANUELLEMENT",
                icon=QMessageBox.Critical,
            )
        return False

    def voltage_evaluate_plausibility(self, mode):
        mode = str(mode).upper()
        info = self.voltage_ramp_info if self.voltage_ramp_info and self.voltage_ramp_info.get("mode") == mode else None
        measured = self.voltage_results.get(mode, {}).get("GLOBAL")
        elapsed_us = self.voltage_time_results.get(mode, {}).get("GLOBAL")
        result = {
            "status": "NON_VERIFIE",
            "mode": mode,
            "measured_v": measured,
            "elapsed_s": None if elapsed_us is None else float(elapsed_us) / 1_000_000.0,
            "expected_elapsed_s": None,
            "elapsed_error_s": None,
            "tolerance_s": None,
            "detail": "Données insuffisantes pour le contrôle de plausibilité.",
        }
        if not info or measured is None or elapsed_us is None:
            self.voltage_plausibility[mode] = result
            return result

        start_v = float(info["start_v"])
        end_v = float(info["end_v"])
        duration_s = max(0.0001, float(info["duration_s"]))
        span_v = end_v - start_v
        if abs(span_v) < 1e-12:
            result["detail"] = "Rampe sans écart de tension : plausibilité temporelle impossible."
            self.voltage_plausibility[mode] = result
            return result

        fraction = (float(measured) - start_v) / span_v
        expected_elapsed_s = duration_s * fraction
        elapsed_s = float(elapsed_us) / 1_000_000.0
        elapsed_error_s = elapsed_s - expected_elapsed_s
        tolerance_s = max(VOLTAGE_PLAUSIBILITY_MIN_TOL_S, duration_s * VOLTAGE_PLAUSIBILITY_REL_TOL)
        voltage_margin_v = max(0.100, abs(span_v) * 0.010)
        low_v = min(start_v, end_v) - voltage_margin_v
        high_v = max(start_v, end_v) + voltage_margin_v
        in_voltage_range = low_v <= float(measured) <= high_v
        time_ok = -tolerance_s <= elapsed_error_s <= tolerance_s

        result.update({
            "expected_elapsed_s": expected_elapsed_s,
            "elapsed_error_s": elapsed_error_s,
            "tolerance_s": tolerance_s,
            "fraction": fraction,
            "start_v": start_v,
            "end_v": end_v,
            "duration_s": duration_s,
        })
        if in_voltage_range and time_ok:
            result["status"] = "OK"
            result["detail"] = (
                f"Plausibilité OK : t mesuré {elapsed_s:.3f} s, t théorique {expected_elapsed_s:.3f} s, "
                f"écart {elapsed_error_s:+.3f} s, tolérance ±{tolerance_s:.3f} s."
            )
        else:
            result["status"] = "INCOHERENT"
            causes = []
            if not in_voltage_range:
                causes.append(f"tension {float(measured):.3f} V hors rampe {start_v:.3f}→{end_v:.3f} V")
            if not time_ok:
                causes.append(
                    f"écart temporel {elapsed_error_s:+.3f} s supérieur à ±{tolerance_s:.3f} s"
                )
            result["detail"] = "Plausibilité incohérente : " + "; ".join(causes) + "."
        self.voltage_plausibility[mode] = result
        return result

    def voltage_metadata(self):
        data = {
            "lot": self.lineEdit_voltage_lot.text().strip(),
            "relais": self.lineEdit_voltage_relais.text().strip(),
            "sn": self.lineEdit_voltage_sn.text().strip(),
            "date_test": self.lineEdit_voltage_date.text().strip() or QDate.currentDate().toString("dd/MM/yyyy"),
            "ambiance_c": self.lineEdit_voltage_ambiance.text().strip(),
            "nom_test": self.lineEdit_voltage_test.text().strip(),
        }
        missing = [label for key, label in (("lot", "Lot"), ("relais", "Relais"), ("sn", "SN"), ("nom_test", "Nom du test")) if not data[key]]
        if missing:
            raise ValueError("Champs obligatoires manquants : " + ", ".join(missing))
        return data

    def voltage_on_nb_inverseurs_changed(self):
        """Rafraîchit immédiatement la table et les LED après changement de N."""
        self.voltage_refresh_results_table()
        self.voltage_refresh_contact_leds()

    def voltage_refresh_contact_leds(self):
        """Affiche les 8 états de contact dans l'onglet tension.

        Les inverseurs au-delà du nombre sélectionné restent visibles mais sont
        volontairement grisés afin de ne pas les confondre avec un état inconnu.
        """
        if not hasattr(self, "label_voltage_led_r1"):
            return
        nb = int(self.spinBox_voltage_nb_inverseurs.value()) if hasattr(self, "spinBox_voltage_nb_inverseurs") else 0
        values = list(getattr(self, "contacts_known_values", [None] * 8))
        while len(values) < 8:
            values.append(None)
        labels_r = [self.label_voltage_led_r1, self.label_voltage_led_r2, self.label_voltage_led_r3, self.label_voltage_led_r4]
        labels_t = [self.label_voltage_led_t1, self.label_voltage_led_t2, self.label_voltage_led_t3, self.label_voltage_led_t4]
        for index in range(4):
            inv = index + 1
            if inv > nb:
                self.voltage_set_contact_indicator(labels_r[index], f"R{inv}", None, "green", selected=False)
                self.voltage_set_contact_indicator(labels_t[index], f"T{inv}", None, "red", selected=False)
                continue
            self.voltage_set_contact_indicator(labels_r[index], f"R{inv}", values[index], "green")
            self.voltage_set_contact_indicator(labels_t[index], f"T{inv}", values[4 + index], "red")

    @staticmethod
    def voltage_set_contact_indicator(label, title, value, color, selected=True):
        """LED circulaire réelle pour l'état direct des contacts de l'onglet tensions."""
        # Ne pas utiliser le caractère Unicode « ● » : sa taille visible dépend
        # de la police et reste beaucoup plus petite que le QLabel. Le fond du
        # QLabel constitue directement le rond, avec un diamètre réel de 18 px.
        diameter = 18
        label.setFixedSize(diameter, diameter)
        label.setAlignment(Qt.AlignCenter)
        label.setText("")
        if not selected:
            rgb = "165,165,165"
            border_rgb = "125,125,125"
            tooltip = f"{title} : inverseur non sélectionné"
        elif value is None:
            rgb = "95,95,95"
            border_rgb = "55,55,55"
            tooltip = f"{title} : état inconnu"
        elif str(value) == "1":
            rgb = "0,205,75" if color == "green" else "225,0,0"
            border_rgb = "0,125,45" if color == "green" else "145,0,0"
            tooltip = f"{title} : contact fermé détecté"
        else:
            rgb = "25,85,45" if color == "green" else "90,25,25"
            border_rgb = "15,55,30" if color == "green" else "55,15,15"
            tooltip = f"{title} : contact ouvert / non détecté"
        label.setToolTip(tooltip)
        label.setStyleSheet(
            f"background-color: rgb({rgb}); "
            f"border: 1px solid rgb({border_rgb}); "
            "border-radius: 9px; padding: 0px; margin: 0px;"
        )

    @staticmethod
    def voltage_contact_state_text(value):
        if value is None:
            return "inconnu"
        return "fermé" if str(value) == "1" else "ouvert"

    def voltage_timeout_contact_diagnostic(self):
        """Construit le message opérateur lorsqu'aucun passage global n'est validé."""
        mode = str(getattr(self, "voltage_active_scan", "") or "").upper()
        if mode not in ("PICKUP", "DROPOUT"):
            mode = "PICKUP" if str(getattr(self, "voltage_requested_mode", "")).upper() == "PICKUP" else "DROPOUT"
        relay_type = self.voltage_relay_type()
        nb = int(self.spinBox_voltage_nb_inverseurs.value())
        vmax = self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax)
        stable_ms = int(self.spinBox_voltage_stable_ms.value())
        values = list(getattr(self, "contacts_known_values", [None] * 8))
        while len(values) < 8:
            values.append(None)

        target_work = mode == "PICKUP"
        expected_r = "0" if target_work else "1"
        expected_t = "1" if target_work else "0"
        failures = []
        states = []
        for index in range(nb):
            inv = index + 1
            r_value = None if values[index] is None else str(values[index])
            t_value = None if values[4 + index] is None else str(values[4 + index])
            states.append(
                f"I{inv}: R{inv}={self.voltage_contact_state_text(r_value)}, "
                f"T{inv}={self.voltage_contact_state_text(t_value)}"
            )
            issues = []
            if r_value is None:
                issues.append(f"R{inv} inconnu")
            elif r_value != expected_r:
                issues.append(f"R{inv} encore fermé" if expected_r == "0" else f"R{inv} non fermé")
            if t_value is None:
                issues.append(f"T{inv} inconnu")
            elif t_value != expected_t:
                issues.append(f"T{inv} non fermé" if expected_t == "1" else f"T{inv} encore fermé")
            if issues:
                failures.append(f"Inverseur {inv} : " + ", ".join(issues))

        if mode == "PICKUP":
            title = f"ÉCHEC COLLAGE / BE À VMAX ({vmax:.3f} V)"
            target_text = "position travail attendue : R ouverts et T fermés"
        elif relay_type == "BISTABLE":
            title = f"ÉCHEC DÉCLENCHEMENT / BR À VMAX ({vmax:.3f} V)"
            target_text = "position repos attendue : R fermés et T ouverts"
        else:
            title = "ÉCHEC DÉCOLLAGE À 0 V"
            target_text = "position repos attendue : R fermés et T ouverts"

        if failures:
            detail = "Contacts bloquants :\n- " + "\n- ".join(failures)
        else:
            detail = (
                f"Les contacts sont actuellement dans la position demandée, mais la position globale "
                f"n'a pas été stable pendant {stable_ms} ms avant la fin du délai. "
                "Cause possible : rebonds persistants ou basculement trop tardif."
            )
        state_text = "États finaux : " + " | ".join(states) if states else "États finaux : aucun contact sélectionné."
        return title, f"{target_text}.\n\n{detail}\n\n{state_text}"

    def voltage_contacts_expected(self, mode):
        nb = int(self.spinBox_voltage_nb_inverseurs.value())
        values = list(getattr(self, "contacts_known_values", []))
        if len(values) < 8 or any(values[i] is None for i in list(range(nb)) + list(range(4, 4 + nb))):
            return False, "Contacts inconnus"
        rest = all(int(values[i]) == 1 for i in range(nb)) and all(int(values[4 + i]) == 0 for i in range(nb))
        work = all(int(values[i]) == 0 for i in range(nb)) and all(int(values[4 + i]) == 1 for i in range(nb))
        expected = rest if mode == "PICKUP" else work
        return expected, "repos R fermés/T ouverts" if mode == "PICKUP" else "travail R ouverts/T fermés"

    def voltage_send_config(self):
        ok, reason = self.voltage_active_calibration_required()
        if not ok:
            raise RuntimeError(reason)
        ratio_u6 = int(round(float(self.active_voltage_calibration["divider_ratio"]) * 1_000_000.0))
        offset_mv = int(self.active_voltage_calibration["offset_mv"])
        stable_us = int(self.spinBox_voltage_stable_ms.value()) * 1000
        self.send_command(f"VOLTAGE_CFG;{ratio_u6};{offset_mv};{stable_us}")
        self.send_command("ADS?")

    def voltage_start_test(self, requested_mode):
        if self.voltage_test_running:
            QMessageBox.warning(self.window, "Collage / décollage", "Une mesure est déjà en cours.")
            return
        if not self.is_connected():
            QMessageBox.warning(self.window, "Collage / décollage", "RP2040 non connecté.")
            return
        if not self.ea_psu.connected:
            QMessageBox.warning(self.window, "Tension de fonctionnement", "Alimentation EA non connectée.")
            return
        calibration_ok, calibration_reason = self.voltage_active_calibration_required()
        if not calibration_ok:
            QMessageBox.warning(self.window, "Tension de fonctionnement", calibration_reason + "\n\nOuvrir l'onglet Étalonnage tension.")
            return
        try:
            # Force la prise en compte d'une valeur encore en cours de saisie dans un QDoubleSpinBox.
            for spin in (
                self.doubleSpinBox_voltage_vmax,
                self.doubleSpinBox_voltage_ramp_up_s,
                self.doubleSpinBox_voltage_ramp_down_s,
                self.doubleSpinBox_voltage_current_limit,
                self.doubleSpinBox_voltage_interphase_s,
            ):
                spin.interpretText()
            self.voltage_metadata()
            if float(self.doubleSpinBox_voltage_vmax.value()) <= 0:
                raise ValueError("La tension maximale doit être supérieure à 0 V.")
            if float(self.doubleSpinBox_voltage_ramp_up_s.value()) < 0.3:
                raise ValueError("La durée de montée BE doit être au moins 0,3 s.")
            if float(self.doubleSpinBox_voltage_ramp_down_s.value()) < 0.3:
                raise ValueError("La durée de retour BE/BR doit être au moins 0,3 s.")
            if float(self.doubleSpinBox_voltage_interphase_s.value()) < 3.0:
                raise ValueError("L'attente entre opérations doit être au moins 3,0 s pour laisser l'EA valider la seconde rampe.")
            max_duration_s = self.voltage_update_ramp_limits()
            for label, spin in (
                ("montée BE", self.doubleSpinBox_voltage_ramp_up_s),
                ("retour BE/BR", self.doubleSpinBox_voltage_ramp_down_s),
            ):
                if float(spin.value()) > max_duration_s + 1e-9:
                    raise ValueError(
                        f"Durée de {label} trop longue pour la pente minimale EA : "
                        f"maximum {max_duration_s:.3f} s avec Vmax={self.doubleSpinBox_voltage_vmax.value():.3f} V."
                    )
        except Exception as exc:
            QMessageBox.warning(self.window, "Collage / décollage", str(exc))
            return

        self.voltage_capture_run_settings()
        self.voltage_requested_mode = str(requested_mode).upper()
        self.voltage_active_scan = ""
        self._voltage_bistable_pickup_prepositioned = False
        self.voltage_pending_ramp = None
        self.voltage_ramp_info = None
        self.voltage_ramp_started_monotonic = None
        self.voltage_interphase_target_monotonic = None
        self.voltage_interphase_origin_monotonic = None
        self.voltage_interphase_actual_s = None
        self.voltage_waiting_for_rp_arm = False
        self.voltage_time_results = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_first_passage = {"PICKUP": {}, "DROPOUT": {}}
        self.voltage_capture_policy = ""
        self.voltage_effective_ramp_s = {"PICKUP": None, "DROPOUT": None}
        self.voltage_ramp_readbacks = {"PICKUP": {}, "DROPOUT": {}}
        # Chaque essai repart avec un état métrologique et sécurité indépendant.
        # Sans cette remise à zéro, un verdict de l'essai précédent pourrait
        # contaminer la sauvegarde suivante.
        self.voltage_reset_assessment_state()
        if self.voltage_requested_mode in ("PICKUP", "CYCLE"):
            self.voltage_results = {"PICKUP": {}, "DROPOUT": {}}
            self.voltage_raw_results = {"PICKUP": {}, "DROPOUT": {}}
        elif self.voltage_requested_mode == "DROPOUT":
            self.voltage_results["DROPOUT"] = {}
            self.voltage_raw_results["DROPOUT"] = {}

        self.voltage_test_running = True
        self.voltage_ads_ok = False
        self.voltage_refresh_results_table()
        self.voltage_update_button_states()
        try:
            self.ea_psu.set_remote()
            if not self.voltage_stop_ea_and_confirm("préparation avant mesure", show_alert=False):
                raise RuntimeError("L'arrêt initial de l'alimentation EA n'est pas confirmé.")
            self.ea_psu.send(f"CURR {self.voltage_run_setting('current_limit_a', self.doubleSpinBox_voltage_current_limit):.6f}")
            self.ea_psu.send("POW MAX")
            self.ea_psu.send("VOLT 0")
            self.send_command("VOLTAGE_SCAN;CANCEL")
            self.voltage_set_coil_hold("OFF")
            self.voltage_send_config()
            if self.voltage_requested_mode in ("PICKUP", "CYCLE"):
                self.label_voltage_status.setText("Préparation collage : mise à 0 V et vérification de la position repos.")
                QTimer.singleShot(400, self.voltage_begin_pickup)
            else:
                self.label_voltage_status.setText("Préconditionnement décollage : alimentation du relais à V maxi.")
                QTimer.singleShot(300, self.voltage_precondition_dropout)
        except Exception as exc:
            self.voltage_abort_test(f"Préparation impossible : {exc}")

    def voltage_set_coil_hold(self, coil):
        coil = str(coil or "OFF").upper()
        if coil not in ("BE", "BR", "OFF"):
            raise ValueError(f"Bobine invalide : {coil}")
        self.send_command(f"COIL_HOLD;{coil}")

    def voltage_preposition_bistable(self, target_position, callback):
        """Positionne avec l'EA à Vmax, puis coupe et revient à 0 V avant la rampe de mesure."""
        if not self.voltage_test_running:
            return
        target_position = str(target_position).upper()
        coil = "BR" if target_position == "REST" else "BE"
        try:
            vmax = self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax)
            # L'arrêt initial a déjà confirmé SEL=NONE. Pour ce prépositionnement
            # statique, ne pas envoyer WAVE:STAT STOP hors mode générateur.
            self.ea_psu.send("FUNC:GEN:SEL NONE")
            self.ea_psu.send(f"CURR {self.voltage_run_setting('current_limit_a', self.doubleSpinBox_voltage_current_limit):.6f}")
            self.ea_psu.send(f"VOLT {vmax:.6f}")
            self.ea_psu.output(True)
            self.voltage_set_coil_hold(coil)
            self.label_voltage_status.setText(f"Prépositionnement bistable {target_position} par bobine {coil} à {vmax:.3f} V.")
            def finish_preposition():
                if not self.voltage_test_running:
                    return
                self.voltage_set_coil_hold("OFF")
                if not self.voltage_stop_ea_and_confirm("fin prépositionnement bistable", show_alert=False):
                    self.voltage_abort_test("Prépositionnement : arrêt EA non confirmé.")
                    return
                QTimer.singleShot(350, callback)
            QTimer.singleShot(350, finish_preposition)
        except Exception as exc:
            self.voltage_abort_test(f"Prépositionnement bistable impossible : {exc}")

    def voltage_begin_pickup(self):
        if not self.voltage_test_running:
            return
        if self.voltage_relay_type() == "BISTABLE" and not getattr(self, "_voltage_bistable_pickup_prepositioned", False):
            self._voltage_bistable_pickup_prepositioned = True
            self.voltage_preposition_bistable("REST", self.voltage_begin_pickup)
            return
        ok, description = self.voltage_contacts_expected("PICKUP")
        if not ok:
            self.voltage_abort_test(f"Position initiale incorrecte : attendu {description}.")
            return
        try:
            nb = int(self.spinBox_voltage_nb_inverseurs.value())
            self.voltage_set_coil_hold("BE")
            self.voltage_active_scan = "PICKUP"
            self.voltage_configure_ramp(
                0.0,
                self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax),
                self.voltage_run_setting("ramp_up_s", self.doubleSpinBox_voltage_ramp_up_s),
                mode="PICKUP",
                nb_inverseurs=nb,
            )
        except Exception as exc:
            self.voltage_abort_test(f"Démarrage collage/BE impossible : {exc}")

    def voltage_precondition_dropout(self):
        if not self.voltage_test_running:
            return
        if self.voltage_relay_type() == "BISTABLE":
            self.voltage_preposition_bistable("WORK", self.voltage_begin_dropout)
            return
        try:
            vmax = self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax)
            self.voltage_set_coil_hold("BE")
            self.ea_psu.set_remote()
            self.ea_psu.send(f"CURR {self.voltage_run_setting('current_limit_a', self.doubleSpinBox_voltage_current_limit):.6f}")
            self.ea_psu.send(f"VOLT {vmax:.6f}")
            self.ea_psu.output(True)
            self.label_voltage_status.setText(f"Préconditionnement monostable à {vmax:.3f} V - attente position travail.")
            QTimer.singleShot(700, self.voltage_begin_dropout)
        except Exception as exc:
            self.voltage_abort_test(f"Préconditionnement impossible : {exc}")

    def voltage_begin_dropout(self, not_before=None):
        if not self.voltage_test_running:
            return
        ok, description = self.voltage_contacts_expected("DROPOUT")
        if not ok:
            self.voltage_abort_test(f"Relais non en position travail avant mesure retour : attendu {description}.")
            return
        try:
            nb = int(self.spinBox_voltage_nb_inverseurs.value())
            self.voltage_active_scan = "DROPOUT"
            if self.voltage_relay_type() == "BISTABLE":
                self.voltage_set_coil_hold("BR")
                # En bistable le second champ de l'IHM est bien la durée de montée BR.
                self.voltage_configure_ramp(
                    0.0,
                    self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax),
                    self.voltage_run_setting("ramp_down_s", self.doubleSpinBox_voltage_ramp_down_s),
                    mode="DROPOUT",
                    nb_inverseurs=nb,
                    not_before=not_before,
                )
            else:
                self.voltage_set_coil_hold("BE")
                self.voltage_configure_ramp(
                    self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax),
                    0.0,
                    self.voltage_run_setting("ramp_down_s", self.doubleSpinBox_voltage_ramp_down_s),
                    mode="DROPOUT",
                    nb_inverseurs=nb,
                    not_before=not_before,
                )
        except Exception as exc:
            self.voltage_abort_test(f"Démarrage décollage/BR impossible : {exc}")

    def voltage_configure_ramp(self, start_v, end_v, duration_s, mode, nb_inverseurs, not_before=None):
        start_v = float(start_v)
        end_v = float(end_v)
        duration_s = float(duration_s)
        mode = str(mode).upper()
        hold_s = max(1.0, duration_s + 1.0)
        self.label_voltage_status.setText(
            f"Configuration rampe EA {mode} : {start_v:.3f} → {end_v:.3f} V en {duration_s:.3f} s."
        )
        self.ea_psu.configure_voltage_ramp(
            start_v, end_v, duration_s, hold_s,
            self.voltage_run_setting("current_limit_a", self.doubleSpinBox_voltage_current_limit),
        )
        self.voltage_pending_ramp = {
            "start_v": start_v,
            "end_v": end_v,
            "duration_s": duration_s,
            "hold_s": hold_s,
            "mode": mode,
            "nb_inverseurs": int(nb_inverseurs),
            "not_before": float(not_before) if not_before is not None else None,
            "verified": False,
        }
        # L'EA doit terminer le SUBMIT avant toute interrogation ou commande suivante.
        self.voltage_phase_timer.start(2200)

    def voltage_validate_configured_ramp(self):
        if not self.voltage_test_running or not self.voltage_pending_ramp:
            return
        pending = self.voltage_pending_ramp
        try:
            if not pending.get("verified"):
                readback = self.ea_psu.verify_voltage_ramp(
                    pending["start_v"], pending["end_v"], pending["duration_s"]
                )
                pending["verified"] = True
                pending["readback"] = dict(readback)
                self.voltage_ramp_readbacks[pending["mode"]] = dict(readback)
            target = pending.get("not_before")
            if target is not None:
                remaining_s = float(target) - time.monotonic()
                if remaining_s > 0.001:
                    self.label_voltage_status.setText(
                        f"Rampe {pending['mode']} vérifiée par relecture EA : "
                        f"{pending['duration_s']:.3f} s. Démarrage dans {remaining_s:.2f} s."
                    )
                    self.voltage_phase_timer.start(max(1, int(round(remaining_s * 1000.0))))
                    return
            self.voltage_arm_configured_ramp()
        except Exception as exc:
            self.voltage_abort_test(f"Validation de la rampe EA impossible : {exc}")

    def voltage_arm_configured_ramp(self):
        if not self.voltage_test_running or not self.voltage_pending_ramp:
            return
        pending = self.voltage_pending_ramp
        try:
            self.voltage_waiting_for_rp_arm = True
            self.send_command(
                f"VOLTAGE_SCAN;ARM;{pending['mode']};{pending['nb_inverseurs']}"
            )
            self.label_voltage_status.setText(
                f"Rampe EA relue et conforme ({pending['duration_s']:.3f} s). "
                f"Armement RP2040 {pending['mode']}..."
            )
            self.voltage_arm_timeout_timer.start(1500)
        except Exception as exc:
            self.voltage_abort_test(f"Armement RP2040 impossible : {exc}")

    def voltage_arm_timeout(self):
        if self.voltage_waiting_for_rp_arm and self.voltage_test_running:
            self.voltage_abort_test("Le RP2040 n'a pas confirmé l'armement de la mesure tension.")

    def voltage_start_configured_ramp(self):
        if not self.voltage_test_running or not self.voltage_pending_ramp:
            return
        pending = dict(self.voltage_pending_ramp)
        self.voltage_pending_ramp = None
        try:
            self.ea_psu.start_generator()
            self.voltage_ramp_started_monotonic = time.monotonic()
            interphase_text = ""
            if pending.get("not_before") is not None and self.voltage_interphase_origin_monotonic is not None:
                self.voltage_interphase_actual_s = self.voltage_ramp_started_monotonic - self.voltage_interphase_origin_monotonic
                requested_interphase = self.voltage_run_setting("interphase_s", self.doubleSpinBox_voltage_interphase_s)
                interphase_text = f" Attente réelle : {self.voltage_interphase_actual_s:.3f} s (demandée {requested_interphase:.3f} s)."
            self.voltage_ramp_info = pending
            self.voltage_progress_timer.start()
            self.voltage_ea_monitor_timer.start()
            self.label_voltage_status.setText(
                f"Rampe {pending['mode']} EN COURS : {pending['start_v']:.3f} → "
                f"{pending['end_v']:.3f} V en {pending['duration_s']:.3f} s.{interphase_text}"
            )
            timeout_ms = int((pending["duration_s"] + pending["hold_s"] + 3.0) * 1000)
            self.voltage_timeout_timer.start(timeout_ms)
            QTimer.singleShot(150, self.voltage_check_generator_running)
        except Exception as exc:
            self.voltage_abort_test(f"Départ rampe EA impossible : {exc}")

    def voltage_check_generator_running(self):
        if not self.voltage_test_running or not self.voltage_ramp_info:
            self.voltage_ea_monitor_timer.stop()
            return
        try:
            state = self.ea_psu.generator_state()
            if not self.ea_psu.generator_state_is_running(state):
                self.voltage_abort_test(f"L'alimentation EA ne confirme pas la rampe RUN : {state or 'réponse vide'}")
        except Exception as exc:
            self.voltage_abort_test(f"Contrôle état générateur EA impossible : {exc}")

    def voltage_update_ramp_progress(self):
        info = self.voltage_ramp_info
        started = self.voltage_ramp_started_monotonic
        if not self.voltage_test_running or not info or started is None:
            self.voltage_progress_timer.stop()
            return
        elapsed = max(0.0, time.monotonic() - started)
        duration = max(0.0001, float(info["duration_s"]))
        fraction = max(0.0, min(1.0, elapsed / duration))
        expected = float(info["start_v"]) + (float(info["end_v"]) - float(info["start_v"])) * fraction
        measured = "--"
        if self.voltage_last_adc_mv is not None:
            measured = f"{self.voltage_last_adc_mv / 1000.0:.3f} V"
        self.label_voltage_live.setText(
            f"ADS1115 : {'OK' if self.voltage_ads_ok else '--'} | RAW : "
            f"{'--' if self.voltage_last_adc_raw is None else self.voltage_last_adc_raw} | "
            f"Mesuré : {measured} | Rampe {info['mode']} : {elapsed:.2f}/{duration:.2f} s | "
            f"Attendu linéaire : {expected:.3f} V"
        )
        self.voltage_calibration_update_live()

    def voltage_calculate_effective_ramp(self, mode):
        mode = str(mode).upper()
        info = self.voltage_ramp_info if self.voltage_ramp_info and self.voltage_ramp_info.get("mode") == mode else None
        elapsed_us = self.voltage_time_results.get(mode, {}).get("GLOBAL")
        measured = self.voltage_results.get(mode, {}).get("GLOBAL")
        if not info or elapsed_us is None or measured is None:
            return None
        span = float(info["end_v"]) - float(info["start_v"])
        progressed = float(measured) - float(info["start_v"])
        if abs(span) < 1e-12 or abs(progressed) < 0.010:
            return None
        effective = (float(elapsed_us) / 1_000_000.0) * abs(span / progressed)
        self.voltage_effective_ramp_s[mode] = effective
        return effective

    def voltage_parse_fields(self, line):
        parts = str(line).split(";")
        data = {"frame": parts[0], "kind": ""}
        index = 1
        if len(parts) > 1 and "=" not in parts[1]:
            data["kind"] = parts[1].upper()
            index = 2
        for item in parts[index:]:
            if "=" in item:
                key, value = item.split("=", 1)
                data[key.upper()] = value
        return data

    def voltage_parse_frame(self, line):
        data = self.voltage_parse_fields(line)
        frame = data.get("frame", "")
        kind = data.get("kind", "")
        if frame == "ADS":
            self.voltage_ads_ok = data.get("STATUS", "").upper() == "OK"
            try:
                self.voltage_last_adc_mv = int(data.get("COIL_MV", data.get("MV", "0")))
            except Exception:
                pass
            try:
                self.voltage_last_adc_raw = int(data.get("RAW", "0"))
                self.voltage_calibration_accept_ads_frame(self.voltage_last_adc_raw)
            except Exception:
                pass
            self.voltage_update_live(data.get("MODE", self.voltage_active_scan or "PRÊT"))
            return
        if frame == "COIL":
            self.voltage_update_live(self.voltage_active_scan or kind or "PRÊT")
            return
        if frame != "VSCAN":
            return
        if kind == "ERROR":
            self.voltage_abort_test("RP2040 : " + data.get("REASON", "erreur VSCAN"))
            return
        if kind == "BEGIN":
            self.voltage_active_scan = data.get("MODE", self.voltage_active_scan).upper()
            capture_policy = data.get("CAPTURE", "").upper()
            validation_policy = data.get("VALIDATION", "").upper()
            if capture_policy != "FIRST_PASSAGE" or validation_policy != "STABLE_AFTER_CAPTURE":
                self.voltage_waiting_for_rp_arm = False
                self.voltage_arm_timeout_timer.stop()
                self.voltage_abort_test(
                    "Firmware RP2040 incompatible : la mesure tension doit utiliser "
                    "CAPTURE=FIRST_PASSAGE et VALIDATION=STABLE_AFTER_CAPTURE. "
                    "Téléverser le firmware V2.12.3 R8."
                )
                return
            self.voltage_capture_policy = capture_policy
            if self.voltage_waiting_for_rp_arm:
                self.voltage_waiting_for_rp_arm = False
                self.voltage_arm_timeout_timer.stop()
                self.label_voltage_status.setText(
                    f"RP2040 armé pour {self.voltage_active_scan} : capture au premier passage, "
                    "puis validation de stabilité. Démarrage immédiat de la rampe EA."
                )
                self.voltage_start_configured_ramp()
            return
        if kind == "FIRST":
            mode = data.get("MODE", self.voltage_active_scan).upper()
            inv_txt = str(data.get("INV", "")).upper()
            try:
                mv = int(data.get("MV", "-1"))
                raw = int(data.get("RAW", "-1"))
                t_us = int(data.get("T_US", "-1"))
            except Exception:
                return
            key = "GLOBAL" if inv_txt == "GLOBAL" else int(inv_txt) if inv_txt.isdigit() else inv_txt
            self.voltage_first_passage.setdefault(mode, {})[key] = {
                "mv": mv,
                "raw": raw,
                "t_us": t_us,
            }
            if key == "GLOBAL" and mv >= 0:
                self.label_voltage_status.setText(
                    f"Premier passage complet {mode} capturé à {mv / 1000.0:.3f} V "
                    f"(t={max(0, t_us) / 1_000_000.0:.3f} s). "
                    "Validation de la stabilité en cours ; les rebonds ne modifieront pas cette tension."
                )
            return
        if kind == "INV":
            mode = data.get("MODE", self.voltage_active_scan).upper()
            try:
                inv = int(data.get("INV", "0"))
                mv = int(data.get("MV", "0"))
                if 1 <= inv <= 4:
                    self.voltage_results.setdefault(mode, {})[inv] = mv / 1000.0
                    try:
                        self.voltage_raw_results.setdefault(mode, {})[inv] = int(data.get("RAW", "-1"))
                    except Exception:
                        pass
                    try:
                        self.voltage_time_results.setdefault(mode, {})[inv] = int(data.get("T_US", "-1"))
                    except Exception:
                        pass
                    self.voltage_refresh_results_table()
            except Exception:
                pass
            return
        if kind == "RESULT":
            self.voltage_timeout_timer.stop()
            self.voltage_progress_timer.stop()
            self.voltage_ea_monitor_timer.stop()
            mode = data.get("MODE", self.voltage_active_scan).upper()
            if data.get("CAPTURE", "").upper() != "FIRST_PASSAGE":
                self.voltage_abort_test(
                    "Résultat RP2040 refusé : politique de capture au premier passage non confirmée."
                )
                return
            results = {}
            times = {}
            for inv in range(1, 5):
                value_valid = False
                try:
                    value_mv = int(data.get(f"I{inv}_MV", "-1"))
                    if value_mv >= 0:
                        results[inv] = value_mv / 1000.0
                        value_valid = True
                except Exception:
                    pass
                if value_valid:
                    try:
                        value_us = int(data.get(f"I{inv}_T_US", "-1"))
                        if value_us >= 0:
                            times[inv] = value_us
                    except Exception:
                        pass
            try:
                global_mv = int(data.get("GLOBAL_MV", "-1"))
                if global_mv >= 0:
                    results["GLOBAL"] = global_mv / 1000.0
            except Exception:
                pass
            try:
                global_us = int(data.get("GLOBAL_T_US", data.get("ELAPSED_US", "-1")))
                if global_us >= 0:
                    times["GLOBAL"] = global_us
            except Exception:
                pass
            self.voltage_results[mode] = results
            self.voltage_time_results[mode] = times
            raw_results = {}
            for inv in range(1, 5):
                try:
                    raw_value = int(data.get(f"I{inv}_RAW", "-1"))
                    if raw_value >= 0:
                        raw_results[inv] = raw_value
                except Exception:
                    pass
            try:
                global_raw = int(data.get("GLOBAL_RAW", "-1"))
                if global_raw >= 0:
                    raw_results["GLOBAL"] = global_raw
            except Exception:
                pass
            self.voltage_raw_results[mode] = raw_results
            try:
                self.voltage_last_adc_mv = int(data.get("LAST_MV", data.get("GLOBAL_MV", "0")))
            except Exception:
                pass
            self.voltage_calculate_effective_ramp(mode)
            self.voltage_refresh_results_table()
            self.voltage_handle_scan_complete(mode)

    def voltage_handle_scan_complete(self, mode):
        if not self.voltage_test_running:
            return
        mode = str(mode).upper()
        self.voltage_progress_timer.stop()
        self.voltage_ea_monitor_timer.stop()
        requested_s = None
        if self.voltage_ramp_info and self.voltage_ramp_info.get("mode") == mode:
            requested_s = float(self.voltage_ramp_info.get("duration_s", 0.0))
        effective_s = self.voltage_effective_ramp_s.get(mode)
        global_v = self.voltage_results.get(mode, {}).get("GLOBAL")
        elapsed_us = self.voltage_time_results.get(mode, {}).get("GLOBAL")
        elapsed_s = None if elapsed_us is None else elapsed_us / 1_000_000.0
        plausibility = self.voltage_evaluate_plausibility(mode)
        timing_text = ""
        if requested_s is not None:
            timing_text = f" Rampe demandée {requested_s:.3f} s"
            if effective_s is not None:
                timing_text += f", durée reconstituée {effective_s:.3f} s"
            if elapsed_s is not None and global_v is not None:
                timing_text += f", seuil {global_v:.3f} V à t={elapsed_s:.3f} s."
            else:
                timing_text += "."
        plausibility_text = " " + str(plausibility.get("detail", ""))

        if plausibility.get("status") != "OK":
            self.voltage_result_override = str(plausibility.get("status") or "NON_VERIFIE")
            self.voltage_finish_test(
                False,
                f"Mesure {mode} enregistrée mais classée {self.voltage_result_override}.{timing_text}{plausibility_text}",
            )
            return

        try:
            if mode == "PICKUP" and self.voltage_requested_mode == "CYCLE":
                interphase_s = self.voltage_run_setting("interphase_s", self.doubleSpinBox_voltage_interphase_s)
                self.voltage_interphase_origin_monotonic = time.monotonic()
                target = self.voltage_interphase_origin_monotonic + interphase_s
                self.voltage_interphase_target_monotonic = target
                if self.voltage_relay_type() == "BISTABLE":
                    self.voltage_set_coil_hold("OFF")
                    if not self.voltage_stop_ea_and_confirm("transition BE vers BR", show_alert=False):
                        self.voltage_abort_test("Transition BE vers BR : arrêt EA non confirmé.")
                        return
                    self.label_voltage_status.setText(
                        f"Basculement BE mesuré au premier passage, puis validé stable.{timing_text}{plausibility_text} "
                        f"Préparation BR ; démarrage dans {interphase_s:.2f} s."
                    )
                    self.voltage_begin_dropout(not_before=target)
                else:
                    vmax = self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax)
                    self.ea_psu.stop_generator(leave_mode=True)
                    self.ea_psu.send(f"VOLT {vmax:.6f}")
                    self.ea_psu.output(True)
                    self.label_voltage_status.setText(
                        f"Collage mesuré au premier passage, puis validé stable.{timing_text}{plausibility_text} "
                        f"Maintien à Vmax ; descente dans {interphase_s:.2f} s."
                    )
                    self.voltage_begin_dropout(not_before=target)
                return
            label = "BE" if mode == "PICKUP" and self.voltage_relay_type() == "BISTABLE" else "BR" if mode == "DROPOUT" and self.voltage_relay_type() == "BISTABLE" else mode.lower()
            self.voltage_finish_test(
                True,
                f"Mesure {label} terminée : tension capturée au premier passage, "
                f"puis position validée stable.{timing_text}{plausibility_text}"
            )
        except Exception as exc:
            self.voltage_abort_test(f"Transition de phase impossible : {exc}")

    def voltage_timeout_global_fallback(self, mode):
        """Valeur limite à enregistrer lorsque le global n'a jamais été validé."""
        mode = str(mode or "").upper()
        vmax = self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax)
        if mode == "DROPOUT" and self.voltage_relay_type() == "MONOSTABLE":
            return 0.0, "0 V atteint sans rappel global complet"
        return float(vmax), "Vmax atteint sans basculement global complet"

    def voltage_apply_timeout_global_fallback(self):
        mode = str(self.voltage_active_scan or "").upper()
        if mode not in ("PICKUP", "DROPOUT") and self.voltage_ramp_info:
            mode = str(self.voltage_ramp_info.get("mode", "")).upper()
        if mode not in ("PICKUP", "DROPOUT"):
            mode = "PICKUP" if self.voltage_requested_mode in ("PICKUP", "CYCLE") else "DROPOUT"

        first_global = self.voltage_first_passage.get(mode, {}).get("GLOBAL", {})
        try:
            first_mv = int(first_global.get("mv", -1))
        except Exception:
            first_mv = -1
        if first_mv >= 0:
            value = first_mv / 1000.0
            reason = "premier passage global capturé mais stabilité non confirmée"
            status = "DEFAUT_STABILITE"
            try:
                first_raw = int(first_global.get("raw", -1))
                if first_raw >= 0:
                    self.voltage_raw_results.setdefault(mode, {})["GLOBAL"] = first_raw
            except Exception:
                pass
            try:
                first_t_us = int(first_global.get("t_us", -1))
                if first_t_us >= 0:
                    self.voltage_time_results.setdefault(mode, {})["GLOBAL"] = first_t_us
            except Exception:
                pass
            detail = (
                f"Premier passage global conservé à {value:.3f} V, mais la stabilité finale "
                "n'a pas été confirmée avant la fin de la rampe."
            )
        else:
            value, reason = self.voltage_timeout_global_fallback(mode)
            status = "DEFAUT_CONTACTS"
            detail = (
                f"Contacts incomplets en fin de rampe : valeur globale limite enregistrée "
                f"à {value:.3f} V ({reason})."
            )

        self.voltage_results.setdefault(mode, {})["GLOBAL"] = value
        self.voltage_plausibility[mode] = {
            "status": status,
            "mode": mode,
            "measured_v": value,
            "elapsed_s": None,
            "expected_elapsed_s": None,
            "elapsed_error_s": None,
            "tolerance_s": None,
            "fallback_global_v": value,
            "fallback_reason": reason,
            "detail": detail,
        }
        self.voltage_result_override = status
        self.voltage_refresh_results_table()
        return mode, value, reason, status

    def voltage_test_timeout(self):
        mode, fallback_v, fallback_reason, fallback_status = self.voltage_apply_timeout_global_fallback()
        title, detail = self.voltage_timeout_contact_diagnostic()
        detail = (
            f"{detail}\n\nValeur globale enregistrée : {fallback_v:.3f} V "
            f"({fallback_reason}). Résultat classé {fallback_status}."
        )
        message = f"{title}\n\n{detail}"
        self.voltage_finish_test(
            False,
            message,
            allow_measure_all_continue=True,
        )
        if bool(getattr(self, "voltage_ea_stop_confirmation", {}).get("confirmed", False)):
            self.big_message_box(
                "Échec mesure tension",
                title,
                detail,
                ok_text="COMPRIS",
                icon=QMessageBox.Warning,
            )

    def voltage_abort_test(self, reason="Arrêt sécurité demandé"):
        self.voltage_timeout_timer.stop()
        self.voltage_phase_timer.stop()
        self.voltage_arm_timeout_timer.stop()
        self.voltage_progress_timer.stop()
        self.voltage_ea_monitor_timer.stop()
        self.voltage_pending_ramp = None
        self.voltage_ramp_info = None
        self.voltage_ramp_started_monotonic = None
        self.voltage_waiting_for_rp_arm = False
        try:
            if self.is_connected():
                self.send_command("VOLTAGE_SCAN;CANCEL")
                self.voltage_set_coil_hold("OFF")
        except Exception:
            pass
        stop_confirmed = self.voltage_stop_ea_and_confirm("arrêt de sécurité", show_alert=False)
        self.voltage_test_running = False
        self.voltage_active_scan = ""
        message = str(reason)
        if not stop_confirmed:
            detail = str(getattr(self, "voltage_last_stop_diagnostic", "") or "").strip()
            message += " ARRÊT EA NON CONFIRMÉ — COUPER MANUELLEMENT L'ALIMENTATION."
            if detail:
                message += "\n\n" + detail
        self.label_voltage_status.setText(message)
        self.label_voltage_status.setStyleSheet("background-color: rgb(190,0,0); color: white; font-weight: bold; border: 2px solid black;")
        self.voltage_update_button_states()
        combined_aborted = getattr(self, "measure_all_active", False) and self.measure_all_phase == "VOLTAGE"
        if combined_aborted:
            self.voltage_measure_all_fail(str(reason), preserve_status=True)
        if not stop_confirmed:
            self.big_message_box(
                "Sécurité alimentation EA",
                "ARRÊT EA NON CONFIRMÉ",
                message,
                ok_text="J'AI COUPÉ MANUELLEMENT",
                icon=QMessageBox.Critical,
            )

    def voltage_finish_test(self, success, message, allow_measure_all_continue=False):
        self.voltage_timeout_timer.stop()
        self.voltage_phase_timer.stop()
        self.voltage_arm_timeout_timer.stop()
        self.voltage_progress_timer.stop()
        self.voltage_ea_monitor_timer.stop()
        self.voltage_pending_ramp = None
        self.voltage_ramp_info = None
        self.voltage_ramp_started_monotonic = None
        self.voltage_waiting_for_rp_arm = False
        try:
            if self.is_connected():
                self.send_command("VOLTAGE_SCAN;CANCEL")
                self.voltage_set_coil_hold("OFF")
        except Exception:
            pass
        stop_confirmed = self.voltage_stop_ea_and_confirm("fin de mesure", show_alert=False)
        if not stop_confirmed:
            success = False
            detail = str(getattr(self, "voltage_last_stop_diagnostic", "") or "").strip()
            message += " ARRÊT EA NON CONFIRMÉ — COUPER MANUELLEMENT L'ALIMENTATION."
            if detail:
                message += "\n\n" + detail
        self.voltage_test_running = False
        self.voltage_active_scan = ""
        saved = False
        try:
            self.voltage_save_result()
            saved = True
        except Exception as exc:
            message = f"{message} Sauvegarde impossible : {exc}"
            success = False
        self.label_voltage_status.setText(message + (" Résultat sauvegardé." if saved else ""))
        self.label_voltage_status.setStyleSheet(
            "background-color: rgb(0,150,70); color: white; font-weight: bold; border: 2px solid rgb(0,80,35);"
            if success else
            "background-color: rgb(190,0,0); color: white; font-weight: bold; border: 2px solid black;"
        )
        self.voltage_update_button_states()
        combined_voltage_phase = getattr(self, "measure_all_active", False) and self.measure_all_phase == "VOLTAGE"
        if combined_voltage_phase:
            if (success or allow_measure_all_continue) and saved and stop_confirmed:
                self.measure_all_phase = "CHRONO_PENDING"
                self.label_voltage_status.setText(
                    "MESURE TOTALE 1/2 TERMINÉE — tensions sauvegardées. Préparation de la chronométrie contacts..."
                )
                self.label_voltage_status.setStyleSheet(
                    "background-color: rgb(0,90,160); color: white; font-weight: bold; border: 2px solid rgb(0,45,90);"
                )
                QTimer.singleShot(400, self.voltage_measure_all_start_chrono)
            else:
                self.voltage_measure_all_fail(
                    "La phase tension n'est pas conforme ou n'a pas été sauvegardée.",
                    preserve_status=True,
                )
        if not stop_confirmed:
            self.big_message_box(
                "Sécurité alimentation EA",
                "ARRÊT EA NON CONFIRMÉ",
                message,
                ok_text="J'AI COUPÉ MANUELLEMENT",
                icon=QMessageBox.Critical,
            )

    def voltage_update_live(self, phase=""):
        # Pendant une rampe, le timer affiche déjà le temps demandé, le temps écoulé
        # et la tension théorique. Ne pas l'écraser par la trame ADS périodique.
        if self.voltage_progress_timer.isActive() and self.voltage_ramp_info:
            self.voltage_update_ramp_progress()
            return
        adc = "OK" if self.voltage_ads_ok else "--"
        voltage = "--"
        if self.voltage_last_adc_mv is not None:
            voltage = f"{self.voltage_last_adc_mv / 1000.0:.3f}"
        raw = "--" if self.voltage_last_adc_raw is None else str(self.voltage_last_adc_raw)
        self.label_voltage_live.setText(f"ADS1115 : {adc} | RAW : {raw} | Tension corrigée : {voltage} V | Phase : {phase or 'prête'}")
        self.voltage_calibration_update_live()

    def voltage_refresh_results_table(self):
        nb = int(self.spinBox_voltage_nb_inverseurs.value()) if hasattr(self, "spinBox_voltage_nb_inverseurs") else 2
        rows = list(range(1, nb + 1)) + ["GLOBAL"]
        self.tableWidget_voltage_results.setRowCount(len(rows))
        for row_index, inv in enumerate(rows):
            pickup = self.voltage_results.get("PICKUP", {}).get(inv)
            dropout = self.voltage_results.get("DROPOUT", {}).get(inv)
            state = "Complet" if pickup is not None and dropout is not None else "Partiel" if pickup is not None or dropout is not None else "En attente"
            values = [str(inv), "--" if pickup is None else f"{pickup:.3f}", "--" if dropout is None else f"{dropout:.3f}", state]
            for col, value in enumerate(values):
                self.tableWidget_voltage_results.setItem(row_index, col, QTableWidgetItem(value))
        self.tableWidget_voltage_results.resizeColumnsToContents()

    def voltage_update_button_states(self):
        rp_ok = self.is_connected()
        ea_ok = bool(getattr(self, "ea_psu", None) and self.ea_psu.connected)
        running = bool(getattr(self, "voltage_test_running", False))
        measure_all_busy = bool(getattr(self, "measure_all_active", False))
        chrono_running = bool(getattr(self, "chrono_measure_running", False))
        busy = running or measure_all_busy or chrono_running
        calibration_ok = bool(getattr(self, "active_voltage_calibration", None))
        enabled = rp_ok and ea_ok and calibration_ok and not busy and not getattr(self, "auto_neutral_running", False)
        self.pushButton_voltage_pickup.setEnabled(enabled)
        self.pushButton_voltage_dropout.setEnabled(enabled)
        self.pushButton_voltage_cycle.setEnabled(enabled)
        self.pushButton_voltage_measure_all.setEnabled(enabled)
        self.pushButton_voltage_stop.setEnabled(running or measure_all_busy)
        self.pushButton_voltage_ea_connect.setEnabled(not ea_ok and not busy)
        self.pushButton_voltage_ea_disconnect.setEnabled(ea_ok and not busy)
        self.pushButton_voltage_ea_refresh.setEnabled(not busy)
        for widget in (
            self.doubleSpinBox_voltage_vmax, self.doubleSpinBox_voltage_ramp_up_s,
            self.doubleSpinBox_voltage_ramp_down_s, self.doubleSpinBox_voltage_current_limit,
            self.doubleSpinBox_voltage_chrono_v, self.doubleSpinBox_voltage_interphase_s,
            self.spinBox_voltage_nb_inverseurs, self.spinBox_voltage_stable_ms,
            self.comboBox_voltage_relay_type,
        ):
            widget.setEnabled(not busy)
        self.pushButton_voltage_export_xlsx.setEnabled(not busy)
        self.pushButton_voltage_export_pdf.setEnabled(not busy)

    def voltage_init_db(self):
        self.chrono_init_db()
        with self.chrono_connect_db() as con:
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS mesures_tension_fonctionnement (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    lot TEXT NOT NULL DEFAULT '',
                    date_test TEXT NOT NULL DEFAULT '',
                    relais TEXT NOT NULL DEFAULT '',
                    ambiance_c TEXT NOT NULL DEFAULT '',
                    nom_test TEXT NOT NULL DEFAULT '',
                    sn TEXT NOT NULL DEFAULT '',
                    relay_type TEXT NOT NULL DEFAULT 'MONOSTABLE',
                    nb_inverseurs INTEGER NOT NULL DEFAULT 0,
                    vmax_v REAL NOT NULL DEFAULT 0,
                    ramp_up_s REAL NOT NULL DEFAULT 0,
                    ramp_down_s REAL NOT NULL DEFAULT 0,
                    interphase_s REAL NOT NULL DEFAULT 0,
                    interphase_actual_s REAL,
                    current_limit_a REAL NOT NULL DEFAULT 0,
                    chrono_supply_v REAL NOT NULL DEFAULT 0,
                    stable_ms INTEGER NOT NULL DEFAULT 0,
                    capture_policy TEXT NOT NULL DEFAULT 'FIRST_PASSAGE',
                    validation_policy TEXT NOT NULL DEFAULT 'STABLE_AFTER_CAPTURE',
                    divider_ratio REAL NOT NULL DEFAULT 0,
                    offset_mv INTEGER NOT NULL DEFAULT 0,
                    calibration_id INTEGER,
                    calibration_date TEXT NOT NULL DEFAULT '',
                    calibration_error_v REAL,
                    pickup_global_raw INTEGER,
                    dropout_global_raw INTEGER,
                    pickup_raw_json TEXT NOT NULL DEFAULT '{}',
                    dropout_raw_json TEXT NOT NULL DEFAULT '{}',
                    pickup_global_v REAL,
                    dropout_global_v REAL,
                    pickup_json TEXT NOT NULL DEFAULT '{}',
                    dropout_json TEXT NOT NULL DEFAULT '{}',
                    pickup_elapsed_s REAL,
                    dropout_elapsed_s REAL,
                    pickup_effective_ramp_s REAL,
                    dropout_effective_ramp_s REAL,
                    pickup_time_json TEXT NOT NULL DEFAULT '{}',
                    dropout_time_json TEXT NOT NULL DEFAULT '{}',
                    ea_readback_json TEXT NOT NULL DEFAULT '{}',
                    pickup_plausibility_status TEXT NOT NULL DEFAULT '',
                    dropout_plausibility_status TEXT NOT NULL DEFAULT '',
                    pickup_expected_elapsed_s REAL,
                    dropout_expected_elapsed_s REAL,
                    pickup_elapsed_error_s REAL,
                    dropout_elapsed_error_s REAL,
                    plausibility_json TEXT NOT NULL DEFAULT '{}',
                    ea_stop_confirmed INTEGER NOT NULL DEFAULT -1,
                    ea_final_output_state TEXT NOT NULL DEFAULT '',
                    ea_final_voltage_v REAL,
                    ea_final_generator_state TEXT NOT NULL DEFAULT '',
                    ea_stop_detail TEXT NOT NULL DEFAULT '',
                    resultat TEXT NOT NULL DEFAULT '',
                    timestamp TEXT NOT NULL DEFAULT ''
                )
                """
            )
            con.execute("CREATE INDEX IF NOT EXISTS idx_voltage_lot_sn ON mesures_tension_fonctionnement(lot, sn)")
            columns = {row["name"] for row in con.execute("PRAGMA table_info(mesures_tension_fonctionnement)").fetchall()}
            migrations = {
                "calibration_id": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN calibration_id INTEGER",
                "calibration_date": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN calibration_date TEXT NOT NULL DEFAULT ''",
                "calibration_error_v": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN calibration_error_v REAL",
                "pickup_global_raw": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_global_raw INTEGER",
                "dropout_global_raw": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_global_raw INTEGER",
                "pickup_raw_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_raw_json TEXT NOT NULL DEFAULT '{}'",
                "dropout_raw_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_raw_json TEXT NOT NULL DEFAULT '{}'",
                "interphase_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN interphase_s REAL NOT NULL DEFAULT 0",
                "interphase_actual_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN interphase_actual_s REAL",
                "chrono_supply_v": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN chrono_supply_v REAL NOT NULL DEFAULT 0",
                "pickup_elapsed_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_elapsed_s REAL",
                "dropout_elapsed_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_elapsed_s REAL",
                "pickup_effective_ramp_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_effective_ramp_s REAL",
                "dropout_effective_ramp_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_effective_ramp_s REAL",
                "pickup_time_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_time_json TEXT NOT NULL DEFAULT '{}'",
                "dropout_time_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_time_json TEXT NOT NULL DEFAULT '{}'",
                "ea_readback_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_readback_json TEXT NOT NULL DEFAULT '{}'",
                "capture_policy": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN capture_policy TEXT NOT NULL DEFAULT 'FIRST_PASSAGE'",
                "validation_policy": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN validation_policy TEXT NOT NULL DEFAULT 'STABLE_AFTER_CAPTURE'",
                "pickup_plausibility_status": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_plausibility_status TEXT NOT NULL DEFAULT ''",
                "dropout_plausibility_status": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_plausibility_status TEXT NOT NULL DEFAULT ''",
                "pickup_expected_elapsed_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_expected_elapsed_s REAL",
                "dropout_expected_elapsed_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_expected_elapsed_s REAL",
                "pickup_elapsed_error_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN pickup_elapsed_error_s REAL",
                "dropout_elapsed_error_s": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN dropout_elapsed_error_s REAL",
                "plausibility_json": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN plausibility_json TEXT NOT NULL DEFAULT '{}'",
                "ea_stop_confirmed": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_stop_confirmed INTEGER NOT NULL DEFAULT -1",
                "ea_final_output_state": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_final_output_state TEXT NOT NULL DEFAULT ''",
                "ea_final_voltage_v": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_final_voltage_v REAL",
                "ea_final_generator_state": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_final_generator_state TEXT NOT NULL DEFAULT ''",
                "ea_stop_detail": "ALTER TABLE mesures_tension_fonctionnement ADD COLUMN ea_stop_detail TEXT NOT NULL DEFAULT ''",
            }
            for name, statement in migrations.items():
                if name not in columns:
                    con.execute(statement)

    def voltage_save_result(self):
        meta = self.voltage_metadata()
        pickup = dict(self.voltage_results.get("PICKUP", {}))
        dropout = dict(self.voltage_results.get("DROPOUT", {}))
        pickup_raw = dict(self.voltage_raw_results.get("PICKUP", {}))
        dropout_raw = dict(self.voltage_raw_results.get("DROPOUT", {}))
        pickup_times = dict(self.voltage_time_results.get("PICKUP", {}))
        dropout_times = dict(self.voltage_time_results.get("DROPOUT", {}))
        pickup_global = pickup.get("GLOBAL")
        dropout_global = dropout.get("GLOBAL")
        pickup_elapsed_s = None if pickup_times.get("GLOBAL") is None else pickup_times["GLOBAL"] / 1_000_000.0
        dropout_elapsed_s = None if dropout_times.get("GLOBAL") is None else dropout_times["GLOBAL"] / 1_000_000.0
        pickup_pl = dict(self.voltage_plausibility.get("PICKUP", {}))
        dropout_pl = dict(self.voltage_plausibility.get("DROPOUT", {}))
        stop = dict(self.voltage_ea_stop_confirmation or {})
        stop_errors = "; ".join(str(item) for item in stop.get("errors", []) if str(item).strip())
        stop_detail = f"{str(stop.get('context', '') or 'arrêt')}: {stop_errors or 'confirmation obtenue'}"
        result = str(self.voltage_result_override or "").strip()
        if not result:
            statuses = {str(pickup_pl.get("status", "")), str(dropout_pl.get("status", ""))}
            if "INCOHERENT" in statuses:
                result = "INCOHERENT"
            elif "NON_VERIFIE" in statuses:
                result = "NON_VERIFIE"
            else:
                result = "OK" if (pickup_global is not None or dropout_global is not None) else "INCOMPLET"
        self.voltage_last_saved_result = result
        cal = self.active_voltage_calibration or {}

        record = {
            "lot": meta["lot"], "date_test": meta["date_test"], "relais": meta["relais"],
            "ambiance_c": meta["ambiance_c"], "nom_test": meta["nom_test"], "sn": meta["sn"],
            "relay_type": self.voltage_relay_type(),
            "nb_inverseurs": int(self.spinBox_voltage_nb_inverseurs.value()),
            "chrono_supply_v": self.voltage_run_setting("chrono_supply_v", self.doubleSpinBox_voltage_chrono_v),
            "vmax_v": self.voltage_run_setting("vmax_v", self.doubleSpinBox_voltage_vmax),
            "ramp_up_s": self.voltage_run_setting("ramp_up_s", self.doubleSpinBox_voltage_ramp_up_s),
            "ramp_down_s": self.voltage_run_setting("ramp_down_s", self.doubleSpinBox_voltage_ramp_down_s),
            "interphase_s": self.voltage_run_setting("interphase_s", self.doubleSpinBox_voltage_interphase_s),
            "interphase_actual_s": self.voltage_interphase_actual_s,
            "current_limit_a": self.voltage_run_setting("current_limit_a", self.doubleSpinBox_voltage_current_limit),
            "stable_ms": int(self.spinBox_voltage_stable_ms.value()),
            "capture_policy": self.voltage_capture_policy or "FIRST_PASSAGE",
            "validation_policy": "STABLE_AFTER_CAPTURE",
            "divider_ratio": float(cal.get("divider_ratio", 0)),
            "offset_mv": int(cal.get("offset_mv", 0)),
            "calibration_id": cal.get("id"),
            "calibration_date": str(cal.get("calibration_date", "")),
            "calibration_error_v": cal.get("check_error_v"),
            "pickup_global_raw": pickup_raw.get("GLOBAL"),
            "dropout_global_raw": dropout_raw.get("GLOBAL"),
            "pickup_raw_json": json.dumps(pickup_raw, ensure_ascii=False),
            "dropout_raw_json": json.dumps(dropout_raw, ensure_ascii=False),
            "pickup_global_v": pickup_global,
            "dropout_global_v": dropout_global,
            "pickup_json": json.dumps(pickup, ensure_ascii=False),
            "dropout_json": json.dumps(dropout, ensure_ascii=False),
            "pickup_elapsed_s": pickup_elapsed_s,
            "dropout_elapsed_s": dropout_elapsed_s,
            "pickup_effective_ramp_s": self.voltage_effective_ramp_s.get("PICKUP"),
            "dropout_effective_ramp_s": self.voltage_effective_ramp_s.get("DROPOUT"),
            "pickup_time_json": json.dumps(pickup_times, ensure_ascii=False),
            "dropout_time_json": json.dumps(dropout_times, ensure_ascii=False),
            "ea_readback_json": json.dumps(self.voltage_ramp_readbacks, ensure_ascii=False),
            "pickup_plausibility_status": str(pickup_pl.get("status", "")),
            "dropout_plausibility_status": str(dropout_pl.get("status", "")),
            "pickup_expected_elapsed_s": pickup_pl.get("expected_elapsed_s"),
            "dropout_expected_elapsed_s": dropout_pl.get("expected_elapsed_s"),
            "pickup_elapsed_error_s": pickup_pl.get("elapsed_error_s"),
            "dropout_elapsed_error_s": dropout_pl.get("elapsed_error_s"),
            "plausibility_json": json.dumps(self.voltage_plausibility, ensure_ascii=False),
            "ea_stop_confirmed": 1 if stop.get("confirmed") is True else 0 if stop.get("confirmed") is False else -1,
            "ea_final_output_state": str(stop.get("output_state", "")),
            "ea_final_voltage_v": stop.get("measured_voltage_v"),
            "ea_final_generator_state": str(stop.get("generator_state", "")),
            "ea_stop_detail": stop_detail,
            "resultat": result,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        self.voltage_init_db()
        columns = list(record.keys())
        sql = (
            "INSERT INTO mesures_tension_fonctionnement (" + ", ".join(columns) + ") "
            "VALUES (" + ", ".join("?" for _ in columns) + ")"
        )
        with self.chrono_connect_db() as con:
            con.execute(sql, tuple(record[name] for name in columns))

    def voltage_records_for_lot(self, lot):
        self.voltage_init_db()
        with self.chrono_connect_db() as con:
            return con.execute(
                """
                SELECT id, lot, date_test, relais, ambiance_c, nom_test, sn, relay_type,
                       nb_inverseurs, vmax_v, ramp_up_s, ramp_down_s, interphase_s, interphase_actual_s, current_limit_a, chrono_supply_v,
                       stable_ms, capture_policy, validation_policy, divider_ratio, offset_mv, calibration_id, calibration_date, calibration_error_v,
                       pickup_global_raw, dropout_global_raw, pickup_raw_json, dropout_raw_json,
                       pickup_global_v, dropout_global_v, pickup_json, dropout_json,
                       pickup_elapsed_s, dropout_elapsed_s, pickup_effective_ramp_s, dropout_effective_ramp_s,
                       pickup_time_json, dropout_time_json, ea_readback_json,
                       pickup_plausibility_status, dropout_plausibility_status,
                       pickup_expected_elapsed_s, dropout_expected_elapsed_s,
                       pickup_elapsed_error_s, dropout_elapsed_error_s, plausibility_json,
                       ea_stop_confirmed, ea_final_output_state, ea_final_voltage_v,
                       ea_final_generator_state, ea_stop_detail, resultat, timestamp
                FROM mesures_tension_fonctionnement
                WHERE lot = ?
                ORDER BY timestamp ASC, id ASC
                """, (lot,)
            ).fetchall()

    def voltage_export_headers(self):
        return [
            ("lot", "Lot"), ("date_test", "Date"), ("relais", "Relais"),
            ("sn", "SN"), ("relay_type", "Type relais"), ("nb_inverseurs", "Inv."), ("vmax_v", "V max"),
            ("ramp_up_s", "Montée BE demandée s"), ("ramp_down_s", "Retour BE/BR demandé s"),
            ("interphase_s", "Attente demandée s"),
            ("interphase_actual_s", "Attente réelle s"),
            ("current_limit_a", "Limite A"), ("chrono_supply_v", "Tension chronométrie EA V"), ("stable_ms", "Validation ms"),
            ("capture_policy", "Méthode capture"), ("validation_policy", "Méthode validation"),
            ("calibration_id", "ID étalonnage"), ("calibration_date", "Date étalonnage"),
            ("divider_ratio", "Rapport étalonné"), ("offset_mv", "Offset mV"), ("calibration_error_v", "Erreur contrôle V"),
            ("pickup_global_raw", "RAW global collage/BE"), ("dropout_global_raw", "RAW global décollage/BR"),
            ("pickup_global_v", "Collage/BE global V"), ("dropout_global_v", "Décollage/BR global V"),
            ("pickup_elapsed_s", "Temps seuil BE s"), ("dropout_elapsed_s", "Temps seuil retour s"),
            ("pickup_effective_ramp_s", "Rampe BE reconstituée s"), ("dropout_effective_ramp_s", "Rampe retour reconstituée s"),
            ("pickup_json", "Collage par inverseur"), ("dropout_json", "Décollage par inverseur"),
            ("pickup_time_json", "Temps collage par inverseur µs"), ("dropout_time_json", "Temps retour par inverseur µs"),
            ("ea_readback_json", "Relecture paramètres EA"),
            ("pickup_plausibility_status", "Plausibilité collage/BE"),
            ("dropout_plausibility_status", "Plausibilité retour/BR"),
            ("pickup_expected_elapsed_s", "Temps théorique collage/BE s"),
            ("dropout_expected_elapsed_s", "Temps théorique retour/BR s"),
            ("pickup_elapsed_error_s", "Écart temps collage/BE s"),
            ("dropout_elapsed_error_s", "Écart temps retour/BR s"),
            ("ea_stop_confirmed", "Arrêt EA confirmé"),
            ("ea_final_output_state", "Sortie EA finale"),
            ("ea_final_voltage_v", "Tension EA finale V"),
            ("ea_final_generator_state", "Générateur EA final"),
            ("ea_stop_detail", "Détail arrêt EA"),
            ("resultat", "Résultat"), ("timestamp", "Horodatage"),
        ]

    def voltage_export_lot_xlsx(self):
        lot = self.lineEdit_voltage_lot.text().strip()
        if not lot:
            QMessageBox.warning(self.window, "Export mesures", "Renseigner le lot.")
            return
        try:
            chrono_records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "Export mesures", f"Lecture base impossible : {exc}")
            return
        if not chrono_records and not voltage_records:
            QMessageBox.information(self.window, "Export mesures", f"Aucune mesure pour le lot {lot}.")
            return
        path = self.ask_export_path(
            "Exporter les mesures complètes",
            f"mesures_completes_lot_{self.filename_safe(lot)}.xlsx",
            "Excel (*.xlsx)",
            ".xlsx",
        )
        if not path:
            return
        summary_cards, detail_cards = self.chrono_export_measure_sheets(chrono_records, voltage_records)
        self.write_chrono_lot_xlsx(path, lot, summary_cards, detail_cards)
        QMessageBox.information(self.window, "Export mesures", f"Export XLSX créé :\n{path}")

    def voltage_export_lot_pdf(self):
        lot = self.lineEdit_voltage_lot.text().strip()
        if not lot:
            QMessageBox.warning(self.window, "Export mesures", "Renseigner le lot.")
            return
        try:
            chrono_records = self.chrono_records_for_lot(lot)
            voltage_records = self.voltage_records_for_lot(lot)
        except Exception as exc:
            QMessageBox.warning(self.window, "Export mesures", f"Lecture base impossible : {exc}")
            return
        if not chrono_records and not voltage_records:
            QMessageBox.information(self.window, "Export mesures", f"Aucune mesure pour le lot {lot}.")
            return
        path = self.ask_export_path(
            "Exporter les mesures complètes",
            f"mesures_completes_lot_{self.filename_safe(lot)}.pdf",
            "PDF (*.pdf)",
            ".pdf",
        )
        if not path:
            return
        summary_cards, detail_cards = self.chrono_export_measure_sheets(chrono_records, voltage_records)
        self.write_chrono_lot_pdf(path, lot, chrono_records, summary_cards, detail_cards, voltage_records)
        QMessageBox.information(self.window, "Export mesures", f"Export PDF créé :\n{path}")

    def show(self):
        self.window.show()

    def close(self):
        try:
            self.voltage_disconnect_ea()
        except Exception:
            pass
        self.disconnect_serial()


def main():
    require_license()
    app = QApplication(sys.argv)
    ihm = IhmRelaisRp2040()
    app.aboutToQuit.connect(ihm.close)
    ihm.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
